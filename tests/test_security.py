import contextlib
import io
import json
import os
import shutil
import sqlite3
import sys
import tempfile
import time
import unittest
import uuid
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest import mock

import openpyxl
from flask import session as flask_session
from werkzeug.datastructures import FileStorage

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import app as endemias_app
import etl
from app_core import audit as audit_core
from app_core import auth as auth_core
from app_core import backup as backup_core
from app_core import backup_completo as backup_completo_core
from app_core import amostras_animais as amostras_animais_core
from app_core import esporotricose as esporotricose_core
from app_core import db as db_core
from app_core import diagnostico as diagnostico_core
from app_core import dbml as dbml_core
from app_core import kobo_api as kobo_api_core
from app_core import laboratorio_lancamentos as laboratorio_lancamentos_core
from app_core import larvas as larvas_core
from app_core.excel import excel_safe
from app_core import modules as modules_core
from app_core import ovitrampas as ovitrampas_core
from app_core import bri as bri_core
from app_core import pontos_estrategicos as pe_core
from app_core import recolhimentos as recolhimentos_core
from app_core import sispncd as sispncd_core
from app_core import sispncd_indice as sispncd_indice_core
from app_core import registro_geografico as registro_geografico_core
from app_core import version as version_core
from app_core import work_types
from blueprints import processar as processar_bp
from blueprints import relatorio_agente as relatorio_agente_bp


def _usuario_teste(nivel=None):
    conn = sqlite3.connect(endemias_app.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        params = []
        filtro_nivel = ""
        if nivel:
            filtro_nivel = "AND nivel=?"
            params.append(nivel)
        row = conn.execute(
            """SELECT id_usuario, nome, nivel FROM usuarios
               WHERE ativo=1
               {filtro_nivel}
               ORDER BY CASE nivel
                   WHEN 'admin' THEN 1
                   WHEN 'operador' THEN 2
                   ELSE 3
               END
               LIMIT 1""".format(filtro_nivel=filtro_nivel),
            params,
        ).fetchone()
    finally:
        conn.close()
    if not row:
        alvo = f" com nivel {nivel}" if nivel else ""
        raise RuntimeError(f"Nenhum usuario ativo{alvo} encontrado para os testes.")
    return dict(row)


def _client_logado(nivel=None):
    endemias_app.app.config["TESTING"] = True
    client = endemias_app.app.test_client()
    usuario = _usuario_teste(nivel=nivel)
    with client.session_transaction() as sess:
        sess["uid"] = usuario["id_usuario"]
        sess["nome"] = usuario["nome"]
        sess["nivel"] = usuario["nivel"]
    return client


def _agente_relatorio_teste():
    conn = sqlite3.connect(endemias_app.DB_PATH)
    try:
        row = conn.execute(
            """SELECT ag.nome
               FROM agentes ag
               JOIN esporotricose_visita_agentes va ON va.id_agente=ag.id_agente
               LIMIT 1"""
        ).fetchone()
        if not row:
            row = conn.execute(
                """SELECT a.nome
                   FROM agentes a
                   JOIN visita_agentes va ON va.id_agente=a.id_agente
                   LIMIT 1"""
            ).fetchone()
    finally:
        conn.close()
    if not row:
        raise RuntimeError("Nenhum agente encontrado para os testes do relatorio.")
    return row[0]


def _executar_criar_banco_em(tmpdir, vezes=1):
    import criar_banco

    cwd_original = os.getcwd()
    os.chdir(tmpdir)
    try:
        with contextlib.redirect_stdout(io.StringIO()):
            for _ in range(vezes):
                criar_banco.main()
    finally:
        os.chdir(cwd_original)
    return str(Path(tmpdir) / "endemias.db")


def _login_client_com_usuario(client, usuario):
    with client.session_transaction() as sess:
        sess["uid"] = usuario["id_usuario"]
        sess["nome"] = usuario["nome"]
        sess["nivel"] = usuario["nivel"]


def _client_admin_com_banco_temporario(tmpdir):
    db_path = _executar_criar_banco_em(tmpdir)
    app_temp = endemias_app.create_app({
        "TESTING": True,
        "DB_PATH": db_path,
        "WTF_CSRF_ENABLED": False,
    })
    client = app_temp.test_client()

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        usuario = conn.execute(
            "SELECT id_usuario, nome, nivel FROM usuarios WHERE usuario='admin'"
        ).fetchone()
    finally:
        conn.close()
    if not usuario:
        raise RuntimeError("Usuario admin nao encontrado no banco temporario.")

    _login_client_com_usuario(client, dict(usuario))
    return app_temp, client, db_path


class LoginRateLimitTests(unittest.TestCase):
    def _get_db_temp(self, db_path):
        def get_db():
            return db_core.connect(db_path)

        return get_db

    def test_bloqueia_apos_limite_de_falhas(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            get_db = self._get_db_temp(str(Path(tmpdir) / "login.db"))
            chave = "127.0.0.1:admin"

            for i in range(endemias_app.LOGIN_MAX_TENTATIVAS):
                self.assertFalse(auth_core.login_bloqueado_db(get_db, chave, agora=100 + i))
                auth_core.registrar_login_falha_db(get_db, chave, agora=100 + i)

            self.assertTrue(auth_core.login_bloqueado_db(get_db, chave, agora=120))

    def test_expira_bloqueio_apos_janela(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "login.db")
            get_db = self._get_db_temp(db_path)
            chave = "127.0.0.1:admin"

            for i in range(endemias_app.LOGIN_MAX_TENTATIVAS):
                auth_core.registrar_login_falha_db(get_db, chave, agora=100 + i)

            depois_da_janela = 100 + endemias_app.LOGIN_JANELA_SEG + 1
            self.assertFalse(auth_core.login_bloqueado_db(get_db, chave, agora=depois_da_janela))
            conn = sqlite3.connect(db_path)
            try:
                total = conn.execute("SELECT COUNT(*) FROM login_tentativas WHERE chave=?", (chave,)).fetchone()[0]
            finally:
                conn.close()
            self.assertEqual(total, 0)

    def test_sucesso_limpa_falhas(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "login.db")
            get_db = self._get_db_temp(db_path)
            chave = "127.0.0.1:admin"
            auth_core.registrar_login_falha_db(get_db, chave, agora=100)

            auth_core.limpar_login_falhas_db(get_db, chave)

            conn = sqlite3.connect(db_path)
            try:
                total = conn.execute("SELECT COUNT(*) FROM login_tentativas WHERE chave=?", (chave,)).fetchone()[0]
            finally:
                conn.close()
            self.assertEqual(total, 0)

    def test_rate_limit_persistente_em_sqlite(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "login.db")

            def get_db():
                return db_core.connect(db_path)

            chave = "127.0.0.1:admin"
            for i in range(endemias_app.LOGIN_MAX_TENTATIVAS):
                self.assertFalse(auth_core.login_bloqueado_db(get_db, chave, agora=100 + i))
                auth_core.registrar_login_falha_db(get_db, chave, agora=100 + i)

            self.assertTrue(auth_core.login_bloqueado_db(get_db, chave, agora=120))
            auth_core.limpar_login_falhas_db(get_db, chave)
            self.assertFalse(auth_core.login_bloqueado_db(get_db, chave, agora=121))

    def test_senha_minima_centralizada(self):
        self.assertFalse(auth_core.senha_valida("123456"))
        self.assertTrue(auth_core.senha_valida("1234567890"))
        self.assertIn(str(auth_core.PASSWORD_MIN_LENGTH), auth_core.mensagem_senha_invalida())

    def test_chave_login_ignora_x_forwarded_for_sem_proxy_confiavel(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
            "TRUST_PROXY_HEADERS": False,
        })
        with app_temp.test_request_context(
            "/login",
            environ_base={"REMOTE_ADDR": "10.0.0.5"},
            headers={"X-Forwarded-For": "203.0.113.9"},
        ):
            self.assertEqual(auth_core.chave_login("Admin"), "10.0.0.5:admin")

    def test_chave_login_usa_x_forwarded_for_quando_proxy_confiavel(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
            "TRUST_PROXY_HEADERS": True,
        })
        with app_temp.test_request_context(
            "/login",
            environ_base={"REMOTE_ADDR": "10.0.0.5"},
            headers={"X-Forwarded-For": "203.0.113.9, 10.0.0.5"},
        ):
            self.assertEqual(auth_core.chave_login("Admin"), "203.0.113.9:admin")


class UploadValidationTests(unittest.TestCase):
    def _xlsx_minimo(self):
        stream = io.BytesIO()
        with zipfile.ZipFile(stream, "w") as zf:
            zf.writestr("[Content_Types].xml", "<Types></Types>")
            zf.writestr("xl/workbook.xml", "<workbook></workbook>")
        stream.seek(0)
        return stream

    def test_aceita_xlsx_com_estrutura_zip_excel(self):
        arquivo = FileStorage(
            stream=self._xlsx_minimo(),
            filename="TB_exemplo.xlsx",
        )

        valido, nome, motivo = endemias_app._validar_arquivo_xlsx(arquivo)

        self.assertTrue(valido)
        self.assertEqual(nome, "TB_exemplo.xlsx")
        self.assertEqual(motivo, "")

    def test_rejeita_xlsx_falso(self):
        arquivo = FileStorage(
            stream=io.BytesIO(b"isso nao e xlsx"),
            filename="TB_exemplo.xlsx",
        )

        valido, nome, motivo = endemias_app._validar_arquivo_xlsx(arquivo)

        self.assertFalse(valido)
        self.assertEqual(nome, "TB_exemplo.xlsx")
        self.assertIn("XLSX", motivo)

    def test_rejeita_zip_sem_estrutura_xlsx(self):
        stream = io.BytesIO()
        with zipfile.ZipFile(stream, "w") as zf:
            zf.writestr("conteudo.txt", "nao e excel")
        stream.seek(0)
        arquivo = FileStorage(
            stream=stream,
            filename="TB_exemplo.xlsx",
        )

        valido, nome, motivo = endemias_app._validar_arquivo_xlsx(arquivo)

        self.assertFalse(valido)
        self.assertEqual(nome, "TB_exemplo.xlsx")
        self.assertIn("estrutura XLSX", motivo)

    def test_rejeita_extensao_diferente(self):
        arquivo = FileStorage(
            stream=io.BytesIO(b"PK\x03\x04conteudo"),
            filename="TB_exemplo.csv",
        )

        valido, nome, motivo = endemias_app._validar_arquivo_xlsx(arquivo)

        self.assertFalse(valido)
        self.assertEqual(nome, "")
        self.assertIn("Extens", motivo)


class UploadTempCleanupTests(unittest.TestCase):
    def test_limpa_apenas_jobs_uuid_antigos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            antigo = Path(tmpdir) / str(uuid.uuid4())
            recente = Path(tmpdir) / str(uuid.uuid4())
            nao_job = Path(tmpdir) / "manual"
            antigo.mkdir()
            recente.mkdir()
            nao_job.mkdir()
            velho = time.time() - 48 * 3600
            os.utime(antigo, (velho, velho))
            os.utime(nao_job, (velho, velho))

            app_temp = endemias_app.create_app({
                "TESTING": True,
                "UPLOAD_TEMP": tmpdir,
                "DB_PATH": endemias_app.DB_PATH,
            })
            with app_temp.app_context():
                removidos = processar_bp.limpar_uploads_antigos(max_age_hours=24)

            self.assertEqual(removidos, 1)
            self.assertFalse(antigo.exists())
            self.assertTrue(recente.exists())
            self.assertTrue(nao_job.exists())


class ExcelSafetyTests(unittest.TestCase):
    def test_excel_safe_normaliza_valores_e_escapa_formulas(self):
        self.assertEqual(excel_safe(None), "")
        self.assertEqual(excel_safe(123), "123")
        self.assertEqual(excel_safe("texto"), "texto")
        self.assertEqual(excel_safe("=1+1"), "'=1+1")
        self.assertEqual(excel_safe("+cmd"), "'+cmd")


class LarvasAuditTests(unittest.TestCase):
    def test_audita_larvas_sem_coleta_correspondente(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "larvas.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.executescript(
                    """
                    CREATE TABLE visitas (
                        id_visita TEXT PRIMARY KEY,
                        data TEXT,
                        tipo TEXT,
                        localidade TEXT,
                        quarteirao TEXT,
                        logradouro TEXT,
                        numero TEXT,
                        morador TEXT,
                        tipo_imovel TEXT,
                        observacoes TEXT
                    );
                    CREATE TABLE coletas (
                        id_coleta TEXT PRIMARY KEY,
                        id_visita TEXT,
                        num_tubo TEXT,
                        tipo_deposito TEXT
                    );
                    INSERT INTO visitas (id_visita, data)
                    VALUES ('v1', '2026-06-01');
                    INSERT INTO coletas (id_coleta, id_visita, num_tubo)
                    VALUES ('c1', 'v1', '123');
                    """
                )
                logger = etl.Logger()
                pendentes = etl.auditar_larvas_sem_coleta(
                    conn,
                    [
                        {"tubo": "123", "data": "2026-06-01", "arquivo": "LARVAS_ok.xlsx"},
                        {"tubo": "999", "data": "2026-06-02", "arquivo": "LARVAS_falta.xlsx"},
                    ],
                    logger,
                )
            finally:
                conn.close()

        self.assertEqual([item["tubo"] for item in pendentes], ["999"])
        texto_log = "\n".join(mensagem for mensagem, _ in logger.linhas)
        self.assertIn("Tubo 999", texto_log)
        self.assertIn("LARVAS_falta.xlsx", texto_log)

    def test_etl_audita_pendencias_operacionais(self):
        with sqlite3.connect(":memory:") as conn:
            conn.row_factory = sqlite3.Row
            conn.executescript("""
                CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT);
                CREATE TABLE visitas (
                    id_visita TEXT PRIMARY KEY,
                    data TEXT,
                    tipo TEXT,
                    localidade TEXT,
                    id_localidade INTEGER,
                    quarteirao INTEGER
                );
                CREATE TABLE visita_agentes (id_visita TEXT, id_agente INTEGER);
                CREATE TABLE tratamentos (
                    id INTEGER PRIMARY KEY,
                    id_visita TEXT,
                    tipo TEXT,
                    quantidade_carga REAL,
                    qtd_depositos_tratados INTEGER
                );
                CREATE TABLE depositos_inspecionados (
                    id INTEGER PRIMARY KEY,
                    id_visita TEXT,
                    tipo_deposito TEXT,
                    tratado INTEGER,
                    qtd_carga REAL
                );
            """)
            conn.execute(
                "INSERT INTO visitas(id_visita, data, tipo, localidade, quarteirao) VALUES (?,?,?,?,?)",
                ("v1", "2026-06-10", "PE", "Graziela", 10),
            )
            conn.execute(
                "INSERT INTO tratamentos(id_visita, tipo, quantidade_carga, qtd_depositos_tratados) VALUES (?,?,?,?)",
                ("v1", "BTI (pó)", None, 2),
            )
            logger = etl.Logger()

            avisos = etl.auditar_pendencias_importacao(conn, logger)
            logger_filtrado = etl.Logger()
            avisos_filtrados = etl.auditar_pendencias_importacao(conn, logger_filtrado, ["outra-visita"])

        texto_log = "\n".join(mensagem for mensagem, _ in logger.linhas)
        self.assertEqual(len(avisos), 2)
        self.assertIn("[AVISO OPERACIONAL] Visitas sem agente vinculado: 1", texto_log)
        self.assertIn("[AVISO OPERACIONAL] Tratamentos com deposito tratado, mas sem carga: 1", texto_log)
        self.assertEqual(avisos_filtrados, [])
        self.assertEqual(logger_filtrado.linhas, [])


class RequestParsingTests(unittest.TestCase):
    def test_request_int_arg_usa_default_quando_invalido(self):
        with endemias_app.app.test_request_context("/?pagina=abc"):
            self.assertEqual(endemias_app.request_int_arg("pagina", 1, minimo=1), 1)

    def test_request_int_arg_aplica_limites(self):
        with endemias_app.app.test_request_context("/?pagina=-5&por_pagina=9999"):
            self.assertEqual(endemias_app.request_int_arg("pagina", 1, minimo=1), 1)
            self.assertEqual(
                endemias_app.request_int_arg("por_pagina", 50, minimo=1, maximo=500),
                500,
            )


class DatabaseConnectionTests(unittest.TestCase):
    def test_conexao_sqlite_configura_timeout_e_wal(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "teste.db")
            conn = db_core.connect(db_path)
            try:
                self.assertEqual(conn.execute("PRAGMA busy_timeout").fetchone()[0], 5000)
                self.assertEqual(conn.execute("PRAGMA journal_mode").fetchone()[0].lower(), "wal")
            finally:
                conn.close()


class DbmlGenerationTests(unittest.TestCase):
    def test_gerar_dbml_inclui_tabelas_indices_e_referencias(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "schema.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.executescript(
                    """
                    CREATE TABLE pais (
                        id INTEGER PRIMARY KEY,
                        nome TEXT NOT NULL UNIQUE
                    );
                    CREATE TABLE filhos (
                        id INTEGER PRIMARY KEY,
                        pais_id INTEGER NOT NULL,
                        nome TEXT DEFAULT 'sem nome',
                        FOREIGN KEY (pais_id) REFERENCES pais(id)
                    );
                    CREATE TABLE leituras (
                        tubo TEXT NOT NULL,
                        data TEXT NOT NULL,
                        resultado TEXT,
                        PRIMARY KEY (tubo, data)
                    );
                    CREATE INDEX idx_filhos_nome ON filhos(nome);
                    """
                )
                conn.commit()
            finally:
                conn.close()

            dbml = dbml_core.gerar_dbml(db_path, project_name="Teste")

        self.assertIn("Project Teste", dbml)
        self.assertIn("Table pais", dbml)
        self.assertIn("id INTEGER [pk]", dbml)
        self.assertIn("nome TEXT [not null]", dbml)
        self.assertIn("Table filhos", dbml)
        self.assertIn("default: `'sem nome'`", dbml)
        self.assertIn("idx_filhos_nome", dbml)
        self.assertIn("Ref: filhos.pais_id > pais.id", dbml)
        self.assertIn("Table leituras", dbml)
        self.assertIn("(tubo, data) [pk]", dbml)
        self.assertNotIn("pk_order", dbml)


class BackupTests(unittest.TestCase):
    def _criar_db_minimo(self, db_path):
        conn = sqlite3.connect(db_path)
        try:
            conn.execute("CREATE TABLE dados (id INTEGER PRIMARY KEY, nome TEXT)")
            conn.execute("INSERT INTO dados (nome) VALUES ('registro')")
            conn.commit()
        finally:
            conn.close()

    def test_cria_backup_sqlite_validado_com_metadados(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "origem.db"
            destino = Path(tmpdir) / "backups"
            self._criar_db_minimo(db_path)

            info = backup_core.criar_backup_sqlite(db_path, destino_dir=destino, manter=10)

            backup_path = Path(info["arquivo"])
            meta_path = backup_path.with_suffix(backup_path.suffix + ".json")
            self.assertTrue(backup_path.exists())
            self.assertTrue(meta_path.exists())
            self.assertEqual(info["integridade"], "ok")
            conn = sqlite3.connect(backup_path)
            try:
                self.assertEqual(conn.execute("SELECT nome FROM dados").fetchone()[0], "registro")
            finally:
                conn.close()

    def test_limpa_backups_antigos_respeitando_retencao(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            destino = Path(tmpdir)
            criados = []
            for i in range(4):
                arquivo = destino / f"endemias_20260101_00000{i}.db"
                arquivo.write_text("x", encoding="utf-8")
                arquivo.with_suffix(".db.json").write_text("{}", encoding="utf-8")
                os.utime(arquivo, (100 + i, 100 + i))
                os.utime(arquivo.with_suffix(".db.json"), (100 + i, 100 + i))
                criados.append(arquivo)

            removidos = backup_core.limpar_backups_antigos(destino, manter=2)

            self.assertEqual([p.name for p in removidos], [criados[1].name, criados[0].name])
            self.assertFalse(criados[0].exists())
            self.assertFalse(criados[0].with_suffix(".db.json").exists())
            self.assertTrue(criados[2].exists())
            self.assertTrue(criados[3].exists())

    def test_lista_backups_com_metadados(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            destino = Path(tmpdir)
            arquivo = destino / "endemias_20260101_000000.db"
            arquivo.write_text("x", encoding="utf-8")
            arquivo.with_suffix(".db.json").write_text(
                json.dumps({"integridade": "ok", "validado": True}),
                encoding="utf-8",
            )

            backups = backup_core.listar_backups(destino)

            self.assertEqual(backups[0]["nome"], arquivo.name)
            self.assertEqual(backups[0]["integridade"], "ok")
            self.assertTrue(backups[0]["validado"])

    def test_restaura_backup_sqlite_validado(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "origem.db"
            destino = Path(tmpdir) / "backups"
            self._criar_db_minimo(db_path)
            info = backup_core.criar_backup_sqlite(db_path, destino_dir=destino, manter=10)

            conn = sqlite3.connect(db_path)
            try:
                conn.execute("UPDATE dados SET nome='alterado' WHERE id=1")
                conn.commit()
            finally:
                conn.close()

            restore = backup_core.restaurar_backup_sqlite(db_path, info["arquivo"])

            self.assertEqual(restore["integridade"], "ok")
            conn = sqlite3.connect(db_path)
            try:
                self.assertEqual(conn.execute("SELECT nome FROM dados").fetchone()[0], "registro")
            finally:
                conn.close()

    def test_resolver_backup_rejeita_caminho_fora_da_pasta(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            with self.assertRaises(ValueError):
                backup_core.resolver_backup(Path(tmpdir) / "backups", "../origem.db")

    def test_operacao_exclusiva_rejeita_concorrencia(self):
        with backup_core.operacao_exclusiva():
            with self.assertRaises(RuntimeError):
                with backup_core.operacao_exclusiva():
                    pass


class CriarBancoScriptTests(unittest.TestCase):
    def test_criar_banco_roda_em_base_nova_com_hash_moderno(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir, vezes=2)
            conn = sqlite3.connect(db_path)
            try:
                usuario = conn.execute(
                    "SELECT usuario, senha_hash FROM usuarios WHERE usuario='admin'"
                ).fetchone()
                agenda = conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='agenda_eventos'"
                ).fetchone()
                boletim = conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='boletim_mensal_itens'"
                ).fetchone()
                conn.execute(
                    """INSERT INTO agenda_eventos
                       (titulo, tipo, data_inicio, dia_inteiro, criado_em)
                       VALUES (?, ?, ?, ?, ?)""",
                    ("Planejamento teste", "planejamento", "2026-06-03T08:30", 0, "2026-06-03T08:00"),
                )
                planejamento = conn.execute(
                    "SELECT tipo FROM agenda_eventos WHERE titulo=?",
                    ("Planejamento teste",),
                ).fetchone()
                total_admin = conn.execute(
                    "SELECT COUNT(*) FROM usuarios WHERE usuario='admin'"
                ).fetchone()[0]
            finally:
                conn.close()

        self.assertIsNotNone(usuario)
        self.assertTrue(usuario[1].startswith("pbkdf2:"))
        self.assertIsNotNone(agenda)
        self.assertIsNotNone(boletim)
        self.assertEqual(planejamento[0], "planejamento")
        self.assertEqual(total_admin, 1)


class AdminBackupRoutesTests(unittest.TestCase):
    def _app_e_cliente_admin(self, tmpdir):
        db_path = _executar_criar_banco_em(tmpdir)
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": db_path,
            "INSTANCE_DIR": str(Path(tmpdir)),
            "BACKUP_DIR": str(Path(tmpdir) / "backups_banco"),
            "ANEXOS_DIR": str(Path(tmpdir) / "anexos"),
            "BACKUP_COMPLETO_DIR": str(Path(tmpdir) / "backups_completos"),
            "KOBO_CONFIG_PATH": str(Path(tmpdir) / "kobo_config.json"),
            "SECRET_KEY_PATH": str(Path(tmpdir) / "secret.key"),
            "WTF_CSRF_ENABLED": False,
        })
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        try:
            usuario = conn.execute(
                "SELECT id_usuario, nome, nivel FROM usuarios WHERE usuario='admin'"
            ).fetchone()
        finally:
            conn.close()
        client = app_temp.test_client()
        _login_client_com_usuario(client, dict(usuario))
        return app_temp, client, Path(db_path)

    def test_admin_cria_backup_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            with app_temp.app_context():
                resp = client.post("/admin/sistema/backups/criar")

            self.assertEqual(resp.status_code, 302)
            backups = list(Path(app_temp.config["BACKUP_DIR"]).glob("endemias_*.db"))
            self.assertEqual(len(backups), 1)
            self.assertTrue(backups[0].with_suffix(".db.json").exists())

    def test_admin_restaura_backup_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            backup_dir = Path(app_temp.config["BACKUP_DIR"])
            info = backup_core.criar_backup_sqlite(db_path, destino_dir=backup_dir, manter=10)
            nome_backup = Path(info["arquivo"]).name

            conn = sqlite3.connect(db_path)
            try:
                conn.execute("UPDATE usuarios SET nome='Depois' WHERE usuario='admin'")
                conn.commit()
            finally:
                conn.close()

            with app_temp.app_context():
                resp = client.post("/admin/sistema/backups/restaurar", data={"backup": nome_backup})

            self.assertEqual(resp.status_code, 302)
            conn = sqlite3.connect(db_path)
            try:
                nome = conn.execute("SELECT nome FROM usuarios WHERE usuario='admin'").fetchone()[0]
            finally:
                conn.close()
            self.assertNotEqual(nome, "Depois")
            self.assertTrue(list(backup_dir.glob("pre_restore_*.db")))

    def test_admin_baixa_backup_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            backup_dir = Path(app_temp.config["BACKUP_DIR"])
            info = backup_core.criar_backup_sqlite(db_path, destino_dir=backup_dir, manter=10)
            nome_backup = Path(info["arquivo"]).name

            with app_temp.app_context():
                resp = client.get(f"/admin/sistema/backups/baixar/{nome_backup}")

            self.assertEqual(resp.status_code, 200)
            self.assertIn("attachment", resp.headers.get("Content-Disposition", ""))
            data = resp.get_data()
            resp.close()
            self.assertGreater(len(data), 0)

    def test_admin_cria_backup_completo_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            anexos_dir = Path(app_temp.config["ANEXOS_DIR"])
            anexos_dir.mkdir(parents=True, exist_ok=True)
            (anexos_dir / "teste.txt").write_text("anexo", encoding="utf-8")
            Path(app_temp.config["KOBO_CONFIG_PATH"]).write_text("{}", encoding="utf-8")

            with app_temp.app_context():
                resp = client.post("/admin/sistema/backups-completos/criar")

            self.assertEqual(resp.status_code, 302)
            backups = list(Path(app_temp.config["BACKUP_COMPLETO_DIR"]).glob("endemias_completo_*.zip"))
            self.assertEqual(len(backups), 1)
            itens = backup_completo_core.listar_backups_completos(app_temp.config["BACKUP_COMPLETO_DIR"])
            self.assertEqual(itens[0]["integridade_banco"], "ok")
            with zipfile.ZipFile(backups[0]) as zf:
                nomes = zf.namelist()
            self.assertTrue(any(nome.startswith("banco/") and nome.endswith(".db") for nome in nomes))
            self.assertIn("anexos/teste.txt", nomes)
            self.assertIn("manifesto_backup.json", nomes)

    def test_admin_baixa_backup_completo_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            info = backup_completo_core.criar_backup_completo(
                destino_dir=app_temp.config["BACKUP_COMPLETO_DIR"],
                db_path=db_path,
                raiz=tmpdir,
                secret_key_path=app_temp.config["SECRET_KEY_PATH"],
                kobo_config_path=app_temp.config["KOBO_CONFIG_PATH"],
                anexos_dir=app_temp.config["ANEXOS_DIR"],
            )

            with app_temp.app_context():
                resp = client.get(f"/admin/sistema/backups-completos/baixar/{info['nome']}")

            self.assertEqual(resp.status_code, 200)
            self.assertIn("attachment", resp.headers.get("Content-Disposition", ""))
            data = resp.get_data()
            resp.close()
            self.assertGreater(len(data), 0)

    def test_admin_baixa_dbml_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _db_path = self._app_e_cliente_admin(tmpdir)

            with app_temp.app_context():
                resp = client.get("/admin/sistema/dbml")

            self.assertEqual(resp.status_code, 200)
            self.assertIn("attachment", resp.headers.get("Content-Disposition", ""))
            self.assertIn("schema.dbml", resp.headers.get("Content-Disposition", ""))
            texto = resp.get_data(as_text=True)
            resp.close()
            self.assertIn("Project Endemias", texto)
            self.assertIn("Table usuarios", texto)

    def test_api_diagnostico_retorna_resumo_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)

            with app_temp.app_context():
                resp = client.get("/api/admin/sistema/diagnostico")

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertIn("resumo", data)
            self.assertIn("itens", data)
            self.assertIn(data["resumo"]["status"], {"ok", "atencao", "critico"})
            self.assertEqual(data["resumo"]["modo"], "rapido")
            self.assertTrue(any(item["categoria"] == "Banco" for item in data["itens"]))

            resp_completo = client.get("/api/admin/sistema/diagnostico?completo=1")
            self.assertEqual(resp_completo.status_code, 200)
            data_completo = resp_completo.get_json()
            self.assertEqual(data_completo["resumo"]["modo"], "completo")

    def test_diagnostico_detecta_pendencias_operacionais(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            backup_dir = Path(tmpdir) / "backups"
            backup_dir.mkdir()
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, localidade, processado_em)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    ("v-sem-agente", "uuid-sem-agente", "PE", "2026-06-01", "", "2026-06-01T10:00:00"),
                )
                conn.execute(
                    """INSERT INTO tratamentos (id_visita, tipo, quantidade_carga, qtd_depositos_tratados)
                       VALUES (?, ?, ?, ?)""",
                    ("v-sem-agente", "Natular DT", None, 3),
                )
                conn.commit()

                dados = diagnostico_core.gerar(conn, db_path=db_path, backup_dir=backup_dir)
            finally:
                conn.close()

        titulos = {item["titulo"] for item in dados["itens"]}
        self.assertIn("Visitas sem agente vinculado.", titulos)
        self.assertIn("Tratamentos com deposito tratado, mas sem carga.", titulos)
        self.assertIn("Nenhum backup encontrado.", titulos)

    def test_admin_exclui_backup_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            backup_dir = Path(app_temp.config["BACKUP_DIR"])
            info = backup_core.criar_backup_sqlite(db_path, destino_dir=backup_dir, manter=10)
            backup_path = Path(info["arquivo"])

            with app_temp.app_context():
                resp = client.post("/admin/sistema/backups/excluir", data={"backup": backup_path.name})

            self.assertEqual(resp.status_code, 302)
            self.assertFalse(backup_path.exists())
            self.assertFalse(backup_path.with_suffix(".db.json").exists())

    def test_admin_exclui_backup_completo_pela_central_do_sistema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            info = backup_completo_core.criar_backup_completo(
                destino_dir=app_temp.config["BACKUP_COMPLETO_DIR"],
                db_path=db_path,
                raiz=tmpdir,
                secret_key_path=app_temp.config["SECRET_KEY_PATH"],
                kobo_config_path=app_temp.config["KOBO_CONFIG_PATH"],
                anexos_dir=app_temp.config["ANEXOS_DIR"],
            )
            backup_path = Path(info["arquivo"])

            with app_temp.app_context():
                resp = client.post("/admin/sistema/backups-completos/excluir", data={"backup": backup_path.name})

            self.assertEqual(resp.status_code, 302)
            self.assertFalse(backup_path.exists())

    def test_confirmar_importacao_cria_backup_pre_import(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = self._app_e_cliente_admin(tmpdir)
            upload_temp = Path(tmpdir) / "uploads"
            app_temp.config["UPLOAD_TEMP"] = str(upload_temp)
            job_id = str(uuid.uuid4())
            job_dir = upload_temp / job_id
            job_dir.mkdir(parents=True)
            (job_dir / "TB_teste.xlsx").write_bytes(b"teste")

            with mock.patch("etl.processar_upload", return_value=(True, [])) as processar:
                with app_temp.app_context():
                    resp = client.post(f"/processar/confirmar/{job_id}")
                    data = resp.get_data(as_text=True)
                    resp.close()

            self.assertEqual(resp.status_code, 200)
            self.assertIn("Backup de seguranca criado antes da importacao", data)
            self.assertIn('"done": true, "ok": true', data)
            self.assertNotIn("backup_pre_import", data)
            self.assertNotIn("Erro inesperado ao gravar no banco", data)
            self.assertTrue(processar.call_args.kwargs["backup_confirmado"])
            backups = list(Path(app_temp.config["BACKUP_DIR"]).glob("pre_import_*.db"))
            self.assertEqual(len(backups), 1)
            self.assertTrue(backups[0].with_suffix(".db.json").exists())


class AuditLogTests(unittest.TestCase):
    def test_registra_evento_de_auditoria(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "audit.db")

            def get_db():
                return db_core.connect(db_path)

            app_temp = endemias_app.create_app({
                "TESTING": True,
                "DB_PATH": db_path,
            })
            with app_temp.test_request_context("/", environ_base={"REMOTE_ADDR": "10.0.0.5"}):
                flask_session["uid"] = 7
                flask_session["nome"] = "Audit User"
                audit_core.registrar_evento(
                    get_db,
                    "teste_evento",
                    entidade="usuarios",
                    entidade_id=3,
                    detalhes={"campo": "nivel", "valor_novo": "admin"},
                )

            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                row = conn.execute("SELECT * FROM auditoria_eventos").fetchone()
            finally:
                conn.close()

            self.assertIsNotNone(row)
            self.assertEqual(row["acao"], "teste_evento")
            self.assertEqual(row["entidade"], "usuarios")
            self.assertEqual(row["entidade_id"], "3")
            self.assertIn("valor_novo", row["detalhes_json"])

    def test_admin_auditoria_renderiza_eventos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            app_temp = endemias_app.create_app({"TESTING": True, "DB_PATH": db_path})
            with app_temp.test_request_context("/"):
                flask_session["uid"] = 1
                flask_session["nome"] = "Teste Auditoria"
                audit_core.registrar_evento(
                    endemias_app.get_db,
                    "teste_auditoria_pagina",
                    entidade="usuarios",
                    entidade_id=1,
                    detalhes={"origem": "teste"},
                )

            client = app_temp.test_client()
            _login_client_com_usuario(client, {"id_usuario": 1, "nome": "Administrador", "nivel": "admin"})
            resp = client.get("/admin/auditoria?acao=teste_auditoria_pagina")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("teste_auditoria_pagina", html)
        self.assertIn("Eventos recentes", html)

    def test_admin_auditoria_exporta_xlsx(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            app_temp = endemias_app.create_app({"TESTING": True, "DB_PATH": db_path})
            with app_temp.test_request_context("/"):
                flask_session["uid"] = 1
                flask_session["nome"] = "Teste Auditoria"
                audit_core.registrar_evento(
                    endemias_app.get_db,
                    "teste_auditoria_exportar",
                    entidade="usuarios",
                    entidade_id=1,
                    detalhes={"origem": "teste"},
                )

            client = app_temp.test_client()
            _login_client_com_usuario(client, {"id_usuario": 1, "nome": "Administrador", "nivel": "admin"})
            resp = client.get("/admin/auditoria/exportar?acao=teste_auditoria_exportar")

        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp.content_type,
        )
        self.assertTrue(resp.data.startswith(b"PK"))


class WorkTypesConfigTests(unittest.TestCase):
    def _config(self):
        with open(endemias_app.CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)

    def test_config_json_cobre_tipos_de_trabalho_centrais(self):
        errors = work_types.validate_config_work_types(self._config())
        self.assertEqual(errors, [])

    def test_detecta_tipo_no_config_sem_cadastro_central(self):
        cfg = self._config()
        cfg["tipos_trabalho"]["XYZ"] = dict(next(iter(cfg["tipos_trabalho"].values())))

        errors = work_types.validate_config_work_types(cfg)

        self.assertTrue(any("XYZ" in e for e in errors))

    def test_detecta_tipo_central_sem_config(self):
        cfg = self._config()
        cfg["tipos_trabalho"].pop(work_types.WORK_TYPE_CODES[0])

        errors = work_types.validate_config_work_types(cfg)

        self.assertTrue(any(work_types.WORK_TYPE_CODES[0] in e for e in errors))

    def test_etl_usa_metadados_centrais_para_tratamentos(self):
        row = {
            "O imóvel foi Tratado com Larvicida?": "sim",
            "Tipo L1": "Natular",
            "Quantidade carga (gr)": "3,5",
            "Quantidade depósitos tratados": "2",
        }

        tratamentos = etl.extrair_tratamentos(row, "TB")

        self.assertEqual(tratamentos, [{
            "tipo": "Natular",
            "quantidade_carga": 3.5,
            "qtd_depositos_tratados": 2,
        }])

    def test_kobo_api_traduz_depositos_e_tratamento_aninhados_de_pe(self):
        record = {
            "group_cw8wz69/B": "95",
            "group_cw8wz69/C": "1",
            "group_cw8wz69/Total_de_Dep_sitos_eliminados": "42",
            "group_rb5ho54/Houve_tratamento_quimico": "natular__pastilha",
            "group_rb5ho54/Quantidade_dep_sitos_tratados": "3",
        }

        row = kobo_api_core._ensure_visit_columns({}, "PE", {}, record)
        depositos = etl.extrair_depositos(row, "PE")
        tratamentos = etl.extrair_tratamentos(row, "PE")

        self.assertEqual(
            [(item["tipo_deposito"], item["inspecionado"]) for item in depositos],
            [("B", 95), ("C", 1)],
        )
        self.assertEqual(depositos[0]["eliminado"], 42)
        self.assertEqual(tratamentos, [{
            "tipo": "natular__pastilha",
            "quantidade_carga": None,
            "qtd_depositos_tratados": 3,
        }])

    def test_kobo_api_traduz_depositos_e_tratamento_aninhados_de_pve(self):
        record = {
            "group_cw8wz69/D2": "9",
            "group_cw8wz69/Dep_sitos_Eliminados": "2",
            "group_rb5ho54/O_im_vel_foi_Tratado_com_Larvi": "sim",
            "group_rb5ho54/Tipo_L1": "natular_pastilha",
            "group_rb5ho54/Quantidade_carga_gr": "1",
            "group_rb5ho54/Quantidade_dep_sitos_tratados": "1",
        }

        row = kobo_api_core._ensure_visit_columns({}, "PVE", {}, record)
        depositos = etl.extrair_depositos(row, "PVE")
        tratamentos = etl.extrair_tratamentos(row, "PVE")

        self.assertEqual(depositos, [{
            "tipo_deposito": "D2",
            "inspecionado": 9,
            "eliminado": 2,
            "tratado": None,
            "tipo_tratamento": None,
            "qtd_carga": None,
        }])
        self.assertEqual(tratamentos, [{
            "tipo": "natular_pastilha",
            "quantidade_carga": 1.0,
            "qtd_depositos_tratados": 1,
        }])

    def test_etl_nao_cria_tratamento_vazio_com_nan(self):
        row = {
            "O imÃ³vel foi Tratado com Larvicida?": "sim",
            "Tipo L1": None,
            "Quantidade carga (gr)": float("nan"),
            "Quantidade depÃ³sitos tratados": None,
        }

        self.assertEqual(etl.extrair_tratamentos(row, "TB"), [])

    def test_etl_tbo_mantem_tratamentos_nos_depositos(self):
        self.assertEqual(etl.extrair_tratamentos({}, "TBO"), [])

    def test_regra_de_notificacao_padrao_fica_centralizada(self):
        self.assertEqual(work_types.gera_notificacao_padrao("PE"), 0)
        self.assertEqual(work_types.gera_notificacao_padrao("TB"), 1)
        self.assertEqual(work_types.gera_notificacao_padrao("TIPO_NOVO"), 1)

    def test_tipo_com_duracao_fica_centralizado(self):
        self.assertIn("TBO", work_types.duration_work_type_codes())
        self.assertEqual(work_types.primary_duration_work_type_code(), "TBO")


class ModuleRegistryTests(unittest.TestCase):
    def test_modulos_de_ui_tem_icones_existentes(self):
        for module in modules_core.MODULES:
            with self.subTest(module=module.key):
                self.assertTrue((ROOT / "static" / "icons" / module.icon).exists())

    def test_visualizador_nao_ve_modulos_admin(self):
        visiveis = modules_core.visible_modules({"nivel": "visualizador"})
        keys = {module.key for module in visiveis}

        self.assertNotIn("processar", keys)
        self.assertNotIn("usuarios", keys)

    def test_admin_ve_modulos_admin(self):
        visiveis = modules_core.visible_modules({"nivel": "admin"})
        keys = {module.key for module in visiveis}

        self.assertIn("processar", keys)
        self.assertIn("usuarios", keys)


class EsporotricoseSchemaTests(unittest.TestCase):
    def test_schema_cria_tabelas_de_esporotricose(self):
        with sqlite3.connect(":memory:") as conn:
            esporotricose_core.ensure_schema(conn)
            tabelas = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'esporotricose_%'"
                )
            }

        self.assertIn("esporotricose_visitas", tabelas)
        self.assertIn("esporotricose_animais", tabelas)
        self.assertIn("esporotricose_visita_agentes", tabelas)

    def test_split_agentes_esporotricose_corrige_cecon(self):
        with sqlite3.connect(":memory:") as conn:
            conn.execute("CREATE TABLE agentes (id_agente INTEGER PRIMARY KEY, nome TEXT)")
            conn.executemany(
                "INSERT INTO agentes(nome) VALUES (?)",
                [("Ceccon",), ("cecon",), ("Ana Beatriz",), ("Márcio",)],
            )

            nomes = esporotricose_core._split_agentes(conn, "ana beatriz ana_beatriz cecon m_rcio")

        self.assertEqual(nomes, ["Ana Beatriz", "Ceccon", "Márcio"])

    def test_schema_esporotricose_migra_alias_m_rcio_para_marcio(self):
        with sqlite3.connect(":memory:") as conn:
            conn.row_factory = sqlite3.Row
            conn.executescript("""
                CREATE TABLE agentes (
                    id_agente INTEGER PRIMARY KEY,
                    nome TEXT NOT NULL UNIQUE,
                    ativo INTEGER NOT NULL DEFAULT 1
                );
                INSERT INTO agentes(id_agente, nome) VALUES (18, 'Márcio');
            """)
            esporotricose_core.ensure_schema(conn)
            conn.execute("INSERT INTO agentes(id_agente, nome) VALUES (72, 'm_rcio')")
            conn.execute(
                """INSERT INTO esporotricose_visitas
                   (id_visita, kobo_uuid, data, agentes_texto, processado_em)
                   VALUES ('v1', 'u1', '2026-06-11', 'adilson m_rcio', '2026-06-16')"""
            )
            conn.execute(
                "INSERT INTO esporotricose_visita_agentes(id_visita, id_agente) VALUES ('v1', 72)"
            )
            esporotricose_core.ensure_schema(conn)

            agentes = [row["nome"] for row in conn.execute("SELECT nome FROM agentes ORDER BY nome")]
            vinculos = [
                row["id_agente"]
                for row in conn.execute("SELECT id_agente FROM esporotricose_visita_agentes WHERE id_visita='v1'")
            ]
            texto = conn.execute("SELECT agentes_texto FROM esporotricose_visitas WHERE id_visita='v1'").fetchone()[0]

        self.assertIn("Márcio", agentes)
        self.assertNotIn("m_rcio", agentes)
        self.assertEqual(vinculos, [18])
        self.assertIn("Márcio", texto)

    def test_listar_doentes_csv_exporta_campos_para_qgis(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_qgis.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                esporotricose_core.ensure_schema(conn)
                agora = "2026-06-16T10:00:00"
                conn.execute(
                    """INSERT INTO esporotricose_doentes_animais
                       (chave, tutor, nome, sexo, telefone, localidade, quarteirao, endereco,
                        latitude, longitude, sinan, status, data_notificacao, bloqueio, data_bloqueio,
                        observacoes_entomologica, pedido_zoomed, criado_em, atualizado_em)
                       VALUES ('abc', 'Maria', 'Mimi', 'Fêmea', '5541999999999', 'Graziela',
                               '10', 'Rua Teste, 123', -25.31, -49.30, '12345',
                               'Em tratamento', '2026-06-01', 'Sim', '2026-06-10', 'Observação', 'Sim', ?, ?)""",
                    (agora, agora),
                )
                animal_id = conn.execute("SELECT id_animal_doente FROM esporotricose_doentes_animais").fetchone()[0]
                conn.execute(
                    """INSERT INTO esporotricose_doentes_receitas
                       (id_animal_doente, data_notificacao, data_receita, capsulas_total,
                        status, criado_em, atualizado_em)
                       VALUES (?, '2026-06-01', '2026-06-02', 30, 'Em tratamento', ?, ?)""",
                    (animal_id, agora, agora),
                )
                receita_id = conn.execute("SELECT id_receita FROM esporotricose_doentes_receitas").fetchone()[0]
                conn.execute(
                    """INSERT INTO esporotricose_doentes_entregas
                       (id_receita, quantidade, data_entrega, baixa_zoomed, criado_em)
                       VALUES (?, 30, '2026-06-05', 'Não', ?)""",
                    (receita_id, agora),
                )
                conn.commit()
            finally:
                conn.close()

            rows = esporotricose_core.listar_doentes_csv(str(db_path), {})

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["animal"], "Mimi")
        self.assertEqual(rows[0]["tutor"], "Maria")
        self.assertEqual(rows[0]["latitude"], -25.31)
        self.assertEqual(rows[0]["longitude"], -49.3)
        self.assertEqual(rows[0]["data_notificacao"], "2026-06-01")
        self.assertEqual(rows[0]["primeira_notificacao"], "2026-06-01")
        self.assertEqual(rows[0]["ultima_notificacao"], "2026-06-01")
        self.assertEqual(rows[0]["baixa_zoomed"], "Pendente")
        self.assertEqual(rows[0]["bloqueio"], "Realizado")

    def test_data_notificacao_pertence_ao_doente_e_migra_historico(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_notificacao.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                esporotricose_core.ensure_schema(conn)
                conn.execute(
                    """INSERT INTO esporotricose_doentes_animais
                       (chave, tutor, nome, status, criado_em, atualizado_em)
                       VALUES ('historico', 'ELISA MARIA YAMAGUICHI', 'MAFALDA',
                               'Em tratamento', '2026-06-15T10:00:00', '2026-06-15T10:00:00')"""
                )
                conn.commit()
                esporotricose_core.ensure_schema(conn)
                historico = conn.execute(
                    "SELECT data_notificacao FROM esporotricose_doentes_animais WHERE chave='historico'"
                ).fetchone()[0]
            finally:
                conn.close()

            novo_id = esporotricose_core.salvar_doente(str(db_path), {
                "nome": "Novo animal",
                "tutor": "Novo tutor",
                "status": "Em tratamento",
            })
            receita_id = esporotricose_core.salvar_receita_doente(str(db_path), novo_id, {
                "data_notificacao": "2020-01-01",
                "data_receita": "2026-07-16",
                "capsulas_total": 30,
            })
            esporotricose_core.salvar_doente(str(db_path), {
                "id_animal_doente": novo_id,
                "nome": "Novo animal",
                "tutor": "Novo tutor",
                "status": "Em tratamento",
                "data_notificacao": "2026-07-10",
            })
            conn = sqlite3.connect(db_path)
            try:
                data_novo = conn.execute(
                    "SELECT data_notificacao FROM esporotricose_doentes_animais WHERE id_animal_doente=?",
                    (novo_id,),
                ).fetchone()[0]
                data_receita = conn.execute(
                    "SELECT data_notificacao FROM esporotricose_doentes_receitas WHERE id_receita=?",
                    (receita_id,),
                ).fetchone()[0]
            finally:
                conn.close()

        self.assertEqual(historico, "2026-04-07")
        self.assertEqual(data_novo, "2026-07-10")
        self.assertIsNone(data_receita)

    def test_filtro_doentes_bloqueio_usa_status_operacional(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_bloqueio.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                esporotricose_core.ensure_schema(conn)
                agora = "2026-06-16T10:00:00"
                for chave, nome, bloqueio in (
                    ("a", "Mimi", "REALIZADO"),
                    ("b", "Luna", "NÃO REALIZADO"),
                    ("c", "Nina", "NÃO NECESSÁRIO"),
                ):
                    conn.execute(
                        """INSERT INTO esporotricose_doentes_animais
                           (chave, tutor, nome, status, bloqueio, criado_em, atualizado_em)
                           VALUES (?, 'Tutor', ?, 'Em tratamento', ?, ?, ?)""",
                        (chave, nome, bloqueio, agora, agora),
                    )
                conn.commit()
            finally:
                conn.close()

            realizados = esporotricose_core.listar_doentes(str(db_path), {"bloqueio": "Realizado"})
            nao_realizados = esporotricose_core.listar_doentes(str(db_path), {"bloqueio": "Não realizado"})
            nao_necessarios = esporotricose_core.listar_doentes(str(db_path), {"bloqueio": "Não necessário"})

        self.assertEqual([r["nome"] for r in realizados["registros"]], ["Mimi"])
        self.assertEqual([r["nome"] for r in nao_realizados["registros"]], ["Luna"])
        self.assertEqual([r["nome"] for r in nao_necessarios["registros"]], ["Nina"])

    def test_doente_salva_e_retorna_cpf_do_tutor(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_doente_cpf.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                esporotricose_core.ensure_schema(conn)
            finally:
                conn.close()

            id_animal = esporotricose_core.salvar_doente(str(db_path), {
                "nome": "Paciente CPF",
                "tutor": "Tutor CPF",
                "telefone": "(41) 99999-0000",
                "cpf": "123.456.789-09",
                "status": "Em tratamento",
            })
            animal = esporotricose_core.obter_doente(str(db_path), id_animal)
            filtrados = esporotricose_core.listar_doentes(str(db_path), {"busca": "12345678909"})

        self.assertEqual(animal["cpf"], "12345678909")
        self.assertEqual(filtrados["total"], 1)
        self.assertEqual(filtrados["registros"][0]["cpf"], "12345678909")

    def test_estoque_medicacao_calcula_tratamento_e_movimentos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_estoque.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                esporotricose_core.ensure_schema(conn)
            finally:
                conn.close()

            ativo = esporotricose_core.salvar_doente(str(db_path), {
                "nome": "Ativo",
                "tutor": "Tutor Ativo",
                "status": "Em tratamento",
            })
            falecido = esporotricose_core.salvar_doente(str(db_path), {
                "nome": "Falecido",
                "tutor": "Tutor Falecido",
                "status": "Faleceu",
            })
            receita_ativo = esporotricose_core.salvar_receita_doente(
                str(db_path), ativo, {"capsulas_total": 100, "status": "Em tratamento"},
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path), receita_ativo, {"quantidade": 30, "baixa_zoomed": "Sim"},
            )
            receita_falecido = esporotricose_core.salvar_receita_doente(
                str(db_path), falecido, {"capsulas_total": 200, "status": "Faleceu"},
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path), receita_falecido, {"quantidade": 80, "baixa_zoomed": "Sim"},
            )
            esporotricose_core.salvar_estoque_medicacao(
                str(db_path), {"tipo": "Entrada", "quantidade": 120, "descricao": "Remessa"},
            )
            esporotricose_core.salvar_estoque_medicacao(
                str(db_path), {"tipo": "Sobra", "quantidade": 20, "origem": "Falecido"},
            )
            esporotricose_core.salvar_estoque_medicacao(
                str(db_path), {"tipo": "Sa\u00edda", "quantidade": 10, "descricao": "Uso externo"},
            )
            estoque = esporotricose_core.estoque_medicacao(str(db_path))
            doentes = esporotricose_core.listar_doentes(str(db_path), {})

        self.assertEqual(estoque["totais"]["faltantes_tratamento"], 70)
        self.assertEqual(doentes["totais"]["capsulas_baixa_zoomed"], 110)
        self.assertEqual(estoque["totais"]["saldo_setor"], 20)
        self.assertEqual(estoque["totais"]["saldo_apos_reserva"], -50)
        self.assertEqual(estoque["totais"]["saldo_historico_com_entregas"], 20)
        self.assertEqual(estoque["totais"]["entradas_zoomed"], 120)
        self.assertEqual(estoque["totais"]["sobras_retornadas"], 20)
        self.assertEqual(estoque["totais"]["ajustes_estoque"], 0)
        self.assertEqual(estoque["totais"]["saidas_entregas"], 110)
        self.assertEqual(estoque["totais"]["saidas_manuais"], 10)
        self.assertEqual(estoque["totais"]["saidas_setor"], 120)
        self.assertEqual(estoque["totais"]["sobras_lancadas"], 20)
        self.assertEqual(estoque["totais"]["sobra_potencial_encerrados"], 80)
        self.assertEqual(len(estoque["movimentos"]), 3)
        self.assertEqual(len(estoque["movimentos_automaticos"]), 2)
        self.assertEqual(sum(item["quantidade"] for item in estoque["movimentos_automaticos"]), 110)
        self.assertEqual(estoque["movimentos_automaticos"][0]["tipo"], "Saída automática")

    def test_entrega_historica_converte_acumulado_em_quantidade_da_entrega(self):
        self.assertEqual(esporotricose_core._quantidade_entrega_historica(30, 0), 30)
        self.assertEqual(esporotricose_core._quantidade_entrega_historica(60, 30), 30)
        self.assertEqual(esporotricose_core._quantidade_entrega_historica(90, 60), 30)
        self.assertEqual(esporotricose_core._quantidade_entrega_historica(60, 0), 60)

    def test_animais_permite_filtro_multiplo_e_busca_ferido(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "esporotricose_animais.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.execute(
                    "CREATE TABLE localidades ("
                    "id_localidade INTEGER PRIMARY KEY, nome TEXT UNIQUE, cod_localidade TEXT)"
                )
                esporotricose_core.ensure_schema(conn)
                conn.execute(
                    """INSERT INTO esporotricose_visitas(
                        id_visita, kobo_uuid, data, agentes_texto, localidade,
                        logradouro, numero, morador, origem_estrutura, processado_em
                    ) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    ("v1", "uuid-v1", "2026-07-03", "Agente A", "Centro",
                     "Rua Um", "10", "Tutor", "nova", "2026-07-03T08:00:00"),
                )
                conn.executemany(
                    """INSERT INTO esporotricose_animais(
                        id_animal, id_visita, especie, nome, feridas, processado_em
                    ) VALUES (?,?,?,?,?,?)""",
                    [
                        ("a1", "v1", "Gato", "Mimi", "Sim", "2026-07-03T08:00:00"),
                        ("a2", "v1", "Outro", "Nino", "Desconhecido", "2026-07-03T08:00:00"),
                    ],
                )
                conn.commit()
            finally:
                conn.close()

            esporotricose_core.atualizar_animal(str(db_path), "a1", {
                "busca_ferido_data": "2026-07-04",
                "busca_ferido_agente": "Agente B",
                "busca_ferido_observacoes": "Averiguar lesao",
            })
            animais = esporotricose_core.listar_animais(str(db_path), {
                "especie": ["Gato", "Outro"],
                "feridas": ["Sim", "Desconhecido"],
            })

        self.assertEqual(animais["total"], 2)
        animal = next(item for item in animais["registros"] if item["id_animal"] == "a1")
        self.assertEqual(animal["agentes_texto"], "Agente A")
        self.assertEqual(animal["busca_ferido_data"], "2026-07-04")
        self.assertEqual(animal["busca_ferido_agente"], "Agente B")


class RecolhimentosTests(unittest.TestCase):
    def test_schema_cria_tabelas_de_recolhimentos(self):
        with sqlite3.connect(":memory:") as conn:
            recolhimentos_core.ensure_schema(conn)
            tabelas = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'recolhimento%'"
                )
            }

        self.assertIn("recolhimentos", tabelas)
        self.assertIn("recolhimento_agentes", tabelas)

    def test_parse_recolhimentos_kobo_e_legado(self):
        novo = ROOT / "PAGINA_RECOLHIMENTOS" / "RECOLHIMENTO_2026_05_05-21.xlsx"
        legado = ROOT / "PAGINA_RECOLHIMENTOS" / "RECOLHIMENTO_LEGADO.xlsx"
        if not novo.exists() or not legado.exists():
            self.skipTest("Planilhas temporarias de recolhimento nao estao presentes.")

        registros_novos = recolhimentos_core.parse_workbook(novo)
        registros_legados = recolhimentos_core.parse_workbook(legado)

        self.assertGreaterEqual(len(registros_novos), 1)
        self.assertGreaterEqual(len(registros_legados), 1)
        self.assertEqual(registros_novos[0]["localidade"], "Lamenha")
        self.assertIn("Atagil", registros_novos[0]["agentes_texto"])
        self.assertIn("total_materiais", registros_novos[0])
        self.assertIn("total_materiais", registros_legados[0])

    def test_split_agentes_recolhimento_reconhece_legado_sem_separador(self):
        nomes = recolhimentos_core._split_agentes("Adriana Ana Beatriz Atagil Evaldo Henrique Pedro")

        self.assertEqual(
            nomes,
            ["Adriana", "Ana Beatriz", "Atagil", "Evaldo", "Henrique", "Pedro"],
        )


class AmostrasAnimaisTests(unittest.TestCase):
    def test_schema_cria_tabelas_de_amostras_animais(self):
        with sqlite3.connect(":memory:") as conn:
            amostras_animais_core.ensure_schema(conn)
            tabelas = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'amostra%'"
                )
            }

        self.assertIn("amostras_animais", tabelas)
        self.assertIn("amostra_animais_agentes", tabelas)

    def test_parse_amostras_animais_kobo_e_legado(self):
        novo = ROOT / "AMOSTRA_ANIMAIS_NOVA_PAGINA" / "AMOSTRA_ANIMAIS_KOBO.xlsx"
        legado = ROOT / "AMOSTRA_ANIMAIS_NOVA_PAGINA" / "AMOSTRA_ANIMAIS_LEGADO.xlsx"
        if not novo.exists() or not legado.exists():
            self.skipTest("Planilhas temporarias de amostra de animais nao estao presentes.")

        registros_novos = amostras_animais_core.parse_workbook(novo)
        registros_legados = amostras_animais_core.parse_workbook(legado)

        self.assertGreaterEqual(len(registros_novos), 1)
        self.assertGreaterEqual(len(registros_legados), 1)
        self.assertIn("Atagil", registros_novos[0]["agentes_texto"])
        self.assertEqual(registros_novos[0]["tipo_animal"], "Escorpião")
        self.assertIn("quantidade", registros_novos[0])

    def test_split_agentes_amostras_corrige_fernado(self):
        nomes = amostras_animais_core._split_agentes("Adriana Ana Beatriz Fernado Pedro")

        self.assertEqual(nomes, ["Adriana", "Ana Beatriz", "Fernando", "Pedro"])


class BriTests(unittest.TestCase):
    def test_schema_cria_tabelas_de_bri(self):
        with sqlite3.connect(":memory:") as conn:
            bri_core.ensure_schema(conn)
            tabelas = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'bri%'"
                )
            }

        self.assertIn("bri_registros", tabelas)
        self.assertIn("bri_agentes", tabelas)

    def test_parse_bri_kobo_e_legado(self):
        novo = ROOT / "BRI_NOVA_PAGINA" / "BRI_KOBO.xlsx"
        legado = ROOT / "BRI_NOVA_PAGINA" / "BRI_LEGADO.xlsx"
        if not novo.exists() or not legado.exists():
            self.skipTest("Planilhas temporarias de BRI nao estao presentes.")

        registros_novos = bri_core.parse_workbook(novo)
        registros_legados = bri_core.parse_workbook(legado)

        self.assertGreaterEqual(len(registros_novos), 1)
        self.assertGreaterEqual(len(registros_legados), 1)
        self.assertEqual(registros_novos[0]["destino_tratamento"], "Ponto Estratégico")
        self.assertEqual(registros_novos[1]["destino_tratamento"], "Ovitrampa")
        self.assertEqual(registros_legados[0]["sispncd"], "0065/2026")
        self.assertIn("quantidade_carga", registros_novos[0])

    def test_bri_de_ponto_estrategico_vincula_por_alias(self):
        with sqlite3.connect(":memory:") as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT, cod_localidade TEXT)")
            conn.execute("CREATE TABLE agentes (id_agente INTEGER PRIMARY KEY, nome TEXT UNIQUE)")
            pe_core.ensure_schema(conn)
            bri_core.ensure_schema(conn)
            pe_core.inserir(conn, {
                "codigo_pe": "PE-0009",
                "nome": "Meio Ambiente",
                "localidade": "Paraíso",
                "quarteirao": 481,
            })
            pe_core.ensure_schema(conn)
            conn.execute(
                """INSERT INTO bri_registros (
                       id_bri, data, destino_tratamento, localidade, id_localidade,
                       quarteirao, local_tratamento, logradouro, processado_em
                   ) VALUES (?,?,?,?,?,?,?,?,?)""",
                (
                    "bri-pe",
                    "2026-06-01",
                    "Ponto Estratégico",
                    "Paraíso",
                    1,
                    129,
                    "Meio Ambiente",
                    "Trav. Rio Cachoeirinha",
                    "agora",
                ),
            )

            resultado = bri_core.vincular_registros_pe_por_alias(conn)
            bri = conn.execute("SELECT codigo_pe FROM bri_registros WHERE id_bri='bri-pe'").fetchone()
            row = conn.execute(
                """SELECT pe.codigo_pe,
                          (SELECT COUNT(*) FROM bri_registros b
                            WHERE b.destino_tratamento='Ponto Estratégico'
                              AND (
                                  b.id_pe=pe.id_pe
                                  OR (b.id_pe IS NULL AND b.id_localidade=pe.id_localidade AND b.quarteirao=pe.quarteirao)
                              )) AS bri_total
                     FROM pontos_estrategicos pe
                    WHERE pe.codigo_pe='PE-0009'"""
            ).fetchone()

        self.assertEqual(resultado["atualizados"], 1)
        self.assertEqual(bri["codigo_pe"], "PE-0009")
        self.assertEqual(row["bri_total"], 1)


class PontosEstrategicosTests(unittest.TestCase):
    def test_schema_cria_tabela_de_pontos_estrategicos(self):
        with sqlite3.connect(":memory:") as conn:
            conn.execute("CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT, cod_localidade TEXT)")
            pe_core.ensure_schema(conn)
            tabelas = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='pontos_estrategicos'"
                )
            }

        self.assertIn("pontos_estrategicos", tabelas)

    def test_importa_csv_inicial_com_codigo_sequencial(self):
        csv_path = ROOT / "PONTOS_ESTRATEGICOS.csv"
        if not csv_path.exists():
            self.skipTest("CSV temporario de pontos estrategicos nao esta presente.")

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "pe.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT, cod_localidade TEXT)")
            finally:
                conn.close()

            resultado = pe_core.importar_csv_inicial(csv_path, str(db_path))

            self.assertGreaterEqual(resultado["inseridos"], 1)
            conn = sqlite3.connect(db_path)
            try:
                primeiro = conn.execute(
                    "SELECT codigo_pe, nome FROM pontos_estrategicos ORDER BY id_pe LIMIT 1"
                ).fetchone()
                total = conn.execute("SELECT COUNT(*) FROM pontos_estrategicos").fetchone()[0]
            finally:
                conn.close()

        self.assertEqual(primeiro[0], "PE-0001")
        self.assertEqual(total, resultado["inseridos"])

    def test_exporta_pontos_estrategicos_filtrados_xlsx_e_pdf(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            pe_core.salvar(str(db_path), {
                "nome": "Ferro Velho Central",
                "localidade": "Centro",
                "quarteirao": 10,
                "logradouro": "Rua A",
                "numero": "123",
                "tipo": "Ferro velho",
                "telefone": "41999990000",
                "situacao": 1,
            })
            pe_core.salvar(str(db_path), {
                "nome": "Borracharia Norte",
                "localidade": "Norte",
                "quarteirao": 20,
                "tipo": "Borracharia",
                "situacao": 1,
            })
            app_temp = endemias_app.create_app({"TESTING": True, "DB_PATH": db_path})
            client = app_temp.test_client()
            _login_client_com_usuario(client, {"id_usuario": 1, "nome": "Administrador", "nivel": "admin"})

            resp_xlsx = client.get("/api/pontos-estrategicos/exportar?localidade=Sede&formato=xlsx")
            resp_pdf = client.get("/api/pontos-estrategicos/exportar?localidade=Sede&formato=pdf")

        self.assertEqual(resp_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp_xlsx.content_type,
        )
        self.assertTrue(resp_xlsx.data.startswith(b"PK"))
        wb = openpyxl.load_workbook(io.BytesIO(resp_xlsx.data))
        ws = wb.active
        valores = [cell.value for cell in ws["B"]]
        self.assertIn("Ferro Velho Central", valores)
        self.assertNotIn("Borracharia Norte", valores)

        self.assertEqual(resp_pdf.status_code, 200)
        html = resp_pdf.data.decode("utf-8")
        self.assertIn("Ferro Velho Central", html)
        self.assertNotIn("Borracharia Norte", html)

    def test_aliases_de_pe_resolvem_logradouros_antigos(self):
        with sqlite3.connect(":memory:") as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT, cod_localidade TEXT)")
            pe_core.ensure_schema(conn)
            for codigo, nome, localidade in (
                ("PE-0007", "Borracharia Mauricio", "Tranqueira"),
                ("PE-0020", "Cal Eloi", "Sede"),
                ("PE-0042", "Cemiterio Prado", "Rosana"),
                ("PE-0043", "Condominio Jersey City - LYX", "Rosana"),
            ):
                pe_core.inserir(conn, {"codigo_pe": codigo, "nome": nome, "localidade": localidade})
            pe_core.inserir(conn, {
                "codigo_pe": "PE-0023",
                "nome": "Borracharia (Davi)",
                "localidade": "Roma",
                "logradouro": "Rodovia dos Minerios",
                "numero": "5142",
            })
            pe_core.ensure_schema(conn)

            self.assertEqual(
                pe_core.resolver_alias_visita(
                    conn,
                    "Borracharia (prox celeste) -  Rodovia Dos Minérios",
                    "Tranqueira",
                )["codigo_pe"],
                "PE-0007",
            )
            self.assertEqual(
                pe_core.resolver_alias_visita(conn, "Cal Eloi -  Pedro Teixeira Alves", "Sede")["codigo_pe"],
                "PE-0020",
            )
            self.assertEqual(
                pe_core.resolver_alias_visita(conn, "Pedro Teixeira Alves - Cal Eloi", "Sede")["codigo_pe"],
                "PE-0020",
            )
            self.assertEqual(
                pe_core.resolver_alias_visita(conn, "Rodovia dos Minerios - Borracharia (Davi)", "Sede")["codigo_pe"],
                "PE-0023",
            )
            self.assertEqual(pe_core.resolver_alias_visita(conn, "CEMITERIO", "Rosana")["codigo_pe"], "PE-0042")
            self.assertEqual(pe_core.resolver_alias_visita(conn, "CONDOMINIO", "Rosana")["codigo_pe"], "PE-0043")
            self.assertEqual(pe_core.resolver_alias_visita(conn, "LYX", "Rosana")["codigo_pe"], "PE-0043")
            self.assertIsNone(pe_core.resolver_alias_visita(conn, "cemitério", "São Venâncio"))

    def test_listagem_de_pe_prioriza_vinculo_direto_da_visita(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            pe_core.salvar(str(db_path), {
                "codigo_pe": "PE-9001",
                "nome": "PE Principal",
                "localidade": "Centro",
                "quarteirao": 10,
                "situacao": 1,
            })
            pe_core.salvar(str(db_path), {
                "codigo_pe": "PE-9002",
                "nome": "PE Mesmo Quarteirao",
                "localidade": "Centro",
                "quarteirao": 10,
                "situacao": 1,
            })
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                pe_core.ensure_schema(conn)
                pe1 = conn.execute("SELECT id_pe FROM pontos_estrategicos WHERE codigo_pe='PE-9001'").fetchone()["id_pe"]
                conn.execute(
                    """INSERT INTO visitas (
                           id_visita, kobo_uuid, tipo, data, id_localidade, localidade,
                           quarteirao, logradouro, processado_em, id_pe, codigo_pe
                       ) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    ("v-pe", "uuid-pe", "PE", "2026-06-01", 1, "Centro", 10, "Alias", "agora", pe1, "PE-9001"),
                )
                conn.commit()
            finally:
                conn.close()

            dados = pe_core.listar(str(db_path), limite=None)["registros"]
            por_codigo = {row["codigo_pe"]: row for row in dados if row["codigo_pe"] in ("PE-9001", "PE-9002")}

        self.assertEqual(por_codigo["PE-9001"]["visitas_pe_total"], 1)
        self.assertEqual(por_codigo["PE-9002"]["visitas_pe_total"], 0)

    def test_vincula_visitas_antigas_de_pe_por_alias(self):
        with sqlite3.connect(":memory:") as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT, cod_localidade TEXT)")
            conn.execute(
                """CREATE TABLE visitas (
                       id_visita TEXT PRIMARY KEY,
                       tipo TEXT,
                       data TEXT,
                       localidade TEXT,
                       logradouro TEXT
                   )"""
            )
            pe_core.ensure_schema(conn)
            pe_core.inserir(conn, {"codigo_pe": "PE-0007", "nome": "Borracharia Mauricio", "localidade": "Tranqueira"})
            pe_core.ensure_schema(conn)
            conn.execute(
                """INSERT INTO visitas(id_visita, tipo, data, localidade, logradouro)
                   VALUES ('v1', 'PE', '2026-06-01', 'Tranqueira',
                           'Borracharia (prox celeste) -  Rodovia Dos Minérios')"""
            )

            resultado = pe_core.vincular_visitas_existentes_por_alias(conn)
            visita = conn.execute("SELECT codigo_pe FROM visitas WHERE id_visita='v1'").fetchone()

        self.assertEqual(resultado["atualizadas"], 1)
        self.assertEqual(visita["codigo_pe"], "PE-0007")


class SispncdIndiceTests(unittest.TestCase):
    def _criar_planilha_indice(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "indice"
        ws.append(["TIPO", "SISPNCD", "DATA", "LOCALIDADE"])
        ws.append(["PE", "0000/0000", "2026-02-06", "Lamenha"])
        ws.append(["PE", "0006/2006", "2026-01-06", "São João Batista"])
        ws.append(["TBO", "0163/2026", "2026-04-13", "Graziela"])
        wb.save(path)

    def _criar_banco_visitas(self, path):
        conn = sqlite3.connect(path)
        try:
            conn.executescript("""
                CREATE TABLE localidades (
                    id_localidade INTEGER PRIMARY KEY,
                    nome TEXT NOT NULL
                );
                CREATE TABLE visitas (
                    id_visita TEXT PRIMARY KEY,
                    tipo TEXT NOT NULL,
                    data TEXT NOT NULL,
                    localidade TEXT,
                    id_localidade INTEGER,
                    SISPNCD TEXT
                );
                INSERT INTO localidades(id_localidade, nome) VALUES
                    (1, 'Lamenha'),
                    (2, 'São João Batista'),
                    (3, 'Graziela');
                INSERT INTO visitas(id_visita, tipo, data, localidade, id_localidade, SISPNCD) VALUES
                    ('v1', 'PE', '2026-02-06', 'Lamenha', 1, NULL),
                    ('v2', 'PE', '2026-01-06', 'São João Batista', 2, NULL),
                    ('v3', 'TBO', '2026-04-13', 'Graziela', 3, NULL),
                    ('v4', 'TB', '2026-04-13', 'Graziela', 3, NULL);
            """)
            conn.commit()
        finally:
            conn.close()

    def test_indice_corrige_codigo_e_aplica_0000(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "endemias.db"
            xlsx_path = Path(tmpdir) / "INDICE_SISPNCD.xlsx"
            self._criar_banco_visitas(db_path)
            self._criar_planilha_indice(xlsx_path)

            previa = sispncd_indice_core.previsualizar(str(db_path), str(xlsx_path))
            resultado = sispncd_indice_core.aplicar(str(db_path), str(xlsx_path))

            conn = sqlite3.connect(db_path)
            try:
                valores = dict(conn.execute("SELECT id_visita, SISPNCD FROM visitas").fetchall())
            finally:
                conn.close()

        self.assertEqual(previa["visitas_atualizaveis"], 3)
        self.assertEqual(previa["correcoes"][0]["de"], "0006/2006")
        self.assertEqual(previa["correcoes"][0]["para"], "0006/2026")
        self.assertEqual(resultado["atualizados"], 3)
        self.assertEqual(valores["v1"], "0000/0000")
        self.assertEqual(valores["v2"], "0006/2026")
        self.assertEqual(valores["v3"], "0163/2026")
        self.assertIsNone(valores["v4"])


class ProtectedRouteTests(unittest.TestCase):
    def test_home_sem_login_redireciona_para_login(self):
        endemias_app.app.config["TESTING"] = True
        with endemias_app.app.test_client() as client:
            resp = client.get("/")

        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])

    def test_sessao_com_usuario_inexistente_redireciona_para_login(self):
        endemias_app.app.config["TESTING"] = True
        with endemias_app.app.test_client() as client:
            with client.session_transaction() as sess:
                sess["uid"] = 999999999
                sess["nome"] = "Usuario removido"
                sess["nivel"] = "admin"

            resp = client.get("/")

        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])

    def test_respostas_incluem_headers_basicos_de_seguranca(self):
        endemias_app.app.config["TESTING"] = True
        with endemias_app.app.test_client() as client:
            resp = client.get("/login")

        self.assertEqual(resp.headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(resp.headers["Referrer-Policy"], "same-origin")
        self.assertEqual(resp.headers["X-Frame-Options"], "SAMEORIGIN")
        self.assertIn("camera=()", resp.headers["Permissions-Policy"])
        csp = resp.headers["Content-Security-Policy-Report-Only"]
        self.assertIn("default-src 'self'", csp)
        self.assertIn("frame-ancestors 'self'", csp)
        self.assertIn("form-action 'self'", csp)

    def test_csp_pode_ser_ativada_em_modo_bloqueante(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
            "CSP_REPORT_ONLY": False,
        })
        with app_temp.test_client() as client:
            resp = client.get("/login")

        self.assertIn("default-src 'self'", resp.headers["Content-Security-Policy"])
        self.assertNotIn("Content-Security-Policy-Report-Only", resp.headers)

    def test_csp_pode_remover_inline_quando_configurada(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
            "CSP_ALLOW_INLINE": False,
        })
        with app_temp.test_client() as client:
            resp = client.get("/login")

        csp = resp.headers["Content-Security-Policy-Report-Only"]
        self.assertIn("script-src 'self'", csp)
        self.assertNotIn("'unsafe-inline'", csp)

    def test_cookie_secure_pode_ser_ativado_por_config(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
            "SESSION_COOKIE_SECURE": True,
        })

        self.assertTrue(app_temp.config["SESSION_COOKIE_SECURE"])

    def test_logout_nao_aceita_get(self):
        client = _client_logado()

        resp = client.get("/logout")

        self.assertEqual(resp.status_code, 405)

    def test_minha_senha_exige_csrf(self):
        client = _client_logado()

        resp = client.post("/minha-senha", data={
            "atual": "qualquer",
            "nova": "123456",
            "confirmar": "123456",
        })

        self.assertEqual(resp.status_code, 400)

    def test_confirmar_processamento_nao_aceita_get(self):
        client = _client_logado("admin")

        resp = client.get("/processar/confirmar/job-inexistente")

        self.assertEqual(resp.status_code, 405)

    def test_stream_processamento_nao_aceita_get(self):
        client = _client_logado("admin")

        resp = client.get("/processar/stream/job-inexistente")

        self.assertEqual(resp.status_code, 405)

    def test_rotas_processamento_rejeitam_job_id_invalido(self):
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            client = _client_logado("admin")
            for rota in (
                "/processar/stream/job-invalido",
                "/processar/confirmar/job-invalido",
                "/processar/cancelar/job-invalido",
            ):
                with self.subTest(rota=rota):
                    resp = client.post(rota)
                    self.assertEqual(resp.status_code, 404)
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

    def test_visualizador_nao_gera_consolidados(self):
        client = _client_logado("visualizador")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp = client.post("/saida/gerar-consolidados", json={"tipo": "PE"})
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        self.assertEqual(resp.status_code, 403)

    def test_impressao_html_individual_nao_aceita_get(self):
        client = _client_logado("admin")
        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            row = conn.execute(
                "SELECT id_foco FROM focos_positivos WHERE gera_notificacao=1 LIMIT 1"
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(row)

        resp = client.get(f"/notificacoes/foco/{row[0]}/imprimir-html")

        self.assertEqual(resp.status_code, 405)


class MainPagesSmokeTests(unittest.TestCase):
    def test_favicon_esta_disponivel_no_login_e_no_sistema(self):
        client = endemias_app.app.test_client()
        login = client.get("/login")
        self.assertEqual(login.status_code, 200)
        self.assertIn('/static/img/favicon.png', login.data.decode("utf-8"))

        favicon = client.get("/static/img/favicon.png")
        self.assertEqual(favicon.status_code, 200)
        self.assertEqual(favicon.content_type, "image/png")
        self.assertGreater(len(favicon.data), 0)
        favicon.close()

        client_logado = _client_logado()
        dashboard = client_logado.get("/dashboard")
        self.assertIn('/static/img/favicon.png', dashboard.data.decode("utf-8"))

    def test_paginas_principais_logadas_respondem_200(self):
        client = _client_logado()
        rotas = [
            "/dashboard",
            "/visitas",
            "/laboratorio",
            "/conta-ovos-sispncd",
            "/ovitrampas",
            "/esporotricose",
            "/recolhimentos",
            "/amostras-animais",
            "/bri",
            "/notificacoes",
            "/mapa",
            "/pontos-estrategicos",
            "/acoes-setor",
            "/agenda",
            "/admin/agentes",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)
                self.assertIn("text/html", resp.content_type)

    def test_dashboard_integrado_exibe_esporotricose(self):
        client = _client_logado()
        resp = client.get("/dashboard")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Dashboard Integrado", html)
        self.assertIn("Esporotricose", html)
        self.assertIn("Pendências de Pontos Estratégicos", html)
        self.assertIn("chComparativo", html)
        self.assertIn("chAtividade", html)
        self.assertIn("chEspEvolucao", html)
        self.assertIn("tab-ovitrampas", html)
        self.assertIn("chOviSemana", html)
        self.assertIn("dash-ovi-card", html)
        self.assertIn("filter-choice-grid", html)
        self.assertIn("filter-choice-section", html)
        self.assertIn("abrirAbaOvitrampas", html)
        self.assertIn("t-ovi-semanas", html)
        self.assertIn("t-ovi-localidades-detalhe", html)
        self.assertIn("datalabels:{ display:false }", html)

    def test_pagina_ovitrampas_exibe_edicao_em_lote(self):
        client = _client_logado()
        resp = client.get("/ovitrampas")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Edição em lote das leituras filtradas", html)
        self.assertIn('id="ovi-lote-lab"', html)
        self.assertIn('id="ovi-lote-somente-vazios"', html)
        self.assertIn('data-ovi-tab="diarios"', html)
        self.assertIn('id="ovi-dia-import-form"', html)
        self.assertIn("data-ovi-dia-drag", html)
        self.assertIn("/reordenar", html)
        self.assertIn("addEventListener('pointerdown'", html)
        self.assertIn("addEventListener('pointermove'", html)
        self.assertIn("translate3d", html)
        self.assertIn("layer.style.left = `${rect.left}px`", html)
        self.assertIn("layer.style.top = `${rect.top}px`", html)
        self.assertIn("containerRect.bottom - state.startRect.height", html)
        self.assertIn("e.currentTarget.setPointerCapture", html)
        self.assertNotIn("handle.setPointerCapture", html)
        self.assertIn("velocidadeAutoScrollDiario", html)
        self.assertIn("prefers-reduced-motion", html)
        self.assertNotIn('draggable="true" data-ovi-dia-drag', html)
        self.assertNotIn("addEventListener('dragstart'", html)
        self.assertNotIn('id="ovi-dia-movimento"', html)
        self.assertNotIn('id="ovi-arm-agentes"', html)
        self.assertIn('id="ovi-hist-alt-body"', html)
        self.assertIn("Sem diário definido", html)
        self.assertIn("aplicarOviLote", html)

    def test_pagina_visitas_usa_filtros_modernos(self):
        client = _client_logado()
        resp = client.get("/visitas")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Visitas arboviroses", html)
        self.assertIn("visitas-filter-grid", html)
        self.assertIn("data-multi-picker", html)
        self.assertIn('id="v_agua_sanepar"', html)
        self.assertIn('id="v_observacoes"', html)
        self.assertIn('id="v_deposito"', html)
        self.assertIn('id="v_tratamento"', html)
        self.assertIn('id="v_laboratorio"', html)
        self.assertIn("Água Sanepar", html)
        self.assertIn("visitas-summary", html)
        self.assertIn("Mais filtros", html)
        self.assertIn("/static/js/visitas.js", html)

    def test_relatorio_agente_exibe_esporotricose_e_aviso_de_privacidade(self):
        client = _client_logado()
        resp = client.get("/relatorio-agente")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Central do agente", html)
        self.assertIn("Servidores", html)
        self.assertIn("agent-workspace", html)
        self.assertIn("agente_busca", html)
        self.assertIn("function escapeHtml", html)
        self.assertIn("Agente de Combate a Endemias", html)
        self.assertIn("Produção operacional integrada", html)
        self.assertIn("Relatório do setor", html)
        self.assertIn("gerarPDFSetor", html)
        self.assertIn("Produção de esporotricose", html)
        self.assertIn("médias agregadas dos demais agentes", html)
        self.assertNotIn("pend. SisPNCD", html)

    def test_pdf_relatorio_agente_inclui_esporotricose_sem_identificar_demais(self):
        client = _client_logado()
        agente = _agente_relatorio_teste()
        resp = client.get(
            "/relatorio-agente/pdf",
            query_string={"agente": agente, "d_ini": "2020-01-01", "d_fim": "2030-12-31"},
        )

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Agente de Combate a Endemias", html)
        self.assertIn("Produção operacional integrada", html)
        self.assertIn("Rua Bertolina Kendrik de Oliveira, 681", html)
        self.assertIn("break-inside:avoid", html)
        self.assertIn("page-break-inside:avoid", html)
        self.assertGreaterEqual(html.count('class="bloco-relatorio"'), 6)
        self.assertIn("Produção de esporotricose", html)
        self.assertIn("Comparação agregada de esporotricose", html)
        self.assertIn("Outros agentes não são listados nem identificados", html)
        self.assertIn("Tempo de visita: TBO, esporotricose e total", html)
        self.assertNotIn("Agente selecionado", html)
        for proibido in ("ObservaÃ", "calendÃ", "sÃ£", "â€", "�"):
            self.assertNotIn(proibido, html)
        self.assertNotIn("SisPNCD", html)

    def test_pdf_relatorio_setor_inclui_producao_geral_e_por_agente(self):
        client = _client_logado()
        resp = client.get(
            "/relatorio-agente/setor/pdf",
            query_string={"d_ini": "2020-01-01", "d_fim": "2030-12-31"},
        )

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Relatorio Geral do Setor", html)
        self.assertIn("Resumo do setor", html)
        self.assertIn("Producao por frente de trabalho", html)
        self.assertIn("Ranking geral de producao por agente", html)
        self.assertIn("Producao individual por agente", html)
        self.assertIn("Visitas vetoriais: situacao e tipo de trabalho", html)
        self.assertIn("Tempo das visitas", html)
        self.assertIn("Depositos e laboratorio", html)
        self.assertIn("Graficos analiticos", html)
        self.assertIn("chDuracaoTipo", html)
        self.assertIn("chAgentes", html)
        self.assertIn("Vetores", html)
        self.assertIn("Esporotricose", html)
        self.assertIn("BRI", html)
        self.assertIn("bloco-relatorio", html)
        self.assertNotIn("SisPNCD", html)

    def test_pdf_relatorio_setor_nao_exibe_publico_em_producao_por_frente(self):
        detalhe = relatorio_agente_bp._detalhe_atividade({
            "codigo": "ACOES_SETOR",
            "extras": {"educativas": 2, "limpezas": 1, "publico": 90},
        })

        self.assertEqual(detalhe, "2 educativas, 1 limpezas")
        self.assertNotIn("publico", detalhe.lower())
        self.assertNotIn("público", detalhe.lower())

    def test_relatorios_somam_tratamentos_de_depositos_e_tabela_tratamentos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Tratamento",))
                id_agente = conn.execute(
                    "SELECT id_agente FROM agentes WHERE nome=?",
                    ("Agente Tratamento",),
                ).fetchone()[0]
                conn.executemany(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, visita, processado_em)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    [
                        ("visita-tbo", "uuid-tbo", "TBO", "2026-04-10", "Normal", "2026-04-10T08:00:00"),
                        ("visita-pe", "uuid-pe", "PE", "2026-04-11", "Normal", "2026-04-11T08:00:00"),
                    ],
                )
                conn.executemany(
                    "INSERT INTO visita_agentes (id_visita, id_agente) VALUES (?, ?)",
                    [("visita-tbo", id_agente), ("visita-pe", id_agente)],
                )
                conn.execute(
                    """INSERT INTO depositos_inspecionados
                       (id_visita, tipo_deposito, inspecionado, eliminado, tratado)
                       VALUES (?, ?, ?, ?, ?)""",
                    ("visita-tbo", "A1", 4, 1, 2),
                )
                conn.execute(
                    """INSERT INTO tratamentos
                       (id_visita, tipo, quantidade_carga, qtd_depositos_tratados)
                       VALUES (?, ?, ?, ?)""",
                    ("visita-pe", "BTI", 1.5, 3),
                )
                conn.execute(
                    """INSERT INTO esporotricose_visitas
                       (id_visita, kobo_uuid, data, hora_inicio, hora_fim, visita, processado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        "esporo-tempo",
                        "uuid-esporo-tempo",
                        "2026-04-12",
                        "09:00",
                        "10:30",
                        "Normal",
                        "2026-04-12T10:30:00",
                    ),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.get(
                "/relatorio-agente/setor/pdf",
                query_string={"d_ini": "2026-04-01", "d_fim": "2026-04-30"},
            )
            self.assertEqual(resp.status_code, 200)
            html = resp.data.decode("utf-8")
            self.assertIn('<div class="kpi-val">5</div><div class="kpi-lbl">Dep. tratados</div>', html)
            self.assertIn("Esporotricose", html)
            self.assertIn("90.0 min", html)

            resp = client.get(
                "/api/relatorio-agente",
                query_string={
                    "agente": "Agente Tratamento",
                    "d_ini": "2026-04-01",
                    "d_fim": "2026-04-30",
                },
            )
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.get_json()["totais"]["tratados"], 5)

            resp = client.get("/api/dashboard?d_ini=2026-04-01&d_fim=2026-04-30")
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.get_json()["depositos"]["tratados"], 5)

    def test_pdf_relatorio_setor_inclui_ovitrampas_e_ignora_feriados(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Setor Ovi",))
                conn.commit()
                id_agente = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Agente Setor Ovi",)).fetchone()[0]
            finally:
                conn.close()

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            self.assertEqual(resp.status_code, 200)
            grupos = resp.get_json()["grupos"]

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-08-03",
                "movimento": "instalacao",
                "id_grupo": grupos[0]["id_grupo"],
                "ciclo": "Ciclo 3",
                "agentes": [id_agente],
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-08-10",
                "movimento": "retirada",
                "id_grupo": grupos[1]["id_grupo"],
                "ciclo": "Ciclo 3",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-08-15",
                "movimento": "feriado",
                "titulo": "Feriado ignorado",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get(
                "/relatorio-agente/setor/pdf",
                query_string={"d_ini": "2026-08-01", "d_fim": "2026-08-31"},
            )
            self.assertEqual(resp.status_code, 200)
            html = resp.data.decode("utf-8")
            self.assertIn("Ovitrampas", html)
            self.assertIn("Dias com movimento", html)
            self.assertIn("Agente Setor Ovi", html)
            self.assertIn("Retirada", html)
            self.assertNotIn("Feriado ignorado", html)

    def test_controle_pessoal_exibe_cadastro_e_historico(self):
        client = _client_logado("admin")
        resp = client.get("/admin/agentes")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Controle de Pessoal", html)
        self.assertIn("Servidores cadastrados", html)
        self.assertIn("Novo servidor", html)
        self.assertIn("Nome completo", html)
        self.assertIn("Nome usado no Kobo", html)
        self.assertIn("nome_completo", html)
        self.assertNotIn("Historico de trabalho", html)
        self.assertNotIn("Histórico de trabalho", html)
        self.assertNotIn("Ver histórico", html)
        self.assertNotIn(">Ver<", html)
        self.assertNotIn("data-agent-row", html)
        self.assertNotIn("/api/agentes/", html)

    def test_central_do_sistema_admin_responde_200(self):
        client = _client_logado("admin")
        resp = client.get("/admin/sistema")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Central do Sistema", html)
        self.assertIn("Diagnóstico do sistema", html)
        self.assertIn('id="btn-diag-completo"', html)
        self.assertIn("Saude do ambiente", html)
        self.assertIn("Modo de diario", html)
        self.assertIn("Espera por bloqueio", html)
        self.assertIn("Repeticoes por bloqueio", html)
        self.assertIn("Indices essenciais", html)
        self.assertIn("Backups gerenciados", html)
        self.assertIn("Backups completos", html)
        self.assertIn("/admin/sistema/backups/criar", html)
        self.assertIn("/admin/sistema/backups-completos/criar", html)
        self.assertIn('id="sidebarToggle"', html)
        self.assertIn("data-confirm=", html)
        self.assertNotIn("onclick=", html)
        self.assertNotIn("onsubmit=", html)

    def test_home_operacional_renderiza_blocos_principais(self):
        client = _client_logado()
        resp = client.get("/")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Painel operacional", html)
        self.assertIn("Ritmo dos últimos 14 dias", html)
        self.assertIn("Atalhos do sistema", html)

    def test_home_admin_exibe_blocos_operacionais_restritos(self):
        client = _client_logado("admin")
        resp = client.get("/")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Importações recentes", html)
        self.assertIn("Backups", html)
        self.assertIn("/admin/sistema", html)

    def test_processar_exibe_historico_de_importacoes(self):
        client = _client_logado("admin")
        resp = client.get("/processar")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Ultimas importacoes", html)
        self.assertIn("Importa&ccedil;&atilde;o Kobo", html)
        self.assertIn("Importa&ccedil;&atilde;o por Planilhas", html)
        self.assertIn("Consolidados em planilha", html)
        self.assertIn("Recurso secund&aacute;rio", html)
        self.assertIn('id="area-upload-shell"', html)
        self.assertIn("processar-planilhas-card", html)
        self.assertNotIn('id="btn-toggle-upload"', html)
        self.assertLess(html.index('id="card-kobo-api"'), html.index('id="area-upload-shell"'))
        self.assertIn("LARVAS_ Resultados de Laborat\u00f3rio", html)
        self.assertIn("data-gerar-consolidado", html)
        self.assertIn('id="processar-work-types"', html)
        self.assertIn('id="kobo-config-json"', html)
        self.assertIn('id="card-kobo-api"', html)
        self.assertIn('id="btn-kobo-pendentes"', html)
        self.assertIn('id="kobo-pendentes"', html)
        self.assertIn('id="kobo-preview-limite"', html)
        self.assertIn('value="5000"', html)
        self.assertNotIn('id="btn-kobo-previa"', html)
        self.assertNotIn('id="kobo-preview-apenas-novos"', html)
        self.assertNotIn('id="btn-kobo-lote"', html)
        self.assertNotIn('id="btn-kobo-importar"', html)
        self.assertIn("Importa&ccedil;&atilde;o via API", html)
        self.assertIn("Configurar conexão com Kobo", html)
        self.assertIn("<details", html)
        self.assertIn('data-kobo-asset="PE"', html)
        self.assertIn('data-kobo-asset="ESPOROTRICOSE"', html)
        self.assertIn('data-kobo-asset="BRI"', html)
        self.assertIn('data-kobo-asset="AMOSTRA_ANIMAIS"', html)
        self.assertIn('data-kobo-asset="RECOLHIMENTO"', html)
        self.assertIn("BRI_ Borrifamento residual", html)
        self.assertIn('src="/static/js/processar.js?v=', html)
        self.assertNotIn("configurarAcoesProcessamento()", html)
        self.assertNotIn("onclick=", html)
        self.assertNotIn("onchange=", html)
        proibidos = (
            "\u00c3\u00a7",
            "\u00c3\u00a3",
            "\u00c3\u00a1",
            "\u00c3\u00a9",
            "\u00c3\u00b3",
            "\u00c3\u00ba",
            "\u00c3\u00ad",
            "\u00c3\u00aa",
            "\u00c3\u2021",
            "\u00c3\u0192",
            "\u00e2\u0153",
            "\u00e2\u2020",
            "\u00e2\u20ac\u00a6",
            "\u00f0\u0178",
        )
        self.assertFalse(any(c in html for c in proibidos), html[:500])

        js_resp = client.get("/static/js/processar.js")
        js = js_resp.data.decode("utf-8")
        js_resp.close()
        self.assertFalse(any(c in js for c in proibidos), js[:500])
        self.assertIn("salvarKoboConfig", js)
        self.assertIn("/api/kobo/pendentes", js)
        self.assertIn("buscarKoboPendentes", js)
        self.assertIn("/api/kobo/previa", js)
        self.assertIn("koboDetalhesRegistro", js)
        self.assertIn("prepararKoboImportacao", js)
        self.assertIn("koboImportacaoEmAndamento", js)
        self.assertNotIn("buscarKoboLote", js)
        self.assertNotIn("KOBO_VISIT_TYPES", js)
        self.assertIn("/api/kobo/importar-formulario/iniciar", js)
        self.assertIn("Com atenção", js)
        self.assertIn("Resultados de laboratório", js)
        self.assertIn("s.resultados_novos", js)
        self.assertIn("resultado(s)", js)

    def test_assets_compartilhados_respondem_200(self):
        client = _client_logado()
        for rota in (
            "/static/css/app.css",
            "/static/js/app.js",
            "/static/js/processar.js",
            "/static/js/agenda.js",
            "/static/js/acoes_setor.js",
            "/static/vendor/chartjs/chart.umd.min.js",
            "/static/vendor/chartjs-plugin-datalabels/chartjs-plugin-datalabels.min.js",
            "/static/vendor/leaflet/leaflet.min.css",
            "/static/vendor/leaflet/leaflet.min.js",
        ):
            with self.subTest(rota=rota):
                resp = client.get(rota)
                try:
                    self.assertEqual(resp.status_code, 200)
                finally:
                    resp.close()

    def test_templates_principais_nao_dependem_de_cdn_externo(self):
        for rel in ("templates/base.html", "templates/login.html", "templates/mapa.html"):
            with self.subTest(rel=rel):
                texto = (ROOT / rel).read_text(encoding="utf-8")
                self.assertNotIn("cdnjs.cloudflare.com", texto)
                self.assertNotIn("fonts.googleapis.com", texto)

    def test_tema_claro_escuro_tem_contrato_explicito(self):
        css = (ROOT / "static" / "css" / "app.css").read_text(encoding="utf-8")
        js = (ROOT / "static" / "js" / "app.js").read_text(encoding="utf-8")

        self.assertIn('[data-theme="light"]', css)
        self.assertIn('[data-theme="dark"]', css)
        self.assertIn("localStorage.getItem('theme')", js)
        self.assertIn("document.documentElement.setAttribute('data-theme', currentTheme)", js)
        self.assertIn("currentTheme === 'dark' ? 'light' : 'dark'", js)

    def test_versao_aparece_no_base_e_no_login(self):
        client = _client_logado()
        home = client.get("/")
        login = endemias_app.app.test_client().get("/login")

        self.assertIn(version_core.APP_VERSION_LABEL.encode("utf-8"), home.data)
        self.assertIn(version_core.APP_VERSION_LABEL.encode("utf-8"), login.data)

    def test_iniciar_e_app_usam_versao_centralizada_sem_v3_antigo(self):
        iniciar = (ROOT / "iniciar.bat").read_text(encoding="utf-8")
        app_py = (ROOT / "app.py").read_text(encoding="utf-8")

        self.assertIn("from app_core.version import APP_VERSION_LABEL", iniciar)
        self.assertIn("echo  %APP_VERSION_LABEL%", iniciar)
        self.assertIn("version_core.APP_VERSION_LABEL", app_py)
        self.assertNotIn("Sistema de Gestao Integrado v3", iniciar)
        self.assertNotIn("Sistema de Gestao Integrado v3", app_py)

    def test_detalhe_notificacao_renderiza_icones_sem_escape(self):
        client = _client_logado()
        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            row = conn.execute(
                "SELECT id_foco FROM focos_positivos WHERE gera_notificacao=1 LIMIT 1"
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(row)

        resp = client.get(f"/notificacoes/foco/{row[0]}")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertNotIn("&lt;img", html)
        self.assertIn('src="/static/icons/salvar.svg"', html)

    def test_notificacoes_impressao_padrao_abre_html_e_docx_fica_separado(self):
        client = _client_logado()
        resp = client.get("/notificacoes")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn('id="notif-form" method="POST" action="/notificacoes/imprimir-html" target="_blank"', html)
        self.assertIn('id="btn-imprimir-html"', html)
        self.assertIn('id="btn-imprimir-docx"', html)
        self.assertIn('formaction="/notificacoes/imprimir"', html)
        self.assertIn("function imprimirNotificacaoHtml(id)", html)
        self.assertIn("onclick=\"imprimirNotificacaoHtml('", html)
        self.assertNotIn('action="/notificacoes/foco/', html)

    def test_detalhe_notificacao_imprimir_usa_html_sem_form_aninhado(self):
        client = _client_logado()
        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            row = conn.execute(
                "SELECT id_foco FROM focos_positivos WHERE gera_notificacao=1 LIMIT 1"
            ).fetchone()
        finally:
            conn.close()
        self.assertIsNotNone(row)

        resp = client.get(f"/notificacoes/foco/{row[0]}")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("function imprimirNotificacaoHtml()", html)
        self.assertIn(f"/notificacoes/foco/{row[0]}/imprimir-html", html)
        self.assertNotIn('action="/notificacoes/imprimir"', html)

    def test_agenda_eventos_expoem_cores_para_fullcalendar(self):
        client = _client_logado()
        resp = client.get("/api/agenda/eventos?start=2026-01-01&end=2026-12-31")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        eventos = resp.get_json()
        self.assertTrue(eventos)
        for evento in eventos:
            self.assertIn("backgroundColor", evento)
            self.assertIn("borderColor", evento)
            self.assertIn("textColor", evento)
        fontes_auto = {
            evento.get("extendedProps", {}).get("fonte")
            for evento in eventos
            if evento.get("extendedProps", {}).get("origem") == "auto"
        }
        self.assertIn("VETORES", fontes_auto)
        self.assertIn("BRI", fontes_auto)
        self.assertIn("ESPOROTRICOSE", fontes_auto)
        self.assertIn("RECOLHIMENTO", fontes_auto)
        self.assertIn("AMOSTRA_ANIMAIS", fontes_auto)
        for evento in eventos:
            props = evento.get("extendedProps", {})
            if props.get("fonte") in {"BRI", "ESPOROTRICOSE", "RECOLHIMENTO", "AMOSTRA_ANIMAIS"}:
                self.assertIn("agenda-auto-importado", evento.get("classNames", []))
                self.assertIn("fonteLabel", props)

    def test_agenda_nao_forca_cor_global_nos_eventos(self):
        client = _client_logado()
        resp = client.get("/agenda")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertNotIn("background-color: var(--fc-event-bg-color) !important", html)
        self.assertNotIn("style.setProperty('background-color', bg, 'important')", html)
        js_resp = client.get("/static/js/agenda.js")
        js = js_resp.data.decode("utf-8")
        js_resp.close()
        self.assertIn("style.setProperty('background-color', bg, 'important')", js)
        self.assertIn("eventTimeFormat", js)
        self.assertIn("minute: '2-digit'", js)
        self.assertIn("border-left", js)
        self.assertIn("fonteLabel", js)
        self.assertIn('id="btn-agenda-novo"', html)
        self.assertIn('id="agenda-config"', html)
        self.assertIn("Planejamento", html)
        self.assertIn('id="ev-descricao"', html)
        self.assertIn('rows="7"', html)
        self.assertIn("max-height:120px", js)
        self.assertIn("overflow-y:auto", js)
        self.assertIn("Amostra de animais (auto)", html)
        self.assertIn('src="/static/js/agenda.js"', html)
        self.assertIn('id="ev-recorrencia"', html)
        self.assertIn('id="ev-recorrencia-fim"', html)
        self.assertIn('id="btn-agenda-imprimir"', html)
        self.assertIn('id="agenda-busca"', html)
        self.assertIn('id="btn-agenda-limpar-busca"', html)
        self.assertIn('id="agenda-busca-ano"', html)
        self.assertIn('id="btn-agenda-buscar-ano"', html)
        self.assertIn('id="agenda-busca-status"', html)
        self.assertIn('id="agenda-resultados-ano"', html)
        self.assertIn("Férias", html)
        self.assertIn("Resultados no ano inteiro", html)
        self.assertIn("Buscar no ano inteiro", html)
        self.assertIn("weekNumbers: true", js)
        self.assertIn("weekNumberContent", js)
        self.assertIn("getUTCDay", js)
        self.assertIn("getUTCFullYear", js)
        self.assertIn("SEMANA ${week}", js)
        self.assertIn("fmtDataRange", js)
        self.assertIn("recorrencia_fim", js)
        self.assertIn("normalizarBusca", js)
        self.assertIn("filtrarEventosAgenda", js)
        self.assertIn("filtrarEventosPorTermos", js)
        self.assertIn("buscarAgendaNoAno", js)
        self.assertIn("renderResultadosAno", js)
        self.assertIn("agenda-year-result", js)
        self.assertIn("atualizarAnoBuscaAgenda", js)
        self.assertIn("imprimirAgendaMes", js)
        self.assertIn("/agenda/imprimir?ano=", js)
        self.assertIn("termos.every", js)
        self.assertIn("addEventListener('click', imprimirAgendaMes)", js)
        self.assertIn("addEventListener('input', atualizarBuscaAgenda)", js)
        self.assertIn("addEventListener('click', buscarAgendaNoAno)", js)
        self.assertIn("addEventListener('change', atualizarAnoBuscaAgenda)", js)
        self.assertIn("addEventListener('click', abrirResultadoAno)", js)
        self.assertNotIn("addEventListener('click', abrirModalNovo)", html)
        self.assertNotIn("onclick=", html)
        self.assertNotIn("onchange=", html)

    def test_agenda_evento_dia_inteiro_usa_fim_exclusivo_no_fullcalendar(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Férias - Vanessa teste",
                "tipo": "outro",
                "data_inicio": "2026-07-13",
                "data_fim": "2026-07-27",
                "dia_inteiro": True,
                "lembrete_min": 0,
                "recorrencia": "nenhuma",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/agenda/eventos?start=2026-07-01&end=2026-08-01")
            self.assertEqual(resp.status_code, 200)
            eventos = [
                evento for evento in resp.get_json()
                if evento.get("title") == "Férias - Vanessa teste"
            ]
            self.assertEqual(len(eventos), 1)
            evento = eventos[0]

            self.assertEqual(evento["start"], "2026-07-13")
            self.assertEqual(evento["end"], "2026-07-28")
            self.assertTrue(evento["allDay"])
            self.assertEqual(evento["extendedProps"]["data_fim"], "2026-07-27")

    def test_agenda_persiste_marcacao_de_atividade_externa(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Vistoria externa",
                "tipo": "campo",
                "data_inicio": "2026-07-16T08:00",
                "data_fim": "2026-07-16T12:00",
                "atividade_externa": True,
                "recorrencia": "nenhuma",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/agenda/eventos?start=2026-07-01&end=2026-08-01")
            evento = next(item for item in resp.get_json() if item["title"] == "Vistoria externa")
            self.assertTrue(evento["extendedProps"]["atividade_externa"])

    def test_agenda_exibe_aniversario_de_servidor_automaticamente(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            evento_manual = client.post("/api/agenda/eventos", json={
                "titulo": "Compromisso preservado",
                "tipo": "reuniao",
                "data_inicio": "2026-07-20T09:00",
                "data_fim": "2026-07-20T10:00",
                "dia_inteiro": False,
                "recorrencia": "nenhuma",
            })
            self.assertEqual(evento_manual.status_code, 201)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO agentes (nome, nome_completo, ativo, data_nascimento)
                       VALUES (?, ?, 1, ?)""",
                    ("Aniversariante Kobo", "Aniversariante Completo", "1990-07-16"),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.get("/api/agenda/eventos?start=2026-07-01&end=2026-08-01")
            self.assertEqual(resp.status_code, 200)
            evento = next(
                item for item in resp.get_json()
                if item["title"] == "Aniversário - Aniversariante Completo"
            )
            self.assertEqual(evento["start"], "2026-07-16")
            self.assertEqual(evento["extendedProps"]["fonte"], "ANIVERSARIO")
            self.assertFalse(evento["extendedProps"]["atividade_externa"])
            self.assertTrue(any(
                item["title"] == "Compromisso preservado"
                for item in resp.get_json()
            ))

    def test_agenda_expande_eventos_recorrentes_na_janela(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Reunião semanal teste",
                "tipo": "planejamento",
                "data_inicio": "2026-06-01",
                "data_fim": "2026-06-01",
                "dia_inteiro": True,
                "lembrete_min": 0,
                "recorrencia": "semanal",
                "recorrencia_fim": "2026-06-30",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/agenda/eventos?start=2026-06-01&end=2026-07-01")
            self.assertEqual(resp.status_code, 200)
            eventos = sorted(
                (
                    evento for evento in resp.get_json()
                    if evento.get("title") == "Reunião semanal teste"
                ),
                key=lambda evento: evento["start"],
            )

            self.assertEqual(
                [evento["start"] for evento in eventos],
                ["2026-06-01", "2026-06-08", "2026-06-15", "2026-06-22", "2026-06-29"],
            )
            self.assertTrue(all(evento["end"] > evento["start"] for evento in eventos))
            self.assertEqual({evento["extendedProps"]["recorrencia"] for evento in eventos}, {"semanal"})
            self.assertEqual({evento["extendedProps"]["recorrencia_fim"] for evento in eventos}, {"2026-06-30"})

    def test_agenda_aceita_tipo_ferias(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Férias - Vanessa",
                "tipo": "ferias",
                "data_inicio": "2026-07-13",
                "data_fim": "2026-07-27",
                "dia_inteiro": True,
                "lembrete_min": 0,
                "recorrencia": "nenhuma",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/agenda/eventos?start=2026-07-01&end=2026-08-01")
            self.assertEqual(resp.status_code, 200)
            eventos = [
                evento for evento in resp.get_json()
                if evento.get("title") == "Férias - Vanessa"
            ]
            self.assertEqual(len(eventos), 1)
            self.assertEqual(eventos[0]["extendedProps"]["tipo"], "ferias")
            self.assertEqual(eventos[0]["extendedProps"]["tipoLabel"], "Férias")

    def test_agenda_migra_eventos_com_ferias_no_titulo_para_tipo_ferias(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO agenda_eventos
                       (titulo, descricao, tipo, data_inicio, data_fim, dia_inteiro,
                        lembrete_min, cor, criado_por, criado_em, recorrencia, recorrencia_fim)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        "Ferias antigas",
                        None,
                        "outro",
                        "2026-01-12",
                        "2026-01-20",
                        1,
                        0,
                        "#64748b",
                        "teste",
                        "2026-01-01T08:00:00",
                        "nenhuma",
                        None,
                    ),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.get("/api/agenda/eventos?start=2026-01-01&end=2026-02-01")
            self.assertEqual(resp.status_code, 200)
            eventos = [
                evento for evento in resp.get_json()
                if evento.get("title") == "Ferias antigas"
            ]
            self.assertEqual(len(eventos), 1)
            self.assertEqual(eventos[0]["extendedProps"]["tipo"], "ferias")
            self.assertEqual(eventos[0]["extendedProps"]["tipoLabel"], "Férias")

    def test_agenda_impressao_mes_abre_html_com_detalhes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Reunião detalhada",
                "tipo": "reuniao",
                "data_inicio": "2026-08-12T08:30",
                "data_fim": "2026-08-12T10:00",
                "dia_inteiro": False,
                "lembrete_min": 0,
                "descricao": "Observação importante para impressão",
                "recorrencia": "nenhuma",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/agenda/imprimir?ano=2026&mes=8")
            self.assertEqual(resp.status_code, 200)
            html = resp.data.decode("utf-8")
            self.assertIn("Agenda - Agosto de 2026", html)
            self.assertIn("Reunião detalhada", html)
            self.assertIn("Observação importante para impressão", html)
            self.assertIn("08:30 - 10:00", html)
            self.assertIn("window.print()", html)

    def test_acoes_setor_renderiza_cadastro_manual(self):
        client = _client_logado("admin")
        resp = client.get("/acoes-setor")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Ações e Atendimentos do Setor", html)
        self.assertIn("Ação educativa / palestra", html)
        self.assertIn("Ação de limpeza / mutirão", html)
        self.assertIn("Vistoria / atendimento técnico", html)
        self.assertIn("Reunião / planejamento", html)
        self.assertIn('src="/static/icons/acoes_setor.svg"', html)
        self.assertIn('id="acao-agentes"', html)
        self.assertIn('name="acao-periodo"', html)
        self.assertIn('name="acao-tipo-atividade-realizada"', html)
        self.assertIn('name="acao-publico-alvo"', html)
        self.assertIn('name="acao-recurso-utilizado"', html)
        self.assertIn("Tipo de atividade realizada", html)
        self.assertIn("Público alvo", html)
        self.assertIn("Recurso utilizado", html)
        self.assertIn("acoes-educativa-only", html)
        self.assertIn("acoes-educativa-hidden", html)
        self.assertIn("function atualizarCamposEducativosAcoesSetor", html)
        self.assertIn('id="acao-agentes-busca"', html)
        self.assertIn('name="acao-agente"', html)
        self.assertNotIn('id="acao-agentes" multiple', html)
        self.assertIn('id="acao-anexo-selecionar"', html)
        self.assertIn('id="acao-anexos-baixar-todos"', html)
        self.assertIn('id="acao-anexos-lista"', html)
        self.assertIn('id="acao-dialog"', html)
        self.assertIn('id="acao-data-fim"', html)
        self.assertIn('id="acao-situacao"', html)
        self.assertIn('id="acao-caso"', html)
        self.assertIn('id="acao-resultados"', html)
        self.assertIn('id="acao-parceiros"', html)
        self.assertIn('id="acao-anexos-restritos"', html)
        self.assertIn('id="acao-nova"', html)
        self.assertIn('id="acoes-stat-total"', html)
        self.assertIn('id="acoes-stat-outros"', html)
        self.assertIn('id="acoes-filtro-situacao"', html)
        self.assertIn('id="acoes-filtro-caso"', html)
        self.assertIn('id="acoes-filtro-localidade"', html)
        self.assertIn('id="acoes-filtro-agente"', html)
        self.assertIn('id="acoes-data-inicio"', html)
        self.assertIn('id="acoes-data-fim"', html)
        self.assertIn('data-acoes-tab="anexos"', html)
        self.assertIn('data-acoes-tab="relatorio"', html)
        self.assertIn('id="acoes-panel-anexos"', html)
        self.assertIn('id="acoes-panel-relatorio"', html)
        self.assertIn('id="acoes-relatorio-data-inicio"', html)
        self.assertIn('id="acoes-relatorio-data-fim"', html)
        self.assertIn('id="acoes-relatorio-imagens"', html)
        self.assertIn('id="acoes-relatorio-atualizar"', html)
        self.assertIn('id="acoes-relatorio-gerar"', html)
        self.assertIn('id="acoes-anexos-galeria"', html)
        self.assertIn('id="acoes-anexos-tipo-arquivo"', html)
        self.assertIn('id="acoes-anexos-localidade"', html)
        self.assertIn('id="acoes-anexos-caso"', html)
        self.assertIn('id="acoes-anexos-agente"', html)
        self.assertIn('accept=".jpg,.jpeg,.png,.pdf,.doc,.docx,.xls,.xlsx,.txt,.mp4,.mov,.avi,.mkv,.webm,.m4v,.3gp"', html)
        self.assertIn('<option value="video">', html)
        self.assertIn('src="/static/js/acoes_setor.js"', html)
        js_resp = client.get("/static/js/acoes_setor.js")
        js = js_resp.data.decode("utf-8")
        js_resp.close()
        self.assertIn('data-acao-item', js)
        self.assertIn("let registroAberto = null", js)
        self.assertIn("function alternarRegistro", js)
        self.assertIn("function checkedValues", js)
        self.assertIn("function atualizarCamposEducativos", js)
        self.assertIn("classList.toggle('acoes-educativa-hidden'", js)
        self.assertIn("function anexoEhImagem", js)
        self.assertIn("function anexoEhVideo", js)
        self.assertIn("<video src=", js)
        self.assertIn("function carregarGaleriaAnexos", js)
        self.assertIn("acoes-anexo-preview", js)
        self.assertIn("/api/acoes-setor/anexos?", js)
        self.assertIn("/anexos/baixar-todos", js)
        self.assertIn("loading=\"lazy\"", js)
        self.assertIn("Informe o período do registro.", js)
        self.assertIn("function detalhesRegistroHtml", js)
        self.assertIn("function abrirFormularioNovo", js)
        self.assertIn("function atualizarResumo", js)
        self.assertIn("function limparFiltrosAnexos", js)
        self.assertIn("preencherForm(atualizada,false)", js)
        self.assertIn("function alternarRestricaoAnexo", js)
        self.assertIn("function paramsRelatorio", js)
        self.assertIn("function carregarPreviaRelatorio", js)
        self.assertIn("/acoes-setor/relatorio/pdf?", js)

    def test_acoes_setor_relatorio_filtra_detalhes_e_imagens(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            anexos_dir = Path(tmpdir) / "anexos"
            app_temp.config["ANEXOS_DIR"] = str(anexos_dir)

            incluida = client.post("/api/acoes-setor", json={
                "tipo": "educativa",
                "situacao": "realizada",
                "data": "2026-08-12",
                "periodo": "manha",
                "localidade": "Centro",
                "local": "Escola Municipal",
                "tema": "Ação Centro",
                "contexto": "Contexto técnico da atividade.",
                "resultados": "Providências registradas.",
                "publico_aproximado": 35,
            })
            self.assertEqual(incluida.status_code, 201)
            id_acao = incluida.get_json()["id_acao"]

            excluida = client.post("/api/acoes-setor", json={
                "tipo": "limpeza",
                "situacao": "realizada",
                "data": "2026-08-13",
                "periodo": "tarde",
                "localidade": "Outra localidade",
                "tema": "Ação fora do filtro",
            })
            self.assertEqual(excluida.status_code, 201)

            upload = client.post(
                f"/api/acoes-setor/{id_acao}/anexos",
                data={"arquivos": (io.BytesIO(b"imagem de teste"), "registro-fotografico.png")},
                content_type="multipart/form-data",
            )
            self.assertEqual(upload.status_code, 201)
            id_anexo = upload.get_json()["anexos"][0]["id_anexo"]

            resposta = client.get(
                "/acoes-setor/relatorio/pdf",
                query_string={
                    "localidade": "Centro",
                    "data_inicio": "2026-08-01",
                    "data_fim": "2026-08-31",
                    "imagens": "1",
                },
            )
            self.assertEqual(resposta.status_code, 200)
            html = resposta.data.decode("utf-8")
            self.assertIn("Relatório Técnico de Ações e Atendimentos", html)
            self.assertIn("Síntese dos registros", html)
            self.assertIn("Ação Centro", html)
            self.assertIn("Contexto técnico da atividade.", html)
            self.assertIn("Providências registradas.", html)
            self.assertNotIn("Ação fora do filtro", html)
            self.assertIn("Registro fotográfico", html)
            self.assertIn(
                f"/acoes-setor/anexos/{id_anexo}/download?inline=1",
                html,
            )

            sem_imagens = client.get(
                "/acoes-setor/relatorio/pdf",
                query_string={"localidade": "Centro", "imagens": "0"},
            )
            html_sem_imagens = sem_imagens.data.decode("utf-8")
            self.assertIn("Imagens: não incluídas", html_sem_imagens)
            self.assertNotIn("Registro fotográfico", html_sem_imagens)

    def test_acoes_setor_crud_e_busca_sem_acento(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("João Silva",))
                conn.execute("INSERT INTO localidades (nome) VALUES (?)", ("Centro",))
                conn.commit()
                id_agente = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("João Silva",)).fetchone()[0]
            finally:
                conn.close()

            sem_periodo = client.post("/api/acoes-setor", json={
                "tipo": "educativa",
                "data": "2026-08-12",
                "agentes": [id_agente],
            })
            self.assertEqual(sem_periodo.status_code, 400)
            self.assertIn("período", sem_periodo.get_json()["erro"])

            resp = client.post("/api/acoes-setor", json={
                "tipo": "educativa",
                "data": "2026-08-12",
                "periodo": "manha",
                "hora_inicio": "09:00",
                "hora_fim": "10:30",
                "agentes": [id_agente],
                "localidade": "Centro",
                "endereco": "Rua Principal, 100",
                "local": "Escola Municipal",
                "publico_aproximado": 45,
                "tipo_atividade_realizada": ["palestra", "conversa_educativa"],
                "publico_alvo": ["ensino_fundamental", "professores"],
                "recurso_utilizado": ["banner", "videos_imagens_midia_digital"],
                "tema": "Prevenção da dengue",
                "contexto": "Palestra para alunos",
                "observacoes": "Levar folders",
            })
            self.assertEqual(resp.status_code, 201)
            id_acao = resp.get_json()["id_acao"]

            resp = client.get("/api/acoes-setor?ano=2026&busca=prevencao")
            self.assertEqual(resp.status_code, 200)
            registros = resp.get_json()["registros"]
            self.assertEqual(len(registros), 1)
            self.assertEqual(registros[0]["id_acao"], id_acao)
            self.assertEqual(registros[0]["agentes"][0]["nome"], "João Silva")
            self.assertEqual(registros[0]["periodo"], "manha")
            self.assertIn("Palestra", registros[0]["tipo_atividade_realizada_labels"])
            self.assertIn("Ensino Fundamental", registros[0]["publico_alvo_labels"])
            self.assertIn("Banner", registros[0]["recurso_utilizado_labels"])
            self.assertEqual(registros[0]["total_anexos"], 0)

            filtrado = client.get(
                f"/api/acoes-setor?periodo=manha&localidade=Centro"
                f"&id_agente={id_agente}&data_inicio=2026-08-01&data_fim=2026-08-31"
            )
            self.assertEqual(filtrado.status_code, 200)
            self.assertEqual(filtrado.get_json()["total"], 1)
            fora_periodo = client.get(
                "/api/acoes-setor?data_inicio=2026-09-01&data_fim=2026-09-30"
            )
            self.assertEqual(fora_periodo.get_json()["total"], 0)

            resp = client.put(f"/api/acoes-setor/{id_acao}", json={
                "tipo": "limpeza",
                "data": "2026-08-13",
                "periodo": "tarde",
                "hora_inicio": "08:00",
                "agentes": [id_agente],
                "localidade": "Centro",
                "local": "Praça Central",
                "tipo_atividade_realizada": ["palestra"],
                "publico_alvo": ["professores"],
                "recurso_utilizado": ["banner"],
                "coordenadas": "-25.123, -49.123",
                "observacoes": "Mutirão finalizado",
            })
            self.assertEqual(resp.status_code, 200)

            resp = client.get(f"/api/acoes-setor/{id_acao}")
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.get_json()["tipo"], "limpeza")
            self.assertEqual(resp.get_json()["periodo_label"], "Tarde")
            self.assertEqual(resp.get_json()["tipo_atividade_realizada"], [])
            self.assertEqual(resp.get_json()["publico_alvo"], [])
            self.assertEqual(resp.get_json()["recurso_utilizado"], [])

    def test_acoes_setor_registra_atendimento_agrupado_e_intervalo(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post("/api/acoes-setor", json={
                "tipo": "vistoria",
                "situacao": "em_acompanhamento",
                "data": "2026-08-12",
                "data_fim": "2026-08-14",
                "periodo": "integral",
                "caso": "Acompanhamento sanitário - Centro",
                "localidade": "Centro",
                "local": "Imóvel acompanhado",
                "tema": "Vistoria sanitária",
                "contexto": "Atendimento decorrente de solicitação.",
                "resultados": "Orientações prestadas e retorno agendado.",
                "parceiros": "Meio Ambiente e UBS Centro",
            })
            self.assertEqual(resp.status_code, 201)
            id_acao = resp.get_json()["id_acao"]

            registro = client.get(f"/api/acoes-setor/{id_acao}").get_json()
            self.assertEqual(registro["tipo_label"], "Vistoria / atendimento técnico")
            self.assertEqual(registro["situacao_label"], "Em acompanhamento")
            self.assertEqual(registro["data_fim"], "2026-08-14")
            self.assertEqual(registro["periodo_label"], "Dia inteiro")
            self.assertEqual(registro["caso"], "Acompanhamento sanitário - Centro")
            self.assertIn("retorno", registro["resultados"])
            self.assertIn("Meio Ambiente", registro["parceiros"])

            filtrado = client.get(
                "/api/acoes-setor?tipo=vistoria&situacao=em_acompanhamento"
                "&caso=Acompanhamento%20sanit%C3%A1rio%20-%20Centro"
                "&data_inicio=2026-08-13&data_fim=2026-08-13"
            )
            self.assertEqual(filtrado.status_code, 200)
            self.assertEqual(filtrado.get_json()["total"], 1)

            data_invalida = client.post("/api/acoes-setor", json={
                "tipo": "reuniao",
                "data": "2026-08-14",
                "data_fim": "2026-08-12",
                "periodo": "manha",
            })
            self.assertEqual(data_invalida.status_code, 400)
            self.assertIn("data final", data_invalida.get_json()["erro"].lower())

    def test_acoes_setor_migra_tabela_antiga_sem_perder_vinculos(self):
        from blueprints import acoes_setor as acoes_setor_bp

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "legado.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.executescript("""
                    PRAGMA foreign_keys=ON;
                    CREATE TABLE agentes (
                        id_agente INTEGER PRIMARY KEY,
                        nome TEXT NOT NULL
                    );
                    CREATE TABLE acoes_setor (
                        id_acao INTEGER PRIMARY KEY AUTOINCREMENT,
                        tipo TEXT NOT NULL CHECK(tipo IN ('educativa','limpeza')),
                        data TEXT NOT NULL,
                        hora_inicio TEXT,
                        hora_fim TEXT,
                        localidade TEXT,
                        endereco TEXT,
                        local TEXT,
                        publico_aproximado INTEGER,
                        tema TEXT,
                        contexto TEXT,
                        coordenadas TEXT,
                        observacoes TEXT,
                        criado_por TEXT,
                        criado_em TEXT NOT NULL,
                        atualizado_em TEXT
                    );
                    CREATE TABLE acoes_setor_agentes (
                        id_acao INTEGER NOT NULL REFERENCES acoes_setor(id_acao) ON DELETE CASCADE,
                        id_agente INTEGER NOT NULL REFERENCES agentes(id_agente),
                        PRIMARY KEY (id_acao,id_agente)
                    );
                    CREATE TABLE acoes_setor_anexos (
                        id_anexo INTEGER PRIMARY KEY AUTOINCREMENT,
                        id_acao INTEGER NOT NULL REFERENCES acoes_setor(id_acao) ON DELETE CASCADE,
                        nome_original TEXT NOT NULL,
                        nome_arquivo TEXT NOT NULL,
                        caminho_rel TEXT NOT NULL,
                        mime_type TEXT,
                        tamanho INTEGER NOT NULL DEFAULT 0,
                        criado_por TEXT,
                        criado_em TEXT NOT NULL
                    );
                    INSERT INTO agentes VALUES (1,'João');
                    INSERT INTO acoes_setor
                        (id_acao,tipo,data,tema,criado_por,criado_em)
                    VALUES (7,'educativa','2025-08-15','Planejamento','João','2026-07-01T10:00:00');
                    INSERT INTO acoes_setor_agentes VALUES (7,1);
                    INSERT INTO acoes_setor_anexos
                        (id_anexo,id_acao,nome_original,nome_arquivo,caminho_rel,tamanho,criado_em)
                    VALUES (3,7,'ata.docx','ata.docx','acoes/ata.docx',10,'2026-07-01T10:00:00');
                """)
                acoes_setor_bp.ensure_schema(conn)

                registro = conn.execute(
                    "SELECT * FROM acoes_setor WHERE id_acao=7"
                ).fetchone()
                self.assertEqual(registro["tema"], "Planejamento")
                self.assertEqual(registro["situacao"], "realizada")
                self.assertEqual(
                    conn.execute(
                        "SELECT COUNT(*) FROM acoes_setor_agentes WHERE id_acao=7"
                    ).fetchone()[0],
                    1,
                )
                anexo = conn.execute(
                    "SELECT * FROM acoes_setor_anexos WHERE id_anexo=3"
                ).fetchone()
                self.assertEqual(anexo["restrito"], 0)
                self.assertEqual(conn.execute("PRAGMA foreign_key_check").fetchall(), [])
                conn.execute(
                    """INSERT INTO acoes_setor
                       (tipo,situacao,data,periodo,criado_em)
                       VALUES ('vistoria','realizada','2026-07-01','manha','2026-07-01T10:00:00')"""
                )
                conn.commit()
            finally:
                conn.close()

    def test_acoes_setor_anexos_ficam_em_pasta_e_podem_ser_excluidos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            anexos_dir = Path(tmpdir) / "anexos"
            app_temp.config["ANEXOS_DIR"] = str(anexos_dir)

            resp = client.post("/api/acoes-setor", json={
                "tipo": "limpeza",
                "data": "2026-08-12",
                "periodo": "manha",
                "localidade": "Centro",
                "local": "Terreno particular",
                "observacoes": "Documentação anexada posteriormente",
            })
            self.assertEqual(resp.status_code, 201)
            id_acao = resp.get_json()["id_acao"]

            resp = client.post(
                f"/api/acoes-setor/{id_acao}/anexos",
                data={"arquivos": (io.BytesIO(b"conteudo do documento"), "notificacao.pdf")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp.status_code, 201)
            anexos = resp.get_json()["anexos"]
            self.assertEqual(len(anexos), 1)
            self.assertEqual(anexos[0]["nome_original"], "notificacao.pdf")
            self.assertEqual(anexos[0]["mime_type"], "application/pdf")
            id_anexo = anexos[0]["id_anexo"]

            resp_img = client.post(
                f"/api/acoes-setor/{id_acao}/anexos",
                data={"arquivos": (io.BytesIO(b"imagem fake"), "foto.png")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp_img.status_code, 201)
            self.assertEqual(len(resp_img.get_json()["anexos"]), 2)

            resp_video = client.post(
                f"/api/acoes-setor/{id_acao}/anexos",
                data={"arquivos": (io.BytesIO(b"video fake"), "registro.mp4")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp_video.status_code, 201)
            anexos_video = resp_video.get_json()["anexos"]
            self.assertEqual(len(anexos_video), 3)
            self.assertTrue(any(a["nome_original"] == "registro.mp4" for a in anexos_video))
            lista_acoes = client.get("/api/acoes-setor").get_json()["registros"]
            self.assertEqual(lista_acoes[0]["total_anexos"], 3)

            conn = sqlite3.connect(app_temp.config["DB_PATH"])
            try:
                row = conn.execute(
                    "SELECT caminho_rel FROM acoes_setor_anexos WHERE id_anexo=?",
                    (id_anexo,),
                ).fetchone()
            finally:
                conn.close()
            self.assertIsNotNone(row)
            caminho = anexos_dir / row[0]
            self.assertTrue(caminho.exists())
            self.assertIn(str(id_acao).zfill(6), str(caminho))

            resp = client.get(f"/acoes-setor/anexos/{id_anexo}/download?inline=1")
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.data, b"conteudo do documento")
            resp.close()

            galeria = client.get("/api/acoes-setor/anexos?tipo_arquivo=imagem&busca=foto")
            self.assertEqual(galeria.status_code, 200)
            dados_galeria = galeria.get_json()
            self.assertEqual(dados_galeria["total"], 1)
            self.assertEqual(dados_galeria["anexos"][0]["nome_original"], "foto.png")
            self.assertEqual(dados_galeria["anexos"][0]["acao_tipo"], "limpeza")

            galeria_filtrada = client.get(
                "/api/acoes-setor/anexos?localidade=Centro"
                "&data_inicio=2026-08-01&data_fim=2026-08-31"
            )
            self.assertEqual(galeria_filtrada.status_code, 200)
            self.assertEqual(galeria_filtrada.get_json()["total"], 3)
            galeria_fora_periodo = client.get(
                "/api/acoes-setor/anexos?data_inicio=2026-09-01&data_fim=2026-09-30"
            )
            self.assertEqual(galeria_fora_periodo.get_json()["total"], 0)

            galeria_video = client.get("/api/acoes-setor/anexos?tipo_arquivo=video&busca=registro")
            self.assertEqual(galeria_video.status_code, 200)
            dados_video = galeria_video.get_json()
            self.assertEqual(dados_video["total"], 1)
            self.assertEqual(dados_video["anexos"][0]["nome_original"], "registro.mp4")
            self.assertTrue(dados_video["anexos"][0]["eh_previa"])

            zip_resp = client.get(f"/api/acoes-setor/{id_acao}/anexos/baixar-todos")
            self.assertEqual(zip_resp.status_code, 200)
            self.assertIn("application/zip", zip_resp.content_type)
            self.assertTrue(zip_resp.data.startswith(b"PK"))
            with zipfile.ZipFile(io.BytesIO(zip_resp.data)) as zf:
                self.assertEqual(set(zf.namelist()), {"notificacao.pdf", "foto.png", "registro.mp4"})
            zip_resp.close()

            resp = client.delete(f"/api/acoes-setor/anexos/{id_anexo}")
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(caminho.exists())

    def test_acoes_setor_anexo_restrito_so_aparece_para_admin(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, admin_client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["ANEXOS_DIR"] = str(Path(tmpdir) / "anexos")

            criado = admin_client.post("/api/acoes-setor", json={
                "tipo": "vistoria",
                "data": "2026-08-12",
                "periodo": "tarde",
                "caso": "Caso com documento sensível",
            })
            self.assertEqual(criado.status_code, 201)
            id_acao = criado.get_json()["id_acao"]

            enviado = admin_client.post(
                f"/api/acoes-setor/{id_acao}/anexos",
                data={
                    "restrito": "1",
                    "arquivos": (io.BytesIO(b"documento restrito"), "documento.pdf"),
                },
                content_type="multipart/form-data",
            )
            self.assertEqual(enviado.status_code, 201)
            anexo = enviado.get_json()["anexos"][0]
            self.assertTrue(anexo["restrito"])
            id_anexo = anexo["id_anexo"]

            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.execute(
                    """INSERT INTO usuarios
                       (usuario,nome,senha_hash,nivel,ativo,criado_em)
                       VALUES ('operador_teste','Operador Teste','x','operador',1,?)""",
                    (datetime.now().isoformat(),),
                )
                conn.commit()
                operador = dict(conn.execute(
                    """SELECT id_usuario,nome,nivel FROM usuarios
                        WHERE usuario='operador_teste'"""
                ).fetchone())
            finally:
                conn.close()

            operador_client = app_temp.test_client()
            _login_client_com_usuario(operador_client, operador)
            self.assertEqual(
                operador_client.get(
                    f"/api/acoes-setor/{id_acao}/anexos"
                ).get_json()["anexos"],
                [],
            )
            self.assertEqual(
                operador_client.get("/api/acoes-setor/anexos").get_json()["total"],
                0,
            )
            self.assertEqual(
                operador_client.get(
                    f"/acoes-setor/anexos/{id_anexo}/download"
                ).status_code,
                403,
            )
            self.assertNotIn(
                "documento.pdf",
                operador_client.get(
                    "/acoes-setor/relatorio/pdf"
                ).data.decode("utf-8"),
            )
            self.assertIn(
                "documento.pdf",
                admin_client.get(
                    "/acoes-setor/relatorio/pdf"
                ).data.decode("utf-8"),
            )
            self.assertEqual(
                operador_client.delete(f"/api/acoes-setor/{id_acao}").status_code,
                403,
            )

            liberado = admin_client.put(
                f"/api/acoes-setor/anexos/{id_anexo}",
                json={"restrito": False},
            )
            self.assertEqual(liberado.status_code, 200)
            self.assertFalse(liberado.get_json()["anexo"]["restrito"])
            self.assertEqual(
                len(operador_client.get(
                    f"/api/acoes-setor/{id_acao}/anexos"
                ).get_json()["anexos"]),
                1,
            )

    def test_acoes_setor_aparecem_na_agenda_automatica(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Maria Souza",))
                conn.commit()
                id_agente = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Maria Souza",)).fetchone()[0]
            finally:
                conn.close()

            resp = client.post("/api/acoes-setor", json={
                "tipo": "educativa",
                "data": "2026-08-12",
                "periodo": "manha",
                "hora_inicio": "09:00",
                "hora_fim": "10:00",
                "agentes": [id_agente],
                "localidade": "Centro",
                "local": "Escola Municipal",
                "publico_aproximado": 30,
                "tipo_atividade_realizada": ["oficina"],
                "publico_alvo": ["comunidade_escolar"],
                "recurso_utilizado": ["maquete"],
                "tema": "Saúde ambiental",
                "contexto": "Atividade com estudantes",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/agenda/eventos?start=2026-08-01&end=2026-09-01")
            self.assertEqual(resp.status_code, 200)
            eventos = [
                evento for evento in resp.get_json()
                if evento.get("extendedProps", {}).get("fonte") == "ACAO_EDUCATIVA"
            ]
            self.assertEqual(len(eventos), 1)
            self.assertIn("Saúde ambiental", eventos[0]["title"])
            props = eventos[0]["extendedProps"]
            self.assertEqual(props["origem"], "auto")
            self.assertEqual(props["fonteLabel"], "Ação educativa")
            self.assertIn("Período: Manhã", props["resumo"])
            self.assertIn("Público aproximado: 30", props["resumo"])
            self.assertIn("Tipo de atividade: Oficina", props["resumo"])
            self.assertIn("Público alvo: Comunidade Escolar", props["resumo"])
            self.assertIn("Recurso utilizado: Maquete", props["resumo"])
            self.assertEqual(props["agentes"], "Maria Souza")

    def test_agenda_rejeita_evento_com_fim_antes_do_inicio(self):
        client = _client_logado("admin")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp = client.post("/api/agenda/eventos", json={
                "titulo": "Evento invalido",
                "tipo": "outro",
                "data_inicio": "2026-06-03T14:00",
                "data_fim": "2026-06-03T08:00",
                "dia_inteiro": False,
                "lembrete_min": 0,
            })
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        self.assertEqual(resp.status_code, 400)
        self.assertIn("Data fim", resp.get_json().get("erro", ""))


class MainApisSmokeTests(unittest.TestCase):
    def test_ovitrampas_importa_csv_e_evita_duplicidade(self):
        csv_text = (
            "Ovitrampa ID;Estado;Município;Distrito;Rua;Número;Complemento;Localização;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "Observação;Lat_lng;Quarteirão;Data da instalação;Data de coleta\n"
            "1;Paraná;Almirante Tamandaré;TAMBOARA;Rua A;10;Escola;Parede;"
            "-25,1;-49,2;2026;21;2026-06-01 17:26:10;53;Vanessa;;-25.1,-49.2;"
            "1269;2026-05-25;2026-05-29\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            csv_path = Path(tmpdir) / "3918-2026-21.csv"
            csv_path.write_text(csv_text, encoding="utf-8")

            primeiro = ovitrampas_core.importar_csv(db_path, csv_path)
            segundo = ovitrampas_core.importar_csv(db_path, csv_path)
            resumo = ovitrampas_core.resumo(db_path, {"ano": "2026", "semana": "21"})
            lista = ovitrampas_core.listar(db_path, {"busca": "Escola"})

            self.assertEqual(primeiro["inseridos"], 1)
            self.assertEqual(segundo["duplicados"], 1)
            self.assertEqual(resumo["totais"]["leituras"], 1)
            self.assertEqual(resumo["totais"]["ovos"], 53)
            self.assertEqual(lista["total"], 1)
            self.assertEqual(lista["registros"][0]["distrito"], "Tamboara")

    def test_ovitrampas_importa_cadastro_armadilhas_e_historico(self):
        leitura_csv = (
            "Ovitrampa ID;Estado;MunicÃ­pio;Distrito;Rua;NÃºmero;Complemento;LocalizaÃ§Ã£o;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "ObservaÃ§Ã£o;Lat_lng;QuarteirÃ£o;Data da instalaÃ§Ã£o;Data de coleta\n"
            "1;ParanÃ¡;Almirante TamandarÃ©;TAMBOARA;Rua A;10;Escola;Parede;"
            "-25,1;-49,2;2026;21;2026-06-01 17:26:10;53;Vanessa;;-25.1,-49.2;"
            "1269;2026-05-25;2026-05-29\n"
        )
        cadastro_csv = (
            "ID;Rua;Número do logradouro;Complemento;Bairro;Localização da ovitrampa;"
            "Setor/Distrito da ovitrampa;Responsável;Quarteirão;Latitude;Longitude\n"
            "1;Rua A;10;Escola;TAMBOARA;Parede;lamenha;Joel;1269;-25,1;-49,2\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            leitura_path = Path(tmpdir) / "3918-2026-21.csv"
            cadastro_path = Path(tmpdir) / "almirante-tamandare-pr-ovitraps.csv"
            cadastro_atualizado_path = Path(tmpdir) / "almirante-tamandare-pr-ovitraps-2.csv"
            leitura_path.write_text(leitura_csv, encoding="utf-8")
            cadastro_path.write_text(cadastro_csv, encoding="utf-8")
            cadastro_atualizado_path.write_text(
                cadastro_csv
                .replace("Rua A;10;Escola", "Rua B;20;Garagem")
                .replace("Parede;lamenha", "Portao lateral;lamenha")
                .replace("Joel;1269", "Vanessa;1270"),
                encoding="utf-8",
            )

            ovitrampas_core.importar_csv(db_path, leitura_path)
            primeiro = ovitrampas_core.importar_armadilhas_csv(db_path, cadastro_path)
            segundo = ovitrampas_core.importar_armadilhas_csv(db_path, cadastro_atualizado_path)
            armadilhas = ovitrampas_core.listar_armadilhas(db_path, {"busca": "Vanessa"})
            historico = ovitrampas_core.historico_armadilha(db_path, "1")

            self.assertEqual(primeiro["inseridos"], 1)
            self.assertEqual(segundo["atualizados"], 1)
            self.assertEqual(armadilhas["total"], 1)
            self.assertEqual(armadilhas["registros"][0]["leituras"], 1)
            self.assertEqual(armadilhas["registros"][0]["localidade"], "Lamenha")
            self.assertEqual(historico["armadilha"]["rua"], "Rua B")
            self.assertEqual(historico["armadilha"]["numero"], "20")
            self.assertEqual(historico["armadilha"]["quarteirao"], "1270")
            self.assertEqual(historico["armadilha"]["localizacao"], "Portao lateral")
            self.assertEqual(historico["armadilha"]["responsavel"], "Vanessa")
            self.assertEqual(historico["leituras"][0]["ovos"], 53)
            campos = {item["campo"] for item in historico["alteracoes"]}
            self.assertTrue({"rua", "numero", "complemento", "localizacao", "responsavel", "quarteirao"}.issubset(campos))

    def test_ovitrampas_importa_diarios_xlsx_com_ordem_e_realocar(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            cadastro_csv = (
                "ID;Rua;NÃºmero do logradouro;Complemento;Bairro;LocalizaÃ§Ã£o da ovitrampa;"
                "Setor/Distrito da ovitrampa;ResponsÃ¡vel;QuarteirÃ£o;Latitude;Longitude\n"
                "1;Rua A;10;Escola;TAMBOARA;Parede;TAMBOARA;Joel;1269;-25,1;-49,2\n"
                "2;REALOCAR;20;Loja;;Canto;REALOCAR;Vanessa;1270;-25,3;-49,4\n"
                "3;Rua C;30;Mercado;;Fundo;TAMBOARA;Maria;1271;-25,5;-49,6\n"
            )
            cadastro_path = Path(tmpdir) / "cadastro.csv"
            cadastro_path.write_text(cadastro_csv, encoding="utf-8")
            ovitrampas_core.importar_armadilhas_csv(db_path, cadastro_path, motivo="Teste")

            xlsx_path = Path(tmpdir) / "diarios.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tamboara e Graziela"
            ws["A21"] = "ID"
            ws["B21"] = "Endereço"
            ws["F21"] = "Q."
            ws["G21"] = "Localização"
            ws["A22"] = 1
            ws["B22"] = "Rua A, 10 - Escola"
            ws["F22"] = "1269"
            ws["G22"] = "Parede"
            ws["B23"] = "(41) 99999-0000"
            ws["D23"] = "Joel"
            ws["A24"] = "1-A"
            ws["B24"] = "Rua B, 20 - Comercio"
            ws["F24"] = "1270"
            ws["G24"] = "Arvore"
            ws["B25"] = "(41) 98888-0000"
            ws["D25"] = "Ana"
            ws["A26"] = "Ocorrências"
            wb.save(xlsx_path)

            result = ovitrampas_core.importar_diarios_xlsx(db_path, xlsx_path, usuario="Teste")
            dados = ovitrampas_core.diarios_dados(db_path)
            detalhe = ovitrampas_core.diario_detalhe(db_path, dados["diarios"][0]["id_diario"])

            self.assertEqual(result["diarios"], 1)
            self.assertEqual(result["vinculos"], 2)
            self.assertEqual(dados["totais"]["realocar"], 1)
            self.assertEqual(dados["totais"]["sem_diario"], 1)
            self.assertEqual([r["ovitrampa_id"] for r in detalhe["registros"]], ["1", "1-A"])
            self.assertEqual(detalhe["registros"][0]["telefone_responsavel"], "(41) 99999-0000")

            reordenado = ovitrampas_core.reordenar_armadilhas_diario(
                db_path,
                dados["diarios"][0]["id_diario"],
                ["1-A", "1"],
            )
            self.assertEqual([r["ovitrampa_id"] for r in reordenado["registros"]], ["1-A", "1"])

            impressao = ovitrampas_core.diario_impressao(db_path, dados["diarios"][0]["id_diario"])
            self.assertEqual([r["ovitrampa_id"] for r in impressao["registros"]], ["1-A", "1"])
            self.assertNotIn("movimento", impressao)

    def test_api_ovitrampas_importa_csv(self):
        csv_bytes = (
            "Ovitrampa ID;Estado;Município;Distrito;Rua;Número;Complemento;Localização;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "Observação;Lat_lng;Quarteirão;Data da instalação;Data de coleta\n"
            "2;Paraná;Almirante Tamandaré;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;22;2026-06-08 10:00:00;0;Vanessa;;-25.3,-49.4;"
            "1300;2026-06-01;2026-06-05\n"
        ).encode("utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post(
                "/api/ovitrampas/importar",
                data={"arquivos": (io.BytesIO(csv_bytes), "3918-2026-22.csv")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data["inseridos"], 1)

            resp = client.get("/api/ovitrampas/listar?busca=Rua+B")
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.get_json()["total"], 1)

    def test_api_ovitrampas_atualiza_laboratorista_e_data_leitura(self):
        csv_bytes = (
            "Ovitrampa ID;Estado;MunicÃ­pio;Distrito;Rua;NÃºmero;Complemento;LocalizaÃ§Ã£o;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "ObservaÃ§Ã£o;Lat_lng;QuarteirÃ£o;Data da instalaÃ§Ã£o;Data de coleta\n"
            "2;ParanÃ¡;Almirante TamandarÃ©;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;22;2026-06-08 10:00:00;0;Vanessa;;-25.3,-49.4;"
            "1300;2026-06-01;2026-06-05\n"
        ).encode("utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post(
                "/api/ovitrampas/importar",
                data={"arquivos": (io.BytesIO(csv_bytes), "3918-2026-22.csv")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp.status_code, 200)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Laboratorista Teste",))
                conn.commit()
                id_leitura = conn.execute("SELECT id_leitura FROM ovitrampas_leituras").fetchone()[0]
                agente = conn.execute("SELECT id_agente, nome FROM agentes ORDER BY id_agente LIMIT 1").fetchone()
            finally:
                conn.close()

            resp = client.put(
                f"/api/ovitrampas/leituras/{id_leitura}",
                json={"id_laboratorista": agente[0], "data_leitura": "2026-06-09"},
            )

            self.assertEqual(resp.status_code, 200)
            registro = resp.get_json()["registro"]
            self.assertEqual(registro["id_laboratorista"], agente[0])
            self.assertEqual(registro["laboratorista"], agente[1])
            self.assertEqual(registro["data_leitura"], "2026-06-09")

    def test_api_ovitrampas_atualiza_lote_somente_campos_vazios(self):
        csv_bytes = (
            "Ovitrampa ID;Estado;MunicÃƒÂ­pio;Distrito;Rua;NÃƒÂºmero;Complemento;LocalizaÃƒÂ§ÃƒÂ£o;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "ObservaÃƒÂ§ÃƒÂ£o;Lat_lng;QuarteirÃƒÂ£o;Data da instalaÃƒÂ§ÃƒÂ£o;Data de coleta\n"
            "2;ParanÃƒÂ¡;Almirante TamandarÃƒÂ©;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;22;2026-06-08 10:00:00;0;Vanessa;;-25.3,-49.4;"
            "1300;2026-06-01;2026-06-05\n"
            "3;ParanÃƒÂ¡;Almirante TamandarÃƒÂ©;GRAZIELA;Rua C;30;Casa;Muro;"
            "-25,4;-49,5;2026;22;2026-06-08 10:10:00;12;Vanessa;;-25.4,-49.5;"
            "1301;2026-06-01;2026-06-05\n"
        ).encode("utf-8")
        with tempfile.TemporaryDirectory() as tmpdir:
            _app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            resp = client.post(
                "/api/ovitrampas/importar",
                data={"arquivos": (io.BytesIO(csv_bytes), "3918-2026-22.csv")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp.status_code, 200)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Lab Antigo",))
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Lab Novo",))
                agentes = conn.execute("SELECT id_agente, nome FROM agentes ORDER BY id_agente").fetchall()
                antigo = agentes[-2][0]
                novo = agentes[-1][0]
                leitura_preenchida = conn.execute(
                    "SELECT id_leitura FROM ovitrampas_leituras WHERE ovitrampa_id='2'"
                ).fetchone()[0]
                conn.execute(
                    "UPDATE ovitrampas_leituras SET id_laboratorista=?, data_leitura=? WHERE id_leitura=?",
                    (antigo, "2026-06-09", leitura_preenchida),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.put(
                "/api/ovitrampas/leituras/lote?ano=2026&semana=22&distrito=Graziela",
                json={"id_laboratorista": novo, "data_leitura": "2026-06-10", "somente_vazios": True},
            )

            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp.get_json()["atualizados"], 1)
            conn = sqlite3.connect(db_path)
            try:
                rows = conn.execute(
                    "SELECT ovitrampa_id, id_laboratorista, data_leitura FROM ovitrampas_leituras ORDER BY ovitrampa_id"
                ).fetchall()
            finally:
                conn.close()
            self.assertEqual(rows[0][1], antigo)
            self.assertEqual(rows[0][2], "2026-06-09")
            self.assertEqual(rows[1][1], novo)
            self.assertEqual(rows[1][2], "2026-06-10")

    def test_ovitrampas_monitoramento_retorna_analises_operacionais(self):
        leitura_csv = (
            "Ovitrampa ID;Estado;MunicÃ­pio;Distrito;Rua;NÃºmero;Complemento;LocalizaÃ§Ã£o;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "ObservaÃ§Ã£o;Lat_lng;QuarteirÃ£o;Data da instalaÃ§Ã£o;Data de coleta;Ocorrencia\n"
            "1;ParanÃ¡;Almirante TamandarÃ©;GRAZIELA;Rua A;10;Casa;Parede;"
            "-25,1;-49,2;2026;21;2026-06-01 17:26:10;53;Vanessa;;-25.1,-49.2;"
            "1269;2026-05-25;2026-05-29;\n"
            "2;ParanÃ¡;Almirante TamandarÃ©;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;21;2026-06-01 17:30:10;0;Vanessa;;-25.3,-49.4;"
            "1270;2026-05-25;2026-05-29;5\n"
            "2;ParanÃ¡;Almirante TamandarÃ©;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;22;2026-06-08 17:30:10;12;Vanessa;;-25.3,-49.4;"
            "1270;2026-06-01;2026-06-05;2\n"
        )
        cadastro_csv = (
            "ID;Rua;NÃºmero do logradouro;Complemento;Bairro;LocalizaÃ§Ã£o da ovitrampa;"
            "Setor/Distrito da ovitrampa;ResponsÃ¡vel;QuarteirÃ£o;Latitude;Longitude\n"
            "1;Rua A;10;Casa;;Parede;GRAZIELA;Joel;1269;-25,1;-49,2\n"
            "2;REALOCAR;20;Loja;;Canto;GRAZIELA;Vanessa;1270;-25,3;-49,4\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            leitura_path = Path(tmpdir) / "3918-2026-21.csv"
            cadastro_path = Path(tmpdir) / "almirante-tamandare-pr-ovitraps.csv"
            leitura_path.write_text(leitura_csv, encoding="utf-8")
            cadastro_path.write_text(cadastro_csv, encoding="utf-8")
            ovitrampas_core.importar_csv(db_path, leitura_path)
            ovitrampas_core.importar_armadilhas_csv(db_path, cadastro_path)

            dados = ovitrampas_core.monitoramento(db_path, {"ano": "2026", "semana_ini": "21", "semana_fim": "22"})

        self.assertEqual(dados["totais"]["leituras"], 3)
        self.assertEqual(dados["totais"]["armadilhas_positivas"], 2)
        self.assertEqual(dados["totais"]["ocorrencias"], 2)
        self.assertEqual(dados["realocar"]["total"], 1)
        self.assertEqual(dados["ranking_positivas"][0]["ovitrampa_id"], "1")
        ocorrencias = {row["codigo"]: row for row in dados["ocorrencias"]}
        self.assertEqual(ocorrencias[5]["descricao"], "Armadilha seca")
        self.assertEqual(ocorrencias[5]["armadilhas_destaque"][0]["ovitrampa_id"], "2")

    def test_ovitrampas_importa_historico_ocorrencias_conta_ovos(self):
        leitura_csv = (
            "Ovitrampa ID;Estado;MunicÃƒÂ­pio;Distrito;Rua;NÃƒÂºmero;Complemento;LocalizaÃƒÂ§ÃƒÂ£o;"
            "Latitude;Longitude;Ano;Semana;Data do envio da contagem;Ovos;Quem enviou;"
            "ObservaÃƒÂ§ÃƒÂ£o;Lat_lng;QuarteirÃƒÂ£o;Data da instalaÃƒÂ§ÃƒÂ£o;Data de coleta\n"
            "2;ParanÃƒÂ¡;Almirante TamandarÃƒÂ©;GRAZIELA;Rua B;20;Loja;Canto;"
            "-25,3;-49,4;2026;21;2026-06-01 17:30:10;0;Vanessa;;-25.3,-49.4;"
            "1270;2026-05-25;2026-05-29\n"
            "3;ParanÃƒÂ¡;Almirante TamandarÃƒÂ©;GRAZIELA;Rua C;30;Casa;Muro;"
            "-25,5;-49,6;2026;21;2026-06-01 17:31:10;0;Vanessa;;-25.5,-49.6;"
            "1271;2026-05-25;2026-05-29\n"
        )
        cadastro_csv = (
            "ID;Rua;NÃƒÂºmero do logradouro;Complemento;Bairro;LocalizaÃƒÂ§ÃƒÂ£o da ovitrampa;"
            "Setor/Distrito da ovitrampa;ResponsÃƒÂ¡vel;QuarteirÃƒÂ£o;Latitude;Longitude\n"
            "2;Rua B;20;Loja;;Canto;GRAZIELA;Vanessa;1270;-25,3;-49,4\n"
            "3;Rua C;30;Casa;;Muro;GRAZIELA;Vanessa;1271;-25,5;-49,6\n"
        )
        ocorrencias_csv = (
            "Ciclo;CRS;Município;Código do IBGE;Ano;Número da armadilha;Ovos;Link;"
            "Latitude;Longitude;Resultado;Semana;Lat_lng;Data;ID da Contagem;"
            "Tempo de Envio da Contagem;Código da Observação;Observação\n"
            "Junho;2;Almirante Tamandaré;4100400;2026;2;0;Almirante Tamandaré_2;"
            "-25,30000;-49,40000;Negativa;21;-25.3,-49.4;2026-05-29;9001;"
            "2026-06-01 10:00:00;6;Ovitrampa seca\n"
            "Junho;2;Almirante Tamandaré;4100400;2026;2;0;Almirante Tamandaré_2;"
            "-25,30000;-49,40000;Negativa;21;-25.3,-49.4;2026-05-29;9002;"
            "2026-06-01 10:00:00;10;Outra Observação\n"
            "Junho;2;Almirante Tamandaré;4100400;2026;3;0;Almirante Tamandaré_3;"
            "-25,50000;-49,60000;Negativa;21;-25.5,-49.6;2026-05-29;9003;"
            "2026-06-01 10:00:00;1;Sem Observações\n"
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            leitura_path = Path(tmpdir) / "3918-2026-21.csv"
            cadastro_path = Path(tmpdir) / "almirante-tamandare-pr-ovitraps.csv"
            ocorrencias_path = Path(tmpdir) / "municipality-3918-2026.csv"
            leitura_path.write_text(leitura_csv, encoding="utf-8")
            cadastro_path.write_text(cadastro_csv, encoding="utf-8")
            ocorrencias_path.write_text(ocorrencias_csv, encoding="utf-8")
            ovitrampas_core.importar_csv(db_path, leitura_path)
            ovitrampas_core.importar_armadilhas_csv(db_path, cadastro_path)

            result = ovitrampas_core.importar_ocorrencias_csv(db_path, ocorrencias_path)
            result_reimportado = ovitrampas_core.importar_ocorrencias_csv(db_path, ocorrencias_path)
            dados = ovitrampas_core.monitoramento(db_path, {"ano": "2026", "semana_ini": "21", "semana_fim": "21"})

        self.assertEqual(result["linhas"], 3)
        self.assertEqual(result["inseridos"], 3)
        self.assertEqual(result["ocorrencias"], 2)
        self.assertEqual(result_reimportado["sem_alteracao"], 3)
        self.assertEqual(dados["totais"]["ocorrencias"], 2)
        self.assertEqual(dados["ocorrencias_fonte"], "Histórico de ocorrências importado do Conta Ovos")
        ocorrencias = {row["codigo"]: row for row in dados["ocorrencias"]}
        self.assertEqual(ocorrencias[5]["total"], 1)
        self.assertEqual(ocorrencias[9]["total"], 1)
        self.assertEqual(ocorrencias[9]["descricao"], "Outros")
        self.assertEqual(ocorrencias[5]["armadilhas_destaque"][0]["observacao"], "Ovitrampa seca")
        self.assertEqual(ocorrencias[5]["armadilhas_destaque"][0]["data"], "2026-05-29")

    def test_ovitrampas_calendario_salva_evento_com_agentes_e_aparece_na_agenda(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Ovitrampa",))
                conn.commit()
                id_agente = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Agente Ovitrampa",)).fetchone()[0]
            finally:
                conn.close()

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            self.assertEqual(resp.status_code, 200)
            grupo = resp.get_json()["grupos"][0]

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-01-06",
                "movimento": "instalacao",
                "id_grupo": grupo["id_grupo"],
                "ciclo": "Ciclo 1",
                "observacoes": "Planejado",
                "agentes": [id_agente],
            })
            self.assertEqual(resp.status_code, 201)
            evento = resp.get_json()["evento"]
            self.assertEqual(evento["movimento_label"], "Instalação")
            self.assertEqual(evento["agentes"][0]["nome"], "Agente Ovitrampa")

            resp = client.get("/api/agenda/eventos?start=2026-01-01&end=2026-01-31")
            self.assertEqual(resp.status_code, 200)
            eventos = resp.get_json()
            ovitrampa = [e for e in eventos if e["extendedProps"].get("fonte") == "OVITRAMPA"]
            self.assertEqual(len(ovitrampa), 1)
            self.assertIn("Instalação de ovitrampas", ovitrampa[0]["title"])
            self.assertEqual(ovitrampa[0]["extendedProps"]["agentes"], "Agente Ovitrampa")

    def test_ovitrampas_calendario_feriado_nao_aparece_na_agenda(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-02-16",
                "movimento": "feriado",
                "titulo": "Carnaval",
                "observacoes": "Sem movimento de campo",
            })
            self.assertEqual(resp.status_code, 201)
            evento = resp.get_json()["evento"]
            self.assertEqual(evento["movimento_label"], "Feriado")
            self.assertEqual(evento["titulo"], "Carnaval")

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            eventos_calendario = resp.get_json()["eventos"]
            self.assertTrue(any(e["movimento"] == "feriado" for e in eventos_calendario))

            resp = client.get("/api/agenda/eventos?start=2026-02-01&end=2026-02-28")
            self.assertEqual(resp.status_code, 200)
            eventos_agenda = resp.get_json()
            self.assertFalse(any(e["extendedProps"].get("fonte") == "OVITRAMPA" for e in eventos_agenda))

    def test_ovitrampas_calendario_impressao_html_com_legenda(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)

            resp = client.get("/ovitrampas")
            self.assertEqual(resp.status_code, 200)
            pagina = resp.get_data(as_text=True)
            self.assertIn('id="ovi-cal-imprimir"', pagina)
            self.assertIn("/ovitrampas/calendario/imprimir?ano=", pagina)
            self.assertIn("renderLegendaMes", pagina)

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            self.assertEqual(resp.status_code, 200)
            grupo = resp.get_json()["grupos"][0]

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-03-10",
                "movimento": "troca",
                "id_grupo": grupo["id_grupo"],
                "observacoes": "Troca semanal",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-03-19",
                "movimento": "feriado",
                "titulo": "Feriado municipal",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-04-01",
                "movimento": "instalacao",
                "id_grupo": grupo["id_grupo"],
            })
            self.assertEqual(resp.status_code, 201)

            dados_impressao = ovitrampas_core.calendario_impressao(
                app_temp.config["DB_PATH"], 2026
            )
            marco = next(m for m in dados_impressao["meses"] if m["numero"] == 3)
            dias_fora_mes = [d for semana in marco["semanas"] for d in semana["dias"] if d["fora_mes"]]
            self.assertTrue(dias_fora_mes)
            self.assertTrue(all(d["evento"] is None and d["dia"] == "" for d in dias_fora_mes))

            resp = client.get("/ovitrampas/calendario/imprimir?ano=2026")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("window.print()", html)
            self.assertIn("print-color-adjust: exact", html)
            self.assertIn("Calendário de ovitrampas - 2026", html)
            self.assertIn(grupo["nome"], html)
            self.assertNotIn(f" - {grupo['localidades']}", html)
            self.assertIn("Feriado municipal", html)

    def test_ovitrampas_calendario_migra_vinculo_agentes_apontando_tabela_antiga(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            ovitrampas_core.calendario_dados(db_path, 2026)
            conn = sqlite3.connect(db_path)
            try:
                conn.executescript(
                    """
                    DROP TABLE ovitrampas_calendario_agentes;
                    CREATE TABLE ovitrampas_calendario_agentes (
                        id_evento INTEGER NOT NULL REFERENCES "ovitrampas_calendario_eventos_old"(id_evento) ON DELETE CASCADE,
                        id_agente INTEGER NOT NULL REFERENCES agentes(id_agente),
                        PRIMARY KEY (id_evento, id_agente)
                    );
                    """
                )
                conn.commit()
            finally:
                conn.close()

            evento = ovitrampas_core.salvar_evento_calendario(db_path, {
                "data": "2026-12-25",
                "movimento": "feriado",
                "titulo": "Natal",
            })

            conn = sqlite3.connect(db_path)
            try:
                destinos = [row[2] for row in conn.execute("PRAGMA foreign_key_list(ovitrampas_calendario_agentes)")]
            finally:
                conn.close()
            self.assertEqual(evento["titulo"], "Natal")
            self.assertIn("ovitrampas_calendario_eventos", destinos)
            self.assertNotIn("ovitrampas_calendario_eventos_old", destinos)

    def test_ovitrampas_fetch_com_csrf_invalido_retorna_json(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            app_temp = endemias_app.create_app({
                "TESTING": True,
                "DB_PATH": db_path,
                "WTF_CSRF_ENABLED": True,
            })
            client = app_temp.test_client()
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                usuario = conn.execute(
                    "SELECT id_usuario, nome, nivel FROM usuarios WHERE usuario='admin'"
                ).fetchone()
            finally:
                conn.close()
            _login_client_com_usuario(client, dict(usuario))

            resp = client.post(
                "/api/ovitrampas/calendario/eventos",
                json={"data": "2026-02-16", "movimento": "feriado", "titulo": "Carnaval"},
                headers={"X-Requested-With": "XMLHttpRequest"},
            )

            self.assertEqual(resp.status_code, 400)
            self.assertTrue(resp.is_json)
            self.assertIn("Token de seguranca", resp.get_json()["erro"])

    def test_create_app_permanece_configuravel(self):
        app_temp = endemias_app.create_app({
            "TESTING": True,
            "DB_PATH": endemias_app.DB_PATH,
        })

        self.assertTrue(app_temp.config["TESTING"])
        self.assertEqual(app_temp.config["DB_PATH"], endemias_app.DB_PATH)
        self.assertGreater(app_temp.config["MAX_CONTENT_LENGTH"], 0)
        endpoints = {str(rule): rule.endpoint for rule in app_temp.url_map.iter_rules()}
        self.assertEqual(endpoints["/"], "home.page")

    def test_wrappers_de_banco_respeitam_db_path_do_app_atual(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = str(Path(tmpdir) / "isolado.db")
            app_temp = endemias_app.create_app({
                "TESTING": True,
                "DB_PATH": db_path,
            })

            with app_temp.app_context():
                conn = endemias_app.get_db()
                try:
                    conn.execute("CREATE TABLE marcador (valor TEXT)")
                    conn.execute("INSERT INTO marcador (valor) VALUES ('ok')")
                    conn.commit()
                finally:
                    conn.close()

                self.assertEqual(endemias_app.qval("SELECT valor FROM marcador"), "ok")

            self.assertTrue(Path(db_path).exists())

    def test_resolve_paths_suporta_instance_dir_e_overrides(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            env = {
                "ENDEMIAS_INSTANCE_DIR": tmpdir,
                "ENDEMIAS_DB_PATH": str(Path(tmpdir) / "dados.db"),
                "ENDEMIAS_UPLOAD_TEMP": str(Path(tmpdir) / "uploads"),
                "ENDEMIAS_ANEXOS_DIR": str(Path(tmpdir) / "anexos"),
                "ENDEMIAS_KOBO_CONFIG_PATH": str(Path(tmpdir) / "kobo_config.json"),
            }

            paths = endemias_app.resolve_paths(env=env, base_dir=str(ROOT))

        self.assertEqual(paths["INSTANCE_DIR"], os.path.abspath(tmpdir))
        self.assertEqual(paths["DB_PATH"], os.path.abspath(env["ENDEMIAS_DB_PATH"]))
        self.assertEqual(paths["UPLOAD_TEMP"], os.path.abspath(env["ENDEMIAS_UPLOAD_TEMP"]))
        self.assertEqual(paths["ANEXOS_DIR"], os.path.abspath(env["ENDEMIAS_ANEXOS_DIR"]))
        self.assertEqual(paths["KOBO_CONFIG_PATH"], os.path.abspath(env["ENDEMIAS_KOBO_CONFIG_PATH"]))
        self.assertEqual(paths["CONFIG_PATH"], os.path.abspath(str(ROOT / "config.json")))
        self.assertTrue(paths["LOG_PATH"].endswith("endemias.log"))
        self.assertTrue(paths["SECRET_KEY_PATH"].endswith("secret.key"))

    def test_apis_principais_logadas_retornam_json(self):
        client = _client_logado()
        rotas = [
            "/api/dashboard",
            "/api/visitas?pagina=1&por_pagina=5",
            "/api/laboratorio?pagina=1&por_pagina=5",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)
                self.assertTrue(resp.is_json)

    def test_api_dashboard_inclui_esporotricose_e_comparativo(self):
        client = _client_logado()
        resp = client.get("/api/dashboard")

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("comparativo_mensal", dados)
        self.assertIn("esporotricose", dados)
        self.assertIn("pontos_estrategicos", dados)
        self.assertIn("producao_operacional", dados)
        self.assertIn("resumo", dados["esporotricose"])
        self.assertIn("dashboard", dados["esporotricose"])
        self.assertIn("evolucao", dados["esporotricose"]["dashboard"])
        self.assertIn("totais", dados["pontos_estrategicos"])
        self.assertIn("por_atividade", dados["producao_operacional"])

    def test_api_producao_operacional_retorna_central_integrada(self):
        client = _client_logado()
        resp = client.get("/api/producao-operacional?d_ini=2026-05-01&d_fim=2026-05-31")

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("totais", dados)
        self.assertIn("por_atividade", dados)
        self.assertIn("por_localidade", dados)
        self.assertIn("por_agente", dados)
        codigos = {item["codigo"] for item in dados["por_atividade"]}
        self.assertIn("VETORES", codigos)
        self.assertIn("ESPOROTRICOSE", codigos)
        self.assertIn("RECOLHIMENTO", codigos)
        self.assertIn("AMOSTRA_ANIMAIS", codigos)
        self.assertIn("BRI", codigos)

    def test_api_dashboard_inclui_ovitrampas_com_leituras_e_calendario(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Lab Ovi",))
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Campo Ovi",))
                conn.commit()
                lab_id = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Lab Ovi",)).fetchone()[0]
                campo_id = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Campo Ovi",)).fetchone()[0]
                ovitrampas_core.ensure_schema(conn)
                conn.execute(
                    """INSERT INTO ovitrampas_armadilhas
                       (ovitrampa_id, localidade, rua, atualizado_em)
                       VALUES (?, ?, ?, ?)""",
                    ("900", "Centro", "Rua Teste", "2026-08-12T08:00:00"),
                )
                conn.execute(
                    """INSERT INTO ovitrampas_leituras
                       (id_leitura, ovitrampa_id, distrito, ano, semana, data_coleta,
                        ovos, id_laboratorista, importado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    ("L-900", "900", "Centro", 2026, 33, "2026-08-12", 42, lab_id, "2026-08-12T10:00:00"),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            self.assertEqual(resp.status_code, 200)
            grupo = resp.get_json()["grupos"][0]
            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-08-13",
                "movimento": "troca",
                "id_grupo": grupo["id_grupo"],
                "agentes": [campo_id],
            })
            self.assertEqual(resp.status_code, 201)
            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-08-14",
                "movimento": "feriado",
                "titulo": "Feriado dashboard",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get("/api/dashboard?d_ini=2026-08-01&d_fim=2026-08-31")
            self.assertEqual(resp.status_code, 200)
            dados = resp.get_json()
            self.assertIn("ovitrampas", dados)
            self.assertEqual(dados["ovitrampas"]["leituras"]["totais"]["leituras"], 1)
            self.assertEqual(dados["ovitrampas"]["leituras"]["totais"]["positivas"], 1)
            self.assertEqual(dados["ovitrampas"]["leituras"]["totais"]["ovos"], 42)
            self.assertEqual(dados["ovitrampas"]["calendario"]["totais"]["movimentos"], 1)
            self.assertEqual(dados["ovitrampas"]["calendario"]["por_agente"][0]["agente"], "Campo Ovi")
            agosto = next(item for item in dados["comparativo_mensal"] if item["mes"] == "2026-08")
            self.assertEqual(agosto["ovitrampas"], 1)
            self.assertNotIn("Feriado dashboard", json.dumps(dados, ensure_ascii=False))

    def test_api_relatorio_agente_inclui_esporotricose_sem_expor_outros_agentes(self):
        client = _client_logado()
        agente = _agente_relatorio_teste()
        resp = client.get(
            "/api/relatorio-agente",
            query_string={"agente": agente, "d_ini": "2020-01-01", "d_fim": "2030-12-31"},
        )

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("esporotricose", dados)
        self.assertIn("producao_operacional", dados)
        self.assertIn("por_atividade", dados["producao_operacional"])
        self.assertIn("totais", dados["esporotricose"])
        self.assertIn("animais", dados["esporotricose"])
        self.assertIn("comparacao_esporotricose", dados)
        self.assertIn("esporotricose", dados["tbo_duracao"])
        self.assertIn("total", dados["tbo_duracao"])
        self.assertIn("n", dados["tbo_duracao"]["esporotricose"])
        self.assertIn("media", dados["tbo_duracao"]["total"])
        self.assertNotIn("recentes", dados["esporotricose"])
        self.assertNotIn(
            "sispncd",
            json.dumps(dados["producao_operacional"], ensure_ascii=False).lower(),
        )

        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            outros = [
                row[0]
                for row in conn.execute(
                    "SELECT nome FROM agentes WHERE nome <> ? ORDER BY nome LIMIT 5",
                    (agente,),
                )
            ]
        finally:
            conn.close()
        payload = json.dumps(dados, ensure_ascii=False)
        for outro in outros:
            self.assertNotIn(outro, payload)

    def test_relatorio_agente_inclui_ovitrampas_vinculadas_e_ignora_feriados(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Ovi",))
                conn.commit()
                id_agente = conn.execute("SELECT id_agente FROM agentes WHERE nome=?", ("Agente Ovi",)).fetchone()[0]
            finally:
                conn.close()

            resp = client.get("/api/ovitrampas/calendario?ano=2026")
            self.assertEqual(resp.status_code, 200)
            grupo = resp.get_json()["grupos"][0]

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-07-14",
                "movimento": "troca",
                "id_grupo": grupo["id_grupo"],
                "ciclo": "Ciclo 2",
                "observacoes": "Equipe completa",
                "agentes": [id_agente],
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.post("/api/ovitrampas/calendario/eventos", json={
                "data": "2026-07-15",
                "movimento": "feriado",
                "titulo": "Feriado teste",
            })
            self.assertEqual(resp.status_code, 201)

            resp = client.get(
                "/api/relatorio-agente",
                query_string={"agente": "Agente Ovi", "d_ini": "2026-07-01", "d_fim": "2026-07-31"},
            )
            self.assertEqual(resp.status_code, 200)
            dados = resp.get_json()
            self.assertEqual(dados["ovitrampas"]["totais"]["eventos"], 1)
            self.assertEqual(dados["ovitrampas"]["totais"]["dias"], 1)
            self.assertEqual(dados["ovitrampas"]["eventos"][0]["movimento_label"], "Troca")
            self.assertNotIn("Feriado teste", json.dumps(dados, ensure_ascii=False))

            resp = client.get(
                "/relatorio-agente/pdf",
                query_string={"agente": "Agente Ovi", "d_ini": "2026-07-01", "d_fim": "2026-07-31"},
            )
            self.assertEqual(resp.status_code, 200)
            html = resp.data.decode("utf-8")
            self.assertIn("Ovitrampas", html)
            self.assertIn("Dias em ovitrampas", html)
            self.assertIn("Equipe completa", html)
            self.assertNotIn("Feriado teste", html)

    def test_relatorio_agente_inclui_acoes_setor_e_registro_geografico(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Integrado",))
                id_agente = conn.execute(
                    "SELECT id_agente FROM agentes WHERE nome=?",
                    ("Agente Integrado",),
                ).fetchone()["id_agente"]
                loc = conn.execute("SELECT id_localidade, nome FROM localidades ORDER BY id_localidade LIMIT 1").fetchone()
                if not loc:
                    conn.execute("INSERT INTO localidades (nome) VALUES (?)", ("Teste",))
                    loc = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome=?", ("Teste",)).fetchone()

                cur = conn.execute(
                    """INSERT INTO acoes_setor
                       (tipo, data, hora_inicio, hora_fim, localidade, endereco, local,
                        publico_aproximado, tema, contexto, coordenadas, observacoes,
                        criado_por, criado_em, atualizado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        "educativa",
                        "2026-06-10",
                        "08:00",
                        "09:00",
                        loc["nome"],
                        "Rua Teste",
                        "Escola Teste",
                        35,
                        "Dengue",
                        "Palestra",
                        None,
                        None,
                        "teste",
                        "2026-06-10T08:00:00",
                        "2026-06-10T08:00:00",
                    ),
                )
                conn.execute(
                    "INSERT INTO acoes_setor_agentes (id_acao, id_agente) VALUES (?, ?)",
                    (cur.lastrowid, id_agente),
                )

                registro_geografico_core.ensure_schema(conn)
                cur_q = conn.execute(
                    """INSERT INTO registro_geografico_quarteiroes
                       (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                       VALUES (?, ?, ?, ?, ?)""",
                    (loc["id_localidade"], loc["nome"], "0007", "2026-06-11T08:00:00", "2026-06-11T08:00:00"),
                )
                for ordem, numero in enumerate(("10", "12"), 1):
                    cur_i = conn.execute(
                        """INSERT INTO registro_geografico_imoveis
                           (id_quarteirao, ordem, id_localidade, localidade, quarteirao, logradouro,
                            numero, sequencia, lado, tipo, condominio, observacao, data_atualizacao,
                            agentes_texto, busca_normalizada, chave_origem, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            cur_q.lastrowid,
                            ordem,
                            loc["id_localidade"],
                            loc["nome"],
                            "0007",
                            "Rua Teste",
                            numero,
                            None,
                            "1",
                            "R",
                            None,
                            None,
                            "2026-06-11",
                            "Agente Integrado",
                            "agente integrado rua teste",
                            f"teste-rg-{ordem}",
                            "2026-06-11T08:00:00",
                            "2026-06-11T08:00:00",
                        ),
                    )
                    conn.execute(
                        "INSERT INTO registro_geografico_imovel_agentes (id_imovel, id_agente) VALUES (?, ?)",
                        (cur_i.lastrowid, id_agente),
                    )
                conn.execute(
                    """INSERT INTO resultados_laboratorio
                       (id_coleta, num_tubo, data_coleta, laboratorista, data_leitura,
                        aegypt_larvas, aegypt_pupas, aegypt_exuvias, aegypt_adulto,
                        albopictus_larvas, albopictus_pupas, albopictus_exuvias, albopictus_adulto,
                        outra_larvas, outra_pupas, outra_exuvias, outra_adulto)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        "coleta-lab-1",
                        "T-100",
                        "2026-06-12",
                        "agente integrado",
                        "2026-06-13",
                        1,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                        0,
                    ),
                )
                ovitrampas_core.ensure_schema(conn)
                conn.execute(
                    """INSERT INTO ovitrampas_leituras
                       (id_leitura, ovitrampa_id, ano, semana, distrito, ovos,
                        id_laboratorista, data_leitura, importado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        "ovi-lab-1",
                        "OVI-100",
                        2026,
                        24,
                        loc["nome"],
                        52,
                        id_agente,
                        "2026-06-14",
                        "2026-06-14T08:00:00",
                    ),
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.get(
                "/api/relatorio-agente",
                query_string={"agente": "Agente Integrado", "d_ini": "2026-06-01", "d_fim": "2026-06-30"},
            )
            self.assertEqual(resp.status_code, 200)
            dados = resp.get_json()
            atividades = {item["codigo"]: item for item in dados["producao_operacional"]["por_atividade"]}
            self.assertEqual(atividades["ACOES_SETOR"]["registros"], 1)
            self.assertEqual(atividades["ACOES_SETOR"]["extras"]["educativas"], 1)
            self.assertEqual(dados["registro_geografico"]["totais"]["imoveis"], 2)
            self.assertEqual(dados["registro_geografico"]["totais"]["quarteiroes"], 1)
            self.assertEqual(dados["laboratorio"]["totais"]["leituras"], 2)
            self.assertEqual(dados["laboratorio"]["totais"]["tubos"], 1)
            self.assertEqual(dados["laboratorio"]["totais"]["ovos"], 52)

            resp = client.get(
                "/relatorio-agente/pdf",
                query_string={"agente": "Agente Integrado", "d_ini": "2026-06-01", "d_fim": "2026-06-30"},
            )
            self.assertEqual(resp.status_code, 200)
            html = resp.data.decode("utf-8")
            self.assertIn("Ações e Atendimentos", html)
            self.assertIn("Registro Geogr&aacute;fico", html)
            self.assertIn("Laborat&oacute;rio", html)

    def test_api_conta_ovos_e_sispncd_consultas_retornam_json(self):
        client = _client_logado()
        padrao = sispncd_core.get_default_conta_ovos(endemias_app.DB_PATH)
        rotas = [
            f"/api/conta-ovos?data={padrao['data']}&quarteirao={padrao['quarteirao']}",
            "/api/sispncd/pesquisar?ano=2026&semana=20&tipo=TB/TBO&tipo=PVE&tipo=PE",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)
                self.assertTrue(resp.is_json)

    def test_api_esporotricose_retorna_json(self):
        client = _client_logado()
        resp = client.get("/api/esporotricose")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        self.assertIn("totais", dados)
        self.assertIn("animais", dados)

    def test_pagina_esporotricose_exibe_abas_principais(self):
        client = _client_logado()
        resp = client.get("/esporotricose")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("esporo-tabbar", html)
        self.assertIn('data-esp-main="visitas"', html)
        self.assertIn('data-esp-main="doentes"', html)
        self.assertIn("esp-visitas-tabs", html)
        self.assertIn("esp-doentes-tabs", html)
        self.assertIn("esp-tab-visitas", html)
        self.assertIn("esp-tab-animais", html)
        self.assertIn("esporo-multi-picker", html)
        self.assertIn("initMultiPickers", html)
        self.assertIn("renderEstoqueMovimentoLinha", html)
        self.assertIn("encontrarEspAnimalDetalhe", html)
        self.assertIn("encontrarEspAnimalLinha", html)
        self.assertIn("toggleEspAnimalDetalhes(this.dataset.id, event)", html)
        self.assertIn("esp-tab-doentes", html)
        self.assertIn("esp-tab-doentes-resumo", html)
        self.assertIn("esp-tab-doentes-estoque", html)
        self.assertIn("Estoque medicação", html)
        self.assertIn("doe-res-caps-entregues", html)
        self.assertIn("doe-res-caps-baixa-zoomed", html)
        self.assertIn("CPF do tutor", html)
        self.assertIn("Cápsulas", html)
        self.assertIn("proxima_entrega", html)
        self.assertIn("rec-capsulas-dia", html)
        self.assertIn("rec-pendente", html)
        self.assertIn("Receita pendente", html)
        self.assertIn("esp-visitas-kpis", html)
        self.assertIn("doe-kpi-total", html)
        self.assertIn("gato_doente.svg", html)
        self.assertIn("Pedir documentos", html)
        self.assertIn("esp-tab-localidades", html)
        self.assertIn("esp-tab-dashboard", html)
        self.assertIn("esp-animal-especie", html)
        self.assertIn("esp-animal-motivo", html)
        self.assertIn("imprimirEspAnimais", html)
        self.assertIn("renderEspAnimalDetalheImpressao", html)
        self.assertIn("detail-row", html)
        self.assertIn("prioritarios", html)

    def test_pagina_recolhimentos_exibe_controles_principais(self):
        client = _client_logado()
        resp = client.get("/recolhimentos")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Recolhimento de Materiais", html)
        self.assertIn("rec-registros", html)
        self.assertIn("rec-body", html)

    def test_api_recolhimentos_retorna_json(self):
        client = _client_logado()
        resp = client.get("/api/recolhimentos?d_ini=2099-01-01&d_fim=2099-01-02")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("totais", dados)
        self.assertIn("por_localidade", dados)

    def test_pagina_amostras_animais_exibe_controles_principais(self):
        client = _client_logado()
        resp = client.get("/amostras-animais")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Amostra de Animais", html)
        self.assertIn("am-registros", html)
        self.assertIn("am-body", html)

    def test_api_amostras_animais_retorna_json(self):
        client = _client_logado()
        resp = client.get("/api/amostras-animais?d_ini=2099-01-01&d_fim=2099-01-02")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("totais", dados)
        self.assertIn("por_tipo", dados)
        self.assertIn("por_localidade", dados)

    def test_pagina_bri_exibe_controles_principais(self):
        client = _client_logado()
        resp = client.get("/bri")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Borrifamento Residual Intradomiciliar", html)
        self.assertIn("bri-registros", html)
        self.assertIn("bri-body", html)

    def test_api_bri_retorna_json(self):
        client = _client_logado()
        resp = client.get("/api/bri?d_ini=2099-01-01&d_fim=2099-01-02")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("totais", dados)
        self.assertIn("por_destino", dados)
        self.assertIn("por_localidade", dados)
        self.assertIn("vinculados_pe", dados["totais"])
        self.assertIn("ambiguos_pe", dados["totais"])

    def test_pagina_pontos_estrategicos_exibe_controles_principais(self):
        client = _client_logado()
        resp = client.get("/pontos-estrategicos")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Pontos Estratégicos", html)
        self.assertIn("pe-total", html)
        self.assertIn("pe-body", html)
        self.assertIn("function peData", html)
        self.assertIn("function peIsoData", html)
        self.assertIn('placeholder="DD-MM-AAAA"', html)
        self.assertIn("${partes[2]}-${partes[1]}-${partes[0]}", html)

    def test_api_pontos_estrategicos_retorna_json(self):
        client = _client_logado()
        resp = client.get("/api/pontos-estrategicos")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("totais", dados)
        self.assertIn("registros", dados)

    def test_api_pontos_estrategicos_filtra_pendencias(self):
        client = _client_logado()
        resp = client.get("/api/pontos-estrategicos?pendencias=1")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("totais", dados)
        for registro in dados["registros"]:
            self.assertTrue(registro.get("visita_atrasada") or registro.get("pendencias_cadastro"))

    def test_api_pontos_estrategicos_situacao_invalida_retorna_400(self):
        client = _client_logado("admin")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp = client.post("/api/pontos-estrategicos/1/situacao", json={"situacao": "abc"})
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        self.assertEqual(resp.status_code, 400)
        self.assertTrue(resp.is_json)

    def test_pagina_registro_geografico_renderiza(self):
        client = _client_logado()
        resp = client.get("/registro-geografico")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("Boletim de Registro de Reconhecimento Geogr", html)
        self.assertIn("rg-panel-consulta", html)
        self.assertIn("rg-panel-edicao", html)
        self.assertIn("rg-panel-lotes", html)
        self.assertIn("rg-panel-acompanhamento", html)
        self.assertIn("Edição em lotes", html)
        self.assertIn("REF - Referencia do quarteirao", html)
        self.assertIn("/api/registro-geografico/logradouros-similares", html)
        self.assertIn("/api/registro-geografico/logradouros-sugestoes", html)
        self.assertIn("/api/registro-geografico/lote/preview", html)
        self.assertIn("/api/registro-geografico/lote/aplicar", html)
        self.assertIn("/api/registro-geografico/quarteirao/limpar", html)
        self.assertIn("/api/registro-geografico/quarteirao/excluir", html)
        self.assertIn("/api/registro-geografico/acompanhamento", html)
        self.assertIn("function rgAcompCarregar", html)
        self.assertIn("rg-panel-mapa", html)
        self.assertIn('id="rg-mapa"', html)
        self.assertIn('id="rg-map-detail-title"', html)
        self.assertIn('id="rg-map-reset-colors"', html)
        self.assertIn('class="rg-map-color"', html)
        self.assertIn("/static/vendor/leaflet/leaflet.min.css", html)
        self.assertIn("/static/vendor/leaflet/leaflet.min.js", html)
        self.assertIn("rg-filter-card", html)
        self.assertIn("rg-print-actions", html)
        self.assertIn('id="rg-imp-mini-mapa" checked', html)
        self.assertIn('id="rg-imp-duplex-recto"', html)
        self.assertIn('id="rg-imp-selecionar-todos"', html)
        self.assertIn('id="rg-imp-limpar-selecao"', html)
        self.assertIn('id="rg-imp-status"', html)
        self.assertIn("Incluir mapas sat", html)
        self.assertIn("params.set('mini_mapa', '1')", html)
        self.assertIn("params.set('duplex_recto', '1')", html)
        self.assertIn("function rgMarcarQuarteiroesImpressao", html)
        self.assertIn("rgMarcarQuarteiroesImpressao(true)", html)
        self.assertIn("rgMarcarQuarteiroesImpressao(false)", html)
        self.assertIn("rg-kpi-populacao", html)
        self.assertIn('id="rg-localidade" multiple data-multi-picker', html)
        self.assertIn('id="rg-quarteiroes"', html)
        self.assertIn("rgCarregarChecksQuarteiroes(selecionadas[0], 'rg-quarteiroes')", html)
        self.assertNotIn('id="rg-quarteirao" multiple', html)
        self.assertIn('id="rg-ed-copy-up"', html)
        self.assertIn('id="rg-ed-fill-down"', html)
        self.assertIn('id="rg-ed-fill-empty"', html)
        self.assertIn('id="rg-ed-limpar-quarteirao"', html)
        self.assertIn('id="rg-ed-excluir-quarteirao"', html)
        self.assertIn("Limpar quarteir&atilde;o", html)
        self.assertIn("Excluir quarteir&atilde;o", html)
        self.assertIn('id="rg-logradouro-suggest"', html)
        self.assertIn("function rgBuscarLogradouroSugestoes", html)
        self.assertIn("function rgLimparQuarteirao", html)
        self.assertIn("function rgExcluirQuarteirao", html)
        self.assertIn("function rgFocusCell", html)
        self.assertIn("rg-row-active", html)
        self.assertIn("nth-child(even)", html)
        self.assertIn("ev.key === 'ArrowRight'", html)
        self.assertIn("rgFillColumnBelow", html)
        self.assertIn("function rgMapInit", html)
        self.assertIn("/api/registro-geografico/mapa-resumo", html)
        self.assertIn("rgMapToggleSelection", html)
        self.assertIn("function rgMapRenderSelectionDetail", html)
        self.assertIn("Quarteirões selecionados", html)
        self.assertIn("rg-map-selected-list", html)
        self.assertIn("max-height:230px;overflow:auto", html)
        self.assertIn("World_Imagery/MapServer", html)
        self.assertIn("Satélite + ruas", html)
        self.assertIn("/static/ubs_setor_endemias.geojson", html)
        self.assertIn("Equipamentos da Saúde", html)
        self.assertIn("function rgMapRenderPontos", html)
        self.assertIn("rg_mapa_localidade_cores", html)
        self.assertIn("function rgMapResetColors", html)
        self.assertIn("rg-carregar-todos", html)
        self.assertIn("limite', rgState.limiteTodos ? 'todos' : '200'", html)
        self.assertIn("IBGE Censo 2022", html)

    def test_registro_geografico_geojson_pontos_saude(self):
        client = _client_logado()
        resp = client.get("/static/ubs_setor_endemias.geojson")

        self.assertEqual(resp.status_code, 200)
        dados = json.loads(resp.data.decode("utf-8"))
        resp.close()
        self.assertEqual(dados["type"], "FeatureCollection")
        self.assertGreaterEqual(len(dados["features"]), 1)
        nomes = {f.get("properties", {}).get("Nomes") for f in dados["features"]}
        self.assertIn("Setor de Endemias", nomes)

    def test_api_registro_geografico_mapa_resumo(self):
        client = _client_logado()
        resp = client.get("/api/registro-geografico/mapa-resumo")

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("quarteiroes", dados)
        self.assertIn("total", dados)
        self.assertIn("fonte_populacao", dados)
        self.assertIn("tipos", dados)
        self.assertIsInstance(dados["quarteiroes"], dict)
        self.assertGreaterEqual(dados["total"]["quarteiroes"], 1)
        primeiro = next(iter(dados["quarteiroes"].values()))
        self.assertIn("imoveis", primeiro)
        self.assertIn("imoveis_reais", primeiro)
        self.assertIn("populacao_aproximada", primeiro)
        self.assertIn("tipos", primeiro)

    def test_api_registro_geografico_acompanhamento(self):
        client = _client_logado()
        resp = client.get("/api/registro-geografico/acompanhamento")

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("totais", dados)
        self.assertIn("registros", dados)
        self.assertIsInstance(dados["registros"], list)
        for campo in ("quarteiroes", "atualizado", "parcial", "pendente", "sem_cadastro"):
            self.assertIn(campo, dados["totais"])
        if dados["registros"]:
            primeiro = dados["registros"][0]
            self.assertIn("situacao", primeiro)
            self.assertIn("atualizadas", primeiro)
            self.assertIn("linhas", primeiro)

    def test_api_registro_geografico_retorna_registros_e_salva_edicao(self):
        client = _client_logado("admin")
        resp = client.get("/api/registro-geografico?limite=1")
        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("totais", dados)
        self.assertIn("registros", dados)
        self.assertIn("populacao_aproximada", dados["totais"])
        self.assertEqual(
            dados["totais"]["populacao_aproximada"],
            round((dados["totais"].get("residencias_reais") or 0) * 2.93),
        )
        localidades = [
            str(r["id_localidade"])
            for r in dados["registros"]
            if r.get("id_localidade") is not None
        ]
        if len(set(localidades)) >= 2:
            selecionadas = sorted(set(localidades))[:2]
            resp_multi = client.get(
                "/api/registro-geografico?limite=5&"
                + "&".join(f"localidade={item}" for item in selecionadas)
            )
            self.assertEqual(resp_multi.status_code, 200)
            self.assertTrue(
                set(str(r["id_localidade"]) for r in resp_multi.get_json()["registros"]).issubset(set(selecionadas))
            )
        if not dados["registros"]:
            self.skipTest("Sem registros de RG para testar edicao.")

        registro = dados["registros"][0]
        self.assertEqual(registro.get("ordem"), 1)
        self.assertEqual(registro.get("quarteirao"), "1")
        detalhe = client.get(f"/api/registro-geografico/{registro['id_imovel']}").get_json()
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp_post = client.post(
                f"/api/registro-geografico/{registro['id_imovel']}",
                json={
                    "id_localidade": detalhe["id_localidade"],
                    "quarteirao": detalhe["quarteirao"],
                    "logradouro": detalhe["logradouro"],
                    "numero": detalhe["numero"],
                    "sequencia": detalhe.get("sequencia") or "",
                    "lado": detalhe.get("lado") or "",
                    "tipo": detalhe.get("tipo") or "",
                    "condominio": detalhe.get("condominio") or "",
                    "data_atualizacao": detalhe.get("data_atualizacao") or "",
                    "observacao": detalhe.get("observacao") or "",
                    "agentes_ids": detalhe.get("agentes_ids") or [],
                },
            )
            self.assertEqual(resp_post.status_code, 200)
            resp_criar = client.post(
                "/api/registro-geografico",
                json={
                    "after_id": registro["id_imovel"],
                    "id_localidade": detalhe["id_localidade"],
                    "quarteirao": detalhe["quarteirao"],
                    "logradouro": detalhe["logradouro"],
                    "numero": "TESTE",
                    "sequencia": "TESTE",
                    "lado": detalhe.get("lado") or "",
                    "tipo": detalhe.get("tipo") or "R",
                    "condominio": "",
                    "data_atualizacao": "",
                    "observacao": "Linha temporaria de teste",
                    "agentes_ids": [],
                },
            )
            self.assertEqual(resp_criar.status_code, 201)
            criado = resp_criar.get_json()["registro"]
            resp_ordem = client.get("/api/registro-geografico?limite=3")
            ids = [r["id_imovel"] for r in resp_ordem.get_json()["registros"]]
            self.assertEqual(ids[1], criado["id_imovel"])
            resp_delete = client.delete(f"/api/registro-geografico/{criado['id_imovel']}")
            self.assertEqual(resp_delete.status_code, 200)
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original
        self.assertTrue(resp_post.is_json)

    def test_registro_geografico_edita_quarteirao_e_imprime(self):
        client = _client_logado("admin")
        primeiro = client.get("/api/registro-geografico?limite=1").get_json()["registros"][0]
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp_q = client.get(
                f"/api/registro-geografico/quarteirao"
                f"?localidade={primeiro['id_localidade']}&quarteirao={primeiro['quarteirao']}"
            )
            self.assertEqual(resp_q.status_code, 200)
            dados = resp_q.get_json()
            linhas = dados["registros"][:]
            linhas.append({
                "logradouro": linhas[-1]["logradouro"],
                "numero": "TESTE-Q",
                "sequencia": "T",
                "lado": linhas[-1].get("lado") or "",
                "tipo": "R",
                "condominio": "",
                "observacao": "Linha temporaria de teste por quarteirao",
            })
            resp_save = client.post(
                "/api/registro-geografico/quarteirao",
                json={
                    "id_localidade": primeiro["id_localidade"],
                    "quarteirao": primeiro["quarteirao"],
                    "data_atualizacao": "2026-06-15",
                    "agentes_ids": dados.get("agentes_ids") or [],
                    "linhas": linhas,
                    "deleted_ids": [],
                },
            )
            self.assertEqual(resp_save.status_code, 200)
            salvo = resp_save.get_json()["quarteirao"]
            self.assertEqual(salvo["registros"][-1]["numero"], "TESTE-Q")
            apagado_id = salvo["registros"][-1]["id_imovel"]
            linhas_limpa = [r for r in salvo["registros"] if r["id_imovel"] != apagado_id]
            resp_clean = client.post(
                "/api/registro-geografico/quarteirao",
                json={
                    "id_localidade": primeiro["id_localidade"],
                    "quarteirao": primeiro["quarteirao"],
                    "data_atualizacao": dados.get("data_atualizacao") or "",
                    "agentes_ids": dados.get("agentes_ids") or [],
                    "linhas": linhas_limpa,
                    "deleted_ids": [apagado_id],
                },
            )
            self.assertEqual(resp_clean.status_code, 200)
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        impressao = client.get(
            f"/registro-geografico/imprimir?localidade={primeiro['id_localidade']}&quarteirao={primeiro['quarteirao']}"
        )
        self.assertEqual(impressao.status_code, 200)
        html = impressao.data.decode("utf-8")
        self.assertIn("Boletim de Registro de Reconhecimento Geográfico", html)
        self.assertNotIn("Reconhecimento Geográfico Digital", html)
        self.assertIn("Quarteirão: 1", html)
        self.assertIn("@page{size:A4 portrait", html)
        self.assertIn("C&oacute;digo da localidade", html)
        self.assertIn("qrcode_mapa.svg", html)
        self.assertIn("Total geral", html)
        self.assertIn("IBGE Censo 2022", html)
        self.assertIn("Popula&ccedil;&atilde;o aproximada", html)
        self.assertIn("considerando condomínios", html)
        self.assertNotIn("<th>Sem condomínios</th><th>Com condomínios</th>", html)

        self.assertNotIn("function rgPrintInitMaps", html)

        impressao_mapa = client.get(
            f"/registro-geografico/imprimir?localidade={primeiro['id_localidade']}"
            f"&quarteirao={primeiro['quarteirao']}&mini_mapa=1"
        )
        self.assertEqual(impressao_mapa.status_code, 200)
        html_mapa = impressao_mapa.data.decode("utf-8")
        self.assertLess(html_mapa.index('<div class="qr">'), html_mapa.index('<div class="rg-mini-map-stack">'))
        self.assertIn("/static/vendor/leaflet/leaflet.min.css", html_mapa)
        self.assertIn("/static/vendor/leaflet/leaflet.min.js", html_mapa)
        self.assertIn("rg-mini-map-stack", html_mapa)
        self.assertIn("rg-mini-map", html_mapa)
        self.assertIn(".rg-mini-map-stack{grid-column:1/-1;display:grid;grid-template-columns:1fr 1fr;gap:8pt;margin-top:4pt;break-inside:avoid;page-break-inside:avoid;}", html_mapa)
        self.assertIn(".rg-bottom,.rg-mini-map-stack,.rg-mini-map-box,.rg-mini-map{break-inside:avoid;page-break-inside:avoid;}", html_mapa)
        self.assertIn('data-mapa-base="satelite"', html_mapa)
        self.assertIn('data-mapa-base="osm"', html_mapa)
        self.assertIn("data-localidade", html_mapa)
        self.assertIn("data-quarteirao", html_mapa)
        self.assertIn("World_Imagery/MapServer", html_mapa)
        self.assertIn("tile.openstreetmap.org", html_mapa)
        self.assertNotIn("tileSize:128", html_mapa)
        self.assertNotIn("zoomOffset:1", html_mapa)
        self.assertIn("keepBuffer:2", html_mapa)
        self.assertIn("zoomSnap:0.1", html_mapa)
        self.assertIn("height:5.2cm", html_mapa)
        self.assertIn("bounds.pad(0.02)", html_mapa)
        self.assertIn("function rgPrintFitBounds", html_mapa)
        self.assertIn("function rgPrintStabilizeMap", html_mapa)
        self.assertIn("weight:isOsm ? 1 : 4", html_mapa)
        self.assertIn("fillOpacity:0.08", html_mapa)
        self.assertNotIn("rg-print-map-label", html_mapa)
        self.assertIn("function rgPrintInitMaps", html_mapa)

        impressao_duplex = client.get(
            f"/registro-geografico/imprimir?localidade={primeiro['id_localidade']}"
            f"&quarteirao={primeiro['quarteirao']}&quarteirao={primeiro['quarteirao']}&duplex_recto=1"
        )
        self.assertEqual(impressao_duplex.status_code, 200)
        html_duplex = impressao_duplex.data.decode("utf-8")
        self.assertIn("const RG_PRINT_DUPLEX_RECTO = true;", html_duplex)
        self.assertIn(".rg-duplex-blank{display:block!important;height:281mm;break-after:page;page-break-after:always;}", html_duplex)
        self.assertIn("function rgPrintPrepareDuplexRecto", html_duplex)
        self.assertIn("function rgPrintEstimateFolhaPages", html_duplex)
        self.assertIn("insertAdjacentHTML('afterend', '<div class=\"rg-duplex-blank\" aria-hidden=\"true\"></div>')", html_duplex)

        quarteiroes = client.get(f"/api/registro-geografico/quarteiroes?localidade={primeiro['id_localidade']}")
        self.assertEqual(quarteiroes.status_code, 200)
        self.assertGreaterEqual(len(quarteiroes.get_json()["registros"]), 1)

    def test_registro_geografico_lista_quarteirao_sem_imoveis(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            with app_temp.app_context():
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    registro_geografico_core.ensure_schema(conn)
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Teste RG", 9999))
                    loc = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Teste RG'").fetchone()
                    self.assertIsNotNone(loc)
                    conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (loc["id_localidade"], loc["nome"], "1406", "2026-07-01T08:00:00", "2026-07-01T08:00:00"),
                    )
                    conn.commit()
                    id_localidade = loc["id_localidade"]
                finally:
                    conn.close()

                resp = client.get(f"/api/registro-geografico/quarteiroes?localidade={id_localidade}")
                self.assertEqual(resp.status_code, 200)
                registros = resp.get_json()["registros"]
                item = next((r for r in registros if r["quarteirao"] == "1406"), None)
                self.assertIsNotNone(item)
                self.assertEqual(item["imoveis"], 0)

    def test_registro_geografico_ref_nao_contabiliza_como_imovel(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            with app_temp.app_context():
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    registro_geografico_core.ensure_schema(conn)
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Praca RG", 9201))
                    loc = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Praca RG'").fetchone()
                    agora = "2026-07-08T08:00:00"
                    cur_q = conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (loc["id_localidade"], loc["nome"], "0003", agora, agora),
                    )
                    linhas = [
                        ("Rua A", "SN", "O", "Praca acessivel por tres ruas"),
                        ("Rua B", "", "REF", "Lado do quarteirao"),
                        ("Rua C", "", "REF", "Lado do quarteirao"),
                    ]
                    for ordem, (logradouro, numero, tipo, observacao) in enumerate(linhas, 1):
                        item = {
                            "localidade": loc["nome"],
                            "quarteirao": "0003",
                            "logradouro": logradouro,
                            "numero": numero or "SN",
                            "sequencia": "",
                            "lado": "",
                            "tipo": tipo,
                            "observacao": observacao,
                            "agentes_texto": "",
                        }
                        conn.execute(
                            """INSERT INTO registro_geografico_imoveis
                               (id_quarteirao, ordem, id_localidade, localidade, quarteirao,
                                logradouro, numero, tipo, observacao, busca_normalizada, chave_origem, criado_em, atualizado_em)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                cur_q.lastrowid,
                                ordem,
                                loc["id_localidade"],
                                loc["nome"],
                                "0003",
                                logradouro,
                                numero or "SN",
                                tipo,
                                observacao,
                                registro_geografico_core._busca_normalizada(item),
                                f"teste-rg-ref-{ordem}",
                                agora,
                                agora,
                            ),
                        )
                    conn.commit()
                    loc_id = loc["id_localidade"]
                finally:
                    conn.close()

                lista = client.get(f"/api/registro-geografico?localidade={loc_id}&limite=todos")
                self.assertEqual(lista.status_code, 200)
                dados_lista = lista.get_json()
                self.assertEqual(len([r for r in dados_lista["registros"] if r["quarteirao_raw"] == "0003"]), 3)
                self.assertEqual(dados_lista["totais"]["imoveis"], 1)
                self.assertEqual(dados_lista["totais"]["imoveis_reais"], 1)

                quarteiroes = client.get(f"/api/registro-geografico/quarteiroes?localidade={loc_id}")
                self.assertEqual(quarteiroes.status_code, 200)
                self.assertEqual(quarteiroes.get_json()["registros"][0]["imoveis"], 1)

                detalhe = client.get(f"/api/registro-geografico/quarteirao?localidade={loc_id}&quarteirao=0003")
                self.assertEqual(detalhe.status_code, 200)
                dados = detalhe.get_json()
                self.assertEqual(len(dados["registros"]), 3)
                self.assertEqual(dados["resumo"]["total_sem_condominio"], 1)
                self.assertEqual(dados["resumo"]["linhas"][-1]["codigo"], "O")

                mapa = client.get("/api/registro-geografico/mapa-resumo")
                self.assertEqual(mapa.status_code, 200)
                item_mapa = mapa.get_json()["quarteiroes"][f"{loc_id}:3"]
                self.assertEqual(item_mapa["imoveis"], 1)
                self.assertNotIn("REF", {v.get("codigo") for v in item_mapa["tipos"].values()})

    def test_registro_geografico_move_quarteirao_para_outra_localidade_e_numero(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            with app_temp.app_context():
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    registro_geografico_core.ensure_schema(conn)
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Origem RG", 9001))
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Destino RG", 9002))
                    origem = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Origem RG'").fetchone()
                    destino = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Destino RG'").fetchone()
                    agora = "2026-07-01T09:00:00"
                    cur = conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (origem["id_localidade"], origem["nome"], "0100", agora, agora),
                    )
                    id_quarteirao = cur.lastrowid
                    conn.execute(
                        """INSERT INTO registro_geografico_imoveis
                           (id_quarteirao, ordem, id_localidade, localidade, quarteirao,
                            logradouro, numero, tipo, busca_normalizada, chave_origem, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            id_quarteirao, 1, origem["id_localidade"], origem["nome"], "0100",
                            "Rua Teste", "10", "R", "origem rg 0100 rua teste 10", "teste-move-rg", agora, agora,
                        ),
                    )
                    conn.commit()
                    origem_id = origem["id_localidade"]
                    destino_id = destino["id_localidade"]
                finally:
                    conn.close()

                resp = client.get(f"/api/registro-geografico/quarteirao?localidade={origem_id}&quarteirao=0100")
                self.assertEqual(resp.status_code, 200)
                dados = resp.get_json()
                linhas = dados["registros"]
                resp_save = client.post(
                    "/api/registro-geografico/quarteirao",
                    json={
                        "origem_id_localidade": origem_id,
                        "origem_quarteirao": "0100",
                        "id_localidade": destino_id,
                        "quarteirao": "0200",
                        "data_atualizacao": "2026-07-01",
                        "agentes_ids": [],
                        "linhas": linhas,
                        "deleted_ids": [],
                    },
                )
                self.assertEqual(resp_save.status_code, 200)
                salvo = resp_save.get_json()["quarteirao"]
                self.assertEqual(salvo["localidade"]["id_localidade"], destino_id)
                self.assertEqual(salvo["quarteirao_raw"], "0200")

                origem_vazia = client.get(f"/api/registro-geografico/quarteirao?localidade={origem_id}&quarteirao=0100")
                self.assertEqual(origem_vazia.status_code, 200)
                self.assertEqual(origem_vazia.get_json()["registros"], [])

    def test_registro_geografico_edicao_lotes_detecta_e_substitui_logradouro(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            with app_temp.app_context():
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    registro_geografico_core.ensure_schema(conn)
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Lote RG", 9101))
                    loc = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Lote RG'").fetchone()
                    agora = "2026-07-07T08:00:00"
                    cur_q = conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (loc["id_localidade"], loc["nome"], "0001", agora, agora),
                    )
                    for ordem, logradouro in enumerate(("R. São Gabriel", "Rua Sao Gabriel", "Rua Semelhante"), 1):
                        item = {
                            "localidade": loc["nome"],
                            "quarteirao": "0001",
                            "logradouro": logradouro,
                            "numero": str(ordem),
                            "sequencia": "",
                            "lado": "",
                            "tipo": "R",
                            "observacao": "",
                            "agentes_texto": "",
                        }
                        conn.execute(
                            """INSERT INTO registro_geografico_imoveis
                               (id_quarteirao, ordem, id_localidade, localidade, quarteirao,
                                logradouro, numero, tipo, busca_normalizada, chave_origem, criado_em, atualizado_em)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                cur_q.lastrowid,
                                ordem,
                                loc["id_localidade"],
                                loc["nome"],
                                "0001",
                                logradouro,
                                str(ordem),
                                "R",
                                registro_geografico_core._busca_normalizada(item),
                                f"teste-rg-lote-{ordem}",
                                agora,
                                agora,
                            ),
                        )
                    conn.commit()
                    loc_id = loc["id_localidade"]
                finally:
                    conn.close()

                similares = client.get(
                    f"/api/registro-geografico/logradouros-similares?localidade={loc_id}&score_min=78"
                )
                self.assertEqual(similares.status_code, 200)
                pares = similares.get_json()["pares"]
                encontrados = {
                    frozenset((par["a"]["logradouro"], par["b"]["logradouro"]))
                    for par in pares
                }
                self.assertIn(frozenset(("R. São Gabriel", "Rua Sao Gabriel")), encontrados)

                payload = {
                    "campo": "logradouro",
                    "modo": "exato",
                    "busca": "R. São Gabriel",
                    "novo": "Rua Sao Gabriel",
                    "filtros": {"localidade": [loc_id]},
                }
                preview = client.post("/api/registro-geografico/lote/preview", json=payload)
                self.assertEqual(preview.status_code, 200)
                self.assertEqual(preview.get_json()["total"], 1)

                aplicado = client.post("/api/registro-geografico/lote/aplicar", json=payload)
                self.assertEqual(aplicado.status_code, 200)
                self.assertEqual(aplicado.get_json()["atualizados"], 1)

                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    antigos = conn.execute(
                        "SELECT COUNT(*) FROM registro_geografico_imoveis WHERE logradouro=?",
                        ("R. São Gabriel",),
                    ).fetchone()[0]
                    novos = conn.execute(
                        "SELECT COUNT(*) FROM registro_geografico_imoveis WHERE logradouro=?",
                        ("Rua Sao Gabriel",),
                    ).fetchone()[0]
                    busca = conn.execute(
                        "SELECT busca_normalizada FROM registro_geografico_imoveis WHERE numero='1'"
                    ).fetchone()["busca_normalizada"]
                finally:
                    conn.close()
                self.assertEqual(antigos, 0)
                self.assertEqual(novos, 2)
                self.assertIn("rua sao gabriel", busca)

    def test_registro_geografico_sugere_e_limpa_quarteirao(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            with app_temp.app_context():
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                try:
                    registro_geografico_core.ensure_schema(conn)
                    conn.execute("INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)", ("Sugestao RG", 9301))
                    loc = conn.execute("SELECT id_localidade, nome FROM localidades WHERE nome='Sugestao RG'").fetchone()
                    agora = "2026-07-08T10:00:00"
                    cur_q1 = conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (loc["id_localidade"], loc["nome"], "0001", agora, agora),
                    )
                    cur_q2 = conn.execute(
                        """INSERT INTO registro_geografico_quarteiroes
                           (id_localidade, localidade, quarteirao, criado_em, atualizado_em)
                           VALUES (?, ?, ?, ?, ?)""",
                        (loc["id_localidade"], loc["nome"], "0002", agora, agora),
                    )
                    for ordem, (id_q, q, logradouro) in enumerate(
                        (
                            (cur_q1.lastrowid, "0001", "Rua Sao Gabriel"),
                            (cur_q1.lastrowid, "0001", "Rua Sao Gabriel"),
                            (cur_q2.lastrowid, "0002", "Rua Outra"),
                        ),
                        1,
                    ):
                        item = {
                            "localidade": loc["nome"],
                            "quarteirao": q,
                            "logradouro": logradouro,
                            "numero": str(ordem),
                            "sequencia": "",
                            "lado": "",
                            "tipo": "R",
                            "observacao": "",
                            "agentes_texto": "",
                        }
                        conn.execute(
                            """INSERT INTO registro_geografico_imoveis
                               (id_quarteirao, ordem, id_localidade, localidade, quarteirao,
                                logradouro, numero, tipo, busca_normalizada, chave_origem, criado_em, atualizado_em)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                id_q,
                                ordem,
                                loc["id_localidade"],
                                loc["nome"],
                                q,
                                logradouro,
                                str(ordem),
                                "R",
                                registro_geografico_core._busca_normalizada(item),
                                f"teste-rg-sugestao-{ordem}",
                                agora,
                                agora,
                            ),
                        )
                    conn.commit()
                    loc_id = loc["id_localidade"]
                finally:
                    conn.close()

                sugestoes = client.get(f"/api/registro-geografico/logradouros-sugestoes?q=sao%20gab&localidade={loc_id}")
                self.assertEqual(sugestoes.status_code, 200)
                dados_sugestoes = sugestoes.get_json()
                self.assertGreaterEqual(dados_sugestoes["total"], 1)
                self.assertEqual(dados_sugestoes["sugestoes"][0]["logradouro"], "Rua Sao Gabriel")
                self.assertTrue(dados_sugestoes["sugestoes"][0]["mesma_localidade"])

                limpar = client.post(
                    "/api/registro-geografico/quarteirao/limpar",
                    json={"id_localidade": loc_id, "quarteirao": "0001"},
                )
                self.assertEqual(limpar.status_code, 200)
                dados_limpar = limpar.get_json()["quarteirao"]
                self.assertEqual(dados_limpar["removidos"], 2)
                self.assertEqual(dados_limpar["registros"], [])

                q1 = client.get(f"/api/registro-geografico/quarteirao?localidade={loc_id}&quarteirao=0001")
                self.assertEqual(q1.status_code, 200)
                self.assertEqual(q1.get_json()["registros"], [])

                q2 = client.get(f"/api/registro-geografico/quarteirao?localidade={loc_id}&quarteirao=0002")
                self.assertEqual(q2.status_code, 200)
                self.assertEqual(len(q2.get_json()["registros"]), 1)

                excluir = client.post(
                    "/api/registro-geografico/quarteirao/excluir",
                    json={"id_localidade": loc_id, "quarteirao": "0002"},
                )
                self.assertEqual(excluir.status_code, 200)
                dados_excluir = excluir.get_json()["quarteirao"]
                self.assertEqual(dados_excluir["removidos"], 1)
                self.assertEqual(dados_excluir["quarteirao"], "2")

                lista = client.get(f"/api/registro-geografico/quarteiroes?localidade={loc_id}")
                self.assertEqual(lista.status_code, 200)
                quarteiroes = {item["quarteirao_raw"] for item in lista.get_json()["registros"]}
                self.assertIn("0001", quarteiroes)
                self.assertNotIn("0002", quarteiroes)

                conn = sqlite3.connect(db_path)
                try:
                    self.assertEqual(
                        conn.execute(
                            "SELECT COUNT(*) FROM registro_geografico_quarteiroes WHERE id_localidade=? AND quarteirao=?",
                            (loc_id, "0002"),
                        ).fetchone()[0],
                        0,
                    )
                    self.assertEqual(
                        conn.execute(
                            "SELECT COUNT(*) FROM registro_geografico_imoveis WHERE id_localidade=? AND quarteirao=?",
                            (loc_id, "0002"),
                        ).fetchone()[0],
                        0,
                    )
                finally:
                    conn.close()

    def test_api_esporotricose_animais_retorna_detalhes(self):
        client = _client_logado()
        resp = client.get("/api/esporotricose/animais?feridas=Sim")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        self.assertIn("total", dados)
        self.assertIn("registros", dados)
        if dados["registros"]:
            self.assertIn("nome", dados["registros"][0])
            self.assertIn("morador", dados["registros"][0])
            self.assertIn("agentes_texto", dados["registros"][0])
            self.assertIn("busca_ferido_data", dados["registros"][0])

    def test_api_esporotricose_animais_filtra_motivo_atencao_e_data(self):
        client = _client_logado()
        resp = client.get(
            "/api/esporotricose/animais"
            "?prioritarios=1"
            "&d_ini=2026-06-02"
            "&d_fim=2026-06-02"
            "&especie=C%C3%A3o"
            "&especie=Gato"
            "&motivo_atencao=Ferida%20informada"
        )

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        self.assertIn("registros", dados)
        for registro in dados["registros"]:
            self.assertEqual(registro.get("data"), "2026-06-02")
            self.assertIn(registro.get("especie"), {"Cão", "Gato"})
            self.assertEqual(registro.get("motivo_atencao"), "Ferida informada")

    def test_api_esporotricose_visitas_retorna_agentes_e_busca(self):
        client = _client_logado()
        resp = client.get("/api/esporotricose/visitas?busca=Graziela")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        self.assertIn("total", dados)
        self.assertIn("registros", dados)
        if dados["registros"]:
            self.assertIn("agentes", dados["registros"][0])
            self.assertIn("morador", dados["registros"][0])

    def test_api_esporotricose_localidades_retorna_resumo(self):
        client = _client_logado()
        resp = client.get("/api/esporotricose/localidades")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        self.assertIn("registros", dados)
        if dados["registros"]:
            self.assertIn("localidade", dados["registros"][0])
            self.assertIn("visitas", dados["registros"][0])

    def test_api_esporotricose_dashboard_retorna_series(self):
        client = _client_logado()
        resp = client.get("/api/esporotricose/dashboard")

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.is_json)
        dados = resp.get_json()
        for chave in ("evolucao", "status", "especies", "ambiente", "localidades", "saude"):
            self.assertIn(chave, dados)

    def test_api_esporotricose_doentes_retorna_registros_e_status(self):
        client = _client_logado()
        resp_status = client.get("/api/esporotricose/doentes/status")
        self.assertEqual(resp_status.status_code, 200)
        status_dados = resp_status.get_json()
        self.assertIn("registros", status_dados)
        status_nomes = [r["nome"] for r in status_dados["registros"]]
        self.assertIn("Em tratamento", status_nomes)
        self.assertIn("Aguardando documentos", status_nomes)
        self.assertIn("Não é esporotricose", status_nomes)
        self.assertNotIn("Em Tratamento", status_nomes)
        self.assertNotIn("Não mandou documentos", status_nomes)

        resp = client.get("/api/esporotricose/doentes")
        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("total", dados)
        self.assertIn("registros", dados)
        self.assertIn("totais", dados)
        self.assertIn("capsulas_baixa_zoomed", dados["totais"])
        if dados["registros"]:
            self.assertIn("whatsapp_documentos", dados["registros"][0])
            self.assertIn("receitas", dados["registros"][0])
            self.assertIn("receita_pendente", dados["registros"][0])
            self.assertIn("entregas_zoomed_pendentes", dados["registros"][0])
        notificacoes = [r.get("ultima_notificacao") for r in dados["registros"] if r.get("ultima_notificacao")]
        self.assertEqual(notificacoes, sorted(notificacoes, reverse=True))

        resp_estoque = client.get("/api/esporotricose/doentes/estoque")
        self.assertEqual(resp_estoque.status_code, 200)
        estoque = resp_estoque.get_json()
        self.assertIn("totais", estoque)
        self.assertIn("movimentos", estoque)
        self.assertIn("faltantes_tratamento", estoque["totais"])

    def test_download_esporotricose_doentes_csv(self):
        client = _client_logado()
        resp = client.get("/esporotricose/doentes/casos.csv")

        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp.headers.get("Content-Type", ""))
        self.assertIn("Casos-esporotricose.csv", resp.headers.get("Content-Disposition", ""))
        texto = resp.data.decode("utf-8-sig")
        header = texto.splitlines()[0]
        self.assertIn("latitude", header)
        self.assertIn("longitude", header)
        self.assertIn("tutor", header)
        self.assertIn("animal", header)

    def test_paginas_esporotricose_doentes_renderizam_fluxo_proprio(self):
        client = _client_logado()
        novo = client.get("/esporotricose/doentes/novo")
        self.assertEqual(novo.status_code, 200)
        html_novo = novo.data.decode("utf-8")
        self.assertIn("Novo paciente", html_novo)
        self.assertIn("Pedido ZooMed", html_novo)
        self.assertIn("Data notificação", html_novo)
        self.assertIn('id="data_notificacao"', html_novo)
        self.assertNotIn('id="rec_data_notificacao"', html_novo)
        self.assertIn("Não é esporotricose", html_novo)

        dados = client.get("/api/esporotricose/doentes").get_json()
        if dados["registros"]:
            id_animal = dados["registros"][0]["id_animal_doente"]
            detalhe = client.get(f"/esporotricose/doentes/{id_animal}")
            self.assertEqual(detalhe.status_code, 200)
            html_detalhe = detalhe.data.decode("utf-8")
            self.assertIn("Dados do paciente", html_detalhe)
            self.assertIn("Data notificação", html_detalhe)
            self.assertIn("Receitas e entregas", html_detalhe)
            self.assertNotIn('id="rec-data-notificacao"', html_detalhe)
            self.assertIn("Entrega cadastrada na ZOOMED", html_detalhe)
            self.assertIn("whatsapp.svg", html_detalhe)
            self.assertIn("excluirReceita", html_detalhe)
            self.assertIn("Editar receita", html_detalhe)
            self.assertIn("salvarReceitaEdit", html_detalhe)
            self.assertIn("/api/esporotricose/doentes/receitas/", html_detalhe)
            detalhe_json = client.get(f"/api/esporotricose/doentes/{id_animal}").get_json()
            entregas = [
                entrega
                for receita in detalhe_json.get("receitas", [])
                for entrega in receita.get("entregas", [])
            ]
            if entregas:
                entrega = entregas[0]
                csrf_original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
                endemias_app.app.config["WTF_CSRF_ENABLED"] = False
                try:
                    resp_put = client.put(
                        f"/api/esporotricose/doentes/entregas/{entrega['id_entrega']}",
                        json={
                            "data_entrega": entrega.get("data_entrega"),
                            "quantidade": entrega.get("quantidade") or 1,
                            "baixa_zoomed": entrega.get("baixa_zoomed") or "Sim",
                        },
                    )
                    self.assertEqual(resp_put.status_code, 200)
                finally:
                    endemias_app.app.config["WTF_CSRF_ENABLED"] = csrf_original

    def test_excluir_receita_doente_remove_entregas_vinculadas(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "doentes_receitas.db"
            conn = db_core.connect(db_path)
            try:
                esporotricose_core.ensure_schema(conn)
            finally:
                conn.close()

            id_animal = esporotricose_core.salvar_doente(
                str(db_path),
                {
                    "nome": "Teste",
                    "especie": "Gato",
                    "sexo": "Fêmea",
                    "tutor": "Tutor",
                    "endereco": "Rua A",
                    "localidade": "Centro",
                    "status": "Em tratamento",
                    "pedido_zoomed": "Sim",
                },
            )
            id_receita = esporotricose_core.salvar_receita_doente(
                str(db_path),
                id_animal,
                {"data_receita": "2026-06-16", "capsulas_total": 30, "status": "Em tratamento"},
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path),
                id_receita,
                {"data_entrega": "2026-06-16", "quantidade": 30, "baixa_zoomed": "Sim"},
            )
            removido_de = esporotricose_core.excluir_receita_doente(str(db_path), id_receita)
            conn = db_core.connect(db_path)
            try:
                receitas = conn.execute("SELECT COUNT(*) FROM esporotricose_doentes_receitas").fetchone()[0]
                entregas = conn.execute("SELECT COUNT(*) FROM esporotricose_doentes_entregas").fetchone()[0]
                animais = conn.execute("SELECT COUNT(*) FROM esporotricose_doentes_animais").fetchone()[0]
            finally:
                conn.close()

        self.assertEqual(removido_de, id_animal)
        self.assertEqual(receitas, 0)
        self.assertEqual(entregas, 0)
        self.assertEqual(animais, 1)

    def test_receita_doente_exibe_saldo_e_nao_duplica_entregas_por_anexo(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "doentes_saldo.db"
            conn = db_core.connect(db_path)
            try:
                esporotricose_core.ensure_schema(conn)
            finally:
                conn.close()

            id_animal = esporotricose_core.salvar_doente(
                str(db_path),
                {
                    "nome": "Mimi",
                    "especie": "Gato",
                    "tutor": "Maria",
                    "endereco": "Rua A",
                    "localidade": "Centro",
                    "status": "Em tratamento",
                    "pedido_zoomed": "Sim",
                },
            )
            id_receita = esporotricose_core.salvar_receita_doente(
                str(db_path),
                id_animal,
                {
                    "data_receita": "2026-06-16",
                    "capsulas_total": 180,
                    "posologia": "2 capsulas ao dia",
                    "capsulas_por_dia": 2,
                    "status": "Em tratamento",
                },
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path),
                id_receita,
                {"data_entrega": "2026-06-16", "quantidade": 30, "baixa_zoomed": "Sim"},
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path),
                id_receita,
                {"data_entrega": "2026-07-01", "quantidade": 30, "baixa_zoomed": "Sim"},
            )
            conn = db_core.connect(db_path)
            try:
                agora = "2026-07-01T10:00:00"
                for idx in range(2):
                    conn.execute(
                        """INSERT INTO esporotricose_doentes_anexos
                           (id_animal_doente, id_receita, nome_original, nome_arquivo,
                            caminho_rel, mime_type, tamanho, criado_por, criado_em)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            id_animal,
                            id_receita,
                            f"receita-{idx}.pdf",
                            f"receita-{idx}.pdf",
                            f"teste/receita-{idx}.pdf",
                            "application/pdf",
                            10,
                            "teste",
                            agora,
                        ),
                    )
                conn.commit()
            finally:
                conn.close()

            lista = esporotricose_core.listar_doentes(str(db_path), {})["registros"][0]
            self.assertEqual(lista["capsulas_entregues"], 60)
            self.assertEqual(lista["capsulas_receitadas"], 180)
            self.assertEqual(lista["capsulas_restantes"], 120)
            self.assertEqual(lista["proxima_entrega"], "2026-07-16")

            csv_row = esporotricose_core.listar_doentes_csv(str(db_path), {})[0]
            self.assertEqual(csv_row["capsulas_entregues"], 60)
            self.assertEqual(csv_row["capsulas_receitadas"], 180)
            self.assertEqual(csv_row["capsulas_restantes"], 120)
            self.assertEqual(csv_row["proxima_entrega"], "2026-07-16")

            detalhe = esporotricose_core.obter_doente(str(db_path), id_animal)
            receita = detalhe["receitas"][0]
            self.assertEqual(receita["capsulas_entregues"], 60)
            self.assertEqual(receita["capsulas_restantes"], 120)
            self.assertEqual(receita["capsulas_por_dia"], 2)
            self.assertEqual(receita["entregas_count"], 2)
            self.assertEqual(receita["entregas_observacao"], "Duas entregas feitas")
            self.assertIn("faltam 120", receita["saldo_observacao"])

            id_animal_atualizado = esporotricose_core.atualizar_receita_doente(
                str(db_path),
                id_receita,
                {
                    "data_receita": "2026-07-01",
                    "capsulas_total": 210,
                    "posologia": "1 capsula ao dia",
                    "capsulas_por_dia": 1,
                    "status": "Em tratamento",
                    "observacoes": "Receita atualizada",
                },
            )
            self.assertEqual(id_animal_atualizado, id_animal)
            receita_atualizada = esporotricose_core.obter_doente(str(db_path), id_animal)["receitas"][0]
            self.assertEqual(receita_atualizada["data_receita"], "2026-07-01")
            self.assertEqual(receita_atualizada["capsulas_total"], 210)
            self.assertEqual(receita_atualizada["capsulas_por_dia"], 1)
            self.assertEqual(receita_atualizada["capsulas_restantes"], 150)
            self.assertIn("faltam 150", receita_atualizada["saldo_observacao"])

            with self.assertRaises(esporotricose_core.ValidationError):
                esporotricose_core.salvar_entrega_doente(
                    str(db_path),
                    id_receita,
                    {"data_entrega": "2026-07-01", "quantidade": -30, "baixa_zoomed": "Sim"},
                )

    def test_receita_pendente_com_entrega_nao_gera_excedente_falso(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "doentes_receita_pendente.db"
            conn = db_core.connect(db_path)
            try:
                esporotricose_core.ensure_schema(conn)
            finally:
                conn.close()

            id_animal = esporotricose_core.salvar_doente(
                str(db_path),
                {
                    "nome": "Lobo",
                    "especie": "Gato",
                    "tutor": "Calina",
                    "endereco": "Rua A",
                    "localidade": "São João Batista",
                    "status": "Em tratamento",
                },
            )
            id_receita = esporotricose_core.salvar_receita_doente(
                str(db_path),
                id_animal,
                {"status": "Em tratamento", "receita_pendente": True},
            )
            esporotricose_core.salvar_entrega_doente(
                str(db_path),
                id_receita,
                {"data_entrega": "2026-07-07", "quantidade": 15, "baixa_zoomed": "Não"},
            )

            lista = esporotricose_core.listar_doentes(str(db_path), {})["registros"][0]
            self.assertTrue(lista["receita_pendente"])
            self.assertEqual(lista["capsulas_receitadas"], 0)
            self.assertEqual(lista["capsulas_entregues"], 15)
            self.assertEqual(lista["capsulas_restantes"], 0)
            self.assertEqual(lista["capsulas_excedentes"], 0)
            self.assertEqual(lista["ultima_receita"], None)
            self.assertEqual(lista["ultima_entrega"], "2026-07-07")
            self.assertEqual(lista["proxima_entrega"], "2026-07-22")

            detalhe = esporotricose_core.obter_doente(str(db_path), id_animal)
            receita = detalhe["receitas"][0]
            self.assertEqual(receita["receita_pendente"], 1)
            self.assertEqual(receita["capsulas_excedentes"], 0)
            self.assertIn("receita pendente", receita["saldo_observacao"])

    def test_consultas_sispncd_nao_alteram_coluna_sispncd(self):
        client = _client_logado()
        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            antes = conn.execute(
                "SELECT COUNT(*), COUNT(SISPNCD), COUNT(DISTINCT SISPNCD) FROM visitas"
            ).fetchone()
        finally:
            conn.close()

        resp = client.get("/api/sispncd/pesquisar?ano=2026&semana=20&tipo=TB/TBO&tipo=PVE&tipo=PE")
        self.assertEqual(resp.status_code, 200)

        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            depois = conn.execute(
                "SELECT COUNT(*), COUNT(SISPNCD), COUNT(DISTINCT SISPNCD) FROM visitas"
            ).fetchone()
        finally:
            conn.close()

        self.assertEqual(antes, depois)

    def test_conta_ovos_filtra_status_pendente_e_salva_status(self):
        client = _client_logado("admin")
        padrao = sispncd_core.get_default_conta_ovos(endemias_app.DB_PATH)

        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            pendentes_antes = conn.execute(
                "SELECT COUNT(*) FROM visitas WHERE tipo='TBO' AND CONTAOVOS_STATUS=0"
            ).fetchone()[0]
            ja_registrados = conn.execute(
                "SELECT COUNT(*) FROM visitas WHERE tipo='TBO' AND CONTAOVOS_STATUS=1 AND data=? AND quarteirao=?",
                (padrao["data"], padrao["quarteirao"]),
            ).fetchone()[0]
            esperado_periodo = conn.execute(
                """SELECT COUNT(*) FROM visitas
                   WHERE tipo='TBO'
                     AND CONTAOVOS_STATUS=0
                     AND data=?
                     AND quarteirao=?
                     AND LOWER(COALESCE(visita,'')) IN ('normal','recuperado','fechado','recusa')""",
                (padrao["data"], padrao["quarteirao"]),
            ).fetchone()[0]
        finally:
            conn.close()

        resp = client.get(f"/api/conta-ovos?data={padrao['data']}&quarteirao={padrao['quarteirao']}")
        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertEqual(dados["total_visitas"], esperado_periodo)
        if ja_registrados:
            self.assertLessEqual(dados["total_visitas"], pendentes_antes)

        with tempfile.TemporaryDirectory() as tmpdir:
            db_tmp = Path(tmpdir) / "endemias.db"
            shutil.copy2(endemias_app.DB_PATH, db_tmp)
            original_db = endemias_app.app.config["DB_PATH"]
            csrf_original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
            endemias_app.app.config["DB_PATH"] = str(db_tmp)
            endemias_app.app.config["WTF_CSRF_ENABLED"] = False
            try:
                padrao_tmp = sispncd_core.get_default_conta_ovos(str(db_tmp))
                conn = sqlite3.connect(db_tmp)
                esperado_salvar = conn.execute(
                    """SELECT COUNT(*) FROM visitas
                       WHERE tipo='TBO' AND CONTAOVOS_STATUS=0 AND data=? AND quarteirao=?""",
                    (padrao_tmp["data"], padrao_tmp["quarteirao"]),
                ).fetchone()[0]
                conn.close()
                if esperado_salvar == 0:
                    self.skipTest("Banco atual nao possui pendencias TBO para salvar status de conta ovos.")

                client_tmp = _client_logado("admin")
                salvar = client_tmp.post(
                    "/api/conta-ovos/salvar-status",
                    json={"data": padrao_tmp["data"], "quarteirao": padrao_tmp["quarteirao"]},
                )
                self.assertEqual(salvar.status_code, 200)
                self.assertEqual(salvar.get_json()["atualizados"], esperado_salvar)

                conn = sqlite3.connect(db_tmp)
                restantes = conn.execute(
                    """SELECT COUNT(*) FROM visitas
                       WHERE tipo='TBO' AND CONTAOVOS_STATUS=0 AND data=? AND quarteirao=?""",
                    (padrao_tmp["data"], padrao_tmp["quarteirao"]),
                ).fetchone()[0]
                conn.close()
                self.assertEqual(restantes, 0)
            finally:
                endemias_app.app.config["DB_PATH"] = original_db
                endemias_app.app.config["WTF_CSRF_ENABLED"] = csrf_original

    def test_sispncd_salva_codigo_sem_sobrescrever_existentes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_tmp = Path(tmpdir) / "endemias.db"
            shutil.copy2(endemias_app.DB_PATH, db_tmp)
            original_db = endemias_app.app.config["DB_PATH"]
            csrf_original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
            endemias_app.app.config["DB_PATH"] = str(db_tmp)
            endemias_app.app.config["WTF_CSRF_ENABLED"] = False
            try:
                conn = sqlite3.connect(db_tmp)
                row = conn.execute(
                    """SELECT data, tipo
                        FROM visitas
                        WHERE SISPNCD IS NULL OR TRIM(SISPNCD)=''
                        ORDER BY data DESC
                        LIMIT 1"""
                ).fetchone()
                conn.close()
                self.assertIsNotNone(row)
                ano, semana = sispncd_core.epidemiological_week_for_date(row[0])

                client = _client_logado("admin")
                codigo = "9999/2026"
                salvar = client.post(
                    "/api/sispncd/salvar",
                    json={"ano": ano, "semana": semana, "tipo": [row[1]], "codigo": codigo},
                )
                self.assertEqual(salvar.status_code, 200)
                atualizados = salvar.get_json()["atualizados"]
                self.assertGreater(atualizados, 0)

                conn = sqlite3.connect(db_tmp)
                try:
                    gravados = conn.execute(
                        "SELECT COUNT(*) FROM visitas WHERE SISPNCD=?",
                        (codigo,),
                    ).fetchone()[0]
                    pendentes_restantes_no_codigo = conn.execute(
                        "SELECT COUNT(*) FROM visitas WHERE SISPNCD IS NULL OR TRIM(SISPNCD)=''"
                    ).fetchone()[0]
                finally:
                    conn.close()
                self.assertEqual(gravados, atualizados)
                self.assertGreaterEqual(pendentes_restantes_no_codigo, 0)
            finally:
                endemias_app.app.config["DB_PATH"] = original_db
                endemias_app.app.config["WTF_CSRF_ENABLED"] = csrf_original

    def test_api_pendencias_conta_ovos_sispncd_retorna_resumo(self):
        client = _client_logado()
        resp = client.get("/api/conta-ovos-sispncd/pendencias")

        self.assertEqual(resp.status_code, 200)
        dados = resp.get_json()
        self.assertIn("conta_ovos", dados)
        self.assertIn("sispncd", dados)
        self.assertGreaterEqual(dados["sispncd"]["total"], 0)
        for grupo in dados["sispncd"]["grupos"]:
            self.assertIn("ano", grupo)
            self.assertIn("semana", grupo)
            self.assertIn("id_localidade", grupo)
        self.assertIn("total", dados["conta_ovos"])
        self.assertIn("grupos", dados["sispncd"])

        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            esperado_graziela = conn.execute(
                """SELECT COUNT(*) FROM visitas
                   WHERE tipo='TBO'
                     AND localidade='Graziela'
                     AND data='2026-05-04'
                     AND CONTAOVOS_STATUS=0"""
            ).fetchone()[0]
        finally:
            conn.close()

        if esperado_graziela:
            grupos = dados["conta_ovos"]["grupos"]
            self.assertTrue(any(
                g["data"] == "2026-05-04" and g["localidade"] == "Graziela"
                for g in grupos
            ))
            self.assertTrue(all("id_localidade" in g for g in grupos))

    def test_sispncd_0000_nao_conta_como_pendente(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "sispncd.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.executescript("""
                    CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT);
                    CREATE TABLE visitas (
                        id_visita TEXT PRIMARY KEY,
                        tipo TEXT,
                        data TEXT,
                        quarteirao INTEGER,
                        localidade TEXT,
                        id_localidade INTEGER,
                        SISPNCD TEXT,
                        CONTAOVOS_STATUS INTEGER
                    );
                    INSERT INTO localidades(id_localidade, nome) VALUES (1, 'Centro');
                    INSERT INTO visitas VALUES
                        ('a', 'PE', '2026-05-03', 1, 'Centro', 1, '0000/0000', NULL),
                        ('b', 'PE', '2026-05-04', 1, 'Centro', 1, NULL, NULL),
                        ('c', 'PE', '2026-05-05', 1, 'Centro', 1, '', NULL);
                """)
                conn.commit()
            finally:
                conn.close()

            pendencias = sispncd_core.pendencias_envio(str(db_path))
            resultado = sispncd_core.salvar_sispncd(str(db_path), 2026, 18, ["PE"], "0123/2026", id_localidade=1)

            conn = sqlite3.connect(db_path)
            try:
                valores = dict(conn.execute("SELECT id_visita, SISPNCD FROM visitas").fetchall())
            finally:
                conn.close()

        self.assertEqual(pendencias["sispncd"]["total"], 2)
        self.assertEqual(resultado["atualizados"], 2)
        self.assertEqual(valores["a"], "0000/0000")
        self.assertEqual(valores["b"], "0123/2026")
        self.assertEqual(valores["c"], "0123/2026")

    def test_sispncd_usa_semana_epidemiologica_domingo_sabado(self):
        self.assertEqual(
            sispncd_core.epidemiological_week_range(2026, 1),
            ("2026-01-04", "2026-01-10"),
        )
        self.assertEqual(
            sispncd_core.epidemiological_week_range(2026, 18),
            ("2026-05-03", "2026-05-09"),
        )
        self.assertEqual(
            sispncd_core.epidemiological_week_range(2026, 52),
            ("2026-12-27", "2027-01-02"),
        )
        self.assertEqual(
            sispncd_core.epidemiological_week_for_date("2026-01-03"),
            (2025, 53),
        )
        self.assertEqual(
            sispncd_core.epidemiological_week_for_date("2026-01-04"),
            (2026, 1),
        )

    def test_sispncd_inclui_bri_pendente_e_salva_codigo(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "sispncd_bri.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.executescript("""
                    CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT);
                    CREATE TABLE visitas (
                        id_visita TEXT PRIMARY KEY,
                        tipo TEXT,
                        data TEXT,
                        quarteirao INTEGER,
                        localidade TEXT,
                        id_localidade INTEGER,
                        tipo_imovel TEXT,
                        visita TEXT,
                        SISPNCD TEXT,
                        CONTAOVOS_STATUS INTEGER
                    );
                    CREATE TABLE depositos_inspecionados (
                        id_visita TEXT,
                        tipo_deposito TEXT,
                        inspecionado INTEGER,
                        eliminado INTEGER,
                        tratado INTEGER,
                        qtd_carga REAL
                    );
                    CREATE TABLE tratamentos (
                        id INTEGER PRIMARY KEY,
                        id_visita TEXT,
                        tipo TEXT,
                        quantidade_carga REAL
                    );
                    CREATE TABLE coletas (
                        id_coleta TEXT PRIMARY KEY,
                        id_visita TEXT,
                        tipo_deposito TEXT
                    );
                    CREATE TABLE resultados_laboratorio (
                        id_coleta TEXT,
                        aegypt_larvas INTEGER,
                        aegypt_pupas INTEGER,
                        aegypt_exuvias INTEGER,
                        aegypt_adulto INTEGER,
                        albopictus_larvas INTEGER,
                        albopictus_pupas INTEGER,
                        albopictus_exuvias INTEGER,
                        albopictus_adulto INTEGER,
                        outra_larvas INTEGER,
                        outra_pupas INTEGER,
                        outra_exuvias INTEGER,
                        outra_adulto INTEGER
                    );
                    CREATE TABLE visita_agentes (
                        id_visita TEXT,
                        id_agente INTEGER
                    );
                    INSERT INTO localidades(id_localidade, nome) VALUES (1, 'Centro');
                """)
                bri_core.ensure_schema(conn)
                conn.executescript("""
                    INSERT INTO bri_registros (
                        id_bri, sispncd, data, id_localidade, localidade, destino_tratamento,
                        quantidade_carga, quantidade_carga_extra, origem_estrutura, processado_em
                    ) VALUES
                        ('bri1', NULL, '2026-05-24', 1, 'Centro', 'Ovitrampa', 10, 0, 'nova', '2026-05-24T08:00:00'),
                        ('bri2', '', '2026-05-25', 1, 'Centro', 'Ponto Estratégico', 20, 5, 'nova', '2026-05-25T08:00:00'),
                        ('bri3', '0000/0000', '2026-05-26', 1, 'Centro', 'Outro', 30, 0, 'nova', '2026-05-26T08:00:00');
                """)
                conn.commit()
            finally:
                conn.close()

            pendencias = sispncd_core.pendencias_envio(str(db_path))
            consulta = sispncd_core.sispncd(str(db_path), 2026, 21, ["BRI"], id_localidade=1)
            resultado = sispncd_core.salvar_sispncd(str(db_path), 2026, 21, ["BRI"], "0456/2026", id_localidade=1)

            conn = sqlite3.connect(db_path)
            try:
                valores = dict(conn.execute("SELECT id_bri, sispncd FROM bri_registros").fetchall())
            finally:
                conn.close()

        grupos_bri = [g for g in pendencias["sispncd"]["grupos"] if g["tipo"] == "BRI"]
        self.assertEqual(pendencias["sispncd"]["total"], 2)
        self.assertEqual(grupos_bri[0]["semana"], 21)
        self.assertEqual(consulta["dados_gerais"]["bri"]["registros"], 3)
        self.assertEqual(consulta["dados_gerais"]["bri"]["pendentes_sispncd"], 2)
        self.assertEqual(resultado["atualizados"], 2)
        self.assertEqual(resultado["bri_atualizados"], 2)
        self.assertEqual(valores["bri1"], "0456/2026")
        self.assertEqual(valores["bri2"], "0456/2026")
        self.assertEqual(valores["bri3"], "0000/0000")

    def test_sispncd_e_conta_ovos_normalizam_depositos_por_codigo(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "sispncd_depositos.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.executescript("""
                    CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT);
                    CREATE TABLE visitas (
                        id_visita TEXT PRIMARY KEY,
                        tipo TEXT,
                        data TEXT,
                        quarteirao INTEGER,
                        localidade TEXT,
                        id_localidade INTEGER,
                        tipo_imovel TEXT,
                        visita TEXT,
                        SISPNCD TEXT,
                        CONTAOVOS_STATUS INTEGER
                    );
                    CREATE TABLE depositos_inspecionados (
                        id_visita TEXT,
                        tipo_deposito TEXT,
                        inspecionado INTEGER,
                        eliminado INTEGER,
                        tratado INTEGER,
                        qtd_carga REAL
                    );
                    CREATE TABLE tratamentos (
                        id INTEGER PRIMARY KEY,
                        id_visita TEXT,
                        tipo TEXT,
                        quantidade_carga REAL
                    );
                    CREATE TABLE coletas (
                        id_coleta TEXT PRIMARY KEY,
                        id_visita TEXT,
                        codigo_deposito TEXT,
                        tipo_deposito TEXT
                    );
                    CREATE TABLE resultados_laboratorio (
                        id_coleta TEXT,
                        aegypt_larvas INTEGER,
                        aegypt_pupas INTEGER,
                        aegypt_exuvias INTEGER,
                        aegypt_adulto INTEGER,
                        albopictus_larvas INTEGER,
                        albopictus_pupas INTEGER,
                        albopictus_exuvias INTEGER,
                        albopictus_adulto INTEGER,
                        outra_larvas INTEGER,
                        outra_pupas INTEGER,
                        outra_exuvias INTEGER,
                        outra_adulto INTEGER
                    );
                    CREATE TABLE visita_agentes (
                        id_visita TEXT,
                        id_agente INTEGER
                    );
                    INSERT INTO localidades(id_localidade, nome) VALUES (1, 'Graziela');
                    INSERT INTO visitas (
                        id_visita, tipo, data, quarteirao, localidade, id_localidade,
                        tipo_imovel, visita, SISPNCD, CONTAOVOS_STATUS
                    ) VALUES
                        ('v_tbo', 'TBO', '2026-05-05', 10, 'Graziela', 1, 'Residência', 'Normal', NULL, 0),
                        ('v_tb', 'TB', '2026-05-05', 10, 'Graziela', 1, 'Terreno Baldio', 'Normal', NULL, 1);
                    INSERT INTO depositos_inspecionados (
                        id_visita, tipo_deposito, inspecionado, eliminado, tratado, qtd_carga
                    ) VALUES
                        ('v_tbo', 'Pneus', 3, 1, 0, 0),
                        ('v_tb', 'Garrafas, latas e lixo', 5, 2, 2, 3.5),
                        ('v_tb', 'HISTORICO', 0, 4, 4, 0);
                    INSERT INTO coletas (
                        id_coleta, id_visita, codigo_deposito, tipo_deposito
                    ) VALUES
                        ('c1', 'v_tbo', 'A1', 'Caixa d''agua'),
                        ('c2', 'v_tbo', NULL, 'Pneus'),
                        ('c3', 'v_tbo', 'B', 'Depósito móvel'),
                        ('c4', 'v_tbo', NULL, 'Vasos e pratos');
                    INSERT INTO resultados_laboratorio (
                        id_coleta,
                        aegypt_larvas, aegypt_pupas, aegypt_exuvias, aegypt_adulto,
                        albopictus_larvas, albopictus_pupas, albopictus_exuvias, albopictus_adulto,
                        outra_larvas, outra_pupas, outra_exuvias, outra_adulto
                    ) VALUES
                        ('c1', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
                        ('c2', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
                        ('c3', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
                        ('c4', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0);
                """)
                conn.commit()
            finally:
                conn.close()

            conta = sispncd_core.conta_ovos(str(db_path), "2026-05-05", 10, id_localidade=1)
            consulta = sispncd_core.sispncd(str(db_path), 2026, 18, ["TB/TBO"], id_localidade=1)

        self.assertEqual(conta["depositos"]["d1"]["quantidade"], 3)
        self.assertEqual(conta["depositos"]["d1"]["eliminado"], 1)
        self.assertEqual(consulta["dados_gerais"]["depositos"]["d1"], 3)
        self.assertEqual(consulta["dados_gerais"]["depositos"]["d2"], 5)
        self.assertEqual(consulta["dados_gerais"]["total_depositos_inspecionados"], 8)
        self.assertEqual(consulta["dados_gerais"]["total_eliminados"], 7)
        self.assertEqual(consulta["dados_gerais"]["imoveis_tratados"], 1)
        self.assertEqual(consulta["dados_gerais"]["total_tratados"], 2)
        self.assertEqual(
            consulta["dados_gerais"]["tratamentos"],
            [{"tipo": "Sem tipo", "quantidade": 2, "carga_kg": 3.5}],
        )
        depositos_lab = {
            item["tipo_deposito"]: item["quantidade"]
            for item in consulta["laboratorio"]["depositos_aegypti"]
        }
        self.assertEqual(depositos_lab, {"A1": 1, "B": 2, "D1": 1})
        self.assertEqual(consulta["laboratorio"]["quarteiroes_aegypti"], 1)
        self.assertEqual(consulta["laboratorio"]["quarteiroes_aegypti_lista"], [10])
        self.assertEqual(consulta["laboratorio"]["quarteiroes_albopictus_lista"], [])
        self.assertEqual(consulta["laboratorio"]["quarteiroes_ambas_lista"], [])
        self.assertNotIn("Pneus", depositos_lab)
        self.assertNotIn("Caixa d'agua", depositos_lab)

    def test_extrair_depositos_preserva_eliminados_sem_a1(self):
        col_eliminados = work_types.etl_fields_for("PE")["depositos_eliminados_col"]
        row = etl.pd.Series({
            "A1": 0,
            "A2": 0,
            "B": 4,
            col_eliminados: 2,
        })

        depositos = etl.extrair_depositos(row, "PE")

        self.assertEqual(len(depositos), 1)
        self.assertEqual(depositos[0]["tipo_deposito"], "B")
        self.assertEqual(depositos[0]["inspecionado"], 4)
        self.assertEqual(depositos[0]["eliminado"], 2)

    def test_reimportacao_depositos_preenche_eliminados_ausentes(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        try:
            conn.execute("""
                CREATE TABLE depositos_inspecionados (
                    id INTEGER PRIMARY KEY,
                    id_visita TEXT NOT NULL,
                    tipo_deposito TEXT NOT NULL,
                    inspecionado INTEGER,
                    eliminado INTEGER,
                    tratado INTEGER,
                    tipo_tratamento TEXT,
                    qtd_carga REAL,
                    UNIQUE(id_visita, tipo_deposito)
                )
            """)
            conn.execute("""
                INSERT INTO depositos_inspecionados
                    (id_visita, tipo_deposito, inspecionado, eliminado, tratado, tipo_tratamento, qtd_carga)
                VALUES ('v1', 'B', 4, NULL, NULL, NULL, NULL)
            """)

            etl._salvar_deposito_inspecionado(conn.cursor(), "v1", {
                "tipo_deposito": "B",
                "inspecionado": 4,
                "eliminado": 2,
                "tratado": None,
                "tipo_tratamento": None,
                "qtd_carga": None,
            })

            row = conn.execute("""
                SELECT inspecionado, eliminado
                  FROM depositos_inspecionados
                 WHERE id_visita='v1' AND tipo_deposito='B'
            """).fetchone()
            self.assertEqual(row["inspecionado"], 4)
            self.assertEqual(row["eliminado"], 2)
        finally:
            conn.close()

    def test_sispncd_tratamentos_somam_depositos_e_ignoram_linhas_vazias(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "sispncd_tratamentos.db"
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                conn.executescript("""
                    CREATE TABLE localidades (id_localidade INTEGER PRIMARY KEY, nome TEXT);
                    CREATE TABLE visitas (
                        id_visita TEXT PRIMARY KEY,
                        tipo TEXT,
                        data TEXT,
                        quarteirao INTEGER,
                        localidade TEXT,
                        id_localidade INTEGER,
                        tipo_imovel TEXT,
                        visita TEXT,
                        SISPNCD TEXT,
                        CONTAOVOS_STATUS INTEGER
                    );
                    CREATE TABLE depositos_inspecionados (
                        id_visita TEXT,
                        tipo_deposito TEXT,
                        inspecionado INTEGER,
                        eliminado INTEGER,
                        tratado INTEGER,
                        qtd_carga REAL
                    );
                    CREATE TABLE tratamentos (
                        id INTEGER PRIMARY KEY,
                        id_visita TEXT,
                        tipo TEXT,
                        quantidade_carga REAL,
                        qtd_depositos_tratados INTEGER
                    );
                    CREATE TABLE coletas (
                        id_coleta TEXT PRIMARY KEY,
                        id_visita TEXT,
                        tipo_deposito TEXT
                    );
                    CREATE TABLE resultados_laboratorio (
                        id_coleta TEXT,
                        aegypt_larvas INTEGER,
                        aegypt_pupas INTEGER,
                        aegypt_exuvias INTEGER,
                        aegypt_adulto INTEGER,
                        albopictus_larvas INTEGER,
                        albopictus_pupas INTEGER,
                        albopictus_exuvias INTEGER,
                        albopictus_adulto INTEGER,
                        outra_larvas INTEGER,
                        outra_pupas INTEGER,
                        outra_exuvias INTEGER,
                        outra_adulto INTEGER
                    );
                    CREATE TABLE visita_agentes (
                        id_visita TEXT,
                        id_agente INTEGER
                    );
                    INSERT INTO localidades(id_localidade, nome) VALUES (1, 'Graziela');
                    INSERT INTO visitas (
                        id_visita, tipo, data, quarteirao, localidade, id_localidade,
                        tipo_imovel, visita, SISPNCD, CONTAOVOS_STATUS
                    ) VALUES
                        ('v1', 'PVE', '2026-05-11', 10, 'Graziela', 1, 'Residência', 'Normal', NULL, 1),
                        ('v2', 'PVE', '2026-05-12', 11, 'Graziela', 1, 'Residência', 'Normal', NULL, 1),
                        ('v3', 'PVE', '2026-05-13', 12, 'Graziela', 1, 'Residência', 'Normal', NULL, 1),
                        ('v4', 'PVE', '2026-05-14', 13, 'Graziela', 1, 'Residência', 'Normal', NULL, 1);
                    INSERT INTO tratamentos (
                        id_visita, tipo, quantidade_carga, qtd_depositos_tratados
                    ) VALUES
                        ('v1', 'Natular DT', 1.5, 2),
                        ('v2', 'Natular DT', 1.0, 3),
                        ('v3', NULL, NULL, NULL),
                        ('v4', 'Natular DT', 0, 9);
                    INSERT INTO coletas (id_coleta, id_visita, tipo_deposito) VALUES
                        ('c1', 'v1', 'B'),
                        ('c2', 'v2', 'C'),
                        ('c3', 'v3', 'D1');
                    INSERT INTO resultados_laboratorio (
                        id_coleta,
                        aegypt_larvas, aegypt_pupas, aegypt_exuvias, aegypt_adulto,
                        albopictus_larvas, albopictus_pupas, albopictus_exuvias, albopictus_adulto,
                        outra_larvas, outra_pupas, outra_exuvias, outra_adulto
                    ) VALUES
                        ('c1', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
                        ('c2', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0),
                        ('c3', 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0);
                """)
                conn.commit()
            finally:
                conn.close()

            consulta = sispncd_core.sispncd(str(db_path), 2026, 19, ["PVE"], id_localidade=1)

        self.assertEqual(consulta["dados_gerais"]["imoveis_tratados"], 2)
        self.assertEqual(consulta["dados_gerais"]["total_tratados"], 5)
        self.assertEqual(
            consulta["dados_gerais"]["tratamentos"],
            [{"tipo": "Natular DT", "quantidade": 5, "carga_kg": 2.5}],
        )
        self.assertEqual(consulta["laboratorio"]["quarteiroes_aegypti_lista"], [10, 12])
        self.assertEqual(consulta["laboratorio"]["quarteiroes_albopictus_lista"], [11, 12])
        self.assertEqual(consulta["laboratorio"]["quarteiroes_ambas_lista"], [12])

    def test_conta_ovos_pendencias_sao_clicaveis_para_filtrar(self):
        client = _client_logado()
        resp = client.get("/conta-ovos-sispncd")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn("selecionarPendenciaContaOvos", html)
        self.assertIn("selecionarPendenciaSisPNCD", html)
        self.assertIn("data-localidade-id", html)
        self.assertIn("sis-bri", html)
        self.assertNotIn("sis-bri-registros", html)
        self.assertNotIn("sis-bri-pendentes", html)
        self.assertNotIn("sis-bri-carga", html)
        self.assertIn("sis-total-depositos-inspecionados", html)
        self.assertIn("Depósitos tratados", html)
        self.assertIn("await buscarContaOvos()", html)
        self.assertIn("await buscarSisPNCD()", html)

    def test_api_visitas_tem_campos_de_paginacao(self):
        client = _client_logado()
        resp = client.get("/api/visitas?pagina=1&por_pagina=5")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("total", dados)
        self.assertIn("total_paginas", dados)
        self.assertIn("pagina", dados)
        self.assertIn("registros", dados)

    def test_api_visitas_edita_dados_principais_e_agentes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, localidade, logradouro, processado_em)
                       VALUES ('v-edit', 'uuid-v-edit', 'PVE', '2026-06-19', 'Lamenha', 'Rua A', '2026-06-19T10:00:00')"""
                )
                conn.commit()
            finally:
                conn.close()

            resp = client.post("/api/visitas/v-edit/editar", json={
                "data": "2026-06-19",
                "hora_inicio": "09:15",
                "localidade": "grasiela",
                "quarteirao": "123",
                "logradouro": "Rua B",
                "numero": "45",
                "morador": "Maria",
                "tipo_imovel": "Residência",
                "visita": "Normal",
                "agentes": "viviane_1, cecon",
                "observacoes": "corrigido manualmente",
            })

            self.assertEqual(resp.status_code, 200)
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            try:
                visita = conn.execute("SELECT * FROM visitas WHERE id_visita='v-edit'").fetchone()
                agentes = [r[0] for r in conn.execute(
                    """SELECT a.nome FROM visita_agentes va
                       JOIN agentes a ON a.id_agente=va.id_agente
                      WHERE va.id_visita='v-edit' ORDER BY a.nome"""
                )]
            finally:
                conn.close()

            self.assertEqual(visita["localidade"], "Graziela")
            self.assertEqual(visita["quarteirao"], 123)
            self.assertEqual(visita["morador"], "Maria")
            self.assertEqual(agentes, ["Ceccon", "Viviane"])

    def test_api_visitas_expoe_e_filtra_dados_relacionados(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, localidade, logradouro, visita, processado_em)
                       VALUES ('v-completa', 'uuid-v-completa', 'PVE', '2026-07-20',
                               'Lamenha', 'Rua Completa', 'Normal', '2026-07-20T10:00:00')"""
                )
                conn.execute(
                    """INSERT INTO depositos_inspecionados
                       (id_visita, tipo_deposito, inspecionado, eliminado, tratado,
                        tipo_tratamento, qtd_carga)
                       VALUES ('v-completa', 'B', 3, 1, 2, 'Natular', 4)"""
                )
                conn.execute(
                    """INSERT INTO tratamentos
                       (id_visita, tipo, quantidade_carga, qtd_depositos_tratados)
                       VALUES ('v-completa', 'Natular', 4, 2)"""
                )
                conn.execute(
                    """INSERT INTO coletas
                       (id_coleta, id_visita, num_tubo, codigo_deposito, tipo_deposito, deposito_eliminado)
                       VALUES ('coleta-completa', 'v-completa', 'T-500', 'B1', 'B', 0)"""
                )
                conn.execute(
                    """INSERT INTO resultados_laboratorio
                       (id_coleta, num_tubo, data_coleta, data_leitura, aegypt_larvas)
                       VALUES ('coleta-completa', 'T-500', '2026-07-20', '2026-07-20', 2)"""
                )
                conn.commit()
            finally:
                conn.close()

            lista = client.get(
                "/api/visitas?localidade=Lamenha&deposito=B&tratamento=Natular&coleta=com&laboratorio=positivo"
            )
            self.assertEqual(lista.status_code, 200)
            dados_lista = lista.get_json()
            self.assertEqual(dados_lista["total"], 1)
            self.assertEqual(dados_lista["registros"][0]["id_visita"], "v-completa")
            self.assertEqual(dados_lista["registros"][0]["laboratorio_status"], "positivo")

            detalhe = client.get("/api/visitas/v-completa")
            self.assertEqual(detalhe.status_code, 200)
            dados = detalhe.get_json()
            self.assertEqual(dados["depositos"][0]["inspecionado"], 3)
            self.assertEqual(dados["tratamentos"][0]["tipo"], "Natular")
            self.assertEqual(dados["coletas"][0]["aegypt_larvas"], 2)

            edicao = client.post("/api/visitas/v-completa/editar", json={
                "data": "2026-07-20",
                "depositos": [{
                    "tipo_deposito": "A2", "inspecionado": 5, "eliminado": 2,
                    "tratado": 1, "tipo_tratamento": "Pyriproxyfen", "qtd_carga": 3,
                }],
                "tratamentos": [{
                    "tipo": "Pyriproxyfen", "quantidade_carga": 3,
                    "qtd_depositos_tratados": 1,
                }],
                "coletas": [{
                    "id_coleta": "coleta-completa", "num_tubo": "T-501",
                    "codigo_deposito": "A2-1", "tipo_deposito": "A2",
                    "deposito_eliminado": True,
                }],
            })
            self.assertEqual(edicao.status_code, 200)
            detalhe_editado = client.get("/api/visitas/v-completa").get_json()
            self.assertEqual(detalhe_editado["depositos"][0]["tipo_deposito"], "A2")
            self.assertEqual(detalhe_editado["tratamentos"][0]["tipo"], "Pyriproxyfen")
            self.assertEqual(detalhe_editado["coletas"][0]["num_tubo"], "T-501")
            self.assertEqual(detalhe_editado["coletas"][0]["aegypt_larvas"], 2)

    def test_api_laboratorio_tem_campos_de_paginacao(self):
        client = _client_logado()
        resp = client.get("/api/laboratorio?pagina=1&por_pagina=5")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("total", dados)
        self.assertIn("total_paginas", dados)
        self.assertIn("pagina", dados)
        self.assertIn("registros", dados)

    def test_api_mapa_expoe_tipos_dinamicos(self):
        client = _client_logado()
        resp = client.get("/api/mapa")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIsInstance(dados, dict)
        if dados:
            primeiro = next(iter(dados.values()))
            self.assertIn("tipos", primeiro)
            self.assertIsInstance(primeiro["tipos"], dict)
            self.assertIn("esporo_visitas", primeiro)
            self.assertIn("esporo_animais", primeiro)
            self.assertIn("esporo_feridas", primeiro)
            self.assertIn("pes_ativos", primeiro)
            self.assertIn("pes_atrasados", primeiro)
            self.assertIn("ovi_armadilhas", primeiro)
            self.assertIn("ovi_positivas", primeiro)
            self.assertIn("ovi_ovos", primeiro)

    def test_api_mapa_ovitrampas_retorna_pontos_com_coordenadas(self):
        client = _client_logado()
        resp = client.get("/api/mapa/ovitrampas")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("resumo", dados)
        self.assertIn("pontos", dados)
        self.assertIn("armadilhas", dados["resumo"])
        if dados["pontos"]:
            primeiro = dados["pontos"][0]
            self.assertIn("latitude", primeiro)
            self.assertIn("longitude", primeiro)
            self.assertIn("ovos", primeiro)
            self.assertIn("positivas", primeiro)

    def test_mapa_exibe_camadas_de_esporotricose(self):
        client = _client_logado()
        resp = client.get("/mapa")

        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode("utf-8")
        self.assertIn('data-modo="esporotricose"', html)
        self.assertIn('data-modo="pes"', html)
        self.assertIn('data-modo="ovitrampas"', html)
        self.assertIn('data-modo="atencao"', html)
        self.assertIn("ovi-map-visual", html)
        self.assertIn("/api/mapa/ovitrampas", html)
        self.assertIn("formatarDataIso", html)
        self.assertIn("kpi-esporo-visitas", html)
        self.assertIn("kpi-pes", html)
        self.assertIn("kpi-ovi-armadilhas", html)
        self.assertIn("kpi-esporo-feridas", html)

    def test_mapa_usa_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {"/mapa", "/api/mapa"}
        }

        self.assertEqual(endpoints["/mapa"], "mapa.page")
        self.assertEqual(endpoints["/api/mapa"], "mapa.api_mapa")

    def test_apis_consultas_usam_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {"/api/dashboard", "/api/laboratorio", "/api/visitas"}
        }

        self.assertEqual(endpoints["/api/dashboard"], "consultas.api_dashboard")
        self.assertEqual(endpoints["/api/laboratorio"], "consultas.api_laboratorio")
        self.assertEqual(endpoints["/api/visitas"], "consultas.api_visitas")

    def test_exportacoes_usam_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {
                "/api/visitas/exportar",
                "/api/notificacoes/exportar",
                "/api/laboratorio/exportar",
                "/saida/consolidados/status",
                "/saida/gerar-consolidados",
                "/saida/download/<tipo>",
            }
        }

        self.assertEqual(endpoints["/api/visitas/exportar"], "exportacoes.exportar_visitas")
        self.assertEqual(endpoints["/api/notificacoes/exportar"], "exportacoes.exportar_notificacoes")
        self.assertEqual(endpoints["/api/laboratorio/exportar"], "exportacoes.exportar_laboratorio")
        self.assertEqual(endpoints["/saida/consolidados/status"], "exportacoes.consolidados_status")
        self.assertEqual(endpoints["/saida/gerar-consolidados"], "exportacoes.gerar_consolidados")
        self.assertEqual(endpoints["/saida/download/<tipo>"], "exportacoes.saida_download")

    def test_notificacoes_usam_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {
                "/notificacoes",
                "/notificacoes/foco/<id_foco>",
                "/notificacoes/foco/<id_foco>/atualizar",
                "/notificacoes/foco/<id_foco>/status",
                "/notificacoes/imprimir",
                "/notificacoes/foco/<id_foco>/imprimir-html",
                "/notificacoes/imprimir-html",
            }
        }

        self.assertEqual(endpoints["/notificacoes"], "notificacoes.page")
        self.assertEqual(endpoints["/notificacoes/foco/<id_foco>"], "notificacoes.foco_detalhe")
        self.assertEqual(
            endpoints["/notificacoes/foco/<id_foco>/atualizar"],
            "notificacoes.foco_atualizar",
        )
        self.assertEqual(
            endpoints["/notificacoes/foco/<id_foco>/status"],
            "notificacoes.foco_status_rapido",
        )
        self.assertEqual(endpoints["/notificacoes/imprimir"], "notificacoes.imprimir")
        self.assertEqual(
            endpoints["/notificacoes/foco/<id_foco>/imprimir-html"],
            "notificacoes.imprimir_html_single",
        )
        self.assertEqual(
            endpoints["/notificacoes/imprimir-html"],
            "notificacoes.imprimir_html_lote",
        )

    def test_auth_usa_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {"/login", "/logout", "/minha-senha"}
        }

        self.assertEqual(endpoints["/login"], "auth.login")
        self.assertEqual(endpoints["/logout"], "auth.logout")
        self.assertEqual(endpoints["/minha-senha"], "auth.minha_senha")

    def test_admin_auditoria_usa_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) in {"/admin/sistema", "/admin/auditoria", "/admin/auditoria/exportar"}
        }

        self.assertEqual(endpoints["/admin/sistema"], "admin.admin_sistema")
        self.assertEqual(endpoints["/admin/auditoria"], "admin.admin_auditoria")
        self.assertEqual(endpoints["/admin/auditoria/exportar"], "admin.admin_auditoria_exportar")

    def test_home_usa_blueprint_proprio(self):
        endpoints = {
            str(rule): rule.endpoint
            for rule in endemias_app.app.url_map.iter_rules()
            if str(rule) == "/"
        }

        self.assertEqual(endpoints["/"], "home.page")

    def test_exportacoes_retornam_xlsx(self):
        client = _client_logado()
        rotas = [
            "/api/visitas/exportar?d_ini=2099-01-01&d_fim=2099-01-02",
            "/api/notificacoes/exportar?d_ini=2099-01-01&d_fim=2099-01-02",
            "/api/laboratorio/exportar?d_ini=2099-01-01&d_fim=2099-01-02",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                try:
                    self.assertEqual(resp.status_code, 200)
                    self.assertIn(
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        resp.content_type,
                    )
                    self.assertTrue(resp.data.startswith(b"PK"))
                finally:
                    resp.close()

    def test_exportacao_xlsx_escapa_formulas(self):
        rows = [{"campo": "=HYPERLINK(\"http://exemplo\")"}]
        with endemias_app.app.test_request_context("/"):
            from blueprints.exportacoes import _gerar_xlsx

            resposta = _gerar_xlsx(["Campo"], rows, "teste")
            resposta.direct_passthrough = False
            data = resposta.get_data()

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        try:
            self.assertEqual(wb.active["A2"].value, "'=HYPERLINK(\"http://exemplo\")")
        finally:
            wb.close()

    def test_boletim_mensal_api_retorna_indicadores(self):
        client = _client_logado()
        resp = client.get("/api/boletim-mensal?mes=2026-06&modo=auto")
        dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("periodo", dados)
        self.assertIn("linhas", dados)
        self.assertIn("total", dados)
        chaves = {item["chave"] for item in dados["linhas"]}
        self.assertIn("visitas_pve", chaves)
        self.assertIn("visitas_tbo", chaves)
        texto = "\n".join(item["indicador"] for item in dados["linhas"])
        for esperado in ("denúncias", "transmissão", "imóveis", "Depósitos", "laboratório", "louças", "plásticos", "vigilância", "saúde", "reclamações"):
            self.assertIn(esperado, texto)
        for proibido in ("denuncias", "transmissao", "imoveis", "Depositos", "laboratorio", "loucas", "plasticos", "vigilancia", "saude", "reclamacoes", "Ã", "â€", "�"):
            self.assertNotIn(proibido, texto)

    def test_boletim_mensal_soma_acoes_do_setor_educativas_e_limpezas(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            _, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("INSERT INTO agentes (nome, ativo) VALUES (?, 1)", ("Agente Boletim",))
                conn.execute("INSERT INTO localidades (nome) VALUES (?)", ("Centro",))
                conn.commit()
                id_agente = conn.execute(
                    "SELECT id_agente FROM agentes WHERE nome=?",
                    ("Agente Boletim",),
                ).fetchone()[0]
            finally:
                conn.close()

            base = {
                "periodo": "manha",
                "hora_inicio": "09:00",
                "agentes": [id_agente],
                "localidade": "Centro",
                "local": "Unidade de Saude",
            }
            educativa = client.post(
                "/api/acoes-setor",
                json={**base, "tipo": "educativa", "data": "2099-10-05"},
            )
            limpeza = client.post(
                "/api/acoes-setor",
                json={**base, "tipo": "limpeza", "data": "2099-10-12"},
            )

            self.assertEqual(educativa.status_code, 201)
            self.assertEqual(limpeza.status_code, 201)

            resp = client.get("/api/boletim-mensal?mes=2099-10&modo=auto")
            dados = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        linha = next(item for item in dados["linhas"] if item["chave"] == "acoes_setor_total")
        self.assertEqual(linha["quantidade"], 2)
        self.assertEqual(linha["unidade"], "a\u00e7\u00f5es")
        self.assertIn("educativas e limpezas", linha["indicador"])

    def test_boletim_mensal_pagina_pdf_e_xlsx(self):
        client = _client_logado()
        rotas = [
            ("/boletim-mensal", "text/html"),
            ("/boletim-mensal/pdf?mes=2026-06", "text/html"),
            (
                "/api/boletim-mensal/exportar?mes=2026-06",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ]

        for rota, content_type in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                try:
                    self.assertEqual(resp.status_code, 200)
                    self.assertIn(content_type, resp.content_type)
                    if content_type == "text/html":
                        html = resp.data.decode("utf-8")
                        self.assertNotIn("Ã", html)
                        self.assertNotIn("â€", html)
                        self.assertNotIn("�", html)
                        self.assertIn("Boletim Mensal", html)
                finally:
                    resp.close()

    def test_boletim_mensal_salva_linha_manual(self):
        client = _client_logado("admin")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        chave = "manual_teste_unitario"
        try:
            resp = client.post(
                "/api/boletim-mensal",
                json={
                    "mes": "2099-12",
                    "linhas": [{
                        "chave": chave,
                        "origem": "manual",
                        "ordem": 10,
                        "indicador": "Ação educativa teste",
                        "quantidade": 3,
                        "unidade": "acoes",
                        "ativo": True,
                    }],
                },
            )
            dados = resp.get_json()
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original
            conn = sqlite3.connect(endemias_app.DB_PATH)
            try:
                conn.execute(
                    "DELETE FROM boletim_mensal_itens WHERE ano_mes=? AND chave=?",
                    ("2099-12", chave),
                )
                conn.commit()
            finally:
                conn.close()

        self.assertEqual(resp.status_code, 200)
        self.assertTrue(dados["ok"])
        self.assertEqual(dados["total"], 3)

    def test_boletim_mensal_normaliza_indicador_antigo_sem_acento(self):
        from app_core import boletim_mensal as boletim_core

        client = _client_logado()
        boletim_core.ensure_schema(endemias_app.DB_PATH)
        conn = sqlite3.connect(endemias_app.DB_PATH)
        try:
            conn.execute("""
                INSERT OR REPLACE INTO boletim_mensal_itens
                    (ano_mes, chave, origem, ordem, indicador, quantidade, unidade, ativo, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                "2099-11",
                "visitas_pve",
                "auto",
                10,
                "Vistorias de denuncias e focos suspeitos do Aedes aegypti (PVE - Pesquisa Vetorial Especial)",
                7,
                "visitas",
                1,
                "2099-11-01T08:00:00",
            ))
            conn.commit()
            resp = client.get("/api/boletim-mensal?mes=2099-11")
            dados = resp.get_json()
        finally:
            conn.execute(
                "DELETE FROM boletim_mensal_itens WHERE ano_mes=? AND chave=?",
                ("2099-11", "visitas_pve"),
            )
            conn.commit()
            conn.close()

        self.assertEqual(resp.status_code, 200)
        linha = next(item for item in dados["linhas"] if item["chave"] == "visitas_pve")
        self.assertIn("denúncias", linha["indicador"])
        self.assertNotIn("denuncias", linha["indicador"])

    def test_status_consolidados_retorna_tipos(self):
        client = _client_logado()
        resp = client.get("/saida/consolidados/status")
        data = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIn("tipos", data)
        self.assertEqual({item["tipo"] for item in data["tipos"]}, set(work_types.WORK_TYPE_CODES))

    def test_gerar_consolidados_chama_gerador_sob_demanda(self):
        client = _client_logado("admin")
        csrf_original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            with mock.patch("gerar_consolidado.gerar_todos", return_value=[
                {"tipo": "PE", "caminho": "saida/PE_consolidado.xlsx", "visitas": 1, "coletas": 2}
            ]) as gerar:
                resp = client.post("/saida/gerar-consolidados", json={"tipo": "PE"})
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = csrf_original

        data = resp.get_json()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(data["ok"])
        gerar.assert_called_once()
        self.assertEqual(gerar.call_args.kwargs["tipos"], ["PE"])

    def test_gerador_consolidado_cria_abas_visitas_e_coletas(self):
        import gerar_consolidado

        with tempfile.TemporaryDirectory() as tmpdir:
            conn = sqlite3.connect(endemias_app.DB_PATH)
            try:
                cur = conn.cursor()
                caminho, visitas, coletas = gerar_consolidado.gerar_xlsx_tipo(cur, "PE", tmpdir)
            finally:
                conn.close()

            if caminho is None:
                self.skipTest("Banco de teste sem dados PE para consolidado.")

            wb = openpyxl.load_workbook(caminho, read_only=True)
            try:
                self.assertIn("Visitas", wb.sheetnames)
                self.assertIn("Coletas", wb.sheetnames)
                self.assertGreaterEqual(visitas, 1)
                self.assertGreaterEqual(coletas, 1)
            finally:
                wb.close()

    def test_api_agenda_eventos_automaticos_sem_mojibake(self):
        client = _client_logado()
        resp = client.get("/api/agenda/eventos?start=2020-01-01&end=2035-12-31")
        eventos = resp.get_json()

        self.assertEqual(resp.status_code, 200)
        self.assertIsInstance(eventos, list)
        proibidos = ("Ã", "â€", "ðŸ", "\ufffd")
        for evento in eventos:
            props = evento.get("extendedProps", {})
            if props.get("origem") != "auto":
                continue
            textos = [
                evento.get("title", ""),
                props.get("resumo", ""),
                props.get("agentes", ""),
            ]
            for texto in textos:
                self.assertFalse(any(c in texto for c in proibidos), texto)

    def test_api_agenda_lembrete_invalido_retorna_400(self):
        client = _client_logado("admin")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp = client.post(
                "/api/agenda/eventos",
                json={"titulo": "Teste", "data_inicio": "2026-06-01", "lembrete_min": "abc"},
            )
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        self.assertEqual(resp.status_code, 400)
        self.assertTrue(resp.is_json)

    def test_api_agenda_tipo_invalido_retorna_400(self):
        client = _client_logado("admin")
        original = endemias_app.app.config.get("WTF_CSRF_ENABLED", True)
        endemias_app.app.config["WTF_CSRF_ENABLED"] = False
        try:
            resp = client.post(
                "/api/agenda/eventos",
                json={
                    "titulo": "Teste",
                    "tipo": "planejamento-invalido",
                    "data_inicio": "2026-06-01",
                },
            )
        finally:
            endemias_app.app.config["WTF_CSRF_ENABLED"] = original

        self.assertEqual(resp.status_code, 400)
        self.assertTrue(resp.is_json)

    def test_kobo_config_salva_sem_expor_token(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")

            resp = client.post("/api/kobo/config", json={
                "server_url": "kf.kobotoolbox.org/",
                "api_token": "segredo-token",
                "assets": {"PE": "asset-pe", "BRI": "asset-bri"},
            })

            self.assertEqual(resp.status_code, 200)
            dados = resp.get_json()
            self.assertTrue(dados["config"]["has_token"])
            self.assertNotIn("segredo-token", json.dumps(dados))
            salvo = kobo_api_core.load_config(app_temp.config["KOBO_CONFIG_PATH"])
            self.assertEqual(salvo["api_token"], "segredo-token")
            self.assertEqual(salvo["server_url"], "https://kf.kobotoolbox.org")
            self.assertEqual(salvo["assets"]["PE"], "asset-pe")
            self.assertEqual(salvo["assets"]["BRI"], "asset-bri")

    def test_kobo_previa_classifica_novos_e_duplicados(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({
                    "count": 2,
                    "results": [
                        {"_uuid": "uuid-existente", "_id": 1, "_submission_time": "2026-06-10T12:00:00"},
                        {"_uuid": "uuid-novo", "_id": 2, "_submission_time": "2026-06-10T13:00:00", "Localidade": "Centro"},
                    ],
                }).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"PE": "asset-pe"},
            })
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, processado_em)
                       VALUES (?, ?, ?, ?, ?)""",
                    ("v-existente", "uuid-existente", "PE", "2026-06-10", "2026-06-10T10:00:00"),
                )
                conn.commit()
            finally:
                conn.close()

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/previa", json={"tipo": "PE", "limite": 10})

            self.assertEqual(resp.status_code, 200)
            resumo = resp.get_json()["resumo"]
            self.assertEqual(resumo["total"], 2)
            self.assertEqual(resumo["novos"], 1)
            self.assertEqual(resumo["duplicados"], 1)
            self.assertEqual([item["status"] for item in resumo["amostra"]], ["duplicado", "novo"])
            self.assertEqual(resumo["amostra"][1]["detalhes"]["localidade"], "Sede")

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/previa", json={"tipo": "PE", "limite": 10, "apenas_novos": True})

            self.assertEqual(resp.status_code, 200)
            resumo = resp.get_json()["resumo"]
            self.assertEqual(resumo["total"], 2)
            self.assertEqual(resumo["novos"], 1)
            self.assertEqual(resumo["duplicados"], 1)
            self.assertEqual(resumo["amostra_filtro"], "novos")
            self.assertEqual([item["status"] for item in resumo["amostra"]], ["novo"])

    def test_kobo_validacao_de_morador_respeita_tipo_e_situacao_da_visita(self):
        base = {
            "_uuid": "uuid-validacao",
            "_submission_time": "2026-07-14T09:00:00",
            "Data": "2026-07-14",
            "Digite a hora": "09:00",
            "QuarteirÃ£o": "10",
            "NÃºmero": "20",
        }

        pe = kobo_api_core.summarize_submissions([base], tipo="PE")["amostra"][0]
        self.assertFalse(any("morador" in problema for problema in pe["problemas"]))

        fechado = kobo_api_core.summarize_submissions(
            [{**base, "_uuid": "uuid-fechado", "Visita": "fechado"}],
            tipo="TB",
        )["amostra"][0]
        self.assertFalse(any("morador" in problema for problema in fechado["problemas"]))

        recusa = kobo_api_core.summarize_submissions(
            [{**base, "_uuid": "uuid-recusa", "Visita": "recusa"}],
            tipo="PVE",
        )["amostra"][0]
        self.assertFalse(any("morador" in problema for problema in recusa["problemas"]))

        normal = kobo_api_core.summarize_submissions(
            [{**base, "_uuid": "uuid-normal", "Visita": "normal"}],
            tipo="TB",
        )["amostra"][0]
        self.assertTrue(any("morador" in problema for problema in normal["problemas"]))

    def test_larvas_resolvem_lote_com_data_divergente_sem_confundir_tubo_repetido(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        try:
            conn.executescript("""
                CREATE TABLE visitas (
                    id_visita TEXT PRIMARY KEY, tipo TEXT, data TEXT, localidade TEXT,
                    quarteirao INTEGER, logradouro TEXT, numero TEXT, morador TEXT,
                    tipo_imovel TEXT, observacoes TEXT
                );
                CREATE TABLE coletas (
                    id_coleta TEXT PRIMARY KEY, id_visita TEXT, num_tubo TEXT,
                    tipo_deposito TEXT
                );
                INSERT INTO visitas (id_visita, tipo, data) VALUES
                    ('visita-lote', 'PVE', '2026-03-31'),
                    ('visita-outra', 'PVE', '2026-06-18');
                INSERT INTO coletas (id_coleta, id_visita, num_tubo) VALUES
                    ('c693', 'visita-lote', '693'),
                    ('c674', 'visita-lote', '674'),
                    ('c360-lote', 'visita-lote', '360'),
                    ('c360-outra', 'visita-outra', '360');
            """)

            resolvidas = larvas_core.resolver_coletas(conn, {
                ("693", "2026-07-03"),
                ("674", "2026-07-03"),
                ("360", "2026-07-03"),
            })
        finally:
            conn.close()

        self.assertEqual(len(resolvidas), 3)
        self.assertEqual(resolvidas[("360", "2026-07-03")]["id_coleta"], "c360-lote")
        self.assertEqual(resolvidas[("360", "2026-07-03")]["estrategia"], "lote_de_tubos")

    def test_kobo_pendentes_resume_formularios_configurados(self):
        class FakeResponse:
            def __init__(self, payload):
                self.payload = payload
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps(self.payload).encode("utf-8")

        def fake_urlopen(req, timeout=30):
            url = req.full_url
            if "/asset-pe/" in url:
                return FakeResponse({"results": [
                    {"_uuid": "uuid-existente", "_id": 1, "_submission_time": "2026-07-10T12:00:00"},
                    {"_uuid": "uuid-novo", "_id": 2, "_submission_time": "2026-07-10T13:00:00"},
                ]})
            if "/asset-bri/" in url:
                return FakeResponse({"results": [
                    {"_uuid": "bri-novo", "_id": 3, "_submission_time": "2026-07-10T14:00:00"},
                ]})
            return FakeResponse({"results": []})

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"PE": "asset-pe", "BRI": "asset-bri"},
            })
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, processado_em)
                       VALUES (?, ?, ?, ?, ?)""",
                    ("v-existente", "uuid-existente", "PE", "2026-07-10", "2026-07-10T10:00:00"),
                )
                conn.commit()
            finally:
                conn.close()

            with mock.patch("app_core.kobo_api.request.urlopen", side_effect=fake_urlopen):
                resp = client.post("/api/kobo/pendentes", json={"limite": 10})

            self.assertEqual(resp.status_code, 200)
            dados = resp.get_json()
            por_tipo = {item["tipo"]: item for item in dados["itens"]}
            self.assertEqual(por_tipo["PE"]["novos"], 1)
            self.assertEqual(por_tipo["PE"]["duplicados"], 1)
            self.assertEqual(por_tipo["BRI"]["novos"], 1)
            self.assertEqual(dados["total_novos"], 2)

    def test_kobo_pendentes_aplica_vinculo_resolvido_as_larvas(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({"results": [{
                    "_uuid": "larva-pendente",
                    "_submission_time": "2026-07-03T17:01:50",
                    "NÃºmero do tubito": "693",
                    "Data da coleta": "2026-07-03",
                }]}).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"LARVAS": "asset-larvas"},
            })
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, processado_em)
                       VALUES (?, ?, ?, ?, ?)""",
                    ("visita-antiga", "visita-antiga-uuid", "PVE", "2026-03-31", "2026-07-07T10:00:00"),
                )
                conn.execute(
                    "INSERT INTO coletas (id_coleta, id_visita, num_tubo) VALUES (?, ?, ?)",
                    ("coleta-693", "visita-antiga", "693"),
                )
                conn.commit()
            finally:
                conn.close()

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/pendentes", json={"limite": 10})

            self.assertEqual(resp.status_code, 200)
            larvas = next(item for item in resp.get_json()["itens"] if item["tipo"] == "LARVAS")
            self.assertEqual(larvas["novos"], 1)
            self.assertEqual(larvas["pendencias"], 0)

    def test_kobo_api_marca_agentes_para_etl(self):
        record = {
            "_uuid": "uuid-agente",
            "_id": 33,
            "_submission_time": "2026-06-10T12:00:00",
            "Data": "2026-06-10",
            "Localidade": "Graziela",
            "Nome do(s) agente(s)": "ana_beatriz",
            "Nome do(s) agente(s)/m_arcio": "1",
        }
        cfg_tipo = {
            "col_data": "Data",
            "col_localidade": "Localidade",
            "prefixo_agente": "Nome do(s) agente(s)/",
            "col_sequencia": "Sequência",
        }

        detalhes = kobo_api_core.record_details("PVE", record)
        row = kobo_api_core._ensure_visit_columns({}, "PVE", cfg_tipo, record)

        self.assertEqual(detalhes["agentes"], "Ana Beatriz, Márcio")
        self.assertEqual(row["Nome do(s) agente(s)/Ana Beatriz"], 1)
        self.assertEqual(row["Nome do(s) agente(s)/Márcio"], 1)

    def test_kobo_api_pve_aproveita_campos_aninhados_para_etl(self):
        record = {
            "_uuid": "uuid-pve-nested",
            "_submission_time": "2026-06-19T12:00:00",
            "grupo_endereco": {
                "Localidade": "lamenha",
                "Quarteirão": "321",
                "Logradouro": "Rua das Flores",
                "Número": "55",
                "Morador": "Joana",
                "Tipo do imóvel": "Residência",
            },
            "grupo_visita": {
                "Hora": "08:40",
                "Visita": "Normal",
            },
        }
        cfg_tipo = {
            "col_data": "Data",
            "col_localidade": "Localidade",
            "prefixo_agente": "Nome do(s) agente(s)/",
            "col_sequencia": "Sequência",
        }

        detalhes = kobo_api_core.record_details("PVE", record)
        row = kobo_api_core._ensure_visit_columns({}, "PVE", cfg_tipo, record)

        self.assertEqual(detalhes["localidade"], "Lamenha")
        self.assertEqual(row["Quarteirão"], "321")
        self.assertEqual(row["Morador"], "Joana")
        self.assertEqual(row["Tipo do imóvel"], "Residência")
        self.assertEqual(row["Hora"], "08:40")

    def test_kobo_previa_normaliza_codigos_de_recolhimento(self):
        record = {
            "_uuid": "uuid-recolhimento",
            "_submission_time": "2026-06-17T18:29:15",
            "Data": "2026-06-17",
            "Localidade": "s_o_francisco",
            "Agentes": "cecon evaldo",
        }

        detalhes = kobo_api_core.record_details("RECOLHIMENTO", record)

        self.assertEqual(detalhes["localidade"], "S\u00e3o Francisco")
        self.assertEqual(detalhes["agentes"], "Ceccon, Evaldo")

    def test_etl_normaliza_nome_de_agente_ao_extrair(self):
        row = etl.pd.Series({
            "Nome do(s) agente(s)/m_arcio": 1,
            "Nome do(s) agente(s)/ana_beatriz": "1",
            "Nome do(s) agente(s)/Ceccon": 0,
        })

        nomes = etl.extrair_agentes(row, {"prefixo_agente": "Nome do(s) agente(s)/"})

        self.assertEqual(nomes, ["Márcio", "Ana Beatriz"])

    def test_kobo_previa_usa_tabela_do_modulo_extra(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({
                    "results": [
                        {"_uuid": "bri-existente", "_id": 10, "_submission_time": "2026-06-10T12:00:00"},
                        {"_uuid": "bri-novo", "_id": 11, "_submission_time": "2026-06-10T13:00:00"},
                    ],
                }).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"BRI": "asset-bri"},
            })
            conn = sqlite3.connect(db_path)
            try:
                conn.execute("CREATE TABLE IF NOT EXISTS bri_registros (id_bri TEXT PRIMARY KEY, kobo_uuid TEXT UNIQUE)")
                conn.execute("INSERT INTO bri_registros (id_bri, kobo_uuid) VALUES (?, ?)", ("bri1", "bri-existente"))
                conn.commit()
            finally:
                conn.close()

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/previa", json={"tipo": "BRI", "limite": 10})

            self.assertEqual(resp.status_code, 200)
            resumo = resp.get_json()["resumo"]
            self.assertEqual(resumo["novos"], 1)
            self.assertEqual(resumo["duplicados"], 1)

    def test_kobo_previa_larvas_mostra_tubo_sem_visita(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({
                    "results": [
                        {
                            "_uuid": "larva-nova",
                            "_id": 20,
                            "_submission_time": "2026-06-10T12:00:00",
                            "Número do tubito": "123",
                            "Data da coleta": "2026-06-10",
                        },
                    ],
                }).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"LARVAS": "asset-larvas"},
            })

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/previa", json={"tipo": "LARVAS", "limite": 10})

            self.assertEqual(resp.status_code, 200)
            row = resp.get_json()["resumo"]["amostra"][0]
            self.assertEqual(row["detalhes"]["tubo"], "123")
            self.assertEqual(row["detalhes"]["vinculo_visita"], "pendente")
            self.assertIn("Tubo sem visita/coleta correspondente", row["problemas"][0])

    def test_kobo_lote_vetores_larvas_vincula_larva_no_proprio_lote(self):
        class FakeResponse:
            def __init__(self, payload):
                self.payload = payload
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps(self.payload).encode("utf-8")

        def fake_urlopen(req, timeout=30):
            url = req.full_url
            if "/asset-tbo/" in url:
                return FakeResponse({"results": [{
                    "_uuid": "visita-tbo-1",
                    "_id": 1,
                    "_submission_time": "2026-06-10T08:00:00",
                    "Data": "2026-06-10",
                    "Localidade": "Centro",
                    "coletas": [{"Número do tubito": "T-001"}],
                }]})
            if "/asset-larvas/" in url:
                return FakeResponse({"results": [{
                    "_uuid": "larva-1",
                    "_id": 2,
                    "_submission_time": "2026-06-10T12:00:00",
                    "Número do tubito": "T-001",
                    "Data da coleta": "2026-06-10",
                }]})
            return FakeResponse({"results": []})

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {
                    "PE": "asset-pe",
                    "TB": "asset-tb",
                    "TBO": "asset-tbo",
                    "PVE": "asset-pve",
                    "LARVAS": "asset-larvas",
                },
            })

            with mock.patch("app_core.kobo_api.request.urlopen", side_effect=fake_urlopen):
                resp = client.post("/api/kobo/lote-vetores-larvas", json={"limite": 100})

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data["tubos_lote"], 1)
            self.assertEqual(data["larvas_vinculadas_lote"], 1)
            self.assertEqual(data["larvas_pendentes"], 0)
            larva = data["resumos"]["LARVAS"]["amostra"][0]
            self.assertEqual(larva["detalhes"]["vinculo_visita"], "lote")
            self.assertEqual(larva["problemas"], [])

    def test_kobo_importacao_prepara_job_com_xlsx_temporarios(self):
        class FakeResponse:
            def __init__(self, payload):
                self.payload = payload
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps(self.payload).encode("utf-8")

        def fake_urlopen(req, timeout=30):
            url = req.full_url
            if "/asset-tbo/" in url:
                return FakeResponse({"results": [{
                    "_uuid": "visita-tbo-1",
                    "_id": 1,
                    "_submission_time": "2026-06-10T08:00:00",
                    "Data": "2026-06-10",
                    "Localidade": "Centro",
                    "Logradouro": "Rua A",
                    "Número": "10",
                    "Quarteirão": "5",
                    "Visita": "Normal",
                    "group_jr1vc40": [{
                        "N_mero_do_tubito": "T-002",
                        "C_digo_do_dep_sito": "D1",
                        "Dep_sito": "Pneu",
                        "O_Dep_sito_onde_foi_coleta_foi_eliminado": "sim",
                    }],
                    "coletas": [{"Número do tubito": "T-001"}],
                }]})
            if "/asset-larvas/" in url:
                return FakeResponse({"results": [{
                    "_uuid": "larva-1",
                    "_id": 2,
                    "_submission_time": "2026-06-10T12:00:00",
                    "Número do tubito": "T-001",
                    "Data da coleta": "2026-06-10",
                    "Aegypt Larvas": "1",
                }]})
            return FakeResponse({"results": []})

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            app_temp.config["UPLOAD_TEMP"] = str(Path(tmpdir) / "uploads_temp")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {
                    "PE": "asset-pe",
                    "TB": "asset-tb",
                    "TBO": "asset-tbo",
                    "PVE": "asset-pve",
                    "LARVAS": "asset-larvas",
                },
            })

            with mock.patch("app_core.kobo_api.request.urlopen", side_effect=fake_urlopen):
                resp = client.post("/api/kobo/importar-vetores-larvas/iniciar", json={"limite": 100})

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertTrue(data["ok"])
            self.assertEqual(data["por_tipo"]["TBO"], 1)
            self.assertEqual(data["por_tipo"]["LARVAS"], 1)
            job_dir = Path(app_temp.config["UPLOAD_TEMP"]) / data["job_id"]
            self.assertTrue((job_dir / data["arquivos"][0]).exists())
            self.assertTrue(any(nome.startswith("TBO_") for nome in data["arquivos"]))
            self.assertTrue(any(nome.startswith("LARVAS_") for nome in data["arquivos"]))
            tbo_path = next(job_dir / nome for nome in data["arquivos"] if nome.startswith("TBO_"))
            wb = openpyxl.load_workbook(tbo_path, read_only=True)
            try:
                self.assertIn("dados", wb.sheetnames)
                self.assertIn("coletas", wb.sheetnames)
                coletas = wb["coletas"]
                headers = [cell.value for cell in next(coletas.iter_rows(max_row=1))]
                values = [cell.value for cell in next(coletas.iter_rows(min_row=2, max_row=2))]
                row = dict(zip(headers, values))
                self.assertIn(row["Número do tubito"], {"T-001", "T-002"})
                self.assertEqual(row["submission__uuid"], "visita-tbo-1")
                rows = [
                    dict(zip(headers, [cell.value for cell in linha]))
                    for linha in coletas.iter_rows(min_row=2)
                ]
                config = json.loads((ROOT / "config.json").read_text(encoding="utf-8"))
                cfg_tbo = config["tipos_trabalho"]["TBO"]
                row_deposito = next(item for item in rows if item[cfg_tbo["col_numero_tubo_coletas"]] == "T-002")
                self.assertEqual(row_deposito[cfg_tbo["col_codigo_deposito_coletas"]], "D1")
                self.assertEqual(row_deposito[cfg_tbo["col_nome_deposito_coletas"]], "Pneu")
                self.assertEqual(row_deposito[cfg_tbo["col_deposito_eliminado_coletas"]], "sim")
            finally:
                wb.close()

            caminhos = [str(job_dir / nome) for nome in data["arquivos"]]
            arquivos_larvas = [c for c in caminhos if Path(c).name.upper().startswith("LARVAS")]
            arquivos_trabalho = [c for c in caminhos if c not in arquivos_larvas]
            ok, sumario = etl.processar_upload(
                arquivos_trabalho,
                arquivos_larvas,
                db_path,
                str(ROOT / "config.json"),
                etl.Logger(),
                dry_run=True,
            )
            self.assertTrue(ok)
            self.assertEqual(sumario[0]["tipo"], "TBO")
            self.assertEqual(sumario[0]["visitas_novas"], 1)
            self.assertEqual(sumario[0]["coletas_novas"], 2)
            self.assertEqual(sumario[0]["resultados_novos"], 1)

    def test_kobo_importacao_larvas_prepara_formulario_selecionado(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({"results": [{
                    "_uuid": "larva-api-1",
                    "_id": 2,
                    "_submission_time": "2026-06-10T12:00:00",
                    "start": "2026-06-11T08:00:00",
                    "group_we1tn02/Nome_do_laboratorista": "azimir",
                    "group_we1tn02/Data_leitura": "2026-06-11",
                    "group_we1tn02/Numero_tubo": "T-001",
                    "group_we1tn02/Data_da_coleta": "2026-06-10",
                    "group_nr95y76/group_mw55v39/Aegypt_Larvas": "1",
                    "group_nr95y76/group_pe3hs09/Outra_Esp_cie_Larvas": "3",
                }]}).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            app_temp.config["UPLOAD_TEMP"] = str(Path(tmpdir) / "uploads_temp")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"LARVAS": "asset-larvas"},
            })

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/importar-formulario/iniciar", json={"tipo": "LARVAS", "limite": 100})

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data["por_tipo"]["LARVAS"], 1)
            job_dir = Path(app_temp.config["UPLOAD_TEMP"]) / data["job_id"]
            self.assertEqual(len(data["arquivos"]), 1)
            self.assertTrue(data["arquivos"][0].startswith("LARVAS_"))
            caminho = job_dir / data["arquivos"][0]
            self.assertTrue(caminho.exists())
            wb = openpyxl.load_workbook(caminho, read_only=True)
            try:
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                values = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
                row = dict(zip(headers, values))
            finally:
                wb.close()
            self.assertEqual(str(row["Número do tubito"]), "T-001")
            self.assertEqual(str(row["Data da coleta"])[:10], "2026-06-10")
            self.assertEqual(str(row["Data da leitura"])[:10], "2026-06-11")
            self.assertEqual(str(row["Nome do laboratorista"]), "azimir")
            self.assertEqual(int(row["Aegypt Larvas"]), 1)
            config = json.loads((ROOT / "config.json").read_text(encoding="utf-8"))
            col_outra_larvas = next(
                col for col in config["larvas"]["colunas_resultado"]
                if "Outra" in col and "Larvas" in col
            )
            self.assertEqual(int(row[col_outra_larvas]), 3)

    def test_kobo_importacao_extra_prepara_formulario_bri(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({"results": [{
                    "_uuid": "bri-api-1",
                    "_id": 10,
                    "_submission_time": "2026-06-10T12:00:00",
                    "start": "2026-06-10T08:00:00",
                    "end": "2026-06-10T08:30:00",
                    "Digite a data": "2026-06-10",
                    "Digite a hora": "08:00",
                    "Onde vai ser realizado o tratamento?": "Outro",
                    "Localidade": "Centro",
                    "Logradouro": "Rua A",
                    "Número": "10",
                    "Quantidade de carga": "12",
                }]}).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, _ = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            app_temp.config["UPLOAD_TEMP"] = str(Path(tmpdir) / "uploads_temp")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"BRI": "asset-bri"},
            })

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/importar-formulario/iniciar", json={"tipo": "BRI", "limite": 100})

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data["por_tipo"]["BRI"], 1)
            job_dir = Path(app_temp.config["UPLOAD_TEMP"]) / data["job_id"]
            self.assertEqual(len(data["arquivos"]), 1)
            self.assertTrue(data["arquivos"][0].startswith("BRI_"))
            caminho = job_dir / data["arquivos"][0]
            self.assertTrue(caminho.exists())
            wb = openpyxl.load_workbook(caminho, read_only=True)
            try:
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                values = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
                row = dict(zip(headers, values))
                self.assertEqual(row["_uuid"], "bri-api-1")
                self.assertEqual(row["Digite a data"], "2026-06-10")
            finally:
                wb.close()

    def test_kobo_importacao_esporotricose_prepara_planilha_com_animais(self):
        class FakeResponse:
            def __enter__(self):
                return self
            def __exit__(self, exc_type, exc, tb):
                return False
            def read(self):
                return json.dumps({"results": [{
                    "_uuid": "esporo-api-1",
                    "_id": 20,
                    "_submission_time": "2026-06-10T12:00:00",
                    "start": "2026-06-10T08:00:00",
                    "end": "2026-06-10T08:30:00",
                    "group_su8jh28/Data": "2026-06-10",
                    "group_su8jh28/Hora_inicio": "08:00",
                    "Hora_fim": "08:30",
                    "group_su8jh28/Agentes": "Agente Teste",
                    "group_su8jh28/Localidade": "S_O_Jo_O_Batista",
                    "group_su8jh28/Quarteir_o": "12",
                    "group_su8jh28/Tipo_do_im_vel": "resid_ncia",
                    "group_su8jh28/Logradouro": "Rua A",
                    "group_su8jh28/N_mero": "10",
                    "group_su8jh28/Morador": "Pessoa Teste",
                    "group_su8jh28/Visita": "fechado",
                    "Deseja_cadastrar_um_animal": "n_o",
                    "animais": [{
                        "group_tl0nq13/Escolha_o_animal_a_ser_cadastr": "c_o",
                        "group_tl0nq13/Nome_do_animal": "Rex",
                        "group_tl0nq13/Sexo": "macho",
                        "group_tl0nq13/Classifica_o_quanto_em_que_o_animal_vive": "domiciliado",
                        "group_tl0nq13/Vacinado": "sim",
                        "group_tl0nq13/Castrado": "nao",
                        "group_tl0nq13/Apresenta_feridas_pelo_corpo": "nao",
                        "group_tl0nq13/Evolu_o_do_caso": "sem_tratamento",
                    }],
                }]}).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            app_temp.config["KOBO_CONFIG_PATH"] = str(Path(tmpdir) / "kobo_config.json")
            app_temp.config["UPLOAD_TEMP"] = str(Path(tmpdir) / "uploads_temp")
            kobo_api_core.save_config(app_temp.config["KOBO_CONFIG_PATH"], {
                "server_url": "https://kf.kobotoolbox.org",
                "api_token": "token",
                "assets": {"ESPOROTRICOSE": "asset-esporo"},
            })

            with mock.patch("app_core.kobo_api.request.urlopen", return_value=FakeResponse()):
                resp = client.post("/api/kobo/importar-formulario/iniciar", json={"tipo": "ESPOROTRICOSE", "limite": 100})

            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data["por_tipo"]["ESPOROTRICOSE"], 1)
            job_dir = Path(app_temp.config["UPLOAD_TEMP"]) / data["job_id"]
            caminho = job_dir / data["arquivos"][0]
            self.assertTrue(caminho.exists())
            wb = openpyxl.load_workbook(caminho, read_only=True)
            try:
                self.assertIn("dados", wb.sheetnames)
                self.assertIn("Dados do animal", wb.sheetnames)
            finally:
                wb.close()

            import etl

            ok, sumario = etl.processar_upload(
                [str(caminho)],
                [],
                db_path,
                str(ROOT / "config.json"),
                etl.Logger(),
                dry_run=True,
            )
            self.assertTrue(ok)
            self.assertEqual(sumario[0]["tipo"], "ESPOROTRICOSE")
            self.assertEqual(sumario[0]["visitas_novas"], 1)
            self.assertEqual(sumario[0]["animais_novos"], 1)
            visitas, _ = esporotricose_core.parse_workbook(str(caminho), "nova")
            self.assertEqual(visitas[0]["localidade"], "S\u00e3o Jo\u00e3o Batista")
            self.assertEqual(visitas[0]["tipo_imovel"], "Residência")
            self.assertEqual(visitas[0]["visita"], "Fechado")
            self.assertEqual(visitas[0]["deseja_cadastrar_animal"], "N\u00e3o")
            _, animais = esporotricose_core.parse_workbook(str(caminho), "nova")
            self.assertEqual(animais[0]["especie"], "Cão")
            self.assertEqual(animais[0]["sexo"], "Macho")
            self.assertEqual(animais[0]["ambiente"], "Domiciliado")
            self.assertEqual(animais[0]["vacinado"], "Sim")
            self.assertEqual(animais[0]["castrado"], "Não")
            self.assertEqual(animais[0]["feridas"], "Não")

            self.assertEqual(animais[0]["evolucao_caso"], "Sem tratamento")

    def test_etl_normaliza_codigos_tecnicos_da_api_kobo(self):
        import etl
        from app_core import normalizadores

        self.assertEqual(etl.normalizar_localidade("grasiela"), "Graziela")
        self.assertEqual(normalizadores.normalizar_localidade("lamenha"), "Lamenha")
        self.assertEqual(normalizadores.normalizar_localidade("Centro"), "Sede")
        self.assertEqual(normalizadores.normalizar_localidade("s_o_francisco"), "S\u00e3o Francisco")
        self.assertEqual(normalizadores.normalizar_localidade("S_O_Ven_Ncio"), "S\u00e3o Ven\u00e2ncio")
        self.assertEqual(normalizadores.normalizar_localidade("S_O_Jo_O_Batista"), "S\u00e3o Jo\u00e3o Batista")
        self.assertEqual(normalizadores.normalizar_localidade("para_so"), "Para\u00edso")
        self.assertEqual(normalizadores.normalizar_localidade("Tangu"), "Tangu\u00e1")
        self.assertEqual(etl.normalizar_categoria("resid_ncia"), "Residência")
        self.assertEqual(etl.normalizar_categoria("com_rcio"), "Comércio")
        self.assertEqual(etl.normalizar_categoria("fechado"), "Fechado")
        self.assertEqual(etl.normalizar_categoria("normal"), "Normal")

    def test_etl_reimporta_pe_editado_no_kobo_sem_duplicar_visita(self):
        uuid_antigo = "393d21e4-9a66-4877-88b9-3483ec4e5ee0"
        uuid_atual = "c796a9db-cbe2-4513-abec-5a92f6cdeec8"
        kobo_id = 804650646

        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            conn = db_core.connect(db_path)
            id_visita = etl.gerar_id_visita(uuid_antigo, None)
            conn.execute(
                """INSERT INTO visitas
                   (id_visita, kobo_uuid, kobo_id, tipo, data, localidade,
                    logradouro, numero, quarteirao, processado_em)
                   VALUES (?, ?, ?, 'PE', '2026-07-07', 'Lamenha',
                           'WADISLAU BUGALSKI - ORPEC', '4250', 148, ?)""",
                (id_visita, uuid_antigo, kobo_id, "2026-07-07T18:40:00"),
            )
            agente_antigo = etl.obter_ou_criar_agente(conn.cursor(), "Agente antigo")
            conn.execute(
                "INSERT INTO visita_agentes(id_visita, id_agente) VALUES (?, ?)",
                (id_visita, agente_antigo),
            )
            conn.commit()
            conn.close()

            record = {
                "_uuid": uuid_atual,
                "_id": kobo_id,
                "_submission_time": "2026-07-07T18:32:37",
                "Digite a data": "2026-07-07",
                "Nome_do_s_agente_s": "adilson ana_beatriz ceccon fernando",
                "group_zn1kq42/localidade": "Lamenha",
                "group_zn1kq42/logradouro": "RUA SANTOS DUMONT",
                "group_zn1kq42/quarteirao": "0257",
                "group_zn1kq42/numero": "630",
                "group_zn1kq42/Visita": "normal",
            }
            caminhos = kobo_api_core.write_etl_workbooks(
                {"PE": [record]}, str(ROOT / "config.json"), tmpdir, prefix="corrigido"
            )
            with mock.patch(
                "etl.pe_core.resolver_alias_visita",
                return_value={"id_pe": 29, "codigo_pe": "PE-0029"},
            ):
                logs = []
                ok, sumario = etl.processar_upload(
                    caminhos, [], db_path, str(ROOT / "config.json"),
                    etl.Logger(callback=lambda msg, tag: logs.append((tag, msg))),
                    dry_run=False, backup_confirmado=True,
                )

            self.assertTrue(ok, logs)
            self.assertEqual(sumario[0]["visitas_novas"], 0)
            self.assertEqual(sumario[0]["visitas_atualizadas"], 1)
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            visitas = conn.execute(
                "SELECT * FROM visitas WHERE tipo='PE' AND kobo_id=?", (kobo_id,)
            ).fetchall()
            self.assertEqual(len(visitas), 1)
            self.assertEqual(visitas[0]["id_visita"], id_visita)
            self.assertEqual(visitas[0]["kobo_uuid"], uuid_atual)
            self.assertEqual(visitas[0]["logradouro"], "RUA SANTOS DUMONT")
            self.assertEqual(visitas[0]["quarteirao"], 257)
            self.assertEqual(visitas[0]["codigo_pe"], "PE-0029")
            agentes = {
                row[0] for row in conn.execute(
                    """SELECT a.nome FROM visita_agentes va
                       JOIN agentes a ON a.id_agente=va.id_agente
                       WHERE va.id_visita=?""",
                    (id_visita,),
                )
            }
            conn.close()
            self.assertNotIn("Agente antigo", agentes)
            self.assertIn("Adilson", agentes)
            self.assertIn("Fernando", agentes)

            with mock.patch(
                "etl.pe_core.resolver_alias_visita",
                return_value={"id_pe": 29, "codigo_pe": "PE-0029"},
            ):
                ok, repeticao = etl.processar_upload(
                    caminhos, [], db_path, str(ROOT / "config.json"),
                    etl.Logger(), dry_run=False, backup_confirmado=True,
                )
            self.assertTrue(ok)
            self.assertEqual(repeticao[0]["visitas_novas"], 0)
            self.assertEqual(repeticao[0]["visitas_atualizadas"], 0)

    def test_etl_prepara_planilhas_antes_de_abrir_transacao(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            eventos = []
            original_connect = db_core.connect

            def conectar(*args, **kwargs):
                eventos.append("connect")
                return original_connect(*args, **kwargs)

            def preparar(*args, **kwargs):
                eventos.append("prepare")
                return {"preparado": True}

            def gravar(caminho, tipo, cfg_tipo, cfg_larvas, larvas, conn,
                       logger, agora_iso, dry_run=False, preparado=None):
                self.assertTrue(conn.in_transaction)
                self.assertEqual(preparado, {"preparado": True})
                eventos.append("write")
                return {"ok": True, "visitas_novas": 0, "coletas_novas": 0}

            with mock.patch("etl.preparar_arquivo", side_effect=preparar), \
                    mock.patch("etl.processar_arquivo", side_effect=gravar), \
                    mock.patch("etl.db_core.connect", side_effect=conectar):
                ok, _ = etl.processar_upload(
                    [str(Path(tmpdir) / "TB_teste.xlsx")],
                    [],
                    db_path,
                    str(ROOT / "config.json"),
                    etl.Logger(),
                    dry_run=True,
                )

            self.assertTrue(ok)
            self.assertLess(eventos.index("prepare"), eventos.index("connect"))
            self.assertLess(eventos.index("connect"), eventos.index("write"))

    def test_importadores_especializados_preservam_transacao_do_lote(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            casos = (
                (recolhimentos_core, ("nova", [])),
                (amostras_animais_core, ("nova", [])),
                (bri_core, ("nova", [])),
            )
            for modulo, preparado in casos:
                conn = db_core.connect(db_path)
                try:
                    modulo.ensure_schema(conn)
                    conn.commit()
                    conn.execute("BEGIN")
                    modulo.processar_arquivo(
                        "ignorado.xlsx", conn, etl.Logger(),
                        "2026-07-14T12:00:00", preparado=preparado,
                        schema_ready=True,
                    )
                    self.assertTrue(conn.in_transaction, modulo.__name__)
                    conn.rollback()
                finally:
                    conn.close()

    def test_etl_desfaz_primeiro_arquivo_quando_segundo_falha(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            chamadas = 0

            def preparar(*args, **kwargs):
                return {"preparado": True}

            def gravar(caminho, tipo, cfg_tipo, cfg_larvas, larvas, conn,
                       logger, agora_iso, dry_run=False, preparado=None):
                nonlocal chamadas
                chamadas += 1
                self.assertTrue(conn.in_transaction)
                if chamadas == 1:
                    conn.execute(
                        "INSERT INTO localidades(nome, cod_localidade) VALUES (?, ?)",
                        ("Localidade transitoria", "TESTE-ROLLBACK"),
                    )
                    return {"ok": True, "visitas_novas": 0, "coletas_novas": 0}
                raise RuntimeError("falha simulada no segundo arquivo")

            with mock.patch("etl.preparar_arquivo", side_effect=preparar), \
                    mock.patch("etl.processar_arquivo", side_effect=gravar):
                ok, _ = etl.processar_upload(
                    [
                        str(Path(tmpdir) / "TB_primeiro.xlsx"),
                        str(Path(tmpdir) / "TB_segundo.xlsx"),
                    ],
                    [],
                    db_path,
                    str(ROOT / "config.json"),
                    etl.Logger(),
                    dry_run=False,
                )

            self.assertFalse(ok)
            conn = sqlite3.connect(db_path)
            try:
                total = conn.execute(
                    "SELECT COUNT(*) FROM localidades WHERE cod_localidade=?",
                    ("TESTE-ROLLBACK",),
                ).fetchone()[0]
            finally:
                conn.close()
            self.assertEqual(total, 0)

    def test_etl_importa_larvas_sozinhas_em_coletas_existentes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = _executar_criar_banco_em(tmpdir)
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """INSERT INTO visitas
                       (id_visita, kobo_uuid, tipo, data, localidade, quarteirao,
                        logradouro, numero, processado_em)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    ("visita-1", "uuid-visita-1", "TBO", "2026-06-10", "Centro", 5, "Rua A", "10", "2026-06-10T08:00:00"),
                )
                conn.execute(
                    "INSERT INTO coletas (id_coleta, id_visita, num_tubo, tipo_deposito) VALUES (?, ?, ?, ?)",
                    ("coleta-1", "visita-1", "T-001", "A1"),
                )
                conn.commit()
            finally:
                conn.close()

            larvas_path = Path(tmpdir) / "LARVAS_teste.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                "_uuid",
                "Número do tubito",
                "Data da coleta",
                "Nome do laboratorista",
                "Data da leitura",
                "Aegypt Larvas",
                "Aegypt Pupas",
                "Aegypt Exúvias",
                "Aegypt Adulto",
                "Albopictus Larvas",
                "Albopictus Pupas",
                "Albopictus Exúvias",
                "Albopictus Adulto",
                "Outra Espécie Larvas",
                "Outra Espécie Pupas",
                "Outra Espécie Exúvias",
                "Outra Espécie Adulto",
            ])
            ws.append(["uuid-larva-1", "T-001", "2026-06-10", "Lab", "2026-06-11", 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
            wb.save(larvas_path)

            ok, sumario = etl.processar_upload(
                [],
                [str(larvas_path)],
                db_path,
                str(ROOT / "config.json"),
                etl.Logger(),
                dry_run=False,
            )
            self.assertTrue(ok)
            self.assertEqual(sumario[0]["tipo"], "LARVAS")
            self.assertEqual(sumario[0]["resultados_novos"], 1)

            conn = sqlite3.connect(db_path)
            try:
                resultado = conn.execute(
                    "SELECT num_tubo, aegypt_larvas, kobo_uuid FROM resultados_laboratorio WHERE id_coleta=?",
                    ("coleta-1",),
                ).fetchone()
                foco = conn.execute(
                    "SELECT id_visita, num_tubo, tipo_trabalho, gera_notificacao FROM focos_positivos WHERE id_visita=?",
                    ("visita-1",),
                ).fetchone()
            finally:
                conn.close()
            self.assertEqual(resultado, ("T-001", 2, "uuid-larva-1"))
            self.assertEqual(foco, ("visita-1", "T-001", "TBO", 1))


class PermissionMatrixTests(unittest.TestCase):
    def test_fluxo_api_leitura_ovitrampa_ate_envio_conta_ovos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = db_core.connect(db_path)
            ovitrampas_core.ensure_schema(conn)
            agora = datetime.now().isoformat(timespec="seconds")
            grupo = conn.execute(
                "SELECT id_grupo FROM ovitrampas_calendario_grupos WHERE localidades LIKE '%Sede%' ORDER BY id_grupo LIMIT 1"
            ).fetchone()[0]
            diario = conn.execute(
                """INSERT INTO ovitrampas_diarios(nome,ativo,criado_em,atualizado_em)
                   VALUES ('Rota laboratório',1,?,?)""",
                (agora, agora),
            ).lastrowid
            conn.execute(
                """INSERT INTO ovitrampas_armadilhas
                   (ovitrampa_id,complemento,localidade,ativo,atualizado_em)
                   VALUES ('9001','Unidade de teste','Sede',1,?)""",
                (agora,),
            )
            conn.execute(
                """INSERT INTO ovitrampas_diario_armadilhas
                   (id_diario,ovitrampa_id,ordem,criado_em,atualizado_em)
                   VALUES (?,'9001',1,?,?)""",
                (diario, agora, agora),
            )
            conn.execute(
                """INSERT INTO ovitrampas_calendario_eventos
                   (data,movimento,id_grupo,ciclo,criado_em,atualizado_em)
                   VALUES (?,'troca',?,'Teste',?,?)""",
                (date.today().isoformat(), grupo, agora, agora),
            )
            conn.commit()
            conn.close()

            pagina_laboratorio = client.get("/laboratorio/lancamentos").data
            self.assertIn("Leitura Ovitrampa".encode(), pagina_laboratorio)
            self.assertIn("Diários aguardando leitura".encode(), pagina_laboratorio)
            self.assertIn("Selecionar ocorrência".encode(), pagina_laboratorio)
            self.assertNotIn("Leituras desta semana".encode(), pagina_laboratorio)
            self.assertIn(b".ovi-lab-body[hidden]{display:none;}", pagina_laboratorio)
            self.assertIn('data-ovi-tab="laboratorio"'.encode(), client.get("/ovitrampas").data)
            pendentes = client.get("/api/laboratorio/ovitrampas/lotes")
            self.assertEqual(pendentes.status_code, 200)
            lote = pendentes.get_json()["registros"][0]
            detalhe = client.get(
                f"/api/laboratorio/ovitrampas/lotes/{lote['id_lote']}"
            ).get_json()
            leitura = [{
                "id_item": detalhe["itens"][0]["id_item"],
                "ovos": 8,
                "ocorrencia": 7,
            }]

            self.assertEqual(client.put(
                f"/api/laboratorio/ovitrampas/lotes/{lote['id_lote']}/rascunho",
                json={"leituras": leitura},
            ).status_code, 200)
            self.assertEqual(client.post(
                f"/api/laboratorio/ovitrampas/lotes/{lote['id_lote']}/concluir",
                json={"leituras": leitura},
            ).status_code, 200)

            administrativo = client.get("/api/ovitrampas/laboratorio?status=pendente")
            self.assertEqual(administrativo.status_code, 200)
            self.assertEqual(administrativo.get_json()["registros"][0]["ovos"], 8)
            detalhe_administrativo = client.get(
                f"/api/ovitrampas/laboratorio/{lote['id_lote']}"
            ).get_json()
            self.assertEqual(detalhe_administrativo["itens"][0]["ocorrencia"], 7)
            self.assertEqual(
                detalhe_administrativo["itens"][0]["ocorrencia_label"],
                "Ovitrampa cheia de água",
            )
            self.assertEqual(client.post(
                f"/api/ovitrampas/laboratorio/{lote['id_lote']}/enviado",
                json={},
            ).status_code, 200)
            bloqueado = client.put(
                f"/api/laboratorio/ovitrampas/lotes/{lote['id_lote']}/rascunho",
                json={"leituras": leitura},
            )
            self.assertEqual(bloqueado.status_code, 400)

    def test_lancamentos_laboratorio_exige_permissao_especifica(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            cur = conn.execute(
                """INSERT INTO usuarios
                   (usuario,nome,senha_hash,nivel,ativo,criado_em,acesso_laboratorio)
                   VALUES ('laboratorista','Laboratorista','teste','visualizador',1,?,0)""",
                ("2026-07-14T12:00:00",),
            )
            uid = cur.lastrowid
            conn.commit()
            conn.close()
            _login_client_com_usuario(client, {
                "id_usuario": uid, "nome": "Laboratorista", "nivel": "visualizador"
            })
            self.assertEqual(client.get("/laboratorio/lancamentos").status_code, 403)

            conn = sqlite3.connect(db_path)
            conn.execute(
                "UPDATE usuarios SET acesso_laboratorio=1 WHERE id_usuario=?", (uid,)
            )
            conn.commit()
            conn.close()
            self.assertEqual(client.get("/laboratorio/lancamentos").status_code, 200)

    def test_usuario_exclusivo_laboratorio_nao_acessa_outros_modulos(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            cur = conn.execute(
                """INSERT INTO usuarios
                   (usuario,nome,senha_hash,nivel,ativo,criado_em,
                    acesso_laboratorio,somente_laboratorio)
                   VALUES ('lab.exclusivo','Laboratorista','teste','visualizador',1,?,1,1)""",
                ("2026-07-14T12:00:00",),
            )
            uid = cur.lastrowid
            conn.commit()
            conn.close()
            _login_client_com_usuario(client, {
                "id_usuario": uid, "nome": "Laboratorista", "nivel": "visualizador"
            })

            pagina = client.get("/laboratorio/lancamentos")
            self.assertEqual(pagina.status_code, 200)
            self.assertNotIn(b'href="/dashboard"', pagina.data)

            bloqueada = client.get("/dashboard")
            self.assertEqual(bloqueada.status_code, 302)
            self.assertTrue(bloqueada.headers["Location"].endswith("/laboratorio/lancamentos"))

            api_bloqueada = client.get("/api/dashboard")
            self.assertEqual(api_bloqueada.status_code, 403)
            self.assertEqual(
                api_bloqueada.get_json()["erro"],
                "Esta conta possui acesso exclusivo ao laboratorio.",
            )
            self.assertEqual(client.get("/minha-senha").status_code, 200)

    def test_lancamento_manual_sai_da_fila_e_bloqueia_duplicata_kobo(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = sqlite3.connect(db_path)
            conn.execute(
                "INSERT INTO agentes(nome,ativo) VALUES ('Márcio',1)"
            )
            id_marcio = conn.execute(
                "SELECT id_agente FROM agentes WHERE nome='Márcio'"
            ).fetchone()[0]
            conn.execute(
                "INSERT INTO agentes(nome,ativo) VALUES ('Azimir',1)"
            )
            id_laboratorista = conn.execute(
                "SELECT id_agente FROM agentes WHERE nome='Azimir'"
            ).fetchone()[0]
            uid_laboratorista = conn.execute(
                """INSERT INTO usuarios
                   (usuario,nome,senha_hash,nivel,ativo,criado_em,
                    acesso_laboratorio,somente_laboratorio)
                   VALUES ('azimir','Azimir','teste','visualizador',1,?,1,1)""",
                (datetime.now().isoformat(),),
            ).lastrowid
            conn.execute(
                """INSERT INTO visitas
                   (id_visita,kobo_uuid,kobo_id,tipo,data,localidade,logradouro,
                    numero,quarteirao,processado_em)
                   VALUES ('visita-lab','uuid-lab',1,'PVE','2026-07-07','Graziela',
                           'Rua dos Lírios','29',1383,'2026-07-07T12:00:00')"""
            )
            conn.execute(
                "INSERT INTO visita_agentes(id_visita,id_agente) VALUES ('visita-lab',?)",
                (id_marcio,),
            )
            conn.execute(
                """INSERT INTO coletas
                   (id_coleta,id_visita,num_tubo,codigo_deposito,tipo_deposito)
                   VALUES ('coleta-70','visita-lab','70','D2','Balde')"""
            )
            conn.commit()
            conn.close()
            _login_client_com_usuario(client, {
                "id_usuario": uid_laboratorista,
                "nome": "Azimir",
                "nivel": "visualizador",
            })

            pagina = client.get("/laboratorio/lancamentos")
            self.assertNotIn(b"form-laboratorista", pagina.data)
            self.assertNotIn(b"form-data", pagina.data)
            self.assertNotIn("Resultado negativo".encode(), pagina.data)
            self.assertNotIn("Salvar e próximo".encode(), pagina.data)
            self.assertIn("Tem certeza?".encode(), pagina.data)

            fila = client.get("/api/laboratorio/lancamentos/pendentes").get_json()
            self.assertEqual(fila["total"], 1)
            self.assertEqual(fila["pendentes"][0]["agentes"], "Márcio")
            self.assertEqual(fila["pendentes"][0]["tipo"], "PVE")

            resp = client.post(
                "/api/laboratorio/lancamentos/coleta-70/resultado",
                json={
                    "id_laboratorista": -1,
                    "data_leitura": "2000-01-01",
                    "albopictus_larvas": 2,
                },
            )
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(
                client.get("/api/laboratorio/lancamentos/pendentes").get_json()["total"], 0
            )
            historico = client.get(
                "/api/laboratorio/lancamentos/historico"
            ).get_json()["registros"]
            self.assertEqual(historico[0]["origem"], "sistema")
            self.assertEqual(historico[0]["laboratorista"], "Azimir")
            self.assertEqual(historico[0]["data_leitura"], date.today().isoformat())
            self.assertEqual(historico[0]["id_laboratorista"], id_laboratorista)
            self.assertEqual(historico[0]["tipo"], "PVE")
            self.assertFalse(historico[0]["positivo_aegypti"])
            self.assertTrue(historico[0]["editavel"])

            editado = client.post(
                f"/api/laboratorio/lancamentos/resultados/{historico[0]['id_resultado']}/editar",
                json={"aegypt_larvas": 3, "albopictus_larvas": 1},
            )
            self.assertEqual(editado.status_code, 200)
            corrigido = client.get(
                "/api/laboratorio/lancamentos/historico"
            ).get_json()["registros"][0]
            self.assertEqual(corrigido["aegypt_larvas"], 3)
            self.assertEqual(corrigido["albopictus_larvas"], 1)
            self.assertTrue(corrigido["positivo_aegypti"])

            conn = sqlite3.connect(db_path)
            conn.execute(
                "UPDATE resultados_laboratorio SET criado_em=? WHERE id_resultado=?",
                ((date.today() - timedelta(days=4)).isoformat(), historico[0]["id_resultado"]),
            )
            conn.commit()
            conn.close()
            expirado = client.post(
                f"/api/laboratorio/lancamentos/resultados/{historico[0]['id_resultado']}/editar",
                json={"aegypt_larvas": 4},
            )
            self.assertEqual(expirado.status_code, 403)

            conn = db_core.connect(db_path)
            inserido, _ = etl.inserir_resultado_larva(
                conn.cursor(),
                "coleta-70",
                etl.pd.Series({
                    "Número do tubito": "70",
                    "Data da coleta": "2026-07-07",
                    "Nome do laboratorista": "Kobo",
                    "Data da leitura": "2026-07-15",
                    "_uuid": "larva-kobo-70",
                }),
            )
            conn.close()
            self.assertFalse(inserido)

    def test_coleta_encerrada_sem_resultado_nao_aparece_na_fila(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            app_temp, client, db_path = _client_admin_com_banco_temporario(tmpdir)
            conn = db_core.connect(db_path)
            conn.execute(
                """INSERT INTO visitas
                   (id_visita,kobo_uuid,tipo,data,localidade,processado_em)
                   VALUES ('visita-sem','uuid-sem','PE','2026-05-28','Lamenha',?)""",
                ("2026-05-28T12:00:00",),
            )
            conn.execute(
                "INSERT INTO coletas(id_coleta,id_visita,num_tubo) VALUES ('coleta-359','visita-sem','359')"
            )
            laboratorio_lancamentos_core.encerrar_sem_resultado(
                conn, "coleta-359", "Pendência histórica sem possibilidade de correção"
            )
            conn.commit()
            self.assertEqual(
                client.get("/api/laboratorio/lancamentos/pendentes").get_json()["total"], 0
            )
            historico = client.get(
                "/api/laboratorio/lancamentos/historico"
            ).get_json()["registros"]
            self.assertEqual(historico, [])

            inserido, _ = etl.inserir_resultado_larva(
                conn.cursor(),
                "coleta-359",
                etl.pd.Series({
                    "Número do tubito": "359",
                    "Data da coleta": "2026-05-28",
                    "Nome do laboratorista": "Azimir",
                    "Data da leitura": "2026-07-14",
                    "_uuid": "larva-359",
                }),
            )
            conn.commit()
            status = conn.execute(
                "SELECT 1 FROM laboratorio_coletas_status WHERE id_coleta='coleta-359'"
            ).fetchone()
            conn.close()
            self.assertTrue(inserido)
            self.assertIsNone(status)

    def test_visualizador_acessa_paginas_de_consulta(self):
        client = _client_logado("visualizador")
        rotas = [
            "/",
            "/dashboard",
            "/visitas",
            "/laboratorio",
            "/conta-ovos-sispncd",
            "/ovitrampas",
            "/esporotricose",
            "/notificacoes",
            "/mapa",
            "/agenda",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)

    def test_visualizador_nao_acessa_areas_admin(self):
        client = _client_logado("visualizador")
        rotas = [
            "/processar",
            "/admin/sistema",
            "/admin/usuarios",
            "/admin/auditoria",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 403)

    def test_admin_acessa_areas_admin(self):
        client = _client_logado("admin")
        rotas = [
            "/processar",
            "/admin/sistema",
            "/admin/usuarios",
            "/admin/auditoria",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)

    def test_visualizador_acessa_apis_de_consulta(self):
        client = _client_logado("visualizador")
        rotas = [
            "/api/dashboard",
            "/api/visitas?pagina=1&por_pagina=5",
            "/api/laboratorio?pagina=1&por_pagina=5",
            "/api/mapa",
        ]

        for rota in rotas:
            with self.subTest(rota=rota):
                resp = client.get(rota)
                self.assertEqual(resp.status_code, 200)
                self.assertTrue(resp.is_json)


class ImportHistoryTests(unittest.TestCase):
    def test_registra_e_atualiza_importacao_em_banco_temporario(self):
        db_original = endemias_app.DB_PATH
        with tempfile.TemporaryDirectory() as tmp:
            db_temp = str(Path(tmp) / "teste.db")
            endemias_app.DB_PATH = db_temp
            try:
                endemias_app.registrar_importacao(
                    "job-teste",
                    ["TB_teste.xlsx", "LARVAS_teste.xlsx"],
                    usuario="Teste",
                )
                endemias_app.atualizar_importacao(
                    "job-teste",
                    "dry_run_ok",
                    dry_run_ok=True,
                    sumario=[{"arquivo": "TB_teste.xlsx", "visitas_novas": 2}],
                )

                conn = sqlite3.connect(db_temp)
                conn.row_factory = sqlite3.Row
                row = conn.execute(
                    "SELECT * FROM importacoes WHERE job_id=?",
                    ("job-teste",),
                ).fetchone()
                conn.close()
            finally:
                endemias_app.DB_PATH = db_original

        self.assertIsNotNone(row)
        self.assertEqual(row["usuario"], "Teste")
        self.assertEqual(row["status"], "dry_run_ok")
        self.assertEqual(row["dry_run_ok"], 1)
        self.assertEqual(json.loads(row["arquivos_json"]), ["TB_teste.xlsx", "LARVAS_teste.xlsx"])
        self.assertEqual(json.loads(row["sumario_json"])[0]["visitas_novas"], 2)


if __name__ == "__main__":
    unittest.main()
