import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FORM_PATH = ROOT / "formularios_kobo" / "RG_Atualizacao_Completa_Kobo.xlsx"


class RegistroGeograficoKoboFormTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.workbook = load_workbook(FORM_PATH, read_only=True, data_only=False)

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()

    def test_xlsform_tem_abas_obrigatorias(self):
        self.assertEqual(self.workbook.sheetnames, ["survey", "choices", "settings"])

    def test_formulario_representa_quarteirao_completo_em_repeticoes(self):
        survey = list(self.workbook["survey"].iter_rows(values_only=True))
        headers = survey[0]
        type_idx = headers.index("type")
        name_idx = headers.index("name")
        rows = {(row[type_idx], row[name_idx]) for row in survey[1:]}

        self.assertIn(("select_one localidades", "localidade"), rows)
        self.assertIn(("select_one quarteiroes", "quarteirao"), rows)
        self.assertIn(("begin_repeat", "trechos"), rows)
        self.assertIn(("begin_repeat", "imoveis"), rows)
        self.assertIn(("acknowledge", "confirmar_completo"), rows)

    def test_ref_nao_esta_disponivel_como_tipo_de_campo(self):
        choices = list(self.workbook["choices"].iter_rows(values_only=True))
        headers = choices[0]
        list_idx = headers.index("list_name")
        name_idx = headers.index("name")
        tipos = {
            row[name_idx]
            for row in choices[1:]
            if row[list_idx] == "tipos_rg"
        }

        self.assertEqual(tipos, {"R", "C", "O", "TB", "PE", "A"})
        self.assertNotIn("REF", tipos)

    def test_configuracao_identifica_formulario_e_idioma(self):
        settings = list(self.workbook["settings"].iter_rows(values_only=True))
        values = dict(zip(settings[0], settings[1]))

        self.assertEqual(values["form_id"], "rg_atualizacao_completa")
        self.assertEqual(values["default_language"], "Português (pt)")
        self.assertEqual(str(values["version"]), "2026071601")

    def test_confirmacao_final_exige_trecho_e_imovel(self):
        survey = list(self.workbook["survey"].iter_rows(values_only=True))
        headers = survey[0]
        name_idx = headers.index("name")
        constraint_idx = headers.index("constraint")
        row = next(item for item in survey[1:] if item[name_idx] == "confirmar_completo")

        self.assertIn("${total_trechos} > 0", row[constraint_idx])
        self.assertIn("${total_linhas} > 0", row[constraint_idx])


if __name__ == "__main__":
    unittest.main()
