# =============================================================================
#  GERADOR DE CONSOLIDADOS — SISTEMA DE ENDEMIAS
#  Chamado automaticamente pelo unificar_endemias.py ao final de cada execução.
#  Pode também ser rodado manualmente: python gerar_consolidado.py
# =============================================================================

import os
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os as _os
BANCO_DADOS  = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "endemias.db")
PASTA_SAIDA  = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "saida")

# =============================================================================
#  QUERIES POR TIPO
# =============================================================================

# Colunas comuns a todos os tipos
_CABECALHO_BASE = [
    "Data", "Hora Início", "Localidade", "Quarteirão",
    "Logradouro", "Número", "Visita", "Morador", "Agentes",
    # depósitos inspecionados
    "Dep. A1", "Dep. A2", "Dep. B", "Dep. C", "Dep. D1", "Dep. D2", "Dep. E",
    "Dep. Eliminados",
    # tratamento
    "Tratamento Tipo", "Tratamento Carga (g)", "Dep. Tratados",
    # coleta
    "Nº Tubo", "Código Depósito", "Tipo Depósito", "Depósito Eliminado",
    # laboratório
    "Laboratorista", "Data Leitura",
    "Ae. Larvas", "Ae. Pupas", "Ae. Exúvias", "Ae. Adulto",
    "Alb. Larvas", "Alb. Pupas", "Alb. Exúvias", "Alb. Adulto",
    "Outra Larvas", "Outra Pupas", "Outra Exúvias", "Outra Adulto",
]

_QUERY_BASE = """
    SELECT
        v.data, v.hora_inicio, v.localidade, v.quarteirao,
        v.logradouro, v.numero, v.visita, v.morador,
        (SELECT GROUP_CONCAT(a2.nome, ', ')
         FROM (SELECT DISTINCT a2.nome FROM agentes a2
               JOIN visita_agentes va2 ON va2.id_agente = a2.id_agente
               WHERE va2.id_visita = v.id_visita ORDER BY a2.nome) a2
        ) AS agentes,
        /*EXTRAS*/MAX(CASE WHEN d.tipo_deposito='A1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.inspecionado END),
        MAX(d.eliminado),
        (SELECT GROUP_CONCAT(tp, ", ") FROM (SELECT DISTINCT t2.tipo AS tp FROM tratamentos t2 WHERE t2.id_visita=v.id_visita)) AS trat_tipo,
        MAX(t.quantidade_carga)        AS trat_carga,
        MAX(t.qtd_depositos_tratados)  AS trat_dep,
        c.num_tubo, c.codigo_deposito, c.tipo_deposito,
        CASE c.deposito_eliminado WHEN 1 THEN 'Sim' WHEN 0 THEN 'Não' ELSE NULL END,
        r.laboratorista, r.data_leitura,
        r.aegypt_larvas, r.aegypt_pupas, r.aegypt_exuvias, r.aegypt_adulto,
        r.albopictus_larvas, r.albopictus_pupas, r.albopictus_exuvias, r.albopictus_adulto,
        r.outra_larvas, r.outra_pupas, r.outra_exuvias, r.outra_adulto
    FROM visitas v
    LEFT JOIN visita_agentes va       ON va.id_visita  = v.id_visita
    LEFT JOIN agentes a               ON a.id_agente   = va.id_agente
    LEFT JOIN coletas c               ON c.id_visita   = v.id_visita
    LEFT JOIN depositos_inspecionados d ON d.id_visita = v.id_visita
    LEFT JOIN tratamentos t           ON t.id_visita   = v.id_visita
    LEFT JOIN resultados_laboratorio r ON r.id_coleta  = c.id_coleta
    WHERE v.tipo = ?
    GROUP BY v.id_visita, c.id_coleta
    ORDER BY v.data, v.localidade, v.quarteirao, v.logradouro, v.numero
"""

_CABECALHO_VISITAS_BASE = [
    "Data", "Hora Início", "Localidade", "Quarteirão",
    "Logradouro", "Número", "Visita", "Morador", "Agentes",
    "Dep. A1", "Dep. A2", "Dep. B", "Dep. C", "Dep. D1", "Dep. D2", "Dep. E",
    "Dep. Eliminados",
    "Tratamento Tipo", "Tratamento Carga (g)", "Dep. Tratados",
    "Qtde Coletas", "Nº Tubos", "Depósitos Coletados", "Depósitos Eliminados Coleta",
    "Data Última Leitura", "Laboratoristas",
    "Ae. Larvas", "Ae. Pupas", "Ae. Exúvias", "Ae. Adulto",
    "Alb. Larvas", "Alb. Pupas", "Alb. Exúvias", "Alb. Adulto",
    "Outra Larvas", "Outra Pupas", "Outra Exúvias", "Outra Adulto",
]

_QUERY_VISITAS_BASE = """
    SELECT
        v.data, v.hora_inicio, v.localidade, v.quarteirao,
        v.logradouro, v.numero, v.visita, v.morador,
        (SELECT GROUP_CONCAT(a2.nome, ', ')
         FROM (SELECT DISTINCT a2.nome FROM agentes a2
               JOIN visita_agentes va2 ON va2.id_agente = a2.id_agente
               WHERE va2.id_visita = v.id_visita ORDER BY a2.nome) a2
        ) AS agentes,
        /*EXTRAS*/MAX(CASE WHEN d.tipo_deposito='A1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.inspecionado END),
        MAX(d.eliminado),
        (SELECT GROUP_CONCAT(tp, ", ") FROM (SELECT DISTINCT t2.tipo AS tp FROM tratamentos t2 WHERE t2.id_visita=v.id_visita)) AS trat_tipo,
        MAX(t.quantidade_carga)        AS trat_carga,
        MAX(t.qtd_depositos_tratados)  AS trat_dep,
        (SELECT COUNT(*) FROM coletas c2 WHERE c2.id_visita=v.id_visita) AS qtd_coletas,
        (SELECT GROUP_CONCAT(num_tubo, ', ')
         FROM (SELECT c2.num_tubo FROM coletas c2
               WHERE c2.id_visita=v.id_visita AND c2.num_tubo IS NOT NULL AND TRIM(c2.num_tubo)<>''
               ORDER BY c2.num_tubo)) AS tubos,
        (SELECT GROUP_CONCAT(dep, '; ')
         FROM (SELECT COALESCE(c2.codigo_deposito || ' - ', '') || COALESCE(c2.tipo_deposito, '') AS dep
               FROM coletas c2
               WHERE c2.id_visita=v.id_visita
               ORDER BY c2.num_tubo)) AS depositos_coletados,
        (SELECT SUM(CASE WHEN c2.deposito_eliminado=1 THEN 1 ELSE 0 END)
         FROM coletas c2 WHERE c2.id_visita=v.id_visita) AS depositos_eliminados_coleta,
        (SELECT MAX(r2.data_leitura)
         FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta
         WHERE c2.id_visita=v.id_visita) AS data_ultima_leitura,
        (SELECT GROUP_CONCAT(lab, ', ')
         FROM (SELECT DISTINCT r2.laboratorista AS lab
               FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta
               WHERE c2.id_visita=v.id_visita AND r2.laboratorista IS NOT NULL AND TRIM(r2.laboratorista)<>''
               ORDER BY r2.laboratorista)) AS laboratoristas,
        (SELECT COALESCE(SUM(r2.aegypt_larvas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.aegypt_pupas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.aegypt_exuvias),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.aegypt_adulto),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.albopictus_larvas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.albopictus_pupas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.albopictus_exuvias),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.albopictus_adulto),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.outra_larvas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.outra_pupas),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.outra_exuvias),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita),
        (SELECT COALESCE(SUM(r2.outra_adulto),0) FROM coletas c2 JOIN resultados_laboratorio r2 ON r2.id_coleta=c2.id_coleta WHERE c2.id_visita=v.id_visita)
    FROM visitas v
    LEFT JOIN depositos_inspecionados d ON d.id_visita = v.id_visita
    LEFT JOIN tratamentos t             ON t.id_visita = v.id_visita
    WHERE v.tipo = ?
    GROUP BY v.id_visita
    ORDER BY v.data, v.localidade, v.quarteirao, v.logradouro, v.numero
"""

# Extras por tipo inseridos após "Agentes" e antes dos depósitos
_EXTRAS = {
    "PE":  ["Ciclo"],
    "PVE": ["Ciclo", "Lado", "Tipo Imóvel", "Água Sanepar"],
    "TB":  ["Tipo Imóvel", "Água Sanepar"],
    "TBO": ["Hora Fim", "Tipo Imóvel", "Água Sanepar",
            # depósitos TBO detalhados (insp/elim/trat/tipo_trat/carga por tipo)
            "A1 Insp", "A1 Elim", "A1 Trat", "A1 Tipo Trat", "A1 Carga",
            "A2 Insp", "A2 Elim", "A2 Trat", "A2 Tipo Trat", "A2 Carga",
            "B Insp",  "B Elim",  "B Trat",  "B Tipo Trat",  "B Carga",
            "C Insp",  "C Elim",  "C Trat",  "C Tipo Trat",  "C Carga",
            "D1 Insp", "D1 Elim", "D1 Trat", "D1 Tipo Trat", "D1 Carga",
            "D2 Insp", "D2 Elim", "D2 Trat", "D2 Tipo Trat", "D2 Carga",
            "E Insp",  "E Elim",  "E Trat",  "E Tipo Trat",  "E Carga",
            ],
}

_QUERY_EXTRAS = {
    "PE": "v.ciclo,",
    "PVE": "v.ciclo, v.lado, v.tipo_imovel, CASE v.agua_sanepar WHEN 1 THEN 'Sim' WHEN 0 THEN 'Não' ELSE NULL END,",
    "TB":  "v.tipo_imovel, CASE v.agua_sanepar WHEN 1 THEN 'Sim' WHEN 0 THEN 'Não' ELSE NULL END,",
    "TBO": """v.hora_fim, v.tipo_imovel,
        CASE v.agua_sanepar WHEN 1 THEN 'Sim' WHEN 0 THEN 'Não' ELSE NULL END,
        MAX(CASE WHEN d.tipo_deposito='A1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='A1' THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='A1' THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='A1' THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='A1' THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='A2' THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='B'  THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='C'  THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='D1' THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='D2' THEN d.qtd_carga    END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.inspecionado END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.eliminado    END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.tratado      END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.tipo_tratamento END),
        MAX(CASE WHEN d.tipo_deposito='E'  THEN d.qtd_carga    END),""",
}

def montar_query(tipo):
    extras_select = _QUERY_EXTRAS.get(tipo, "")
    if extras_select:
        return _QUERY_BASE.replace("/*EXTRAS*/", extras_select + "\n        ")
    return _QUERY_BASE.replace("/*EXTRAS*/", "")

def montar_query_visitas(tipo):
    extras_select = _QUERY_EXTRAS.get(tipo, "")
    if extras_select:
        return _QUERY_VISITAS_BASE.replace("/*EXTRAS*/", extras_select + "\n        ")
    return _QUERY_VISITAS_BASE.replace("/*EXTRAS*/", "")

def montar_cabecalho(tipo):
    base = list(_CABECALHO_BASE)
    extras = _EXTRAS.get(tipo, [])
    # Inserir extras após "Agentes" (índice 8)
    return base[:9] + extras + base[9:]

def montar_cabecalho_visitas(tipo):
    base = list(_CABECALHO_VISITAS_BASE)
    extras = _EXTRAS.get(tipo, [])
    return base[:9] + extras + base[9:]

# =============================================================================
#  ESTILOS OPENPYXL
# =============================================================================

COR_HEADER = {
    "PE":  "1F4E79",
    "TB":  "1F4E79",
    "TBO": "1F4E79",
    "PVE": "1F4E79",
}

COR_GRUPOS = {
    "visita":    "D6E4F0",
    "depositos": "E2EFDA",
    "tratamento":"FFF2CC",
    "coleta":    "FCE4D6",
    "lab":       "EAD1DC",
}

def estilo_header():
    return Font(name="Arial", bold=True, color="FFFFFF", size=9)

def fill(hex_cor):
    return PatternFill("solid", fgColor=hex_cor)

def borda_fina():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_cabecalho(ws, cabecalho, tipo):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = estilo_header()
        cell.fill      = fill(COR_HEADER[tipo])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda_fina()

    ws.row_dimensions[1].height = 30

def colorir_linha(ws, row_idx, cabecalho, tipo):
    """Alterna cor de fundo por grupo de colunas."""
    grupos = _grupos_cabecalho(cabecalho, tipo)
    for col_idx, titulo in enumerate(cabecalho, start=1):
        grupo = grupos.get(titulo, "visita")
        cor   = COR_GRUPOS.get(grupo, "FFFFFF")
        cell  = ws.cell(row=row_idx, column=col_idx)
        cell.fill   = fill(cor) if row_idx % 2 == 0 else PatternFill()
        cell.border = borda_fina()
        cell.font   = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="center")

def _grupos_cabecalho(cabecalho, tipo):
    grupos = {}
    for t in cabecalho:
        if t in ("Data","Hora Início","Hora Fim","Localidade","Quarteirão",
                 "Logradouro","Número","Visita","Morador","Agentes",
                 "Ciclo","Lado","Tipo Imóvel","Água Sanepar","Sequência"):
            grupos[t] = "visita"
        elif "Dep." in t or "Insp" in t or "Elim" in t:
            grupos[t] = "depositos"
        elif "Trat" in t or "Carga" in t or "Tipo Trat" in t:
            grupos[t] = "tratamento"
        elif t in ("Nº Tubo","Código Depósito","Tipo Depósito","Depósito Eliminado"):
            grupos[t] = "coleta"
        else:
            grupos[t] = "lab"
    return grupos

def ajustar_larguras(ws, cabecalho):
    larguras = {
        "Data": 12, "Hora Início": 10, "Hora Fim": 10,
        "Localidade": 18, "Quarteirão": 10, "Logradouro": 28,
        "Número": 8, "Visita": 10, "Morador": 16, "Agentes": 30,
        "Ciclo": 6, "Lado": 6, "Tipo Imóvel": 14, "Água Sanepar": 12,
        "Nº Tubo": 10, "Código Depósito": 12, "Tipo Depósito": 16,
        "Depósito Eliminado": 14, "Laboratorista": 16, "Data Leitura": 12,
    }
    for col_idx, titulo in enumerate(cabecalho, start=1):
        col_letra = get_column_letter(col_idx)
        ws.column_dimensions[col_letra].width = larguras.get(titulo, 9)

# =============================================================================
#  GERAÇÃO POR TIPO
# =============================================================================

def _preencher_aba(ws, cabecalho, rows, tipo):
    aplicar_cabecalho(ws, cabecalho, tipo)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, valor in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = borda_fina()
            if row_idx % 2 == 0:
                cell.fill = fill("F2F7FB")

        ws.row_dimensions[row_idx].height = 15

    ajustar_larguras(ws, cabecalho)


def gerar_xlsx_tipo(cur, tipo, pasta_saida):
    cabecalho_visitas = montar_cabecalho_visitas(tipo)
    cabecalho_coletas = montar_cabecalho(tipo)

    cur.execute(montar_query_visitas(tipo), (tipo,))
    rows_visitas = cur.fetchall()

    cur.execute(montar_query(tipo), (tipo,))
    rows_coletas = cur.fetchall()

    if not rows_visitas and not rows_coletas:
        return None, 0, 0

    wb = openpyxl.Workbook()
    ws_visitas = wb.active
    ws_visitas.title = "Visitas"
    _preencher_aba(ws_visitas, cabecalho_visitas, rows_visitas, tipo)

    ws_coletas = wb.create_sheet("Coletas")
    _preencher_aba(ws_coletas, cabecalho_coletas, rows_coletas, tipo)

    nome_arquivo = "%s_consolidado.xlsx" % tipo
    caminho      = os.path.join(pasta_saida, nome_arquivo)
    wb.save(caminho)
    return caminho, len(rows_visitas), len(rows_coletas)

# =============================================================================
#  MAIN
# =============================================================================

def gerar_todos(logger=None, banco_dados=None, pasta_saida=None, tipos=None):
    def log(msg, tag="normal"):
        if logger:
            logger.log(msg, tag)
        else:
            print(msg)

    banco_dados = banco_dados or BANCO_DADOS
    pasta_saida = pasta_saida or PASTA_SAIDA
    tipos = tipos or ["PE", "TB", "TBO", "PVE"]

    if not os.path.exists(banco_dados):
        log("[ERRO] Banco '%s' não encontrado." % banco_dados, "erro")
        return

    os.makedirs(pasta_saida, exist_ok=True)
    conn = sqlite3.connect(banco_dados)
    cur  = conn.cursor()

    resultados = []
    log("\nGerando consolidados Excel...", "titulo")

    for tipo in tipos:
        try:
            caminho, qtd_visitas, qtd_coletas = gerar_xlsx_tipo(cur, tipo, pasta_saida)
            if caminho:
                log("  ✓ %s — %d visita(s), %d coleta(s) → %s" % (tipo, qtd_visitas, qtd_coletas, caminho), "ok")
                resultados.append({
                    "tipo": tipo,
                    "caminho": caminho,
                    "visitas": qtd_visitas,
                    "coletas": qtd_coletas,
                })
            else:
                log("  — %s: sem dados no banco." % tipo, "aviso")
                resultados.append({"tipo": tipo, "caminho": None, "visitas": 0, "coletas": 0})
        except Exception as e:
            import traceback
            log("  [ERRO] %s: %s" % (tipo, e), "erro")
            log(traceback.format_exc(), "erro")
            resultados.append({"tipo": tipo, "erro": str(e)})

    conn.close()
    return resultados


if __name__ == "__main__":
    gerar_todos()
