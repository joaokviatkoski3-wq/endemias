import io
import logging
from datetime import datetime

import openpyxl
from flask import Blueprint, current_app, jsonify, render_template, request, send_file
from openpyxl.styles import Alignment, Font, PatternFill

from app_core import audit
from app_core import auth as auth_core
from app_core import db as db_core
from app_core import pontos_estrategicos as pe_core


bp = Blueprint("pontos_estrategicos", __name__)
login_required = auth_core.login_required


def _db_path():
    return current_app.config["DB_PATH"]


def get_db():
    return db_core.connect(_db_path())


def q1(sql, params=()):
    return db_core.query_one(_db_path(), sql, params)


def usuario_atual():
    return auth_core.usuario_atual(q1)


def nivel_min(nivel):
    return auth_core.nivel_min(nivel, usuario_atual)


@bp.route("/pontos-estrategicos")
@login_required
def page():
    return render_template("pontos_estrategicos.html", opcoes=pe_core.opcoes(_db_path()))


@bp.route("/api/pontos-estrategicos")
@login_required
def api_listar():
    return jsonify(pe_core.listar(_db_path(), _filtros()))


@bp.route("/api/pontos-estrategicos/exportar")
@login_required
def api_exportar():
    try:
        filtros = _filtros()
        dados = pe_core.listar(_db_path(), filtros, limite=None)
        registros = dados.get("registros") or []
        formato = (request.args.get("formato") or "xlsx").lower()
        if formato == "pdf":
            return render_template(
                "pontos_estrategicos_pdf.html",
                registros=registros,
                totais=dados.get("totais") or {},
                filtros=_filtros_legiveis(filtros),
                now=datetime.now().strftime("%d/%m/%Y %H:%M"),
            )
        return _gerar_xlsx(registros)
    except Exception:
        logging.exception("Erro em api_exportar pontos estrategicos")
        return jsonify({"erro": "Erro interno. Verifique endemias.log"}), 500


@bp.route("/api/pontos-estrategicos/opcoes")
@login_required
def api_opcoes():
    return jsonify(pe_core.opcoes(_db_path()))


@bp.route("/api/pontos-estrategicos/<int:id_pe>")
@login_required
def api_obter(id_pe):
    registro = pe_core.obter(_db_path(), id_pe)
    if not registro:
        return jsonify({"erro": "Ponto estrategico nao encontrado."}), 404
    return jsonify(registro)


@bp.route("/api/pontos-estrategicos", methods=["POST"])
@login_required
@nivel_min("operador")
def api_criar():
    payload = request.get_json(silent=True) or {}
    if not _payload_valido(payload):
        return jsonify({"erro": "Informe ao menos o nome/local do PE."}), 400
    criado = pe_core.salvar(_db_path(), payload)
    audit.registrar_evento(
        get_db,
        "pe_criado",
        entidade="pontos_estrategicos",
        detalhes={"nome": payload.get("nome"), "localidade": payload.get("localidade")},
    )
    return jsonify({"ok": bool(criado)})


@bp.route("/api/pontos-estrategicos/<int:id_pe>", methods=["POST"])
@login_required
@nivel_min("operador")
def api_atualizar(id_pe):
    payload = request.get_json(silent=True) or {}
    if not _payload_valido(payload):
        return jsonify({"erro": "Informe ao menos o nome/local do PE."}), 400
    atualizado = pe_core.salvar(_db_path(), payload, id_pe=id_pe)
    if not atualizado:
        return jsonify({"erro": "Ponto estrategico nao encontrado."}), 404
    audit.registrar_evento(
        get_db,
        "pe_atualizado",
        entidade="pontos_estrategicos",
        entidade_id=id_pe,
        detalhes={"nome": payload.get("nome"), "localidade": payload.get("localidade")},
    )
    return jsonify({"ok": True})


@bp.route("/api/pontos-estrategicos/<int:id_pe>/situacao", methods=["POST"])
@login_required
@nivel_min("operador")
def api_situacao(id_pe):
    payload = request.get_json(silent=True) or {}
    situacao = int(payload.get("situacao", 1))
    if situacao not in (0, 1):
        return jsonify({"erro": "Situacao invalida."}), 400
    atualizado = pe_core.definir_situacao(_db_path(), id_pe, situacao)
    if not atualizado:
        return jsonify({"erro": "Ponto estrategico nao encontrado."}), 404
    audit.registrar_evento(
        get_db,
        "pe_situacao",
        entidade="pontos_estrategicos",
        entidade_id=id_pe,
        detalhes={"situacao": situacao},
    )
    return jsonify({"ok": True})


def _filtros():
    return {
        "situacao": request.args.get("situacao", ""),
        "localidade": request.args.get("localidade", ""),
        "tipo": request.args.get("tipo", ""),
        "atrasados": request.args.get("atrasados", ""),
        "pendencias": request.args.get("pendencias", ""),
        "busca": request.args.get("busca", "").strip(),
    }


def _filtros_legiveis(filtros):
    partes = []
    if filtros.get("busca"):
        partes.append(f"Pesquisa: {filtros['busca']}")
    if filtros.get("localidade"):
        partes.append(f"Localidade: {filtros['localidade']}")
    if filtros.get("tipo"):
        partes.append(f"Tipo: {filtros['tipo']}")
    if filtros.get("situacao") == "1":
        partes.append("Situacao: Ativos")
    elif filtros.get("situacao") == "0":
        partes.append("Situacao: Inativos")
    if filtros.get("atrasados"):
        partes.append("Revisao: Atrasados")
    if filtros.get("pendencias"):
        partes.append("Revisao: Com pendencia")
    return partes or ["Todos os registros"]


def _valor(row, chave, padrao=""):
    value = row.get(chave)
    return padrao if value is None else value


def _pendencias(row):
    itens = []
    if row.get("visita_atrasada"):
        itens.append("visita atrasada")
    itens.extend(row.get("pendencias_cadastro") or [])
    return ", ".join(itens)


def _gerar_xlsx(registros):
    cabecalho = [
        "Codigo", "Local", "Localidade", "Quarteirao", "Logradouro", "Numero",
        "Tipo", "Situacao", "Telefone", "CNPJ", "Razao social", "Latitude",
        "Longitude", "Data inclusao", "Data desativacao", "Ultima PE",
        "Dias sem visita", "Total visitas PE", "Ultimo BRI", "Total BRI",
        "Focos", "Pendencias", "Observacoes",
    ]
    campos = [
        "codigo_pe", "nome", "localidade", "quarteirao", "logradouro", "numero",
        "tipo", "situacao_label", "telefone", "cnpj", "razao_social", "latitude",
        "longitude", "data_inclusao", "data_desativacao", "ultima_visita_pe",
        "dias_sem_visita", "visitas_pe_total", "ultimo_bri", "bri_total",
        "focos_total", "pendencias", "observacoes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pontos Estrategicos"
    fill = PatternFill("solid", fgColor="166534")
    for ci, col in enumerate(cabecalho, 1):
        cell = ws.cell(1, ci, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(registros, 2):
        linha = dict(row)
        linha["situacao_label"] = "Ativo" if int(linha.get("situacao") or 0) == 1 else "Inativo"
        linha["pendencias"] = _pendencias(linha)
        for ci, campo in enumerate(campos, 1):
            ws.cell(ri, ci, _valor(linha, campo))

    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"pontos_estrategicos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _payload_valido(payload):
    return bool(str(payload.get("nome") or "").strip())
