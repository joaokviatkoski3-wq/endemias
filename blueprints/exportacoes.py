import io
import logging
import os
from datetime import datetime

import openpyxl
from flask import Blueprint, abort, current_app, jsonify, request, send_file
from openpyxl.styles import Alignment, Font, PatternFill

from app_core import auth as auth_core
from app_core import blueprint_helpers as bh
from app_core import db as db_core
from app_core.excel import excel_safe
from app_core import utils as utils_core
from app_core import visitas as visitas_core
from app_core import work_types


bp = Blueprint("exportacoes", __name__)
login_required = auth_core.login_required
nivel_min = bh.nivel_min


def _db_target():
    return db_core.configured_target(current_app.config)


def q(sql, params=()):
    return db_core.query(_db_target(), sql, params)


def _string_aggregates():
    if _db_target().backend == "postgresql":
        return {
            "agentes": "string_agg(nome, ', ' ORDER BY nome)",
            "itens": "string_agg(item, '; ' ORDER BY ordem)",
            "tubos": "string_agg(CAST(num_tubo AS TEXT), ', ' ORDER BY num_tubo)",
        }
    return {
        "agentes": "GROUP_CONCAT(nome, ', ')",
        "itens": "GROUP_CONCAT(item, '; ')",
        "tubos": "GROUP_CONCAT(num_tubo, ', ')",
    }


def _base_dir():
    return current_app.root_path


def _saida_dir():
    return current_app.config.get("SAIDA_DIR") or os.path.join(_base_dir(), "saida")


@bp.route("/api/visitas/exportar")
@login_required
def exportar_visitas():
    try:
        where, params = visitas_core.build_where(request.args)
        aggregates = _string_aggregates()
        rows = q(f"""
            SELECT v.data, v.tipo, COALESCE(l.nome, v.localidade) AS localidade, v.quarteirao,
                   v.logradouro, v.numero, v.visita, v.morador, v.tipo_imovel,
                   v.ciclo, v.sequencia, v.lado, v.hora_inicio, v.hora_fim,
                   CASE v.agua_sanepar WHEN 1 THEN 'Sim' WHEN 0 THEN 'Não' ELSE NULL END,
                   v.observacoes,
                   (SELECT {aggregates['agentes']} FROM (
                        SELECT DISTINCT COALESCE(NULLIF(a.nome_completo,''), a.nome) AS nome FROM visita_agentes va
                        JOIN agentes a ON a.id_agente=va.id_agente
                        WHERE va.id_visita=v.id_visita ORDER BY nome
                   ) agentes_ordenados) AS agentes,
                   COALESCE((SELECT SUM(d.inspecionado) FROM depositos_inspecionados d WHERE d.id_visita=v.id_visita),0),
                   COALESCE((SELECT SUM(d.eliminado) FROM depositos_inspecionados d WHERE d.id_visita=v.id_visita),0),
                   COALESCE((SELECT SUM(d.tratado) FROM depositos_inspecionados d WHERE d.id_visita=v.id_visita),0),
                   (SELECT {aggregates['itens']} FROM (
                        SELECT tipo_deposito || ': ' || CAST(COALESCE(inspecionado,0) AS TEXT) || ' inspec.; ' ||
                               CAST(COALESCE(eliminado,0) AS TEXT) || ' elim.; ' ||
                               CAST(COALESCE(tratado,0) AS TEXT) || ' trat.' AS item,
                               d.id AS ordem
                          FROM depositos_inspecionados d WHERE d.id_visita=v.id_visita ORDER BY d.id
                   ) depositos_ordenados),
                   (SELECT {aggregates['itens']} FROM (
                        SELECT COALESCE(tipo,'Sem produto') || ': ' ||
                               CAST(COALESCE(qtd_depositos_tratados,0) AS TEXT) ||
                               ' depósitos; carga ' || CAST(COALESCE(quantidade_carga,0) AS TEXT) AS item,
                               t.id AS ordem
                          FROM tratamentos t WHERE t.id_visita=v.id_visita ORDER BY t.id
                   ) tratamentos_ordenados),
                   (SELECT COUNT(*) FROM coletas c WHERE c.id_visita=v.id_visita),
                   (SELECT {aggregates['tubos']} FROM (
                        SELECT num_tubo FROM coletas c WHERE c.id_visita=v.id_visita ORDER BY c.num_tubo
                   ) tubos_ordenados),
                   v.SISPNCD, v.CONTAOVOS_STATUS, v.submission_time, v.processado_em
            FROM visitas v
            LEFT JOIN localidades l ON l.id_localidade=v.id_localidade
            {where} ORDER BY {visitas_core.order_sql(request.args.get('ordem'))}
        """, params)
        cabecalho = ["Data", "Tipo", "Localidade", "Quarteirao", "Logradouro", "Numero",
                     "Visita", "Morador", "Tipo Imovel", "Ciclo", "Sequencia",
                     "Lado", "Hora Inicio", "Hora Fim", "Agua Sanepar", "Observacoes", "Agentes",
                     "Depositos Inspecionados", "Depositos Eliminados", "Depositos Tratados",
                     "Detalhes dos Depositos", "Produtos e Tratamentos", "Coletas", "Tubos",
                     "Codigo SisPNCD", "Status Conta Ovos", "Envio Kobo", "Processado em"]
        return _gerar_xlsx(cabecalho, rows, "visitas")
    except Exception:
        logging.exception("Erro em exportar_visitas")
        return jsonify({"erro": "Erro interno. Verifique endemias.log"}), 500


@bp.route("/api/notificacoes/exportar")
@login_required
def exportar_notificacoes():
    try:
        fs = request.args.getlist("status")
        ft = request.args.getlist("tipo")
        fl = request.args.getlist("localidade")
        d_ini = request.args.get("d_ini", "")
        d_fim = request.args.get("d_fim", "")
        busca = request.args.get("busca", "").strip()
        where, params = "WHERE f.gera_notificacao=1", []
        if d_ini:
            where += " AND f.data>=?"
            params.append(d_ini)
        if d_fim:
            where += " AND f.data<=?"
            params.append(d_fim)
        if fs:
            where += f" AND COALESCE(f.status_notificacao,'pendente') IN ({','.join('?' * len(fs))})"
            params += fs
        if ft:
            where += f" AND f.tipo_trabalho IN ({','.join('?' * len(ft))})"
            params += ft
        if fl:
            where += f" AND l.nome IN ({','.join('?' * len(fl))})"
            params += fl
        if busca:
            where += " AND (LOWER(f.logradouro) LIKE LOWER(?) OR LOWER(f.num_tubo) LIKE LOWER(?) OR LOWER(f.nome_morador) LIKE LOWER(?) OR LOWER(f.codigo) LIKE LOWER(?))"
            b = f"%{busca}%"
            params += [b, b, b, b]
        rows = q(f"""
            SELECT f.codigo, f.data, f.tipo_trabalho, l.nome as localidade,
                   f.quarteirao, f.logradouro, f.numero, f.nome_morador,
                   f.num_tubo, f.depositos, f.agentes,
                   COALESCE(f.status_notificacao,'pendente') as status,
                   f.tentativa_1, f.tentativa_2, f.tentativa_3,
                   f.data_entrega, f.observacoes
            FROM focos_positivos f
            LEFT JOIN localidades l ON l.id_localidade=f.id_localidade
            {where} ORDER BY f.data DESC
        """, params)
        cabecalho = ["Codigo", "Data", "Tipo", "Localidade", "Quarteirao", "Logradouro",
                     "Numero", "Morador", "Tubo(s)", "Deposito(s)", "Agentes", "Status",
                     "Tentativa 1", "Tentativa 2", "Tentativa 3", "Data Entrega", "Observacoes"]
        return _gerar_xlsx(cabecalho, rows, "notificacoes")
    except Exception:
        logging.exception("Erro em exportar_notificacoes")
        return jsonify({"erro": "Erro interno. Verifique endemias.log"}), 500


@bp.route("/api/laboratorio/exportar")
@login_required
def exportar_laboratorio():
    try:
        d_ini = request.args.get("d_ini", utils_core.data_n_dias(365))
        d_fim = request.args.get("d_fim", utils_core.hoje())
        tipos = request.args.getlist("tipo")
        locs = request.args.getlist("localidade")
        tubo = request.args.get("tubo", "").strip()
        where = "WHERE v.data BETWEEN ? AND ?"
        params = [d_ini, d_fim]
        if tipos:
            where += f" AND v.tipo IN ({','.join('?' * len(tipos))})"
            params += tipos
        if locs:
            where += f" AND l.nome IN ({','.join('?' * len(locs))})"
            params += locs
        if tubo:
            where += " AND LOWER(c.num_tubo) LIKE LOWER(?)"
            params.append(f"%{tubo}%")
        aggregates = _string_aggregates()
        rows = q(f"""
            SELECT v.data, v.tipo, l.nome as localidade, v.quarteirao,
                   v.logradouro, v.numero, c.num_tubo, c.tipo_deposito,
                   rl.data_leitura, rl.laboratorista,
                   rl.aegypt_larvas, rl.aegypt_pupas, rl.aegypt_exuvias, rl.aegypt_adulto,
                   rl.albopictus_larvas, rl.albopictus_pupas, rl.albopictus_exuvias, rl.albopictus_adulto,
                   rl.outra_larvas, rl.outra_pupas, rl.outra_exuvias, rl.outra_adulto,
                   (SELECT {aggregates['agentes']} FROM (
                        SELECT DISTINCT COALESCE(NULLIF(a.nome_completo,''), a.nome) AS nome
                          FROM visita_agentes va
                          JOIN agentes a ON a.id_agente=va.id_agente
                         WHERE va.id_visita=v.id_visita
                         ORDER BY nome
                   ) agentes_ordenados) AS agentes
            FROM resultados_laboratorio rl
            JOIN coletas c ON c.id_coleta=rl.id_coleta
            JOIN visitas v ON v.id_visita=c.id_visita
            LEFT JOIN localidades l ON l.id_localidade=v.id_localidade
            {where} ORDER BY v.data DESC
        """, params)
        cabecalho = ["Data", "Tipo", "Localidade", "Quarteirao", "Logradouro", "Numero",
                     "Tubo", "Deposito", "Data Leitura", "Laboratorista",
                     "Ae. Larvas", "Ae. Pupas", "Ae. Exuvias", "Ae. Adulto",
                     "Alb. Larvas", "Alb. Pupas", "Alb. Exuvias", "Alb. Adulto",
                     "Outra Larvas", "Outra Pupas", "Outra Exuvias", "Outra Adulto", "Agentes"]
        return _gerar_xlsx(cabecalho, rows, "laboratorio")
    except Exception:
        logging.exception("Erro em exportar_laboratorio")
        return jsonify({"erro": "Erro interno. Verifique endemias.log"}), 500


def _gerar_xlsx(cabecalho, rows, nome):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nome[:31]
    fill = PatternFill("solid", fgColor="1A4FBA")
    for ci, col in enumerate(cabecalho, 1):
        cell = ws.cell(1, ci, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        vals = list(row.values()) if isinstance(row, dict) else list(row)
        for ci, value in enumerate(vals, 1):
            ws.cell(ri, ci, excel_safe(value))
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{nome}_{ts}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _info_consolidado(tipo):
    caminho = os.path.join(_saida_dir(), f"{tipo}_consolidado.xlsx")
    if not os.path.exists(caminho):
        return {
            "tipo": tipo,
            "existe": False,
            "download": f"/saida/download/{tipo}",
            "gerado_em": None,
            "tamanho": 0,
        }
    mtime = os.path.getmtime(caminho)
    return {
        "tipo": tipo,
        "existe": True,
        "download": f"/saida/download/{tipo}",
        "gerado_em": datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M"),
        "tamanho": os.path.getsize(caminho),
    }


@bp.route("/saida/consolidados/status")
@login_required
def consolidados_status():
    return jsonify({
        "tipos": [_info_consolidado(tipo) for tipo in work_types.WORK_TYPE_CODES]
    })


@bp.route("/saida/gerar-consolidados", methods=["POST"])
@login_required
@nivel_min("operador")
def gerar_consolidados():
    try:
        payload = request.get_json(silent=True) or {}
        tipo = (payload.get("tipo") or "TODOS").upper()
        tipos = list(work_types.WORK_TYPE_CODES) if tipo == "TODOS" else [tipo]
        if any(t not in work_types.WORK_TYPE_CODES for t in tipos):
            abort(404)

        from gerar_consolidado import gerar_todos

        logs = []

        class JsonLogger:
            def log(self, texto, tag="normal"):
                logs.append({"msg": texto, "tag": tag})

        resultados = gerar_todos(
            logger=JsonLogger(),
            banco_dados=_db_target(),
            pasta_saida=_saida_dir(),
            tipos=tipos,
        ) or []
        return jsonify({
            "ok": True,
            "logs": logs,
            "resultados": resultados,
            "tipos": [_info_consolidado(t) for t in work_types.WORK_TYPE_CODES],
        })
    except Exception:
        logging.exception("Erro ao gerar consolidados")
        return jsonify({"ok": False, "erro": "Erro interno. Verifique endemias.log"}), 500


@bp.route("/saida/download/<tipo>")
@login_required
def saida_download(tipo):
    tipo = (tipo or "").upper()
    if tipo not in work_types.WORK_TYPE_CODES:
        abort(404)
    caminho = os.path.join(_saida_dir(), f"{tipo}_consolidado.xlsx")
    if not os.path.exists(caminho):
        return f"Arquivo {tipo}_consolidado.xlsx ainda nao gerado. Execute um processamento primeiro.", 404
    return send_file(
        caminho,
        as_attachment=True,
        download_name=f"{tipo}_consolidado_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )
