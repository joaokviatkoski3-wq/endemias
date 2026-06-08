import io
import logging
from datetime import datetime

import openpyxl
from flask import Blueprint, current_app, jsonify, render_template, request, send_file
from openpyxl.styles import Alignment, Font, PatternFill

from app_core import auth as auth_core
from app_core import boletim_mensal as boletim_core
from app_core import blueprint_helpers as bh
from app_core.excel import excel_safe


bp = Blueprint("boletim_mensal", __name__)
login_required = auth_core.login_required
nivel_min = bh.nivel_min


def _db_path():
    return current_app.config["DB_PATH"]


def _mes_arg():
    return request.args.get("mes") or request.form.get("mes") or ""


@bp.route("/boletim-mensal")
@login_required
def page():
    boletim_core.ensure_schema(_db_path())
    return render_template("boletim_mensal.html")


@bp.route("/api/boletim-mensal")
@login_required
def api_boletim():
    usar_salvos = request.args.get("modo") != "auto"
    return jsonify(boletim_core.boletim(_db_path(), _mes_arg(), usar_salvos=usar_salvos))


@bp.route("/api/boletim-mensal", methods=["POST"])
@login_required
@nivel_min("operador")
def api_salvar():
    dados = request.get_json(silent=True) or {}
    try:
        boletim = boletim_core.salvar(_db_path(), dados.get("mes", ""), dados.get("linhas", []))
        return jsonify({"ok": True, **boletim})
    except Exception:
        logging.exception("Erro ao salvar boletim mensal")
        return jsonify({"erro": "Erro ao salvar boletim mensal."}), 500


@bp.route("/boletim-mensal/pdf")
@login_required
def pdf():
    dados = boletim_core.boletim(_db_path(), _mes_arg(), usar_salvos=True)
    return render_template("boletim_mensal_pdf.html", **dados)


@bp.route("/api/boletim-mensal/exportar")
@login_required
def exportar_xlsx():
    dados = boletim_core.boletim(_db_path(), _mes_arg(), usar_salvos=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boletim Mensal"

    fill_title = PatternFill("solid", fgColor="14532D")
    fill_head = PatternFill("solid", fgColor="B6D7A8")
    fill_total = PatternFill("solid", fgColor="D9EAD3")

    ws.merge_cells("A1:C1")
    ws["A1"] = "PREFEITURA MUNICIPAL DE ALMIRANTE TAMANDARÉ"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:C2")
    ws["A2"] = "SECRETARIA MUNICIPAL DE SAÚDE - VIGILÂNCIA AMBIENTAL | BOLETIM MENSAL - ARBOVIROSES E ZOONOSES"
    ws["A2"].font = Font(italic=True, color="1A4FBA")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:C3")
    ws["A3"] = dados["periodo"]["label"]
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = ["Indicador", "Quantidade", "Unidade"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(5, col, header)
        cell.font = Font(bold=True)
        cell.fill = fill_head
        cell.alignment = Alignment(horizontal="center")

    row_idx = 6
    for item in dados["linhas"]:
        if not item.get("ativo"):
            continue
        ws.cell(row_idx, 1, excel_safe(item["indicador"]))
        ws.cell(row_idx, 2, int(item.get("quantidade") or 0))
        ws.cell(row_idx, 3, excel_safe(item.get("unidade") or ""))
        ws.cell(row_idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="right")
        row_idx += 1

    ws.cell(row_idx, 1, "TOTAL")
    ws.cell(row_idx, 2, dados["total"])
    for col in range(1, 4):
        ws.cell(row_idx, col).font = Font(bold=True)
        ws.cell(row_idx, col).fill = fill_total
    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = f"boletim_mensal_{dados['periodo']['ano_mes']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
