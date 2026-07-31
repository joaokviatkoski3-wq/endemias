import io
import json
import secrets
import string
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Blueprint, current_app, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl.styles import Font, PatternFill

from app_core import audit
from app_core import auth as auth_core
from app_core import backup as backup_core
from app_core import backup_completo as backup_completo_core
from app_core import blueprint_helpers as bh
from app_core import diagnostico as diagnostico_core
from app_core import dbml as dbml_core
from app_core import db as db_core
from app_core import import_history
from app_core import sqlite_maintenance
from app_core import usuarios as usuarios_core
from app_core import version as version_core


bp = Blueprint("admin", __name__)
login_required = auth_core.login_required
nivel_min = bh.nivel_min


def _filtros_auditoria():
    return {
        "acao": request.args.get("acao", "").strip(),
        "usuario": request.args.get("usuario", "").strip(),
        "entidade": request.args.get("entidade", "").strip(),
        "d_ini": request.args.get("d_ini", "").strip(),
        "d_fim": request.args.get("d_fim", "").strip(),
    }


def _excel_safe(value):
    text = "" if value is None else str(value)
    return "'" + text if text[:1] in ("=", "+", "-", "@") else text


def _bytes_label(value):
    value = int(value or 0)
    unidades = ("B", "KB", "MB", "GB")
    tamanho = float(value)
    for unidade in unidades:
        if tamanho < 1024 or unidade == unidades[-1]:
            return f"{tamanho:.1f} {unidade}" if unidade != "B" else f"{value} B"
        tamanho /= 1024


def _db_target():
    return db_core.configured_target(current_app.config)


def _sqlite_only_message(operacao):
    if _db_target().backend == "sqlite":
        return None
    return redirect(url_for(
        "admin.admin_sistema",
        backup_erro=(
            f"{operacao} indisponível com PostgreSQL: use o procedimento "
            "operacional homologado para esse banco."
        ),
    ))


def _db_status():
    target = _db_target()
    if target.backend == "postgresql":
        status = {
            "backend": "postgresql",
            "path": target.location,
            "nome": target.location,
            "existe": True,
            "tamanho": "0 B",
            "wal": False,
            "wal_tamanho": "N/A",
            "shm": False,
            "shm_tamanho": "N/A",
            "integridade": "nao verificado",
            "tabelas": 0,
            "journal_mode": "PostgreSQL WAL",
            "synchronous": "Servidor",
            "busy_timeout_ms": 0,
            "wal_autocheckpoint": 0,
            "indices": 0,
            "indices_essenciais": {"total": 0, "presentes": 0, "faltantes": [], "ok": True},
            "metricas": db_core.connection_metrics(),
        }
        conn = None
        try:
            conn = bh.get_db()
            row = conn.execute(
                """SELECT current_database(), pg_database_size(current_database()),
                          current_setting('server_version')"""
            ).fetchone()
            status["nome"] = row[0]
            status["path"] = f"PostgreSQL: {row[0]} (servidor {row[2]})"
            status["tamanho"] = _bytes_label(row[1])
            status["tabelas"] = conn.execute(
                """SELECT COUNT(*) FROM information_schema.tables
                    WHERE table_schema=current_schema() AND table_type='BASE TABLE'"""
            ).fetchone()[0]
            status["indices"] = conn.execute(
                "SELECT COUNT(*) FROM pg_indexes WHERE schemaname=current_schema()"
            ).fetchone()[0]
            status["integridade"] = "ok"
        except Exception as exc:
            status["existe"] = False
            status["integridade"] = "erro"
            status["erro"] = str(exc)
        finally:
            if conn is not None:
                conn.close()
        return status

    db_path = Path(current_app.config["DB_PATH"])
    wal_path = Path(str(db_path) + "-wal")
    shm_path = Path(str(db_path) + "-shm")
    status = {
        "backend": "sqlite",
        "path": str(db_path),
        "nome": db_path.name,
        "existe": db_path.exists(),
        "tamanho": _bytes_label(db_path.stat().st_size) if db_path.exists() else "0 B",
        "wal": wal_path.exists(),
        "wal_tamanho": _bytes_label(wal_path.stat().st_size) if wal_path.exists() else "0 B",
        "shm": shm_path.exists(),
        "shm_tamanho": _bytes_label(shm_path.stat().st_size) if shm_path.exists() else "0 B",
        "integridade": "nao verificado",
        "tabelas": 0,
        "journal_mode": "-",
        "synchronous": "-",
        "busy_timeout_ms": 0,
        "wal_autocheckpoint": 0,
        "indices": 0,
        "indices_essenciais": {"total": 0, "presentes": 0, "faltantes": [], "ok": False},
        "metricas": db_core.connection_metrics(),
    }
    if not db_path.exists():
        return status

    conn = bh.get_db()
    try:
        status["integridade"] = conn.execute("PRAGMA integrity_check").fetchone()[0]
        status["tabelas"] = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchone()[0]
        status["journal_mode"] = conn.execute("PRAGMA journal_mode").fetchone()[0].upper()
        synchronous = conn.execute("PRAGMA synchronous").fetchone()[0]
        status["synchronous"] = {0: "OFF", 1: "NORMAL", 2: "FULL", 3: "EXTRA"}.get(
            synchronous, str(synchronous)
        )
        status["busy_timeout_ms"] = conn.execute("PRAGMA busy_timeout").fetchone()[0]
        status["wal_autocheckpoint"] = conn.execute("PRAGMA wal_autocheckpoint").fetchone()[0]
        status["indices"] = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type='index'"
        ).fetchone()[0]
        status["indices_essenciais"] = sqlite_maintenance.performance_index_status(conn)
    finally:
        conn.close()
    return status


def _contagens_sistema():
    conn = bh.get_db()
    try:
        return {
            "usuarios_ativos": conn.execute("SELECT COUNT(*) FROM usuarios WHERE ativo=1").fetchone()[0],
            "visitas_total": conn.execute("SELECT COUNT(*) FROM visitas").fetchone()[0],
            "focos_pendentes": conn.execute(
                "SELECT COUNT(*) FROM focos_positivos WHERE status_notificacao='pendente' AND gera_notificacao=1"
            ).fetchone()[0],
            "eventos_auditoria": conn.execute(
                "SELECT COUNT(*) FROM auditoria_eventos"
            ).fetchone()[0] if db_core.table_exists(conn, "auditoria_eventos") else 0,
        }
    finally:
        conn.close()


@bp.route("/admin/usuarios")
@login_required
@nivel_min("admin")
def admin_usuarios():
    usuarios = usuarios_core.listar(bh.db_target())
    return render_template("admin_usuarios.html", usuarios=usuarios)


@bp.route("/admin/sistema")
@login_required
@nivel_min("admin")
def admin_sistema():
    target = _db_target()
    db_path = Path(target.location) if target.backend == "sqlite" else target.location
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    backup_completo_dir = Path(current_app.config["BACKUP_COMPLETO_DIR"])
    backups = backup_core.listar_backups(backup_dir, limite=20)
    backups_completos = backup_completo_core.listar_backups_completos(backup_completo_dir, limite=20)
    importacoes = import_history.listar_importacoes_recentes(bh.get_db, limite=5)
    eventos = audit.listar_eventos(bh.get_db, limite=8)
    conn = bh.get_db()
    try:
        diagnostico = diagnostico_core.gerar(conn, db_path=db_path, backup_dir=backup_dir)
    finally:
        conn.close()
    return render_template(
        "admin_sistema.html",
        db_status=_db_status(),
        contagens=_contagens_sistema(),
        backups=backups,
        backups_completos=backups_completos,
        diagnostico=diagnostico,
        backup_dir=str(backup_dir),
        backup_completo_dir=str(backup_completo_dir),
        importacoes=importacoes,
        eventos=eventos,
        app_version=version_core.APP_VERSION_LABEL,
        instance_dir=current_app.config["INSTANCE_DIR"],
        upload_temp=current_app.config["UPLOAD_TEMP"],
        log_path=current_app.config["LOG_PATH"],
        backup_ok=request.args.get("backup_ok", "").strip(),
        backup_erro=request.args.get("backup_erro", "").strip(),
        format_bytes=_bytes_label,
        sqlite_operations=target.backend == "sqlite",
    )


@bp.route("/api/admin/sistema/diagnostico")
@login_required
@nivel_min("admin")
def api_admin_diagnostico():
    target = _db_target()
    db_path = Path(target.location) if target.backend == "sqlite" else target.location
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    completo = request.args.get("completo", "").strip().lower() in {"1", "sim", "true", "completo"}
    conn = bh.get_db()
    try:
        return jsonify(diagnostico_core.gerar(conn, db_path=db_path, backup_dir=backup_dir, completo=completo))
    finally:
        conn.close()


@bp.route("/admin/sistema/backups/criar", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_criar_backup():
    bloqueio = _sqlite_only_message("Backup interno")
    if bloqueio:
        return bloqueio
    db_path = Path(current_app.config["DB_PATH"])
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    try:
        with backup_core.operacao_exclusiva():
            info = backup_core.criar_backup_sqlite(db_path, destino_dir=backup_dir, prefixo="endemias", manter=20)
        audit.registrar_evento(
            bh.get_db,
            "backup_criado",
            entidade="backups",
            entidade_id=Path(info["arquivo"]).name,
            detalhes={
                "arquivo": Path(info["arquivo"]).name,
                "tamanho_bytes": info["tamanho_bytes"],
                "integridade": info["integridade"],
            },
        )
        msg = f"Backup criado: {Path(info['arquivo']).name}"
        return redirect(url_for("admin.admin_sistema", backup_ok=msg))
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao criar backup: {exc}"))


@bp.route("/admin/sistema/backups/restaurar", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_restaurar_backup():
    bloqueio = _sqlite_only_message("Restauração interna")
    if bloqueio:
        return bloqueio
    db_path = Path(current_app.config["DB_PATH"])
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    nome_backup = request.form.get("backup", "").strip()
    try:
        with backup_core.operacao_exclusiva():
            backup_path = backup_core.resolver_backup(backup_dir, nome_backup)
            seguranca = backup_core.criar_backup_sqlite(
                db_path,
                destino_dir=backup_dir,
                prefixo="pre_restore",
                manter=20,
            )
            info = backup_core.restaurar_backup_sqlite(db_path, backup_path)
        audit.registrar_evento(
            bh.get_db,
            "backup_restaurado",
            entidade="backups",
            entidade_id=backup_path.name,
            detalhes={
                "backup_restaurado": backup_path.name,
                "backup_seguranca": Path(seguranca["arquivo"]).name,
                "integridade": info["integridade"],
            },
        )
        msg = f"Backup restaurado: {backup_path.name}. Copia de seguranca criada: {Path(seguranca['arquivo']).name}"
        return redirect(url_for("admin.admin_sistema", backup_ok=msg))
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao restaurar backup: {exc}"))


@bp.route("/admin/sistema/backups/baixar/<nome_backup>")
@login_required
@nivel_min("admin")
def admin_baixar_backup(nome_backup):
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    try:
        backup_path = backup_core.resolver_backup(backup_dir, nome_backup)
        return send_file(backup_path, as_attachment=True, download_name=backup_path.name)
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao baixar backup: {exc}"))


@bp.route("/admin/sistema/dbml")
@login_required
@nivel_min("admin")
def admin_baixar_dbml():
    bloqueio = _sqlite_only_message("Exportação DBML SQLite")
    if bloqueio:
        return bloqueio
    try:
        db_path = Path(current_app.config["DB_PATH"])
        conteudo = dbml_core.gerar_dbml(db_path)
        nome = f"{db_path.stem}_schema.dbml"
        audit.registrar_evento(
            bh.get_db,
            "dbml_baixado",
            entidade="banco",
            entidade_id=db_path.name,
            detalhes={"arquivo": nome},
        )
        return send_file(
            io.BytesIO(conteudo.encode("utf-8")),
            as_attachment=True,
            download_name=nome,
            mimetype="text/plain; charset=utf-8",
        )
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao gerar DBML: {exc}"))


@bp.route("/admin/sistema/backups/excluir", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_excluir_backup():
    backup_dir = Path(current_app.config["BACKUP_DIR"])
    nome_backup = request.form.get("backup", "").strip()
    try:
        with backup_core.operacao_exclusiva():
            backup_path = backup_core.resolver_backup(backup_dir, nome_backup)
            backup_core.excluir_backup(backup_path)
        audit.registrar_evento(
            bh.get_db,
            "backup_excluido",
            entidade="backups",
            entidade_id=backup_path.name,
            detalhes={"arquivo": backup_path.name},
        )
        return redirect(url_for("admin.admin_sistema", backup_ok=f"Backup excluido: {backup_path.name}"))
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao excluir backup: {exc}"))


@bp.route("/admin/sistema/backups-completos/criar", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_criar_backup_completo():
    bloqueio = _sqlite_only_message("Backup completo interno")
    if bloqueio:
        return bloqueio
    try:
        with backup_core.operacao_exclusiva():
            info = backup_completo_core.criar_backup_completo(
                destino_dir=current_app.config["BACKUP_COMPLETO_DIR"],
                manter=10,
                db_path=current_app.config["DB_PATH"],
                raiz=current_app.config["INSTANCE_DIR"],
                anexos_dir=current_app.config["ANEXOS_DIR"],
                kobo_config_path=current_app.config["KOBO_CONFIG_PATH"],
                secret_key_path=current_app.config["SECRET_KEY_PATH"],
            )
        audit.registrar_evento(
            bh.get_db,
            "backup_completo_criado",
            entidade="backups_completos",
            entidade_id=info["nome"],
            detalhes={
                "arquivo": info["nome"],
                "tamanho_bytes": info["tamanho_bytes"],
                "integridade_banco": info["integridade_banco"],
                "destino": current_app.config["BACKUP_COMPLETO_DIR"],
            },
        )
        return redirect(url_for("admin.admin_sistema", backup_ok=f"Backup completo criado: {info['nome']}"))
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao criar backup completo: {exc}"))


@bp.route("/admin/sistema/backups-completos/baixar/<nome_backup>")
@login_required
@nivel_min("admin")
def admin_baixar_backup_completo(nome_backup):
    try:
        backup_path = backup_completo_core.resolver_backup_completo(
            current_app.config["BACKUP_COMPLETO_DIR"],
            nome_backup,
        )
        return send_file(backup_path, as_attachment=True, download_name=backup_path.name)
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao baixar backup completo: {exc}"))


@bp.route("/admin/sistema/backups-completos/excluir", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_excluir_backup_completo():
    nome_backup = request.form.get("backup", "").strip()
    try:
        with backup_core.operacao_exclusiva():
            backup_path = backup_completo_core.resolver_backup_completo(
                current_app.config["BACKUP_COMPLETO_DIR"],
                nome_backup,
            )
            backup_completo_core.excluir_backup_completo(backup_path)
        audit.registrar_evento(
            bh.get_db,
            "backup_completo_excluido",
            entidade="backups_completos",
            entidade_id=backup_path.name,
            detalhes={"arquivo": backup_path.name},
        )
        return redirect(url_for("admin.admin_sistema", backup_ok=f"Backup completo excluido: {backup_path.name}"))
    except Exception as exc:
        return redirect(url_for("admin.admin_sistema", backup_erro=f"Erro ao excluir backup completo: {exc}"))


@bp.route("/admin/auditoria")
@login_required
@nivel_min("admin")
def admin_auditoria():
    filtros = _filtros_auditoria()
    eventos = audit.listar_eventos(bh.get_db, filtros, limite=200)
    return render_template("admin_auditoria.html", eventos=eventos, filtros=filtros)


@bp.route("/admin/auditoria/exportar")
@login_required
@nivel_min("admin")
def admin_auditoria_exportar():
    filtros = _filtros_auditoria()
    eventos = audit.listar_eventos(bh.get_db, filtros, limite=500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    headers = ["Data", "Acao", "Usuario", "IP", "Entidade", "Entidade ID", "Detalhes"]
    fill = PatternFill("solid", fgColor="1A4FBA")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for row_idx, ev in enumerate(eventos, 2):
        detalhes = json.dumps(ev.get("detalhes") or {}, ensure_ascii=False, sort_keys=True)
        values = [
            ev.get("criado_em", ""),
            ev.get("acao", ""),
            ev.get("usuario_nome", ""),
            ev.get("ip", ""),
            ev.get("entidade", ""),
            ev.get("entidade_id", ""),
            detalhes,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row_idx, col, _excel_safe(value))
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"auditoria_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/admin/usuarios/criar", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_criar_usuario():
    dados = request.form.to_dict()
    erro = None
    try:
        novo_id = usuarios_core.criar(bh.db_target(), dados)
        somente_laboratorio = (
            1 if dados.get("somente_laboratorio") == "1" else 0
        )
        acesso_laboratorio = (
            1
            if somente_laboratorio
            or dados.get("acesso_laboratorio") == "1"
            else 0
        )
        audit.registrar_evento(
            bh.get_db,
            "usuario_criado",
            entidade="usuarios",
            entidade_id=novo_id,
            detalhes={
                "usuario": dados.get("usuario", "").strip().lower(),
                "nome": dados.get("nome", "").strip(),
                "nivel": dados.get("nivel", "visualizador"),
                "acesso_laboratorio": acesso_laboratorio,
                "somente_laboratorio": somente_laboratorio,
            },
        )
    except Exception as exc:
        erro = f"Erro: {exc}"
    if erro:
        usuarios = usuarios_core.listar(bh.db_target())
        return render_template("admin_usuarios.html", usuarios=usuarios, erro=erro)
    return redirect(url_for("admin.admin_usuarios"))


@bp.route("/admin/usuarios/<int:uid>/editar", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_editar_usuario(uid):
    campo = request.form.get("campo")
    valor = request.form.get("valor", "").strip()
    try:
        anterior, novo = usuarios_core.editar(
            bh.db_target(),
            uid,
            campo,
            valor,
            usuario_atual_id=session.get("uid"),
        )
    except ValueError as exc:
        return jsonify({"erro": str(exc)}), 400
    detalhes = {"campo": campo}
    detalhes.update({
        "usuario": anterior["usuario"],
        "valor_antigo": anterior.get(campo),
        "valor_novo": novo,
    })
    audit.registrar_evento(
        bh.get_db,
        "usuario_editado",
        entidade="usuarios",
        entidade_id=uid,
        detalhes=detalhes,
    )
    return jsonify({"ok": True})


@bp.route("/admin/usuarios/<int:uid>/resetar-senha", methods=["POST"])
@login_required
@nivel_min("admin")
def admin_resetar_senha(uid):
    tamanho = max(auth_core.PASSWORD_MIN_LENGTH, 12)
    nova = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(tamanho))
    try:
        alvo = usuarios_core.resetar_senha(bh.db_target(), uid, nova)
    except ValueError as exc:
        return jsonify({"erro": str(exc)}), 400
    audit.registrar_evento(
        bh.get_db,
        "usuario_senha_resetada",
        entidade="usuarios",
        entidade_id=uid,
        detalhes={"usuario": alvo["usuario"], "nome": alvo["nome"]},
    )
    return jsonify({"ok": True, "senha": nova})
