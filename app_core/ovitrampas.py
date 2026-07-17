import csv
import hashlib
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from app_core import db as db_core


TABLE = "ovitrampas_leituras"
ARMADILHAS_TABLE = "ovitrampas_armadilhas"
OCORRENCIAS_TABLE = "ovitrampas_ocorrencias_conta_ovos"
CAL_GRUPOS_TABLE = "ovitrampas_calendario_grupos"
CAL_EVENTOS_TABLE = "ovitrampas_calendario_eventos"
CAL_AGENTES_TABLE = "ovitrampas_calendario_agentes"
DIARIOS_TABLE = "ovitrampas_diarios"
DIARIO_ARMADILHAS_TABLE = "ovitrampas_diario_armadilhas"
ARMADILHAS_HISTORICO_TABLE = "ovitrampas_armadilhas_historico"

OCORRENCIAS = {
    1: "Intervalo maior que 7 dias",
    2: "Armadilha ou palheta desaparecida",
    3: "Armadilha ou palheta danificada",
    4: "Armadilha ou palheta removida",
    5: "Armadilha seca",
    6: "Casa fechada",
    7: "Ovitrampa cheia de agua",
    8: "Pouca agua",
    9: "Outros",
}

CONTA_OVOS_OCORRENCIAS = {
    2: 1,
    3: 2,
    4: 3,
    5: 4,
    6: 5,
    7: 6,
    8: 7,
    9: 8,
    10: 9,
}

MOVIMENTOS = {
    "instalacao": "Instalação",
    "troca": "Troca",
    "retirada": "Retirada",
    "feriado": "Feriado",
}

MESES_PT = (
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
)

GRUPOS_PADRAO = (
    ("Tanguá / Paraíso", "Tanguá, Paraíso", "#facc15"),
    ("Sede / São Francisco / Graziela / Tamboara / Rosana", "Sede, São Francisco, Graziela, Tamboara, Rosana", "#22c55e"),
    ("Tranqueira / São João Batista", "Tranqueira, São João Batista", "#3b82f6"),
    ("Cachoeira / São Venâncio / Roma", "Cachoeira, São Venâncio, Roma", "#f97316"),
    ("Lamenha / Santa Maria", "Lamenha, Santa Maria", "#a855f7"),
)


def ensure_schema(conn):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS ovitrampas_armadilhas (
            ovitrampa_id    TEXT PRIMARY KEY,
            rua             TEXT,
            numero          TEXT,
            complemento     TEXT,
            bairro          TEXT,
            localizacao     TEXT,
            localidade      TEXT,
            responsavel     TEXT,
            telefone_responsavel TEXT,
            quarteirao      TEXT,
            latitude        REAL,
            longitude       REAL,
            ativo           INTEGER NOT NULL DEFAULT 1 CHECK(ativo IN (0,1)),
            arquivo_origem  TEXT,
            atualizado_em   TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS ovitrampas_leituras (
            id_leitura          TEXT PRIMARY KEY,
            ovitrampa_id        TEXT NOT NULL,
            estado              TEXT,
            municipio           TEXT,
            distrito            TEXT,
            rua                 TEXT,
            numero              TEXT,
            complemento         TEXT,
            localizacao         TEXT,
            latitude            REAL,
            longitude           REAL,
            ano                 INTEGER NOT NULL,
            semana              INTEGER NOT NULL,
            data_envio_contagem TEXT,
            ovos                INTEGER DEFAULT 0,
            quem_enviou         TEXT,
            observacao          TEXT,
            lat_lng             TEXT,
            quarteirao          TEXT,
            data_instalacao     DATE,
            data_coleta         DATE,
            id_laboratorista    INTEGER REFERENCES agentes(id_agente),
            data_leitura        DATE,
            arquivo_origem      TEXT,
            importado_em        TEXT NOT NULL,
            UNIQUE(ovitrampa_id, ano, semana, data_instalacao, data_coleta, data_envio_contagem)
        );

        CREATE INDEX IF NOT EXISTS idx_ovitrampas_armadilhas_localidade ON ovitrampas_armadilhas(localidade);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_armadilhas_quarteirao ON ovitrampas_armadilhas(quarteirao);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_ano_semana ON ovitrampas_leituras(ano, semana);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_id ON ovitrampas_leituras(ovitrampa_id);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_distrito ON ovitrampas_leituras(distrito);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_coleta ON ovitrampas_leituras(data_coleta);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_ovos ON ovitrampas_leituras(ovos);

        CREATE TABLE IF NOT EXISTS ovitrampas_ocorrencias_conta_ovos (
            id_contagem            TEXT PRIMARY KEY,
            ovitrampa_id           TEXT NOT NULL,
            ano                    INTEGER NOT NULL,
            semana                 INTEGER NOT NULL,
            data                   DATE,
            data_envio_contagem    TEXT,
            ovos                   INTEGER DEFAULT 0,
            resultado              TEXT,
            codigo_conta_ovos      INTEGER,
            observacao_conta_ovos  TEXT,
            ocorrencia_codigo      INTEGER,
            latitude               REAL,
            longitude              REAL,
            lat_lng                TEXT,
            arquivo_origem         TEXT,
            importado_em           TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_ovi_ocorrencias_ano_semana ON ovitrampas_ocorrencias_conta_ovos(ano, semana);
        CREATE INDEX IF NOT EXISTS idx_ovi_ocorrencias_id ON ovitrampas_ocorrencias_conta_ovos(ovitrampa_id);
        CREATE INDEX IF NOT EXISTS idx_ovi_ocorrencias_codigo ON ovitrampas_ocorrencias_conta_ovos(ocorrencia_codigo);

        CREATE TABLE IF NOT EXISTS ovitrampas_calendario_grupos (
            id_grupo      INTEGER PRIMARY KEY AUTOINCREMENT,
            nome          TEXT NOT NULL,
            localidades   TEXT,
            cor           TEXT NOT NULL DEFAULT '#0f766e',
            ativo         INTEGER NOT NULL DEFAULT 1 CHECK(ativo IN (0,1)),
            criado_em     TEXT NOT NULL,
            atualizado_em TEXT
        );

        CREATE TABLE IF NOT EXISTS ovitrampas_calendario_eventos (
            id_evento     INTEGER PRIMARY KEY AUTOINCREMENT,
            data          DATE NOT NULL UNIQUE,
            movimento     TEXT NOT NULL CHECK(movimento IN ('instalacao','troca','retirada','feriado')),
            titulo        TEXT,
            id_grupo      INTEGER REFERENCES ovitrampas_calendario_grupos(id_grupo),
            ciclo         TEXT,
            observacoes   TEXT,
            criado_por    TEXT,
            criado_em     TEXT NOT NULL,
            atualizado_em TEXT
        );

        CREATE TABLE IF NOT EXISTS ovitrampas_calendario_agentes (
            id_evento INTEGER NOT NULL REFERENCES ovitrampas_calendario_eventos(id_evento) ON DELETE CASCADE,
            id_agente INTEGER NOT NULL REFERENCES agentes(id_agente),
            PRIMARY KEY (id_evento, id_agente)
        );

        CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_eventos_data ON ovitrampas_calendario_eventos(data);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_eventos_grupo ON ovitrampas_calendario_eventos(id_grupo);
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_agentes_agente ON ovitrampas_calendario_agentes(id_agente);

        CREATE TABLE IF NOT EXISTS ovitrampas_diarios (
            id_diario     INTEGER PRIMARY KEY AUTOINCREMENT,
            nome          TEXT NOT NULL UNIQUE,
            observacoes   TEXT,
            ativo         INTEGER NOT NULL DEFAULT 1 CHECK(ativo IN (0,1)),
            criado_em     TEXT NOT NULL,
            atualizado_em TEXT
        );

        CREATE TABLE IF NOT EXISTS ovitrampas_diario_armadilhas (
            id_diario     INTEGER NOT NULL REFERENCES ovitrampas_diarios(id_diario) ON DELETE CASCADE,
            ovitrampa_id  TEXT NOT NULL REFERENCES ovitrampas_armadilhas(ovitrampa_id) ON DELETE CASCADE,
            ordem         INTEGER NOT NULL DEFAULT 0,
            observacoes   TEXT,
            criado_em     TEXT NOT NULL,
            atualizado_em TEXT,
            PRIMARY KEY (id_diario, ovitrampa_id)
        );

        CREATE TABLE IF NOT EXISTS ovitrampas_armadilhas_historico (
            id_historico  INTEGER PRIMARY KEY AUTOINCREMENT,
            ovitrampa_id  TEXT NOT NULL,
            campo         TEXT NOT NULL,
            valor_anterior TEXT,
            valor_novo    TEXT,
            motivo        TEXT,
            agentes       TEXT,
            usuario       TEXT,
            arquivo_origem TEXT,
            criado_em     TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_ovi_diario_arm_diario ON ovitrampas_diario_armadilhas(id_diario, ordem);
        CREATE INDEX IF NOT EXISTS idx_ovi_diario_arm_ovitrampa ON ovitrampas_diario_armadilhas(ovitrampa_id);
        CREATE INDEX IF NOT EXISTS idx_ovi_arm_hist_ovitrampa ON ovitrampas_armadilhas_historico(ovitrampa_id, criado_em);
        """
    )
    _ensure_columns(conn, ARMADILHAS_TABLE, {
        "telefone_responsavel": "TEXT",
        "ativo": "INTEGER NOT NULL DEFAULT 1",
    })
    _ensure_columns(conn, TABLE, {
        "id_laboratorista": "INTEGER REFERENCES agentes(id_agente)",
        "data_leitura": "DATE",
        "ocorrencia_codigo": "INTEGER",
    })
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ovitrampas_laboratorista ON ovitrampas_leituras(id_laboratorista)")
    _migrar_calendario_schema(conn)
    _migrar_calendario_agentes_schema(conn)
    _semear_grupos_padrao(conn)
    _normalizar_localidades_existentes(conn)
    _normalizar_ids_armadilhas_existentes(conn)
    _normalizar_quarteiroes_existentes(conn)


def _ensure_columns(conn, table, columns):
    existentes = {row[1] for row in conn.execute(f"PRAGMA table_info({table})")}
    for column, definition in columns.items():
        if column not in existentes:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def _normalizar_quarteiroes_existentes(conn):
    agora = datetime.now().isoformat(timespec="seconds")
    contexto = {
        "motivo": "Padronizacao automatica do quarteirao em quatro digitos",
        "agentes": "",
        "usuario": "sistema",
        "arquivo_origem": None,
        "criado_em": agora,
    }
    armadilhas = conn.execute(
        f"SELECT ovitrampa_id, quarteirao FROM {ARMADILHAS_TABLE} WHERE quarteirao IS NOT NULL"
    ).fetchall()
    for row in armadilhas:
        anterior = row["quarteirao"]
        novo = _normalizar_quarteirao(anterior)
        if novo == anterior:
            continue
        conn.execute(
            f"UPDATE {ARMADILHAS_TABLE} SET quarteirao=?, atualizado_em=? WHERE ovitrampa_id=?",
            (novo, agora, row["ovitrampa_id"]),
        )
        _registrar_alteracao_armadilha(
            conn, row["ovitrampa_id"], "quarteirao", anterior, novo, contexto
        )

    leituras = conn.execute(
        f"SELECT rowid, quarteirao FROM {TABLE} WHERE quarteirao IS NOT NULL"
    ).fetchall()
    for row in leituras:
        novo = _normalizar_quarteirao(row["quarteirao"])
        if novo != row["quarteirao"]:
            conn.execute(
                f"UPDATE {TABLE} SET quarteirao=? WHERE rowid=?",
                (novo, row["rowid"]),
            )


def _semear_grupos_padrao(conn):
    total = conn.execute(f"SELECT COUNT(*) FROM {CAL_GRUPOS_TABLE}").fetchone()[0]
    if total:
        return
    agora = datetime.now().isoformat(timespec="seconds")
    for nome, localidades, cor in GRUPOS_PADRAO:
        conn.execute(
            f"""INSERT INTO {CAL_GRUPOS_TABLE}
                (nome, localidades, cor, ativo, criado_em, atualizado_em)
                VALUES (?, ?, ?, 1, ?, ?)""",
            (nome, localidades, cor, agora, agora),
        )


def _normalizar_localidades_existentes(conn):
    for table, column in ((TABLE, "distrito"), (ARMADILHAS_TABLE, "localidade")):
        rows = conn.execute(
            f"SELECT rowid, {column} FROM {table} WHERE {column} IS NOT NULL AND TRIM({column})<>''"
        ).fetchall()
        for rowid, valor in rows:
            normalizado = _title_distrito(valor)
            if normalizado and normalizado != valor:
                conn.execute(
                    f"UPDATE {table} SET {column}=? WHERE rowid=?",
                    (normalizado, rowid),
                )


def _normalizar_ids_armadilhas_existentes(conn):
    rows = conn.execute(f"SELECT * FROM {ARMADILHAS_TABLE}").fetchall()
    for row in rows:
        antigo = row["ovitrampa_id"]
        novo = _normalizar_ovitrampa_id(antigo)
        if not novo or novo == antigo:
            continue
        alvo = conn.execute(f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (novo,)).fetchone()
        if alvo:
            _fundir_armadilha(conn, antigo, novo, row, alvo)
        else:
            _renomear_ovitrampa(conn, antigo, novo)


def _renomear_ovitrampa(conn, antigo, novo):
    row = conn.execute(f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (antigo,)).fetchone()
    if row:
        cols = list(row.keys())
        payload = [novo if col == "ovitrampa_id" else row[col] for col in cols]
        placeholders = ",".join("?" for _ in cols)
        conn.execute(
            f"INSERT OR IGNORE INTO {ARMADILHAS_TABLE} ({','.join(cols)}) VALUES ({placeholders})",
            payload,
        )
    for table in (TABLE, OCORRENCIAS_TABLE, DIARIO_ARMADILHAS_TABLE, ARMADILHAS_HISTORICO_TABLE):
        conn.execute(f"UPDATE {table} SET ovitrampa_id=? WHERE ovitrampa_id=?", (novo, antigo))
    conn.execute(f"DELETE FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (antigo,))


def _fundir_armadilha(conn, antigo, novo, origem, alvo):
    payload = {}
    for key in origem.keys():
        if key == "ovitrampa_id":
            continue
        origem_val = origem[key]
        alvo_val = alvo[key]
        if key == "telefone_responsavel":
            payload[key] = alvo_val or origem_val
        elif origem_val not in (None, ""):
            payload[key] = origem_val
        else:
            payload[key] = alvo_val
    sets = ", ".join(f"{key}=?" for key in payload)
    conn.execute(
        f"UPDATE {ARMADILHAS_TABLE} SET {sets} WHERE ovitrampa_id=?",
        [payload[key] for key in payload] + [novo],
    )
    conn.execute(f"UPDATE {TABLE} SET ovitrampa_id=? WHERE ovitrampa_id=?", (novo, antigo))
    conn.execute(f"UPDATE {OCORRENCIAS_TABLE} SET ovitrampa_id=? WHERE ovitrampa_id=?", (novo, antigo))
    conn.execute(f"UPDATE {ARMADILHAS_HISTORICO_TABLE} SET ovitrampa_id=? WHERE ovitrampa_id=?", (novo, antigo))
    links = conn.execute(
        f"SELECT id_diario, ordem, observacoes, criado_em, atualizado_em FROM {DIARIO_ARMADILHAS_TABLE} WHERE ovitrampa_id=?",
        (antigo,),
    ).fetchall()
    for link in links:
        existe = conn.execute(
            f"SELECT 1 FROM {DIARIO_ARMADILHAS_TABLE} WHERE id_diario=? AND ovitrampa_id=?",
            (link["id_diario"], novo),
        ).fetchone()
        if not existe:
            conn.execute(
                f"""INSERT INTO {DIARIO_ARMADILHAS_TABLE}
                    (id_diario, ovitrampa_id, ordem, observacoes, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?)""",
                (link["id_diario"], novo, link["ordem"], link["observacoes"], link["criado_em"], link["atualizado_em"]),
            )
    conn.execute(f"DELETE FROM {DIARIO_ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (antigo,))
    conn.execute(f"DELETE FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (antigo,))


def _migrar_calendario_schema(conn):
    sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
        (CAL_EVENTOS_TABLE,),
    ).fetchone()
    if not sql:
        return
    table_sql = sql[0] or ""
    precisa = "feriado" not in table_sql or "id_grupo      INTEGER NOT NULL" in table_sql or "titulo" not in table_sql
    if not precisa:
        return
    conn.executescript(f"""
        ALTER TABLE {CAL_EVENTOS_TABLE} RENAME TO {CAL_EVENTOS_TABLE}_old;
        CREATE TABLE {CAL_EVENTOS_TABLE} (
            id_evento     INTEGER PRIMARY KEY AUTOINCREMENT,
            data          DATE NOT NULL UNIQUE,
            movimento     TEXT NOT NULL CHECK(movimento IN ('instalacao','troca','retirada','feriado')),
            titulo        TEXT,
            id_grupo      INTEGER REFERENCES {CAL_GRUPOS_TABLE}(id_grupo),
            ciclo         TEXT,
            observacoes   TEXT,
            criado_por    TEXT,
            criado_em     TEXT NOT NULL,
            atualizado_em TEXT
        );
        INSERT INTO {CAL_EVENTOS_TABLE}
            (id_evento, data, movimento, titulo, id_grupo, ciclo, observacoes, criado_por, criado_em, atualizado_em)
        SELECT
            id_evento, data, movimento, NULL, id_grupo, ciclo, observacoes, criado_por, criado_em, atualizado_em
        FROM {CAL_EVENTOS_TABLE}_old;
        DROP TABLE {CAL_EVENTOS_TABLE}_old;
    """)
    conn.execute(f"CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_eventos_data ON {CAL_EVENTOS_TABLE}(data)")
    conn.execute(f"CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_eventos_grupo ON {CAL_EVENTOS_TABLE}(id_grupo)")


def _migrar_calendario_agentes_schema(conn):
    sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
        (CAL_AGENTES_TABLE,),
    ).fetchone()
    if not sql:
        return
    table_sql = sql[0] or ""
    if f'REFERENCES "{CAL_EVENTOS_TABLE}_old"' not in table_sql and f"REFERENCES {CAL_EVENTOS_TABLE}_old" not in table_sql:
        return
    conn.executescript(f"""
        ALTER TABLE {CAL_AGENTES_TABLE} RENAME TO {CAL_AGENTES_TABLE}_old;
        CREATE TABLE {CAL_AGENTES_TABLE} (
            id_evento INTEGER NOT NULL REFERENCES {CAL_EVENTOS_TABLE}(id_evento) ON DELETE CASCADE,
            id_agente INTEGER NOT NULL REFERENCES agentes(id_agente),
            PRIMARY KEY (id_evento, id_agente)
        );
        INSERT OR IGNORE INTO {CAL_AGENTES_TABLE} (id_evento, id_agente)
        SELECT old.id_evento, old.id_agente
          FROM {CAL_AGENTES_TABLE}_old old
         WHERE EXISTS (
               SELECT 1 FROM {CAL_EVENTOS_TABLE} e
                WHERE e.id_evento = old.id_evento
         );
        DROP TABLE {CAL_AGENTES_TABLE}_old;
        CREATE INDEX IF NOT EXISTS idx_ovitrampas_cal_agentes_agente ON {CAL_AGENTES_TABLE}(id_agente);
    """)


def importar_pasta(db_path, pasta, logger=None):
    paths = sorted(Path(pasta).glob("*.csv"), key=lambda p: p.name)
    total = {"arquivos": 0, "linhas": 0, "inseridos": 0, "duplicados": 0, "erros": []}
    for path in paths:
        result = importar_csv(db_path, path)
        total["arquivos"] += 1
        total["linhas"] += result["linhas"]
        total["inseridos"] += result["inseridos"]
        total["duplicados"] += result["duplicados"]
        total["erros"].extend(result["erros"])
        if logger:
            logger(f"{path.name}: {result['inseridos']} novo(s), {result['duplicados']} duplicado(s)")
    return total


def importar_csv(db_path, path):
    result = {"arquivo": os.path.basename(path), "linhas": 0, "inseridos": 0, "duplicados": 0, "erros": []}
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for idx, row in enumerate(reader, start=2):
                if not any((v or "").strip() for v in row.values()):
                    continue
                result["linhas"] += 1
                try:
                    registro = _registro(row, result["arquivo"], agora)
                    inserted = _insert(conn, registro)
                    if inserted:
                        result["inseridos"] += 1
                    else:
                        result["duplicados"] += 1
                except Exception as exc:
                    result["erros"].append(f"Linha {idx}: {exc}")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return result


def importar_armadilhas_csv(db_path, path, motivo=None, usuario=None):
    result = {"arquivo": os.path.basename(path), "linhas": 0, "inseridos": 0, "atualizados": 0, "sem_alteracao": 0, "erros": []}
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        contexto = _contexto_historico_importacao(motivo, usuario, result["arquivo"])
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for idx, row in enumerate(reader, start=2):
                if not any((v or "").strip() for v in row.values()):
                    continue
                result["linhas"] += 1
                try:
                    registro = _registro_armadilha(row, result["arquivo"], agora)
                    status = _upsert_armadilha(conn, registro, contexto=contexto)
                    result[status] += 1
                except Exception as exc:
                    result["erros"].append(f"Linha {idx}: {exc}")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return result


def importar_diarios_xlsx(db_path, path, usuario=None):
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("Biblioteca openpyxl nao disponivel para ler XLSX.") from exc

    result = {
        "arquivo": os.path.basename(path),
        "diarios": 0,
        "armadilhas": 0,
        "vinculos": 0,
        "telefones": 0,
        "erros": [],
    }
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            nome = _text(ws.title)
            if not nome:
                continue
            id_diario = _upsert_diario_conn(conn, {"nome": nome, "ativo": True}, agora)
            result["diarios"] += 1
            ordem = 0
            for row in range(22, ws.max_row + 1):
                raw_id = ws.cell(row, 1).value
                ovitrampa_id = _normalizar_ovitrampa_id(raw_id)
                if not ovitrampa_id:
                    marcador = _text(raw_id)
                    if marcador and marcador.upper().startswith("OCORR"):
                        break
                    continue
                ordem += 1
                registro = _registro_armadilha_diario(ws, row, ovitrampa_id, result["arquivo"], agora)
                contexto = {
                    "motivo": "Carga inicial dos diarios de ovitrampas",
                    "agentes": "",
                    "usuario": usuario or "sistema",
                    "arquivo_origem": result["arquivo"],
                    "criado_em": agora,
                }
                status = _upsert_armadilha(conn, registro, contexto=contexto, preservar_cadastro_conta_ovos=True)
                if status == "inseridos":
                    result["armadilhas"] += 1
                if registro.get("telefone_responsavel"):
                    result["telefones"] += 1
                if _vincular_diario_conn(conn, id_diario, ovitrampa_id, ordem, agora):
                    result["vinculos"] += 1
        _reordenar_todos_diarios(conn)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return result


def importar_ocorrencias_csv(db_path, path):
    result = {
        "arquivo": os.path.basename(path),
        "linhas": 0,
        "inseridos": 0,
        "atualizados": 0,
        "sem_alteracao": 0,
        "ocorrencias": 0,
        "erros": [],
    }
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for idx, row in enumerate(reader, start=2):
                if not any((v or "").strip() for v in row.values()):
                    continue
                result["linhas"] += 1
                try:
                    registro = _registro_ocorrencia_conta_ovos(row, result["arquivo"], agora)
                    status = _upsert_ocorrencia(conn, registro)
                    result[status] += 1
                    if registro["ocorrencia_codigo"]:
                        result["ocorrencias"] += 1
                except Exception as exc:
                    result["erros"].append(f"Linha {idx}: {exc}")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return result


def resumo(db_path, filtros=None):
    filtros = filtros or {}
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        where, params = _where(filtros)
        totais = dict(conn.execute(
            f"""SELECT COUNT(*) AS leituras,
                       COUNT(DISTINCT ovitrampa_id) AS ovitrampas,
                       COALESCE(SUM(ovos),0) AS ovos,
                       COALESCE(AVG(ovos),0) AS media_ovos,
                       SUM(CASE WHEN ovos > 0 THEN 1 ELSE 0 END) AS positivas,
                       MAX(ano) AS ultimo_ano,
                       MAX(CASE WHEN ano=(SELECT MAX(ano) FROM ovitrampas_leituras) THEN semana ELSE NULL END) AS ultima_semana
                  FROM ovitrampas_leituras l {where}""",
            params,
        ).fetchone())
        por_distrito = [dict(row) for row in conn.execute(
            f"""SELECT COALESCE(l.distrito,'-') AS distrito, COUNT(*) AS leituras,
                       COUNT(DISTINCT l.ovitrampa_id) AS ovitrampas, COALESCE(SUM(l.ovos),0) AS ovos
                  FROM ovitrampas_leituras l {where}
                 GROUP BY COALESCE(l.distrito,'-')
                 ORDER BY ovos DESC, leituras DESC, distrito
                 LIMIT 12""",
            params,
        )]
        por_semana = [dict(row) for row in conn.execute(
            f"""SELECT l.ano, l.semana, COUNT(*) AS leituras, COALESCE(SUM(l.ovos),0) AS ovos,
                       SUM(CASE WHEN l.ovos > 0 THEN 1 ELSE 0 END) AS positivas
                  FROM ovitrampas_leituras l {where}
                 GROUP BY l.ano, l.semana
                 ORDER BY l.ano DESC, l.semana DESC
                 LIMIT 16""",
            params,
        )]
    finally:
        conn.close()
    return {"totais": totais, "por_distrito": por_distrito, "por_semana": por_semana}


def listar(db_path, filtros=None, limite=500):
    filtros = filtros or {}
    limite = max(1, min(int(limite or 500), 2000))
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        where, params = _where(filtros, busca=True)
        rows = [dict(row) for row in conn.execute(
            f"""SELECT l.*, a.nome AS laboratorista
                  FROM ovitrampas_leituras l
                  LEFT JOIN agentes a ON a.id_agente=l.id_laboratorista
                  {where}
                 ORDER BY ano DESC, semana DESC, ovitrampa_id COLLATE NOCASE
                 LIMIT ?""",
            [*params, limite],
        )]
        total = conn.execute(f"SELECT COUNT(*) FROM ovitrampas_leituras l {where}", params).fetchone()[0]
    finally:
        conn.close()
    return {"total": total, "registros": rows}


def listar_armadilhas(db_path, filtros=None, limite=500):
    filtros = filtros or {}
    limite = max(1, min(int(limite or 500), 2000))
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        where, params = _where_armadilhas(filtros)
        rows = [dict(row) for row in conn.execute(
            f"""SELECT a.*,
                       COUNT(l.id_leitura) AS leituras,
                       COALESCE(SUM(l.ovos),0) AS ovos_total,
                       SUM(CASE WHEN l.ovos > 0 THEN 1 ELSE 0 END) AS positivas,
                       MAX(l.data_coleta) AS ultima_coleta,
                       MAX(l.ano) AS ultimo_ano
                  FROM ovitrampas_armadilhas a
                  LEFT JOIN ovitrampas_leituras l ON l.ovitrampa_id=a.ovitrampa_id
                  {where}
                 GROUP BY a.ovitrampa_id
                 ORDER BY CAST(a.ovitrampa_id AS INTEGER), a.ovitrampa_id COLLATE NOCASE
                 LIMIT ?""",
            [*params, limite],
        )]
        total = conn.execute(f"SELECT COUNT(*) FROM ovitrampas_armadilhas a {where}", params).fetchone()[0]
    finally:
        conn.close()
    return {"total": total, "registros": rows}


def historico_armadilha(db_path, ovitrampa_id):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        armadilha = conn.execute(
            f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?",
            (str(ovitrampa_id),),
        ).fetchone()
        leituras = [dict(row) for row in conn.execute(
            """SELECT l.*, a.nome AS laboratorista
                 FROM ovitrampas_leituras l
                 LEFT JOIN agentes a ON a.id_agente=l.id_laboratorista
                WHERE l.ovitrampa_id=?
                ORDER BY l.ano DESC, l.semana DESC, l.data_coleta DESC""",
            (str(ovitrampa_id),),
        )]
        historico = [dict(row) for row in conn.execute(
            f"""SELECT *
                  FROM {ARMADILHAS_HISTORICO_TABLE}
                 WHERE ovitrampa_id=?
                 ORDER BY criado_em DESC, id_historico DESC
                 LIMIT 80""",
            (str(ovitrampa_id),),
        )]
    finally:
        conn.close()
    return {"armadilha": dict(armadilha) if armadilha else None, "leituras": leituras, "alteracoes": historico}


def diarios_dados(db_path):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        diarios = [_diario_dict(row) for row in conn.execute(
            f"""SELECT d.*,
                       COUNT(da.ovitrampa_id) AS armadilhas
                  FROM {DIARIOS_TABLE} d
                  LEFT JOIN {DIARIO_ARMADILHAS_TABLE} da ON da.id_diario=d.id_diario
                 GROUP BY d.id_diario
                 ORDER BY d.ativo DESC, d.nome COLLATE NOCASE"""
        )]
        realocar = _armadilhas_realocar(conn, {})
        sem_diario = _armadilhas_sem_diario(conn)
        total = conn.execute(
            f"SELECT COUNT(*) FROM {ARMADILHAS_TABLE} WHERE COALESCE(ativo,1)=1"
        ).fetchone()[0]
    finally:
        conn.close()
    return {
        "diarios": diarios,
        "sem_diario": sem_diario,
        "realocar": realocar,
        "totais": {
            "diarios": len(diarios),
            "armadilhas": total or 0,
            "sem_diario": sem_diario["total"],
            "realocar": realocar["total"],
        },
    }


def diario_detalhe(db_path, id_diario, incluir_realocar=True):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        diario = conn.execute(
            f"SELECT * FROM {DIARIOS_TABLE} WHERE id_diario=?",
            (_int(id_diario),),
        ).fetchone()
        if not diario:
            raise ValueError("Diario nao encontrado.")
        rows = [dict(row) for row in conn.execute(
            f"""SELECT da.ordem, da.observacoes AS observacao_rota,
                       a.*,
                       COUNT(l.id_leitura) AS leituras,
                       MAX(l.ano * 100 + l.semana) AS ultima_chave,
                       COALESCE(SUM(l.ovos),0) AS ovos_total,
                       SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas
                  FROM {DIARIO_ARMADILHAS_TABLE} da
                  JOIN {ARMADILHAS_TABLE} a ON a.ovitrampa_id=da.ovitrampa_id
                  LEFT JOIN {TABLE} l ON l.ovitrampa_id=a.ovitrampa_id
                 WHERE da.id_diario=?
                   AND COALESCE(a.ativo,1)=1
                 GROUP BY da.id_diario, da.ovitrampa_id
                 ORDER BY da.ordem, CAST(a.ovitrampa_id AS INTEGER), a.ovitrampa_id COLLATE NOCASE""",
            (_int(id_diario),),
        )]
        registros = []
        for row in rows:
            row["ultima"] = _semana_label_from_key(row.pop("ultima_chave", None))
            row["realocar"] = _armadilha_realocar(row)
            if incluir_realocar or not row["realocar"]:
                registros.append(row)
    finally:
        conn.close()
    return {"diario": _diario_dict(diario), "registros": registros, "total": len(registros)}


def salvar_diario(db_path, dados, id_diario=None):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        id_diario = _upsert_diario_conn(conn, dados, agora, id_diario=id_diario)
        conn.commit()
        row = conn.execute(f"SELECT * FROM {DIARIOS_TABLE} WHERE id_diario=?", (id_diario,)).fetchone()
        return _diario_dict(row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def vincular_armadilha_diario(db_path, dados):
    ovitrampa_id = _normalizar_ovitrampa_id(dados.get("ovitrampa_id"))
    id_diario = _int(dados.get("id_diario"))
    if not ovitrampa_id or not id_diario:
        raise ValueError("Informe o diario e a ovitrampa.")
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        agora = datetime.now().isoformat(timespec="seconds")
        existe = conn.execute(f"SELECT 1 FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (ovitrampa_id,)).fetchone()
        if not existe:
            raise ValueError("Ovitrampa nao encontrada.")
        diario = conn.execute(f"SELECT 1 FROM {DIARIOS_TABLE} WHERE id_diario=?", (id_diario,)).fetchone()
        if not diario:
            raise ValueError("Diario nao encontrado.")
        ordem = _int(dados.get("ordem")) or _proxima_ordem_diario(conn, id_diario)
        _vincular_diario_conn(conn, id_diario, ovitrampa_id, ordem, agora, mover_unico=True)
        _reordenar_diario(conn, id_diario)
        conn.commit()
        return diario_detalhe(db_path, id_diario)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def remover_armadilha_diario(db_path, id_diario, ovitrampa_id):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        cur = conn.execute(
            f"DELETE FROM {DIARIO_ARMADILHAS_TABLE} WHERE id_diario=? AND ovitrampa_id=?",
            (_int(id_diario), str(ovitrampa_id)),
        )
        if cur.rowcount == 0:
            raise ValueError("Vinculo nao encontrado.")
        _reordenar_diario(conn, _int(id_diario))
        conn.commit()
        return diario_detalhe(db_path, id_diario)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def mover_armadilha_diario(db_path, id_diario, ovitrampa_id, direcao):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        id_diario = _int(id_diario)
        atual = conn.execute(
            f"""SELECT ovitrampa_id, ordem
                  FROM {DIARIO_ARMADILHAS_TABLE}
                 WHERE id_diario=?
                 ORDER BY ordem, ovitrampa_id COLLATE NOCASE""",
            (id_diario,),
        ).fetchall()
        ids = [row["ovitrampa_id"] for row in atual]
        if str(ovitrampa_id) not in ids:
            raise ValueError("Ovitrampa nao encontrada neste diario.")
        idx = ids.index(str(ovitrampa_id))
        delta = -1 if direcao == "subir" else 1
        novo = idx + delta
        if novo < 0 or novo >= len(ids):
            return diario_detalhe(db_path, id_diario)
        ids[idx], ids[novo] = ids[novo], ids[idx]
        agora = datetime.now().isoformat(timespec="seconds")
        for ordem, item_id in enumerate(ids, start=1):
            conn.execute(
                f"""UPDATE {DIARIO_ARMADILHAS_TABLE}
                       SET ordem=?, atualizado_em=?
                     WHERE id_diario=? AND ovitrampa_id=?""",
                (ordem, agora, id_diario, item_id),
            )
        conn.commit()
        return diario_detalhe(db_path, id_diario)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def reordenar_armadilhas_diario(db_path, id_diario, ordem):
    id_diario = _int(id_diario)
    ids = []
    for item in ordem or []:
        ovitrampa_id = _normalizar_ovitrampa_id(item)
        if ovitrampa_id and ovitrampa_id not in ids:
            ids.append(ovitrampa_id)
    if not id_diario or not ids:
        raise ValueError("Informe a ordem do diario.")

    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        atual = [
            row["ovitrampa_id"]
            for row in conn.execute(
                f"""SELECT ovitrampa_id
                      FROM {DIARIO_ARMADILHAS_TABLE}
                     WHERE id_diario=?
                     ORDER BY ordem, CAST(ovitrampa_id AS INTEGER), ovitrampa_id COLLATE NOCASE""",
                (id_diario,),
            )
        ]
        if not atual:
            raise ValueError("Diario sem ovitrampas para ordenar.")
        if set(ids) != set(atual) or len(ids) != len(atual):
            raise ValueError("A ordem enviada precisa conter todas as ovitrampas do diario.")

        agora = datetime.now().isoformat(timespec="seconds")
        for posicao, item_id in enumerate(ids, start=1):
            conn.execute(
                f"""UPDATE {DIARIO_ARMADILHAS_TABLE}
                       SET ordem=?, atualizado_em=?
                     WHERE id_diario=? AND ovitrampa_id=?""",
                (posicao, agora, id_diario, item_id),
            )
        conn.commit()
        return diario_detalhe(db_path, id_diario)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def atualizar_armadilha_local(db_path, ovitrampa_id, dados, usuario=None):
    campos = {
        "telefone_responsavel": _text(dados.get("telefone_responsavel")),
        "responsavel": _text(dados.get("responsavel")),
    }
    motivo = _text(dados.get("motivo")) or "Edicao manual da armadilha"
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        atual = conn.execute(f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (str(ovitrampa_id),)).fetchone()
        if not atual:
            raise ValueError("Ovitrampa nao encontrada.")
        agora = datetime.now().isoformat(timespec="seconds")
        contexto = {"motivo": motivo, "agentes": "", "usuario": usuario or "sistema", "arquivo_origem": None, "criado_em": agora}
        for campo, novo in campos.items():
            if atual[campo] != novo:
                conn.execute(
                    f"UPDATE {ARMADILHAS_TABLE} SET {campo}=?, atualizado_em=? WHERE ovitrampa_id=?",
                    (novo, agora, str(ovitrampa_id)),
                )
                _registrar_alteracao_armadilha(conn, str(ovitrampa_id), campo, atual[campo], novo, contexto)
        conn.commit()
        row = conn.execute(f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?", (str(ovitrampa_id),)).fetchone()
        return dict(row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def diario_impressao(db_path, id_diario, filtros=None):
    filtros = filtros or {}
    dados = diario_detalhe(db_path, id_diario, incluir_realocar=False)
    return {
        **dados,
        "ano": _int(filtros.get("ano")) or datetime.now().year,
        "semana": _int(filtros.get("semana")),
        "gerado_em": datetime.now(),
        "ocorrencias": OCORRENCIAS,
    }


def monitoramento(db_path, filtros=None):
    filtros = filtros or {}
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        where, params, periodo = _where_monitoramento(conn, filtros)
        join_base = f"""
            FROM {TABLE} l
            LEFT JOIN {ARMADILHAS_TABLE} am ON am.ovitrampa_id=l.ovitrampa_id
            {where}
        """
        total = dict(conn.execute(
            f"""SELECT COUNT(*) AS leituras,
                       COUNT(DISTINCT l.ovitrampa_id) AS armadilhas_lidas,
                       SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas,
                       COUNT(DISTINCT CASE WHEN COALESCE(l.ovos,0)>0 THEN l.ovitrampa_id END) AS armadilhas_positivas,
                       COALESCE(SUM(l.ovos),0) AS ovos
                  {join_base}""",
            params,
        ).fetchone())

        positivas_recentes = [dict(row) for row in conn.execute(
            f"""
            WITH positivas AS (
                SELECT l.ovitrampa_id,
                       COALESCE(am.localidade, l.distrito, '-') AS localidade,
                       COALESCE(am.rua, l.rua, '-') AS rua,
                       COALESCE(am.numero, l.numero, '') AS numero,
                       COALESCE(am.complemento, l.complemento, am.localizacao, l.localizacao, '') AS complemento,
                       l.quarteirao,
                       l.ano,
                       l.semana,
                       l.ovos,
                       l.data_coleta,
                       ROW_NUMBER() OVER (
                           PARTITION BY l.ovitrampa_id
                           ORDER BY l.ano DESC, l.semana DESC, COALESCE(l.data_coleta,'') DESC
                       ) AS rn,
                       COUNT(*) OVER (PARTITION BY l.ovitrampa_id) AS vezes_positiva,
                       SUM(l.ovos) OVER (PARTITION BY l.ovitrampa_id) AS ovos_periodo
                  {join_base}
                   AND COALESCE(l.ovos,0)>0
            )
            SELECT * FROM positivas
             WHERE rn=1
             ORDER BY ano DESC, semana DESC, ovos DESC, CAST(ovitrampa_id AS INTEGER), ovitrampa_id
             LIMIT 80
            """,
            params,
        )]

        ranking_positivas = [dict(row) for row in conn.execute(
            f"""SELECT l.ovitrampa_id,
                       COALESCE(am.localidade, l.distrito, '-') AS localidade,
                       COALESCE(am.rua, l.rua, '-') AS rua,
                       COALESCE(am.numero, l.numero, '') AS numero,
                       COALESCE(am.complemento, l.complemento, am.localizacao, l.localizacao, '') AS complemento,
                       COUNT(*) AS leituras,
                       SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas,
                       COALESCE(SUM(l.ovos),0) AS ovos,
                       ROUND(100.0 * SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) / COUNT(*), 1) AS positividade,
                       MAX(CASE WHEN COALESCE(l.ovos,0)>0 THEN l.ano * 100 + l.semana ELSE NULL END) AS ultima_chave
                  {join_base}
                 GROUP BY l.ovitrampa_id
                HAVING positivas > 0
                 ORDER BY positivas DESC, ovos DESC, positividade DESC, CAST(l.ovitrampa_id AS INTEGER), l.ovitrampa_id
                 LIMIT 80""",
            params,
        )]
        for row in ranking_positivas:
            row["ultima_positiva"] = _semana_label_from_key(row.pop("ultima_chave", None))

        localidades = [dict(row) for row in conn.execute(
            f"""SELECT COALESCE(am.localidade, l.distrito, '-') AS localidade,
                       COUNT(*) AS leituras,
                       COUNT(DISTINCT l.ovitrampa_id) AS armadilhas_lidas,
                       SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas,
                       COUNT(DISTINCT CASE WHEN COALESCE(l.ovos,0)>0 THEN l.ovitrampa_id END) AS armadilhas_positivas,
                       COALESCE(SUM(l.ovos),0) AS ovos,
                       ROUND(COALESCE(AVG(CASE WHEN COALESCE(l.ovos,0)>0 THEN l.ovos END),0), 1) AS media_ovos_positiva
                  {join_base}
                 GROUP BY COALESCE(am.localidade, l.distrito, '-')
                 ORDER BY ovos DESC, armadilhas_positivas DESC, positivas DESC, localidade
                 LIMIT 40""",
            params,
        )]

        ocorrencias_resumo, total_ocorrencias, fonte_ocorrencias = _monitoramento_ocorrencias(
            conn, filtros, periodo, join_base, params
        )

        realocar = _armadilhas_realocar(conn, filtros)
    finally:
        conn.close()

    total = {key: (value or 0) for key, value in total.items()}
    total["ocorrencias"] = total_ocorrencias
    total["realocar"] = realocar["total"]
    return {
        "periodo": periodo,
        "ocorrencias_fonte": fonte_ocorrencias,
        "ocorrencias_labels": [{"codigo": codigo, "descricao": desc} for codigo, desc in OCORRENCIAS.items()],
        "totais": total,
        "positivas_recentes": positivas_recentes,
        "ranking_positivas": ranking_positivas,
        "localidades": localidades,
        "ocorrencias": ocorrencias_resumo,
        "realocar": realocar,
    }


def atualizar_leitura(db_path, id_leitura, dados):
    id_laboratorista = _int(dados.get("id_laboratorista"))
    data_leitura = _date(dados.get("data_leitura"))
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        if id_laboratorista:
            agente = conn.execute("SELECT id_agente FROM agentes WHERE id_agente=? AND ativo=1", (id_laboratorista,)).fetchone()
            if not agente:
                raise ValueError("Laboratorista nao encontrado.")
        cur = conn.execute(
            f"""UPDATE {TABLE}
                   SET id_laboratorista=?, data_leitura=?
                 WHERE id_leitura=?""",
            (id_laboratorista, data_leitura, id_leitura),
        )
        if cur.rowcount == 0:
            raise ValueError("Leitura nao encontrada.")
        conn.commit()
        row = conn.execute(
            """SELECT l.*, a.nome AS laboratorista
                 FROM ovitrampas_leituras l
                 LEFT JOIN agentes a ON a.id_agente=l.id_laboratorista
                WHERE l.id_leitura=?""",
            (id_leitura,),
        ).fetchone()
        return dict(row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def atualizar_leituras_lote(db_path, filtros=None, dados=None):
    filtros = filtros or {}
    dados = dados or {}
    id_laboratorista = _int(dados.get("id_laboratorista"))
    data_leitura = _date(dados.get("data_leitura"))
    somente_vazios = bool(dados.get("somente_vazios"))
    if not id_laboratorista and not data_leitura:
        raise ValueError("Informe laboratorista ou data da leitura.")

    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        if id_laboratorista:
            agente = conn.execute("SELECT id_agente FROM agentes WHERE id_agente=? AND ativo=1", (id_laboratorista,)).fetchone()
            if not agente:
                raise ValueError("Laboratorista nao encontrado.")
        where, params = _where(filtros, busca=True)
        clauses = []
        sets = []
        update_params = []
        if id_laboratorista:
            if somente_vazios:
                sets.append("id_laboratorista=CASE WHEN id_laboratorista IS NULL THEN ? ELSE id_laboratorista END")
                clauses.append("id_laboratorista IS NULL")
            else:
                sets.append("id_laboratorista=?")
            update_params.append(id_laboratorista)
        if data_leitura:
            if somente_vazios:
                sets.append("data_leitura=CASE WHEN data_leitura IS NULL OR data_leitura='' THEN ? ELSE data_leitura END")
                clauses.append("(data_leitura IS NULL OR data_leitura='')")
            else:
                sets.append("data_leitura=?")
            update_params.append(data_leitura)

        target_where = where
        target_params = list(params)
        if somente_vazios and clauses:
            target_where = (target_where + " AND " if target_where else "WHERE ") + "(" + " OR ".join(clauses) + ")"
        total_alvo = conn.execute(f"SELECT COUNT(*) FROM {TABLE} l {target_where}", target_params).fetchone()[0]
        if not total_alvo:
            return {"atualizados": 0, "total_alvo": 0}
        cur = conn.execute(
            f"""UPDATE {TABLE}
                   SET {', '.join(sets)}
                 WHERE id_leitura IN (
                       SELECT id_leitura FROM {TABLE} l {target_where}
                 )""",
            [*update_params, *target_params],
        )
        conn.commit()
        return {"atualizados": cur.rowcount, "total_alvo": total_alvo}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def agentes(db_path):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        return [dict(row) for row in conn.execute(
            "SELECT id_agente, nome FROM agentes WHERE ativo=1 ORDER BY nome COLLATE NOCASE"
        )]
    finally:
        conn.close()


def calendario_dados(db_path, ano):
    ano = _int(ano) or datetime.now().year
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        grupos = [_grupo_dict(row) for row in conn.execute(
            f"SELECT * FROM {CAL_GRUPOS_TABLE} ORDER BY ativo DESC, nome COLLATE NOCASE"
        )]
        eventos = [_cal_evento_dict(row) for row in conn.execute(
            f"""SELECT e.*, g.nome AS grupo_nome, g.localidades AS grupo_localidades, g.cor AS grupo_cor,
                       GROUP_CONCAT(a.id_agente || ':' || a.nome, '|') AS agentes_raw
                  FROM {CAL_EVENTOS_TABLE} e
                  LEFT JOIN {CAL_GRUPOS_TABLE} g ON g.id_grupo=e.id_grupo
                  LEFT JOIN {CAL_AGENTES_TABLE} ea ON ea.id_evento=e.id_evento
                  LEFT JOIN agentes a ON a.id_agente=ea.id_agente
                 WHERE substr(e.data, 1, 4)=?
                 GROUP BY e.id_evento
                 ORDER BY e.data""",
            (str(ano),),
        )]
    finally:
        conn.close()
    return {"ano": ano, "grupos": grupos, "eventos": eventos, "movimentos": MOVIMENTOS}


def calendario_impressao(db_path, ano):
    dados = calendario_dados(db_path, ano)
    ano = dados["ano"]
    eventos_por_data = {ev["data"]: ev for ev in dados["eventos"]}
    meses = []
    for mes in range(1, 13):
        primeiro = date(ano, mes, 1)
        semanas = []
        dia_atual = primeiro - timedelta(days=(primeiro.weekday() + 1) % 7)
        for _ in range(6):
            semana = {"se": _semana_epi(dia_atual), "dias": []}
            for _dia in range(7):
                key = dia_atual.isoformat()
                fora_mes = dia_atual.month != mes
                semana["dias"].append({
                    "data": dia_atual,
                    "dia": "" if fora_mes else dia_atual.day,
                    "fora_mes": fora_mes,
                    "fim_semana": dia_atual.weekday() in (5, 6),
                    "evento": None if fora_mes else eventos_por_data.get(key),
                })
                dia_atual += timedelta(days=1)
            semanas.append(semana)
        eventos_mes = [ev for ev in dados["eventos"] if ev["data"].startswith(f"{ano}-{mes:02d}-")]
        meses.append({
            "numero": mes,
            "nome": MESES_PT[mes],
            "semanas": semanas,
            "legenda": _legenda_calendario(eventos_mes),
        })
    return {
        "ano": ano,
        "meses": meses,
        "total_eventos": len(dados["eventos"]),
        "gerado_em": datetime.now(),
    }


def salvar_grupo(db_path, dados, id_grupo=None):
    nome = _text(dados.get("nome"))
    if not nome:
        raise ValueError("Informe o nome do grupo.")
    payload = {
        "nome": nome,
        "localidades": _text(dados.get("localidades")),
        "cor": _cor(dados.get("cor")) or "#0f766e",
        "ativo": 1 if dados.get("ativo", True) in (True, 1, "1", "true", "on") else 0,
    }
    agora = datetime.now().isoformat(timespec="seconds")
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        if id_grupo:
            existe = conn.execute(f"SELECT 1 FROM {CAL_GRUPOS_TABLE} WHERE id_grupo=?", (id_grupo,)).fetchone()
            if not existe:
                raise ValueError("Grupo nao encontrado.")
            conn.execute(
                f"""UPDATE {CAL_GRUPOS_TABLE}
                       SET nome=?, localidades=?, cor=?, ativo=?, atualizado_em=?
                     WHERE id_grupo=?""",
                (payload["nome"], payload["localidades"], payload["cor"], payload["ativo"], agora, id_grupo),
            )
        else:
            cur = conn.execute(
                f"""INSERT INTO {CAL_GRUPOS_TABLE}
                    (nome, localidades, cor, ativo, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?)""",
                (payload["nome"], payload["localidades"], payload["cor"], payload["ativo"], agora, agora),
            )
            id_grupo = cur.lastrowid
        conn.commit()
        row = conn.execute(f"SELECT * FROM {CAL_GRUPOS_TABLE} WHERE id_grupo=?", (id_grupo,)).fetchone()
        return _grupo_dict(row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def excluir_grupo(db_path, id_grupo):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        usado = conn.execute(f"SELECT COUNT(*) FROM {CAL_EVENTOS_TABLE} WHERE id_grupo=?", (id_grupo,)).fetchone()[0]
        if usado:
            raise ValueError("Grupo possui eventos vinculados. Desative ou edite os eventos antes de excluir.")
        cur = conn.execute(f"DELETE FROM {CAL_GRUPOS_TABLE} WHERE id_grupo=?", (id_grupo,))
        if cur.rowcount == 0:
            raise ValueError("Grupo nao encontrado.")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def salvar_evento_calendario(db_path, dados, usuario_nome="sistema", id_evento=None):
    payload = _cal_evento_payload(dados)
    agora = datetime.now().isoformat(timespec="seconds")
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        if payload["movimento"] != "feriado":
            grupo = conn.execute(
                f"SELECT id_grupo FROM {CAL_GRUPOS_TABLE} WHERE id_grupo=? AND ativo=1",
                (payload["id_grupo"],),
            ).fetchone()
            if not grupo:
                raise ValueError("Selecione um grupo ativo.")
        if id_evento:
            existe = conn.execute(f"SELECT 1 FROM {CAL_EVENTOS_TABLE} WHERE id_evento=?", (id_evento,)).fetchone()
            if not existe:
                raise ValueError("Evento nao encontrado.")
            conn.execute(
                f"""UPDATE {CAL_EVENTOS_TABLE}
                       SET data=?, movimento=?, titulo=?, id_grupo=?, ciclo=?, observacoes=?, atualizado_em=?
                     WHERE id_evento=?""",
                (
                    payload["data"], payload["movimento"], payload["titulo"], payload["id_grupo"], payload["ciclo"],
                    payload["observacoes"], agora, id_evento,
                ),
            )
        else:
            cur = conn.execute(
                f"""INSERT INTO {CAL_EVENTOS_TABLE}
                    (data, movimento, titulo, id_grupo, ciclo, observacoes, criado_por, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    payload["data"], payload["movimento"], payload["titulo"], payload["id_grupo"], payload["ciclo"],
                    payload["observacoes"], usuario_nome, agora, agora,
                ),
            )
            id_evento = cur.lastrowid
        _salvar_cal_agentes(conn, id_evento, payload["agentes"])
        conn.commit()
    except Exception as exc:
        conn.rollback()
        if isinstance(exc, Exception) and "UNIQUE" in str(exc).upper():
            raise ValueError("Ja existe um movimento de ovitrampa nessa data.") from exc
        raise
    finally:
        conn.close()
    return calendario_evento(db_path, id_evento)


def calendario_evento(db_path, id_evento):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        row = conn.execute(
            f"""SELECT e.*, g.nome AS grupo_nome, g.localidades AS grupo_localidades, g.cor AS grupo_cor,
                       GROUP_CONCAT(a.id_agente || ':' || a.nome, '|') AS agentes_raw
                  FROM {CAL_EVENTOS_TABLE} e
                  LEFT JOIN {CAL_GRUPOS_TABLE} g ON g.id_grupo=e.id_grupo
                  LEFT JOIN {CAL_AGENTES_TABLE} ea ON ea.id_evento=e.id_evento
                  LEFT JOIN agentes a ON a.id_agente=ea.id_agente
                 WHERE e.id_evento=?
                 GROUP BY e.id_evento""",
            (id_evento,),
        ).fetchone()
        if not row:
            raise ValueError("Evento nao encontrado.")
        return _cal_evento_dict(row)
    finally:
        conn.close()


def excluir_evento_calendario(db_path, id_evento):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        conn.execute(f"DELETE FROM {CAL_AGENTES_TABLE} WHERE id_evento=?", (id_evento,))
        cur = conn.execute(f"DELETE FROM {CAL_EVENTOS_TABLE} WHERE id_evento=?", (id_evento,))
        if cur.rowcount == 0:
            raise ValueError("Evento nao encontrado.")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def eventos_agenda(db_path, inicio, fim):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        rows = conn.execute(
            f"""SELECT e.*, g.nome AS grupo_nome, g.localidades AS grupo_localidades, g.cor AS grupo_cor,
                       GROUP_CONCAT(a.nome, ', ') AS agentes
                  FROM {CAL_EVENTOS_TABLE} e
                  JOIN {CAL_GRUPOS_TABLE} g ON g.id_grupo=e.id_grupo
                  LEFT JOIN {CAL_AGENTES_TABLE} ea ON ea.id_evento=e.id_evento
                  LEFT JOIN agentes a ON a.id_agente=ea.id_agente
                 WHERE e.data BETWEEN ? AND ?
                   AND e.movimento <> 'feriado'
                 GROUP BY e.id_evento
                 ORDER BY e.data""",
            (str(inicio)[:10], str(fim)[:10]),
        ).fetchall()
    finally:
        conn.close()
    eventos = []
    for row in rows:
        movimento_label = MOVIMENTOS.get(row["movimento"], row["movimento"])
        titulo = f"{movimento_label} de ovitrampas - {row['grupo_nome']}"
        detalhes = []
        if row["ciclo"]:
            detalhes.append(f"Ciclo: {row['ciclo']}")
        if row["grupo_localidades"]:
            detalhes.append(f"Localidades: {row['grupo_localidades']}")
        if row["observacoes"]:
            detalhes.append(f"Observações: {row['observacoes']}")
        eventos.append({
            "data": row["data"],
            "titulo": titulo,
            "resumo": " | ".join(detalhes),
            "localidades": row["grupo_localidades"] or row["grupo_nome"],
            "agentes": row["agentes"] or "-",
            "cor": row["grupo_cor"] or "#0f766e",
            "id_evento": row["id_evento"],
            "movimento": row["movimento"],
        })
    return eventos


def distritos(db_path):
    conn = db_core.connect(db_path)
    try:
        ensure_schema(conn)
        return [row[0] for row in conn.execute(
            """SELECT nome FROM (
                   SELECT DISTINCT distrito AS nome FROM ovitrampas_leituras WHERE distrito IS NOT NULL AND TRIM(distrito)<>''
                   UNION
                   SELECT DISTINCT localidade AS nome FROM ovitrampas_armadilhas WHERE localidade IS NOT NULL AND TRIM(localidade)<>''
               )
               ORDER BY nome"""
        )]
    finally:
        conn.close()


def _registro(row, arquivo, agora):
    ovitrampa_id = _text(row.get("Ovitrampa ID"))
    ano = _int(row.get("Ano"))
    semana = _int(row.get("Semana"))
    data_instalacao = _date(row.get("Data da instalação"))
    data_coleta = _date(row.get("Data de coleta"))
    data_envio = _datetime(row.get("Data do envio da contagem"))
    if not ovitrampa_id:
        raise ValueError("sem Ovitrampa ID")
    if not ano or not semana:
        raise ValueError("sem ano/semana")
    chave = "|".join([ovitrampa_id, str(ano), str(semana), data_instalacao or "", data_coleta or "", data_envio or ""])
    return {
        "id_leitura": hashlib.md5(chave.encode("utf-8")).hexdigest(),
        "ovitrampa_id": ovitrampa_id,
        "estado": _text(row.get("Estado")),
        "municipio": _text(row.get("Município")),
        "distrito": _title_distrito(row.get("Distrito")),
        "rua": _text(row.get("Rua")),
        "numero": _text(row.get("Número")),
        "complemento": _text(row.get("Complemento")),
        "localizacao": _text(row.get("Localização")),
        "latitude": _real(row.get("Latitude")),
        "longitude": _real(row.get("Longitude")),
        "ano": ano,
        "semana": semana,
        "data_envio_contagem": data_envio,
        "ovos": _int(row.get("Ovos")) or 0,
        "quem_enviou": _text(row.get("Quem enviou")),
        "observacao": _text(_row_value(row, "Observação", "Observacao")),
        "lat_lng": _text(row.get("Lat_lng")),
        "quarteirao": _normalizar_quarteirao(_row_value(row, "Quarteirão", "Quarteirao")),
        "data_instalacao": data_instalacao,
        "data_coleta": data_coleta,
        "ocorrencia_codigo": _ocorrencia_codigo(row),
        "arquivo_origem": arquivo,
        "importado_em": agora,
    }


def _registro_armadilha(row, arquivo, agora):
    ovitrampa_id = _normalizar_ovitrampa_id(row.get("ID"))
    if not ovitrampa_id:
        raise ValueError("sem ID da armadilha")
    return {
        "ovitrampa_id": ovitrampa_id,
        "rua": _text(row.get("Rua")),
        "numero": _text(_row_value(row, "Número do logradouro", "Numero do logradouro")),
        "complemento": _text(row.get("Complemento")),
        "bairro": _text(row.get("Bairro")),
        "localizacao": _text(_row_value(row, "Localização da ovitrampa", "Localizacao da ovitrampa")),
        "localidade": _title_distrito(row.get("Setor/Distrito da ovitrampa")),
        "responsavel": _text(_row_value(row, "Responsável", "Responsavel")),
        "quarteirao": _normalizar_quarteirao(_row_value(row, "Quarteirão", "Quarteirao")),
        "latitude": _real(row.get("Latitude")),
        "longitude": _real(row.get("Longitude")),
        "arquivo_origem": arquivo,
        "atualizado_em": agora,
    }


def _registro_armadilha_diario(ws, row, ovitrampa_id, arquivo, agora):
    endereco = _text(ws.cell(row, 2).value)
    telefone = _text(ws.cell(row + 1, 2).value)
    responsavel = _text(ws.cell(row + 1, 4).value)
    rua, numero, complemento = _separar_endereco_diario(endereco)
    return {
        "ovitrampa_id": ovitrampa_id,
        "rua": rua,
        "numero": numero,
        "complemento": complemento,
        "bairro": None,
        "localizacao": _text(ws.cell(row, 7).value),
        "localidade": _localidade_diario(ws.title),
        "responsavel": responsavel,
        "telefone_responsavel": telefone,
        "quarteirao": _normalizar_quarteirao(ws.cell(row, 6).value),
        "latitude": None,
        "longitude": None,
        "ativo": 1,
        "arquivo_origem": arquivo,
        "atualizado_em": agora,
    }


def _registro_ocorrencia_conta_ovos(row, arquivo, agora):
    id_contagem = _text(_row_value(row, "ID da Contagem"))
    if not id_contagem:
        raise ValueError("sem ID da Contagem")
    ovitrampa_id = _text(_row_value(row, "Número da armadilha", "Numero da armadilha"))
    if not ovitrampa_id:
        raise ValueError("sem numero da armadilha")
    ano = _int(_row_value(row, "Ano"))
    semana = _int(_row_value(row, "Semana"))
    if not ano or not semana:
        raise ValueError("sem ano/semana")
    codigo_conta_ovos = _int(_row_value(row, "Código da Observação", "Codigo da Observacao"))
    return {
        "id_contagem": id_contagem,
        "ovitrampa_id": ovitrampa_id,
        "ano": ano,
        "semana": semana,
        "data": _date(_row_value(row, "Data")),
        "data_envio_contagem": _datetime(_row_value(row, "Tempo de Envio da Contagem")),
        "ovos": _int(_row_value(row, "Ovos")) or 0,
        "resultado": _text(_row_value(row, "Resultado")),
        "codigo_conta_ovos": codigo_conta_ovos,
        "observacao_conta_ovos": _text(_row_value(row, "Observação", "Observacao")),
        "ocorrencia_codigo": CONTA_OVOS_OCORRENCIAS.get(codigo_conta_ovos),
        "latitude": _real(_row_value(row, "Latitude")),
        "longitude": _real(_row_value(row, "Longitude")),
        "lat_lng": _text(_row_value(row, "Lat_lng")),
        "arquivo_origem": arquivo,
        "importado_em": agora,
    }


def _grupo_dict(row):
    item = dict(row)
    item["ativo"] = bool(item.get("ativo"))
    return item


def _cal_evento_dict(row):
    item = dict(row)
    item["movimento_label"] = MOVIMENTOS.get(item.get("movimento"), item.get("movimento") or "")
    if item.get("movimento") == "feriado":
        item["grupo_nome"] = item.get("titulo") or "Feriado"
        item["grupo_localidades"] = ""
        item["grupo_cor"] = "#64748b"
    raw = item.pop("agentes_raw", "") or ""
    item["agentes"] = [
        {"id_agente": int(x.split(":", 1)[0]), "nome": x.split(":", 1)[1]}
        for x in raw.split("|")
        if ":" in x
    ]
    item["agentes_nomes"] = ", ".join(a["nome"] for a in item["agentes"])
    return item


def _semana_epi(dia):
    first = date(dia.year, 1, 1)
    first_sunday = first + timedelta(days=(6 - first.weekday()) % 7)
    if dia < first_sunday:
        return _semana_epi(date(dia.year - 1, 12, 31))
    return ((dia - first_sunday).days // 7) + 1


def _legenda_calendario(eventos):
    legenda = {}
    for ev in eventos:
        if ev.get("movimento") == "feriado":
            chave = f"feriado:{ev.get('grupo_nome') or 'Feriado'}"
            legenda[chave] = {
                "nome": ev.get("grupo_nome") or "Feriado",
                "localidades": "",
                "cor": ev.get("grupo_cor") or "#64748b",
                "feriado": True,
            }
            continue
        chave = f"grupo:{ev.get('id_grupo') or ev.get('grupo_nome')}"
        legenda[chave] = {
            "nome": ev.get("grupo_nome") or "Grupo",
            "localidades": ev.get("grupo_localidades") or "",
            "cor": ev.get("grupo_cor") or "#0f766e",
            "feriado": False,
        }
    return list(legenda.values())


def _cal_evento_payload(dados):
    data = _date(dados.get("data"))
    if not data:
        raise ValueError("Informe a data do movimento.")
    movimento = _text(dados.get("movimento"))
    if movimento not in MOVIMENTOS:
        raise ValueError("Movimento invalido.")
    id_grupo = _int(dados.get("id_grupo"))
    if movimento != "feriado" and not id_grupo:
        raise ValueError("Selecione o grupo.")
    titulo = _text(dados.get("titulo"))
    if movimento == "feriado" and not titulo:
        raise ValueError("Informe o título do feriado.")
    return {
        "data": data,
        "movimento": movimento,
        "titulo": titulo if movimento == "feriado" else None,
        "id_grupo": id_grupo if movimento != "feriado" else None,
        "ciclo": _text(dados.get("ciclo")),
        "observacoes": _text(dados.get("observacoes")),
        "agentes": _parse_ids(dados.get("agentes")),
    }


def _salvar_cal_agentes(conn, id_evento, agentes):
    conn.execute(f"DELETE FROM {CAL_AGENTES_TABLE} WHERE id_evento=?", (id_evento,))
    for id_agente in agentes:
        conn.execute(
            f"INSERT OR IGNORE INTO {CAL_AGENTES_TABLE} (id_evento, id_agente) VALUES (?, ?)",
            (id_evento, id_agente),
        )


def _parse_ids(values):
    ids = []
    for value in values or []:
        numero = _int(value)
        if numero and numero not in ids:
            ids.append(numero)
    return ids


def _cor(value):
    text = _text(value)
    if not text:
        return None
    if len(text) == 7 and text.startswith("#") and all(ch in "0123456789abcdefABCDEF" for ch in text[1:]):
        return text
    return None


def _ocorrencia_codigo(row):
    nomes = (
        "Ocorrência", "Ocorrencia", "Ocorrência da ovitrampa", "Ocorrencia da ovitrampa",
        "Código da ocorrência", "Codigo da ocorrencia", "Código ocorrência", "Codigo ocorrencia",
        "Ocorrência (1 a 9)", "Ocorrencia (1 a 9)",
    )
    for nome in nomes:
        valor = _text(_row_value(row, nome))
        if not valor:
            continue
        numero = _int(valor)
        if numero is None:
            match = re.match(r"\s*([1-9])\b", valor)
            numero = int(match.group(1)) if match else None
        if numero in OCORRENCIAS:
            return numero
    return None


def _semana_label_from_key(value):
    try:
        key = int(value)
    except (TypeError, ValueError):
        return "-"
    ano, semana = divmod(key, 100)
    return f"{ano} / Semana {semana:02d}"


def _where_monitoramento(conn, filtros):
    clauses = ["1=1"]
    params = []
    ano = _int(filtros.get("ano"))
    semana_ini = _int(filtros.get("semana_ini"))
    semana_fim = _int(filtros.get("semana_fim"))
    ultimas = _int(filtros.get("ultimas")) or 8
    ultimas = max(1, min(ultimas, 52))
    if ano:
        clauses.append("l.ano=?")
        params.append(ano)
        if semana_ini:
            clauses.append("l.semana>=?")
            params.append(semana_ini)
        if semana_fim:
            clauses.append("l.semana<=?")
            params.append(semana_fim)
        periodo = {"ano": ano, "semana_ini": semana_ini, "semana_fim": semana_fim, "ultimas": None}
    else:
        latest = conn.execute(
            f"SELECT ano, semana FROM {TABLE} ORDER BY ano DESC, semana DESC LIMIT 1"
        ).fetchone()
        if latest:
            ano = latest["ano"]
            semana_fim = latest["semana"]
            semana_ini = max(1, semana_fim - ultimas + 1)
            clauses.append("l.ano=?")
            clauses.append("l.semana BETWEEN ? AND ?")
            params.extend([ano, semana_ini, semana_fim])
        periodo = {"ano": ano, "semana_ini": semana_ini, "semana_fim": semana_fim, "ultimas": ultimas}
    distrito = _text(filtros.get("distrito"))
    if distrito:
        clauses.append("COALESCE(am.localidade, l.distrito)=?")
        params.append(distrito)
    return "WHERE " + " AND ".join(clauses), params, periodo


def _where_ocorrencias_monitoramento(filtros, periodo):
    clauses = ["1=1"]
    params = []
    ano = _int(filtros.get("ano")) or periodo.get("ano")
    semana_ini = _int(filtros.get("semana_ini")) or periodo.get("semana_ini")
    semana_fim = _int(filtros.get("semana_fim")) or periodo.get("semana_fim")
    if ano:
        clauses.append("o.ano=?")
        params.append(ano)
    if semana_ini:
        clauses.append("o.semana>=?")
        params.append(semana_ini)
    if semana_fim:
        clauses.append("o.semana<=?")
        params.append(semana_fim)
    distrito = _text(filtros.get("distrito"))
    if distrito:
        clauses.append("am.localidade=?")
        params.append(distrito)
    return "WHERE " + " AND ".join(clauses), params


def _monitoramento_ocorrencias(conn, filtros, periodo, legacy_join_base, legacy_params):
    where, params = _where_ocorrencias_monitoramento(filtros, periodo)
    join_base = f"""
        FROM {OCORRENCIAS_TABLE} o
        LEFT JOIN {ARMADILHAS_TABLE} am ON am.ovitrampa_id=o.ovitrampa_id
        {where}
    """
    total_importado = conn.execute(
        f"""SELECT COUNT(*)
              {join_base}
             AND o.ocorrencia_codigo BETWEEN 1 AND 9""",
        params,
    ).fetchone()[0]
    if total_importado:
        resumo = [dict(row) for row in conn.execute(
            f"""SELECT o.ocorrencia_codigo AS codigo,
                       COUNT(*) AS total,
                       COUNT(DISTINCT o.ovitrampa_id) AS armadilhas
                  {join_base}
                   AND o.ocorrencia_codigo BETWEEN 1 AND 9
                 GROUP BY o.ocorrencia_codigo
                 ORDER BY o.ocorrencia_codigo""",
            params,
        )]
        detalhes = _monitoramento_ocorrencias_detalhes_importadas(conn, join_base, params)
        fonte = "Histórico de ocorrências importado do Conta Ovos"
    else:
        resumo = [dict(row) for row in conn.execute(
            f"""SELECT l.ocorrencia_codigo AS codigo,
                       COUNT(*) AS total,
                       COUNT(DISTINCT l.ovitrampa_id) AS armadilhas
                  {legacy_join_base}
                   AND l.ocorrencia_codigo BETWEEN 1 AND 9
                 GROUP BY l.ocorrencia_codigo
                 ORDER BY l.ocorrencia_codigo""",
            legacy_params,
        )]
        detalhes = _monitoramento_ocorrencias_detalhes(conn, legacy_join_base, legacy_params)
        total_importado = sum(row["total"] or 0 for row in resumo)
        fonte = "Ocorrências registradas nas leituras semanais"

    for row in resumo:
        row["descricao"] = OCORRENCIAS.get(row["codigo"], "Ocorrencia")
        row["armadilhas_destaque"] = detalhes.get(row["codigo"], [])
    return resumo, total_importado, fonte


def _monitoramento_ocorrencias_detalhes_importadas(conn, join_base, params):
    rows = [dict(row) for row in conn.execute(
        f"""WITH base AS (
                SELECT o.ocorrencia_codigo AS codigo,
                       o.ovitrampa_id,
                       COALESCE(am.localidade, '-') AS localidade,
                       COALESCE(am.rua, '-') AS rua,
                       COALESCE(am.numero, '') AS numero,
                       COALESCE(am.complemento, am.localizacao, '') AS complemento,
                       COALESCE(am.quarteirao, '') AS quarteirao,
                       o.ano,
                       o.semana,
                       o.data,
                       o.ovos,
                       o.resultado,
                       o.observacao_conta_ovos AS observacao,
                       COUNT(*) OVER (PARTITION BY o.ocorrencia_codigo, o.ovitrampa_id) AS total,
                       MAX(o.ano * 100 + o.semana) OVER (PARTITION BY o.ocorrencia_codigo, o.ovitrampa_id) AS ultima_chave,
                       ROW_NUMBER() OVER (
                           PARTITION BY o.ocorrencia_codigo, o.ovitrampa_id
                           ORDER BY o.ano DESC, o.semana DESC, COALESCE(o.data,'') DESC, o.id_contagem DESC
                       ) AS rn
                  {join_base}
                   AND o.ocorrencia_codigo BETWEEN 1 AND 9
             )
             SELECT codigo, ovitrampa_id, localidade, rua, numero, complemento, quarteirao,
                    ano, semana, data, ovos, resultado, observacao, total, ultima_chave
               FROM base
              WHERE rn=1
              ORDER BY codigo, total DESC, ultima_chave DESC, CAST(ovitrampa_id AS INTEGER), ovitrampa_id""",
        params,
    )]
    por_codigo = {codigo: [] for codigo in OCORRENCIAS}
    for row in rows:
        row["ultima"] = _semana_label_from_key(row.pop("ultima_chave", None))
        codigo = row.get("codigo")
        if codigo in por_codigo and len(por_codigo[codigo]) < 80:
            por_codigo[codigo].append(row)
    return por_codigo


def _monitoramento_ocorrencias_detalhes(conn, join_base, params):
    rows = [dict(row) for row in conn.execute(
        f"""WITH base AS (
                SELECT l.ocorrencia_codigo AS codigo,
                       l.ovitrampa_id,
                       COALESCE(am.localidade, l.distrito, '-') AS localidade,
                       COALESCE(am.rua, l.rua, '-') AS rua,
                       COALESCE(am.numero, l.numero, '') AS numero,
                       COALESCE(am.complemento, l.complemento, am.localizacao, l.localizacao, '') AS complemento,
                       COALESCE(am.quarteirao, l.quarteirao, '') AS quarteirao,
                       l.ano,
                       l.semana,
                       l.data_coleta AS data,
                       l.ovos,
                       CASE WHEN COALESCE(l.ovos,0)>0 THEN 'Positiva' ELSE 'Negativa' END AS resultado,
                       l.observacao,
                       COUNT(*) OVER (PARTITION BY l.ocorrencia_codigo, l.ovitrampa_id) AS total,
                       MAX(l.ano * 100 + l.semana) OVER (PARTITION BY l.ocorrencia_codigo, l.ovitrampa_id) AS ultima_chave,
                       ROW_NUMBER() OVER (
                           PARTITION BY l.ocorrencia_codigo, l.ovitrampa_id
                           ORDER BY l.ano DESC, l.semana DESC, COALESCE(l.data_coleta,'') DESC
                       ) AS rn
                  {join_base}
                   AND l.ocorrencia_codigo BETWEEN 1 AND 9
             )
             SELECT codigo, ovitrampa_id, localidade, rua, numero, complemento, quarteirao,
                    ano, semana, data, ovos, resultado, observacao, total, ultima_chave
               FROM base
              WHERE rn=1
              ORDER BY codigo, total DESC, ultima_chave DESC, CAST(ovitrampa_id AS INTEGER), ovitrampa_id""",
        params,
    )]
    por_codigo = {codigo: [] for codigo in OCORRENCIAS}
    for row in rows:
        row["ultima"] = _semana_label_from_key(row.pop("ultima_chave", None))
        codigo = row.get("codigo")
        if codigo in por_codigo and len(por_codigo[codigo]) < 80:
            por_codigo[codigo].append(row)
    return por_codigo


def _armadilhas_realocar(conn, filtros):
    clauses = [_realocar_sql("a")]
    params = []
    distrito = _text(filtros.get("distrito"))
    if distrito:
        clauses.append("a.localidade=?")
        params.append(distrito)
    where = "WHERE " + " AND ".join(clauses)
    rows = [dict(row) for row in conn.execute(
        f"""SELECT a.*,
                   COUNT(l.id_leitura) AS leituras,
                   MAX(l.ano * 100 + l.semana) AS ultima_chave,
                   COALESCE(SUM(l.ovos),0) AS ovos_total,
                   SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas
              FROM {ARMADILHAS_TABLE} a
              LEFT JOIN {TABLE} l ON l.ovitrampa_id=a.ovitrampa_id
              {where}
             GROUP BY a.ovitrampa_id
             ORDER BY COALESCE(a.localidade,''), CAST(a.ovitrampa_id AS INTEGER), a.ovitrampa_id
             LIMIT 300""",
        params,
    )]
    total = conn.execute(
        f"SELECT COUNT(*) FROM {ARMADILHAS_TABLE} a {where}",
        params,
    ).fetchone()[0]
    for row in rows:
        row["ultima"] = _semana_label_from_key(row.pop("ultima_chave", None))
    return {"total": total or 0, "registros": rows}


def _armadilhas_sem_diario(conn, limite=300):
    where = f"""WHERE COALESCE(a.ativo,1)=1
                  AND NOT ({_realocar_sql("a")})
                  AND NOT EXISTS (
                        SELECT 1 FROM {DIARIO_ARMADILHAS_TABLE} da
                         WHERE da.ovitrampa_id=a.ovitrampa_id
                  )"""
    rows = [dict(row) for row in conn.execute(
        f"""SELECT a.*,
                   COUNT(l.id_leitura) AS leituras,
                   MAX(l.ano * 100 + l.semana) AS ultima_chave,
                   COALESCE(SUM(l.ovos),0) AS ovos_total,
                   SUM(CASE WHEN COALESCE(l.ovos,0)>0 THEN 1 ELSE 0 END) AS positivas
              FROM {ARMADILHAS_TABLE} a
              LEFT JOIN {TABLE} l ON l.ovitrampa_id=a.ovitrampa_id
              {where}
             GROUP BY a.ovitrampa_id
             ORDER BY COALESCE(a.localidade,''), CAST(a.ovitrampa_id AS INTEGER), a.ovitrampa_id COLLATE NOCASE
             LIMIT ?""",
        (limite,),
    )]
    total = conn.execute(
        f"SELECT COUNT(*) FROM {ARMADILHAS_TABLE} a {where}"
    ).fetchone()[0]
    for row in rows:
        row["ultima"] = _semana_label_from_key(row.pop("ultima_chave", None))
    return {"total": total or 0, "registros": rows}


def _diario_dict(row):
    item = dict(row)
    item["ativo"] = bool(item.get("ativo"))
    return item


def _upsert_diario_conn(conn, dados, agora, id_diario=None):
    nome = _text(dados.get("nome"))
    if not nome:
        raise ValueError("Informe o nome do diario.")
    observacoes = _text(dados.get("observacoes"))
    ativo = 1 if dados.get("ativo", True) in (True, 1, "1", "true", "on") else 0
    if id_diario:
        existe = conn.execute(f"SELECT 1 FROM {DIARIOS_TABLE} WHERE id_diario=?", (_int(id_diario),)).fetchone()
        if not existe:
            raise ValueError("Diario nao encontrado.")
        conn.execute(
            f"""UPDATE {DIARIOS_TABLE}
                   SET nome=?, observacoes=?, ativo=?, atualizado_em=?
                 WHERE id_diario=?""",
            (nome, observacoes, ativo, agora, _int(id_diario)),
        )
        return _int(id_diario)
    atual = conn.execute(f"SELECT id_diario FROM {DIARIOS_TABLE} WHERE nome=?", (nome,)).fetchone()
    if atual:
        conn.execute(
            f"""UPDATE {DIARIOS_TABLE}
                   SET observacoes=COALESCE(?, observacoes), ativo=?, atualizado_em=?
                 WHERE id_diario=?""",
            (observacoes, ativo, agora, atual["id_diario"]),
        )
        return atual["id_diario"]
    cur = conn.execute(
        f"""INSERT INTO {DIARIOS_TABLE} (nome, observacoes, ativo, criado_em, atualizado_em)
            VALUES (?, ?, ?, ?, ?)""",
        (nome, observacoes, ativo, agora, agora),
    )
    return cur.lastrowid


def _vincular_diario_conn(conn, id_diario, ovitrampa_id, ordem, agora, mover_unico=False):
    if mover_unico:
        conn.execute(
            f"DELETE FROM {DIARIO_ARMADILHAS_TABLE} WHERE ovitrampa_id=? AND id_diario<>?",
            (ovitrampa_id, id_diario),
        )
    atual = conn.execute(
        f"SELECT ordem FROM {DIARIO_ARMADILHAS_TABLE} WHERE id_diario=? AND ovitrampa_id=?",
        (id_diario, ovitrampa_id),
    ).fetchone()
    if atual:
        conn.execute(
            f"""UPDATE {DIARIO_ARMADILHAS_TABLE}
                   SET ordem=?, atualizado_em=?
                 WHERE id_diario=? AND ovitrampa_id=?""",
            (ordem, agora, id_diario, ovitrampa_id),
        )
        return False
    conn.execute(
        f"""INSERT INTO {DIARIO_ARMADILHAS_TABLE}
            (id_diario, ovitrampa_id, ordem, criado_em, atualizado_em)
            VALUES (?, ?, ?, ?, ?)""",
        (id_diario, ovitrampa_id, ordem, agora, agora),
    )
    return True


def _proxima_ordem_diario(conn, id_diario):
    valor = conn.execute(
        f"SELECT COALESCE(MAX(ordem),0) + 1 FROM {DIARIO_ARMADILHAS_TABLE} WHERE id_diario=?",
        (id_diario,),
    ).fetchone()[0]
    return int(valor or 1)


def _reordenar_todos_diarios(conn):
    ids = [row[0] for row in conn.execute(f"SELECT id_diario FROM {DIARIOS_TABLE}")]
    for id_diario in ids:
        _reordenar_diario(conn, id_diario)


def _reordenar_diario(conn, id_diario):
    rows = conn.execute(
        f"""SELECT ovitrampa_id
              FROM {DIARIO_ARMADILHAS_TABLE}
             WHERE id_diario=?
             ORDER BY ordem, CAST(ovitrampa_id AS INTEGER), ovitrampa_id COLLATE NOCASE""",
        (id_diario,),
    ).fetchall()
    agora = datetime.now().isoformat(timespec="seconds")
    for ordem, row in enumerate(rows, start=1):
        conn.execute(
            f"""UPDATE {DIARIO_ARMADILHAS_TABLE}
                   SET ordem=?, atualizado_em=?
                 WHERE id_diario=? AND ovitrampa_id=?""",
            (ordem, agora, id_diario, row["ovitrampa_id"]),
        )


def _realocar_sql(alias):
    return (
        f"(UPPER(COALESCE({alias}.localidade,'') || ' ' || COALESCE({alias}.rua,'') || ' ' || "
        f"COALESCE({alias}.numero,'') || ' ' || COALESCE({alias}.complemento,'') || ' ' || "
        f"COALESCE({alias}.localizacao,'') || ' ' || COALESCE({alias}.bairro,'') || ' ' || "
        f"COALESCE({alias}.responsavel,'')) LIKE '%REALOCAR%')"
    )


def _armadilha_realocar(row):
    texto = " ".join(str(row.get(campo) or "") for campo in (
        "localidade", "rua", "numero", "complemento", "localizacao", "bairro", "responsavel"
    ))
    return "REALOCAR" in texto.upper()


def _normalizar_ovitrampa_id(value):
    text = _text(value)
    if not text:
        return None
    text = text.replace(".0", "") if re.fullmatch(r"\d+\.0", text) else text
    text = re.sub(r"\s+", "", text).upper()
    if re.fullmatch(r"\d+(?:[-/][A-Z]+)?", text):
        return text.replace("/", "-")
    return None


def _localidade_diario(nome):
    text = _text(nome) or ""
    for sufixo in (" 1", " 2", " 3"):
        if text.endswith(sufixo):
            text = text[:-len(sufixo)]
            break
    return text


def _separar_endereco_diario(endereco):
    text = _text(endereco)
    if not text:
        return None, None, None
    match = re.match(r"(.+?),\s*([0-9A-Za-z\-\/]+)\s*(?:[-(]\s*(.*))?$", text)
    if not match:
        return text, None, None
    rua = _text(match.group(1))
    numero = _text(match.group(2))
    complemento = _text(match.group(3))
    if complemento and complemento.endswith(")"):
        complemento = complemento[:-1].strip()
    return rua, numero, complemento


def _contexto_historico_importacao(motivo, usuario, arquivo):
    return {
        "motivo": _text(motivo) or "Importacao do cadastro Conta Ovos",
        "agentes": "",
        "usuario": usuario or "sistema",
        "arquivo_origem": arquivo,
        "criado_em": datetime.now().isoformat(timespec="seconds"),
    }


def _registrar_alteracao_armadilha(conn, ovitrampa_id, campo, anterior, novo, contexto=None):
    contexto = contexto or {}
    conn.execute(
        f"""INSERT INTO {ARMADILHAS_HISTORICO_TABLE}
            (ovitrampa_id, campo, valor_anterior, valor_novo, motivo, agentes, usuario, arquivo_origem, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            str(ovitrampa_id),
            campo,
            None if anterior is None else str(anterior),
            None if novo is None else str(novo),
            contexto.get("motivo"),
            contexto.get("agentes"),
            contexto.get("usuario"),
            contexto.get("arquivo_origem"),
            contexto.get("criado_em") or datetime.now().isoformat(timespec="seconds"),
        ),
    )


def _insert(conn, registro):
    cols = list(registro.keys())
    placeholders = ",".join("?" for _ in cols)
    cur = conn.execute(
        f"INSERT OR IGNORE INTO {TABLE} ({','.join(cols)}) VALUES ({placeholders})",
        [registro[col] for col in cols],
    )
    return cur.rowcount > 0


def _upsert_armadilha(conn, registro, contexto=None, preservar_cadastro_conta_ovos=False):
    atual = conn.execute(
        f"SELECT * FROM {ARMADILHAS_TABLE} WHERE ovitrampa_id=?",
        (registro["ovitrampa_id"],),
    ).fetchone()
    cols = list(registro.keys())
    if not atual:
        placeholders = ",".join("?" for _ in cols)
        conn.execute(
            f"INSERT INTO {ARMADILHAS_TABLE} ({','.join(cols)}) VALUES ({placeholders})",
            [registro[col] for col in cols],
        )
        _registrar_alteracao_armadilha(conn, registro["ovitrampa_id"], "cadastro", None, "Armadilha cadastrada", contexto)
        return "inseridos"

    if preservar_cadastro_conta_ovos:
        cols_update = [
            col for col in cols
            if col in ("telefone_responsavel", "responsavel", "arquivo_origem", "atualizado_em")
            or (not atual[col] and registro[col] is not None)
        ]
    else:
        cols_update = cols

    mudou = any((atual[col] != registro[col]) for col in cols_update if col not in ("arquivo_origem", "atualizado_em"))
    if not mudou:
        return "sem_alteracao"
    for col in cols_update:
        if col not in ("ovitrampa_id", "arquivo_origem", "atualizado_em") and atual[col] != registro[col]:
            _registrar_alteracao_armadilha(conn, registro["ovitrampa_id"], col, atual[col], registro[col], contexto)
    sets = ",".join(f"{col}=?" for col in cols_update if col != "ovitrampa_id")
    conn.execute(
        f"UPDATE {ARMADILHAS_TABLE} SET {sets} WHERE ovitrampa_id=?",
        [registro[col] for col in cols_update if col != "ovitrampa_id"] + [registro["ovitrampa_id"]],
    )
    return "atualizados"


def _upsert_ocorrencia(conn, registro):
    atual = conn.execute(
        f"SELECT * FROM {OCORRENCIAS_TABLE} WHERE id_contagem=?",
        (registro["id_contagem"],),
    ).fetchone()
    cols = list(registro.keys())
    if not atual:
        placeholders = ",".join("?" for _ in cols)
        conn.execute(
            f"INSERT INTO {OCORRENCIAS_TABLE} ({','.join(cols)}) VALUES ({placeholders})",
            [registro[col] for col in cols],
        )
        return "inseridos"

    mudou = any((atual[col] != registro[col]) for col in cols if col not in ("arquivo_origem", "importado_em"))
    if not mudou:
        return "sem_alteracao"
    sets = ",".join(f"{col}=?" for col in cols if col != "id_contagem")
    conn.execute(
        f"UPDATE {OCORRENCIAS_TABLE} SET {sets} WHERE id_contagem=?",
        [registro[col] for col in cols if col != "id_contagem"] + [registro["id_contagem"]],
    )
    return "atualizados"


def _where(filtros, busca=False):
    clauses = []
    params = []
    if filtros.get("ano"):
        clauses.append("l.ano=?")
        params.append(_int(filtros.get("ano")))
    if filtros.get("semana"):
        clauses.append("l.semana=?")
        params.append(_int(filtros.get("semana")))
    if filtros.get("distrito"):
        clauses.append("l.distrito=?")
        params.append(filtros["distrito"])
    if filtros.get("positivas") == "1":
        clauses.append("l.ovos > 0")
    if busca and filtros.get("busca"):
        term = f"%{filtros['busca'].strip()}%"
        clauses.append("(l.ovitrampa_id LIKE ? OR l.rua LIKE ? OR l.complemento LIKE ? OR l.localizacao LIKE ? OR l.quarteirao LIKE ?)")
        params.extend([term] * 5)
    return ("WHERE " + " AND ".join(clauses)) if clauses else "", params


def _where_armadilhas(filtros):
    clauses = []
    params = []
    if filtros.get("distrito"):
        clauses.append("a.localidade=?")
        params.append(filtros["distrito"])
    if filtros.get("busca"):
        term = f"%{filtros['busca'].strip()}%"
        clauses.append("(a.ovitrampa_id LIKE ? OR a.rua LIKE ? OR a.complemento LIKE ? OR a.localizacao LIKE ? OR a.quarteirao LIKE ? OR a.responsavel LIKE ?)")
        params.extend([term] * 6)
    return ("WHERE " + " AND ".join(clauses)) if clauses else "", params


def _text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text.lower() not in ("nan", "none") else None


def _normalizar_quarteirao(value):
    text = _text(value)
    if text and text.isdigit() and len(text) < 4:
        return text.zfill(4)
    return text


def _row_value(row, *names):
    targets = {_norm_header(name) for name in names}
    for key, value in row.items():
        if _norm_header(key) in targets:
            return value
    return None


def _norm_header(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = "".join(ch.lower() if ch.isalnum() else "_" for ch in text)
    return re.sub(r"_+", "_", text).strip("_")


def _int(value):
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text.replace(".", "").replace(",", ".")))
    except ValueError:
        return None


def _real(value):
    text = _text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _date(value):
    text = _text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:10]).date().isoformat()
    except ValueError:
        return None


def _datetime(value):
    text = _text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text[:19]).isoformat(sep=" ")
    except ValueError:
        return text


def _title_distrito(value):
    text = _text(value)
    if not text:
        return None
    text = " ".join(text.split())
    return text.title() if text.isupper() or text.islower() else text
