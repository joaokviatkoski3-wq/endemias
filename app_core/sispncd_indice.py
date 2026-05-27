import sqlite3
from collections import Counter, defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

from app_core import recolhimentos as normalizadores


COLUNAS_OBRIGATORIAS = ("TIPO", "SISPNCD", "DATA", "LOCALIDADE")
CORRECOES_CODIGO = {
    "0006/2006": "0006/2026",
}


def carregar_indice(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or "").strip() for c in next(rows)]
        idx = {name: pos for pos, name in enumerate(headers)}
        faltantes = [name for name in COLUNAS_OBRIGATORIAS if name not in idx]
        if faltantes:
            raise ValueError(f"Colunas ausentes no indice SISPNCD: {', '.join(faltantes)}")

        registros = []
        for numero_linha, row in enumerate(rows, start=2):
            tipo = _tipo(row[idx["TIPO"]])
            codigo_original = _texto(row[idx["SISPNCD"]])
            data = _data(row[idx["DATA"]])
            localidade = _localidade(row[idx["LOCALIDADE"]])
            if not tipo or not codigo_original or not data or not localidade:
                continue
            codigo = CORRECOES_CODIGO.get(codigo_original, codigo_original)
            registros.append({
                "linha": numero_linha,
                "tipo": tipo,
                "data": data,
                "localidade": localidade,
                "sispncd_original": codigo_original,
                "sispncd": codigo,
                "corrigido": codigo != codigo_original,
            })
    finally:
        wb.close()
    return registros


def preparar_mapeamento(registros):
    por_chave = defaultdict(Counter)
    linhas_por_chave = defaultdict(list)
    for registro in registros:
        chave = _chave(registro)
        por_chave[chave][registro["sispncd"]] += 1
        linhas_por_chave[chave].append(registro["linha"])

    mapeamento = {}
    ambiguidades = []
    for chave, codigos in por_chave.items():
        if len(codigos) == 1:
            mapeamento[chave] = next(iter(codigos))
        else:
            ambiguidades.append({
                "chave": chave,
                "codigos": dict(codigos),
                "linhas": linhas_por_chave[chave],
            })
    return mapeamento, ambiguidades


def previsualizar(db_path, indice_path):
    registros = carregar_indice(indice_path)
    mapeamento, ambiguidades = preparar_mapeamento(registros)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        visitas_por_chave = _visitas_por_chave(conn, somente_vazias=True)
        todas_visitas_por_chave = _visitas_por_chave(conn, somente_vazias=False)
        existentes = conn.execute(
            "SELECT COUNT(*) FROM visitas WHERE SISPNCD IS NOT NULL AND TRIM(SISPNCD)<>''"
        ).fetchone()[0]
    finally:
        conn.close()

    atualizaveis = 0
    por_tipo = Counter()
    por_codigo = Counter()
    chaves_sem_visita = []
    for chave, codigo in mapeamento.items():
        total = visitas_por_chave.get(chave, 0)
        if total:
            atualizaveis += total
            por_tipo[chave[0]] += total
            por_codigo[codigo] += total
        elif not todas_visitas_por_chave.get(chave, 0):
            chaves_sem_visita.append({"chave": chave, "sispncd": codigo})

    sem_indice = sum(
        total for chave, total in visitas_por_chave.items()
        if chave not in mapeamento
    )
    return {
        "linhas_indice": len(registros),
        "chaves_indice": len(mapeamento) + len(ambiguidades),
        "chaves_ambiguas": ambiguidades,
        "chaves_sem_visita": chaves_sem_visita,
        "visitas_atualizaveis": atualizaveis,
        "visitas_sem_indice": sem_indice,
        "visitas_ja_preenchidas": existentes,
        "por_tipo": dict(sorted(por_tipo.items())),
        "por_codigo": dict(sorted(por_codigo.items())),
        "correcoes": [
            {
                "linha": r["linha"],
                "de": r["sispncd_original"],
                "para": r["sispncd"],
                "tipo": r["tipo"],
                "data": r["data"],
                "localidade": r["localidade"],
            }
            for r in registros if r["corrigido"]
        ],
    }


def aplicar(db_path, indice_path):
    previa = previsualizar(db_path, indice_path)
    if previa["chaves_ambiguas"]:
        raise ValueError("Indice SISPNCD possui chaves ambiguas; revise antes de aplicar.")

    registros = carregar_indice(indice_path)
    mapeamento, _ = preparar_mapeamento(registros)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    atualizados = 0
    try:
        conn.execute("BEGIN")
        for chave, codigo in mapeamento.items():
            cur = conn.execute(
                """UPDATE visitas
                      SET SISPNCD=?
                    WHERE tipo=?
                      AND data=?
                      AND COALESCE((
                          SELECT nome FROM localidades WHERE localidades.id_localidade=visitas.id_localidade
                      ), localidade)=?
                      AND (SISPNCD IS NULL OR TRIM(SISPNCD)='')""",
                (codigo, chave[0], chave[1], chave[2]),
            )
            atualizados += cur.rowcount or 0
        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise
    finally:
        conn.close()

    depois = previsualizar(db_path, indice_path)
    return {
        "ok": True,
        "atualizados": atualizados,
        "previa": previa,
        "depois": depois,
    }


def _visitas_por_chave(conn, somente_vazias):
    where = ""
    if somente_vazias:
        where = "WHERE v.SISPNCD IS NULL OR TRIM(v.SISPNCD)=''"
    rows = conn.execute(
        f"""
        SELECT v.tipo,
               v.data,
               COALESCE(l.nome, v.localidade) AS localidade,
               COUNT(*) AS total
          FROM visitas v
          LEFT JOIN localidades l ON l.id_localidade=v.id_localidade
          {where}
         GROUP BY v.tipo, v.data, COALESCE(l.nome, v.localidade)
        """
    )
    return {
        (_tipo(r["tipo"]), r["data"], _localidade(r["localidade"])): r["total"]
        for r in rows
    }


def _chave(registro):
    return (registro["tipo"], registro["data"], registro["localidade"])


def _tipo(value):
    return str(value or "").strip().upper()


def _texto(value):
    if value is None:
        return ""
    return str(value).strip()


def _localidade(value):
    return normalizadores._localidade(value) or ""


def _data(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return ""
    return datetime.fromisoformat(text[:10]).date().isoformat()
